"""
Applies UX/visual polish to the V11 workbook: sheet organization, color
coding (editable vs system-managed), data validation dropdowns, a legend,
a START_HERE guide, and a redesigned ACTIVITY_PLAN with a live timeline +
impact estimate.

Pure presentation layer - does not change any business logic, engine
behavior, or the data model any engine reads/writes by position. Called by
scaffold_workbook.py after all sheets/data are in place.

IMPORTANT SAFETY CONSTRAINT: real Excel "Protect Sheet" blocks Office
Scripts' Range.clear()/setValues() calls unless the script explicitly
unprotects first - none of our engines do that. Enabling real cell
locking + sheet protection is therefore ONLY safe on sheets no engine ever
writes to (pure config). Sheets an engine writes to (POS_MASTER,
MANAGER_PLAN, MANAGER_PLAN_PUBLISHED, PLAN_LIFECYCLE, COMPLIANCE_LOG,
ADVISOR_LOG, VISIT_HISTORY_ACTUAL, DASHBOARD) and pure import-staging sheets
get color-only "please don't hand-edit this" cues, never real protection -
this trade-off is documented in the legend so it isn't a silent gap.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, DataBarRule, IconSetRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.comments import Comment
from openpyxl.chart import LineChart, BarChart, ScatterChart, Series, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
import dashboard_ui
from dashboard_ui import (
    NAVY, WHITE, STATUS_GOOD, STATUS_WARNING, STATUS_SERIOUS, STATUS_CRITICAL,
    FONT_HEADER as HEADER_FONT, FONT_TITLE as TITLE_FONT, FONT_SECTION as SECTION_FONT,
    FONT_NOTE as NOTE_FONT, CARD_BORDER, font_card_value,
    build_nav_rail, build_nav_button, build_dashboard_banner, build_section_header,
    build_filter_bar_background, build_filter_dropdown, build_kpi_card, build_kpi_card_row,
    build_progress_bar, build_status_badge_conditional, apply_severity_conditional_formatting,
    make_bar_chart, make_line_chart, style_dashboard_table_header, apply_table_borders,
)

# ============================================================================
# PALETTE
# ============================================================================
# NAVY/WHITE/STATUS_* above come from dashboard_ui.py - the single source of
# truth for the dashboard color palette (see that file's docstring). The
# fills below are specific to the data-entry/config sheet styling this file
# owns (editable-cell color coding) and have no dashboard-screen equivalent.

EDITABLE_FILL = "FFF2CC"       # warm cream - "you type here"
SYSTEM_FILL = "E7E6E6"         # neutral grey - "the system manages this"
IMPORT_FILL = "DDEBF7"         # light blue - "paste your export here"
OUTPUT_FILL = "E2EFDA"         # light green - "generated results"
LOG_FILL = "F2F2F2"            # very light grey - "append-only history"
WARNING_FILL = "FCE4D6"        # soft orange - inactive/TODO config rows
LOS_FILL = "BDD7EE"            # timeline: LOS campaigns
LOT_FILL = "F8CBAD"            # timeline: LOT campaigns

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

# Sheet grouping -> tab color + intended sheet order (top to bottom in Excel)
SHEET_GROUPS = [
    ("HOME", "404040"),
    ("TECHNICIAN_SCORECARD", "2E75B6"),
    ("PERFORMANCE", "2E75B6"),
    ("EFFICIENCY", "2E75B6"),
    ("MANUAL", "808080"),
    ("WEEK_DASHBOARD", "2E75B6"),
    ("MAP", "2E75B6"),
    ("DASHBOARD", "375623"),
    ("TECHNICIAN_PLAN", "375623"),
    ("POS_MASTER", "7030A0"),
    ("ACTIVITY_PLAN", "BF8F00"),
    ("IMPORT_HUB", "2E75B6"),
    ("RAW_DATA", "2E75B6"),
    ("POS_STATUS_IMPORT", "2E75B6"),
    ("SALESAPP_IMPORT", "2E75B6"),
    ("MANAGER_PLAN", "375623"),
    ("MANAGER_PLAN_PUBLISHED", "375623"),
    ("PLAN_LIFECYCLE", "375623"),
    ("CONTROL", "BF8F00"),
    ("MARKET_RULES", "BF8F00"),
    ("TERMINAL_RULES", "BF8F00"),
    ("BLACKLIST", "C00000"),
    ("POS_ACTIVATE_LIST", "375623"),
    ("CATEGORY_RULES", "BF8F00"),
    ("CADENCE_RULES", "BF8F00"),
    ("PARETO_GROUPS", "BF8F00"),
    ("SCORE_PROFILES", "BF8F00"),
    ("ADVISOR_RULES", "BF8F00"),
    ("CAPACITY_OVERRIDE", "BF8F00"),
    ("COMPLIANCE_LOG", "595959"),
    ("ADVISOR_LOG", "595959"),
    ("VISIT_HISTORY_ACTUAL", "595959"),
    ("VISIT_HISTORY", "595959"),
]

# The 5 sheets a normal user (regional manager) works with day to day.
# Everything else that stays visible (import staging) is a necessary but
# occasional "mailbox", not part of the daily working set - communicated on
# HOME, not hidden, because the user must paste into it weekly.
CORE_DAILY_SHEETS = ["HOME", "TECHNICIAN_SCORECARD", "PERFORMANCE", "WEEK_DASHBOARD", "DASHBOARD", "TECHNICIAN_PLAN", "POS_MASTER", "ACTIVITY_PLAN", "IMPORT_HUB"]
IMPORT_UTILITY_SHEETS = ["RAW_DATA", "POS_STATUS_IMPORT", "SALESAPP_IMPORT"]

# Everything not in CORE_DAILY_SHEETS/IMPORT_UTILITY_SHEETS is implementation
# detail (raw engine data, config, logs) - hidden from the normal user, but
# still fully readable/writable by Office Scripts (hidden sheets are not
# restricted via the API, only invisible in the tab bar).
HIDDEN_SHEETS = {
    "MANAGER_PLAN", "MANAGER_PLAN_PUBLISHED", "PLAN_LIFECYCLE",
    "CONTROL", "MARKET_RULES", "CATEGORY_RULES",
    # TERMINAL_RULES deliberately NOT hidden (product owner, 2026-07-08:
    # wants simple Ano/Ne toggles for VELKY TERMINAL/SMALL TERMINAL/LI on
    # the main control screen, easy to find) - it already has exactly that
    # (a YES/NO dropdown per terminal type, read directly by
    # PlanningEngine.ts's terminalOK() filter), it just needed to be
    # reachable instead of buried in the hidden technical tab group. See
    # the HOME quick-link and add_sheet_purpose_notes() entry below.
    "CADENCE_RULES", "PARETO_GROUPS", "SCORE_PROFILES", "ADVISOR_RULES",
    "CAPACITY_OVERRIDE", "COMPLIANCE_LOG", "ADVISOR_LOG",
    "VISIT_HISTORY_ACTUAL", "VISIT_HISTORY", "PLANNING_HORIZON_RULES",
    # Performance Engine's raw aggregated tables - data sources for the
    # manager UX sheets (docs/MANAGER_UX_ARCHITECTURE.md), not something a
    # manager reads directly, same treatment as COMPLIANCE_LOG/ADVISOR_LOG.
    "TECHNICIAN_PERFORMANCE_LOG", "TECHNICIAN_PERFORMANCE_SUMMARY", "TECHNICIAN_TOP_ISSUES",
    # BUG FIX (found 2026-07-06 during a full test pass): added alongside
    # VISIT_HISTORY_ACTUAL when OTHER_VISIT_LOG was introduced, but never
    # added here - it sat visible in the tab bar as raw unstyled engine data.
    "OTHER_VISIT_LOG",
    # POS_MAP_DATA: ReportingEngine.ts's raw X/Y-per-technician grid feeding
    # the MAP sheet's chart - a manager reads MAP, never this directly.
    "POS_MAP_DATA",
}

# Sheets an engine writes to programmatically - never real-protected, see
# module docstring. Everything else (pure config, user-pasted imports) is
# safe to lock/protect.
ENGINE_WRITABLE = {
    "POS_MASTER", "MANAGER_PLAN", "MANAGER_PLAN_PUBLISHED", "PLAN_LIFECYCLE",
    "COMPLIANCE_LOG", "ADVISOR_LOG", "VISIT_HISTORY_ACTUAL", "DASHBOARD",
    "TECHNICIAN_PERFORMANCE_LOG", "TECHNICIAN_PERFORMANCE_SUMMARY", "TECHNICIAN_TOP_ISSUES",
    "OTHER_VISIT_LOG", "POS_MAP_DATA",
}

# A third category, distinct from both of the above: dashboard screens that
# are neither engine-written nor plain config, but require live user
# interaction (Data Validation dropdowns, a native Excel Table's own
# sort/filter). protect_config_sheet()'s binary "engine-writable = never
# protected, else = locked down" model has no room for "not engine-written,
# but the user still needs to click things" - found as a real bug during a
# post-build QA pass (2026-07-06): TECHNICIAN_SCORECARD/WEEK_DASHBOARD's
# filter dropdowns and PERFORMANCE's Table sort/filter were being silently
# disabled by real Excel sheet protection, defeating the entire point of
# building them as native, interactive Excel elements.
INTERACTIVE_DASHBOARD_SHEETS = {"TECHNICIAN_SCORECARD", "PERFORMANCE", "WEEK_DASHBOARD", "TECHNICIAN_PLAN"}

# Per-sheet: which columns (by header name) are meant for manual editing.
# Everything else on that sheet is shown as system/read-only styling.
EDITABLE_COLUMNS = {
    "CONTROL": ["VALUE"],
    "ACTIVITY_PLAN": ["TYPE", "ACTIVITY", "START_WEEK", "END_WEEK", "PRIORITY", "OVERRIDE_GAP"],
    "TERMINAL_RULES": ["ACTIVE"],
    "MARKET_RULES": ["ACTIVE"],
    "BLACKLIST": ["POS", "NOTES"],
    "POS_ACTIVATE_LIST": ["POS", "NOTES"],
    "CATEGORY_RULES": ["CATEGORY", "RULE"],
    "CADENCE_RULES": ["scope", "matchValue", "minGapWeeks", "maxIntervalWeeks", "intervalType",
                       "guaranteeType", "dedupBy", "campaignChangeOverride", "priority", "active",
                       "validFrom", "validTo", "notes"],
    "PLANNING_HORIZON_RULES": ["appliesFromWeek", "appliesToWeek", "horizonWeeks", "reason", "active", "notes"],
    "PARETO_GROUPS": ["scope", "boundaryType", "boundaryValue", "active", "notes"],
    "SCORE_PROFILES": ["weight", "notes"],
    "ADVISOR_RULES": ["ruleId", "type", "condition", "threshold", "severity", "messageTemplate", "active"],
    "CAPACITY_OVERRIDE": ["technician", "year", "week", "capacity"],
    "POS_STATUS_IMPORT": ["POS", "ACTIVE"],
    "POS_MASTER": ["managerOverrideType", "managerOverridePriority", "managerOverrideTechnician", "plannerNotes"],
}

# Dropdown validations: sheet -> {header name: (list_of_values, allow_blank)}
DROPDOWNS = {
    "TERMINAL_RULES": {"ACTIVE": (["YES", "NO"], False)},
    "MARKET_RULES": {"ACTIVE": (["YES", "NO"], False)},
    "ACTIVITY_PLAN": {"TYPE": (["LOS", "LOT"], False)},
    "CADENCE_RULES": {
        "active": (["YES", "NO"], False),
        "guaranteeType": (["HARD", "SOFT_HIGH_WEIGHT"], True),
        "intervalType": (["RECURRING", "ONCE_PER_CAMPAIGN"], True),
        "dedupBy": (["NONE", "ADDRESS"], True),
        "campaignChangeOverride": (["YES", "NO"], True),
    },
    "PARETO_GROUPS": {
        "active": (["YES", "NO"], False),
        "boundaryType": (["PERCENTILE", "FIXED_VALUE"], True),
        "scope": (["PER_TECHNICIAN", "GLOBAL", "PER_REGION", "PER_MARKET"], True),
    },
    "POS_STATUS_IMPORT": {"ACTIVE": ([1, 0], False)},
    "POS_MASTER": {
        "managerOverrideType": (["", "FORCE_INCLUDE", "FORCE_EXCLUDE"], True),
        "managerOverridePriority": (["", "Low", "Normal", "High", "Critical"], True),
        "status": (["Active", "Closed"], False),
    },
}


def _header_index(ws):
    return {str(c.value): i + 1 for i, c in enumerate(ws[1]) if c.value not in (None, "")}


def style_header_row(ws, freeze_below=True, freeze_col=None):
    for cell in ws[1]:
        if cell.value in (None, ""):
            continue
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    if freeze_below:
        col = freeze_col or "A"
        ws.freeze_panes = f"{col}2"


# Sheets that are wholesale paste-zones (whole sheet = editable import data,
# no per-column editable/system split makes sense) get a flat wash instead.
IMPORT_STAGING_SHEETS = {"RAW_DATA", "POS_STATUS_IMPORT", "SALESAPP_IMPORT"}
OUTPUT_SHEETS = {"MANAGER_PLAN", "MANAGER_PLAN_PUBLISHED", "DASHBOARD", "PLAN_LIFECYCLE"}
LOG_SHEETS = {"COMPLIANCE_LOG", "ADVISOR_LOG", "VISIT_HISTORY_ACTUAL", "VISIT_HISTORY"}

# Import Hub guidance: a cell comment on the header cell of each staging
# sheet, not a banner row - a banner row would shift every data row down by
# one, breaking every engine's row-1-is-header assumption. A comment is pure
# metadata (openpyxl's getValues()-equivalent used by Office Scripts never
# sees it), so it is risk-free. Content documents a capability that already
# works today (ComplianceEngine.ts dedupes SALESAPP_IMPORT rows by UID
# against VISIT_HISTORY_ACTUAL on every run, and ImportEngine.ts upserts
# POS_MASTER by posId), it just wasn't visible/discoverable before.
IMPORT_HUB_GUIDANCE = {
    # RAW_DATA is NOT here deliberately - unlike the other staging sheets it
    # has a legacy instruction row already at row 1 (see
    # fix_raw_data_layout()), rewritten in place as a visible banner instead
    # of a hover-only comment, so a second overlapping comment would be
    # redundant noise on the same cell.
    "POS_STATUS_IMPORT": (
        "Sem vlož export stavu POS (aktivní/uzavřené). Stejná struktura pokaždé - "
        "lze vkládat opakovaně, Import Engine vždy aktualizuje POS_MASTER podle POS_ID."
    ),
    "SALESAPP_IMPORT": (
        "Sem vlož export ze SalesApp. Klidně více exportů najednou (např. 2-3 měsíce) - "
        "přidávej pod poslední řádek. Compliance Engine automaticky odstraní duplicity "
        "podle UID návštěvy a zachová historii - bezpečné i pro překrývající se exporty."
    ),
    "TERMINAL_RULES": (
        "Zapni/vypni Ano (YES)/Ne (NO), jaké typy terminálů se mají v dalším běhu Planning "
        "Engine vůbec uvažovat jako kandidáti na návštěvu. Změna se projeví hned při příštím "
        "spuštění Planning Engine, nic dalšího se nemusí nastavovat."
    ),
    "BLACKLIST": (
        "Vlož sem POS ID provozoven, které chceš úplně vynechat z plánování, bez ohledu na "
        "cokoliv jiného (skóre, kadenční pravidla, filtry). Stačí ID do sloupce POS - engine "
        "je od dalšího běhu Planning Engine ignoruje úplně."
    ),
    "POS_ACTIVATE_LIST": (
        "Opak BLACKLIST: sem vlož POS ID, která chceš AKTIVOVAT i přesto, že je CATEGORY_RULES "
        "aktuálně vyřazuje (např. kategorie 1CD/1POSTA = EXCLUDE). Nemění, kterému technikovi "
        "POS patří - jen mu dovolí vstoupit do plánování. Necháš-li tento seznam prázdný, "
        "spusť místo toho Activate POS Engine s CONTROL.ACTIVATE_COUNT_BY_PPT > 0 - aktivuje "
        "prvních N (podle PPT sestupně) aktuálně vyřazených POS automaticky, bez ručního výpisu. "
        "Náhled aktuálně vyřazených POS a jejich PPT vidíš v tabulce vpravo (sloupec D+)."
    ),
    # Core working screens don't get their own title-banner rows the way
    # HOME/DASHBOARD/IMPORT_HUB do - a banner row would push every real data
    # row down by one and break every engine's "row 1 is the header"
    # assumption for these two specifically (POS_MASTER/ACTIVITY_PLAN are
    # both read positionally by ImportEngine.ts). A header-cell comment gets
    # the same "what is this screen for" clarity without that risk - see the
    # docstring on _nav_button for the same reasoning applied to buttons.
    "POS_MASTER": (
        "POS_MASTER = centrální evidence všech POS. Identifikace vždy podle POS_ID. "
        "Provozovny se stejnou adresou (CORN/9PODNIK) se plánují jako jeden fyzický POS. "
        "Editovat lze jen žlutě podbarvené sloupce (ruční poznámky/výjimky) - zbytek počítají enginy."
    ),
    "ACTIVITY_PLAN": (
        "KROK 4 týdenní rutiny: ACTIVITY_PLAN = plánování kampaní (LOS/LOT). Přidej nebo "
        "uprav kampaň v řádku (typ, název, od týdne, do týdne) a hned vidíš vpravo odhad "
        "dopadu (počet návštěv, časová osa) - žádné přepočítávání ručně. Po úpravě pokračuj "
        "spuštěním Planning Engine (IMPORT_HUB, krok 5)."
    ),
}


def add_sheet_purpose_notes(wb):
    for sheet_name, text in IMPORT_HUB_GUIDANCE.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws["A1"].comment = Comment(text, "Field Force Optimizer")


def build_import_hub(wb, pos_master_tech_col="O"):
    """The actual front door for data entry - foolproof by construction, not
    by trusting the user to remember which of 3 staging sheets a file goes
    to. Product owner's real weekly routine is fixed (SalesApp export -> PPT
    campaign assignment -> small activity tweaks -> run planner -> publish),
    so this screen is built around that exact sequence as numbered steps,
    each answering the same 4 questions on-screen instead of in a doc
    nobody reads: where does this data come from, what file goes here, what
    happens after you run the engine, what's the next step.

    Office Scripts have no OS file-picker API (see docs/ARCHITECTURE.md
    section 15d) so this cannot be a literal drag-and-drop widget - it is
    the closest equivalent achievable on this platform: one screen, live
    row counts per staging sheet (so "did my paste work" is answered
    without opening each sheet), one-click links to each paste target.
    Pure presentation - reads existing sheets, writes nothing any engine
    depends on."""
    if "IMPORT_HUB" in wb.sheetnames:
        del wb["IMPORT_HUB"]
    ws = wb.create_sheet("IMPORT_HUB")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    for col in "CDEF":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["G"].width = 20

    ws.merge_cells("A1:G2")
    ws["A1"] = "IMPORT HUB"
    ws["A1"].font = Font(bold=True, size=22, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A3:G3")
    ws["A3"] = "Týdenní rutina - postupuj podle kroků níže, nic jiného řešit nemusíš"
    ws["A3"].font = Font(italic=True, size=11, color=WHITE)
    ws["A3"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"  # banner stays visible while scrolling through the 5 steps

    r = 5

    def step_card(num, title, source, target_sheet, after, color, extra_note=None):
        nonlocal r
        badge_row = r
        ws.merge_cells(f"B{r}:G{r}")
        ws.cell(r, 2, f"KROK {num} — {title}").font = Font(bold=True, size=13, color=NAVY)
        r += 1
        ws.merge_cells(f"B{r}:G{r}")
        ws.cell(r, 2, f"Odkud: {source}").font = Font(size=10)
        r += 1
        ws.merge_cells(f"B{r}:G{r}")
        ws.cell(r, 2, f"Co se stane po importu: {after}").font = NOTE_FONT
        r += 1
        if target_sheet:
            # RAW_DATA has one extra non-data row at the top (a legacy
            # instruction line inherited from the source workbook, ABOVE its
            # real header - see fix_raw_data_layout()) - COUNTA needs -2
            # there, not -1, to actually count data rows and not off-by-one
            # undercount by exactly one row every time.
            header_rows = 2 if target_sheet == "RAW_DATA" else 1
            ws.cell(r, 2, "Řádků nyní:").font = Font(size=9, color="595959")
            count_cell = ws.cell(r, 3, f'=COUNTA({target_sheet}!A:A)-{header_rows}')
            count_cell.font = Font(bold=True, size=13, color=NAVY)
            status_cell = ws.cell(r, 4, f'=IF(C{r}>0,"✅ Data v systému","⏳ Čeká na vložení")')
            status_cell.font = Font(bold=True, size=9)
            status_cell.alignment = Alignment(horizontal="left")
            ws.conditional_formatting.add(
                status_cell.coordinate,
                FormulaRule(formula=[f'LEFT({status_cell.coordinate},1)="✅"'], font=Font(color="375623", bold=True)),
            )
            ws.conditional_formatting.add(
                status_cell.coordinate,
                FormulaRule(formula=[f'LEFT({status_cell.coordinate},1)="⏳"'], font=Font(color="BF8F00", bold=True)),
            )
            _nav_button(ws, f"F{r}", "Vložit / otevřít →", target_sheet, color=color)
        elif extra_note:
            ws.merge_cells(f"B{r}:G{r}")
            ws.cell(r, 2, extra_note).font = Font(italic=True, size=9, color="808080")
        # Badge spans exactly the 4 rows this card just used (title, source,
        # after, action row) - written last since only now is r_end known.
        ws.merge_cells(f"A{badge_row}:A{r}")
        ws.cell(badge_row, 1, str(num)).font = Font(bold=True, size=18, color=WHITE)
        ws.cell(badge_row, 1).fill = PatternFill("solid", fgColor=color)
        ws.cell(badge_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        r += 2

    step_card(
        1, "Export ze SalesApp (návštěvy)",
        "SalesApp → export realizovaných návštěv za uplynulý týden (klidně i více týdnů/měsíců najednou)",
        "SALESAPP_IMPORT",
        "Compliance Engine sloučí nové řádky s historií, odstraní duplicity podle UID a přepočítá plnění plánu.",
        "2E75B6",
    )
    step_card(
        2, "PPT zadání kampaní (export POS dat)",
        "Export POS dat od zákazníka (PPT zadání) - stejná struktura každý týden, VŽDY kompletní seznam všech POS",
        "RAW_DATA",
        "Import Engine sloučí podle POS_ID; provozovny se stejnou adresou (CORN/9PODNIK) zůstávají jeden fyzický POS. Ruční poznámky u POS se nepřepíší. POS, který v tomto exportu chybí, se automaticky označí jako Closed - žádný zvláštní krok navíc není potřeba.",
        "2E75B6",
        extra_note=None,
    )

    step_card(
        3, "Spusť Import Engine",
        "Záložka Automatizace v Excelu",
        None,
        "POS_MASTER se aktualizuje, historie návštěv se rozšíří (nikdy nepřepisuje ani nemaže).",
        "BF8F00",
        extra_note="⚙ Automatizace → skript \"ImportEngine.ts\" (1. z 3 tento týden: Import → Planning → Publish) - spouští se jako Office Script, ne jako list.",
    )

    # Before Planning Engine: the one check that actually matters here. A
    # POS with no assigned technician never gets planned - Planning Engine
    # has nothing to route it to, and nothing else on this workbook would
    # ever surface that gap. It's invisible until a customer complains their
    # terminal hasn't been serviced in months. Checked right here, between
    # "data just landed" and "run the planner", because this is the one
    # point in the week where fixing it (correcting the assignment in
    # RAW_DATA or via a manager override) is still cheap.
    unassigned_formula = f'=COUNTIFS(POS_MASTER!Q:Q,"Active",POS_MASTER!{pos_master_tech_col}:{pos_master_tech_col},"")'
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(r, 2, "Aktivní POS bez přiřazeného technika:").font = Font(size=10)
    check_cell = ws.cell(r, 5, unassigned_formula)
    check_cell.font = Font(bold=True, size=14, color=NAVY)
    ws.conditional_formatting.add(
        check_cell.coordinate,
        FormulaRule(formula=[f"{check_cell.coordinate}>0"], fill=PatternFill("solid", fgColor=STATUS_SERIOUS), font=Font(bold=True, size=14, color=STATUS_CRITICAL)),
    )
    _nav_button(ws, f"G{r}", "Zkontrolovat →", "POS_MASTER", color="7030A0")
    r += 2

    ws.merge_cells(f"A{r}:G{r}")
    ws.cell(r, 1, "POKRAČUJ DÁL")
    ws.cell(r, 1).font = TITLE_FONT
    r += 1
    for num, title, target, color in [
        (4, "Uprav aktivity/kampaně (ACTIVITY_PLAN), pokud je potřeba - volitelné, jen když se něco mění", "ACTIVITY_PLAN", "BF8F00"),
        (5, "Automatizace → \"PlanningEngine.ts\" (2. skript), zkontroluj MANAGER_PLAN, pak \"PublishEngine.ts\" (3. skript) publikuje nejbližší týden", "TECHNICIAN_PLAN", "375623"),
    ]:
        ws.cell(r, 1, str(num)).font = Font(bold=True, size=14, color=WHITE)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=color)
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"B{r}:E{r}")
        ws.cell(r, 2, f"KROK {num} — {title}").font = Font(size=11)
        ws.cell(r, 2).alignment = Alignment(vertical="center")
        _nav_button(ws, f"G{r}", "Otevřít →", target, color=color)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1

    ws.cell(r, 1, "POSLEDNÍ AKTUALIZACE POS_MASTER").font = Font(size=9, color="595959")
    r += 1
    ws.cell(r, 1, '=IFERROR(TEXT(MAX(POS_MASTER!AM:AM),"DD.MM.YYYY HH:MM"),"zatím žádný import")')
    ws.cell(r, 1).font = Font(bold=True, size=14, color=NAVY)
    r += 2

    _nav_button(ws, f"A{r}", "← Zpět na HOME", "HOME", color="404040")
    return ws


def fix_raw_data_layout(ws, max_rows=500):
    """RAW_DATA is copied verbatim from the legacy V10.5.5 workbook, which
    has a quirk none of the other reference sheets share: row 1 is a stray
    instruction sentence ("Sem vloz export POS"), and the REAL column header
    (CISLO TERMINALU, POS, MARKET, ...) is row 2. ImportEngine.ts already
    knows this (reads raw[1] as the header row, data from raw[2] onward -
    see its own "RAW_DATA COLUMN MAPPING" comment) - this is a confirmed,
    intentional, working data contract, NOT something to "fix" by deleting
    the row (that would be a business-logic-adjacent change to a working
    positional read).

    What WAS actually broken: every generic styling helper in this module
    assumes row 1 is the header, so before this function existed, the
    styling was inverted - the stray instruction sentence got the big navy
    header treatment, and the REAL header row got the pale "this is
    editable data" wash, making the actual column names look like ordinary
    data. Found during a full production review by comparing this sheet
    against the source workbook, not by inspection of the styled workbook
    alone - the mis-styling looked plausible until compared to row 2's
    actual content."""
    # Row 1: a light instruction banner, not a heavy header bar - restyled
    # with a friendlier sentence (ImportEngine.ts never reads this text, so
    # changing it carries no data risk).
    ws["A1"] = "Vlož export POS dat sem (od řádku 3) - lze vkládat i více exportů za sebou"
    ws["A1"].font = Font(italic=True, size=10, color="595959")
    ws["A1"].fill = PatternFill("solid", fgColor=IMPORT_FILL)
    ws.row_dimensions[1].height = 18
    if ws["A1"].comment:
        ws["A1"].comment = None  # redundant with the rewritten instruction text above

    # Row 2: the REAL header - gets the real header treatment.
    for cell in ws[2]:
        if cell.value in (None, ""):
            continue
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Data wash starts at row 3, not row 2.
    capped_max_row = min(ws.max_row, max_rows)
    for row in ws.iter_rows(min_row=3, max_row=max(capped_max_row, 3), max_col=ws.max_column or 1):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=IMPORT_FILL)

    last_col = get_column_letter(ws.max_column or 16)
    ws.auto_filter.ref = f"A2:{last_col}{max(ws.max_row, 2)}"
    ws.freeze_panes = "A3"


def color_editable_columns(ws, sheet_name, max_rows=500):
    if sheet_name in IMPORT_STAGING_SHEETS:
        fill = PatternFill("solid", fgColor=IMPORT_FILL)
    elif sheet_name in OUTPUT_SHEETS:
        fill = PatternFill("solid", fgColor=OUTPUT_FILL)
    elif sheet_name in LOG_SHEETS:
        fill = PatternFill("solid", fgColor=LOG_FILL)
    else:
        fill = None

    if fill is not None:
        # Flat wash for the data area - but capped at a reasonable row count.
        # RAW_DATA can carry 11k+ real rows; individually styling every one
        # of those cells would bloat the file for no visual benefit past
        # what's on screen at once, and the tab color + header already
        # signal "this is an import sheet". Cap applies to all sheets in
        # this bucket for one consistent, predictable rule.
        capped_max_row = min(ws.max_row, max_rows)
        for row in ws.iter_rows(min_row=2, max_row=max(capped_max_row, 2), max_col=ws.max_column or 1):
            for cell in row:
                cell.fill = fill
        return

    editable = EDITABLE_COLUMNS.get(sheet_name)
    if not editable:
        return
    idx = _header_index(ws)
    editable_cols = {idx[h] for h in editable if h in idx}
    fill_editable = PatternFill("solid", fgColor=EDITABLE_FILL)
    fill_system = PatternFill("solid", fgColor=SYSTEM_FILL)
    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, max_rows), max_col=ws.max_column or 1):
        for cell in row:
            if cell.column in editable_cols:
                cell.fill = fill_editable
            else:
                cell.fill = fill_system


def add_dropdowns(ws, sheet_name, max_rows=500):
    rules = DROPDOWNS.get(sheet_name)
    if not rules:
        return
    idx = _header_index(ws)
    for header, (values, allow_blank) in rules.items():
        if header not in idx:
            continue
        col_letter = get_column_letter(idx[header])
        formula = '"' + ",".join(str(v) for v in values) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank, showDropDown=False)
        dv.error = "Vyber prosím hodnotu ze seznamu."
        dv.errorTitle = "Neplatná hodnota"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_rows}")


def protect_config_sheet(ws, sheet_name):
    if sheet_name in ENGINE_WRITABLE or sheet_name in INTERACTIVE_DASHBOARD_SHEETS:
        return
    editable = set(EDITABLE_COLUMNS.get(sheet_name, []))
    idx = _header_index(ws)
    editable_cols = {idx[h] for h in editable if h in idx}
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 300), max_col=ws.max_column or 1):
        for cell in row:
            cell.protection = cell.protection.copy(locked=cell.column not in editable_cols)
    ws.protection.sheet = True
    ws.protection.formatFormulas = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def apply_sheet_order_and_colors(wb):
    order = [name for name, _ in SHEET_GROUPS if name in wb.sheetnames]
    remaining = [n for n in wb.sheetnames if n not in order]
    wb._sheets = [wb[n] for n in order + remaining]
    colors = dict(SHEET_GROUPS)
    for name in wb.sheetnames:
        if name in colors:
            wb[name].sheet_properties.tabColor = colors[name]


def build_legend(ws, start_row):
    ws.cell(start_row, 1, "LEGENDA").font = SECTION_FONT
    legend_rows = [
        (EDITABLE_FILL, "Editovatelné pole - sem zapisuješ hodnoty"),
        (SYSTEM_FILL, "Systémové pole - počítá/zapisuje ho engine, needituj ručně"),
        (IMPORT_FILL, "Import zóna - sem vlož export (Ctrl+A / Ctrl+V přes hlavičku)"),
        (OUTPUT_FILL, "Výstup - generuje Planning/Reporting Engine"),
        (LOG_FILL, "Log - append-only historie, needituj, needituj ani nemaž"),
        (WARNING_FILL, "Neaktivní / čeká na potvrzení hodnoty (viz notes sloupec)"),
    ]
    r = start_row + 1
    for color, text in legend_rows:
        cell = ws.cell(r, 1, "")
        cell.fill = PatternFill("solid", fgColor=color)
        ws.cell(r, 2, text).alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1
    return r


def _nav_button(ws, cell_ref, label, target_sheet, color=NAVY, width=None):
    # width kept as a no-op parameter for existing call-site compatibility;
    # the shared component (dashboard_ui.build_nav_button) doesn't need it.
    build_nav_button(ws, cell_ref, label, target_sheet, color=color)


def build_home(wb, real_control_values, pos_master_tech_col="O"):
    """A real app hub, not a README sheet: a live network-wide KPI card row
    (same component every other dashboard screen uses), a live pipeline
    status strip (each stage checks the actual workbook state - Import/
    Plan/Rozpis/Publikace/Vyhodnocení/Dashboard - and shows Hotovo/Chybí
    with a one-click link), a single "co dělat dál" callout that always
    points at the first incomplete stage, an operational numbers strip,
    quick nav, and the legend - built for someone opening this workbook for
    the first time to understand within 30 seconds where things stand and
    what to do next, without reading any instructions first.

    Uses the same nav rail / banner / KPI card / severity-coloring
    components as TECHNICIAN_SCORECARD/PERFORMANCE/WEEK_DASHBOARD
    (tools/dashboard_ui.py) - HOME is the landing page, so it gets the same
    design language as everything it links to (product owner, 2026-07-06)."""
    TP = "TECHNICIAN_PERFORMANCE_LOG"
    for old_name in ("START_HERE", "HOME"):
        if old_name in wb.sheetnames:
            del wb[old_name]
    ws = wb.create_sheet("HOME", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    for col in "CDEFGHIJKLMNO":
        ws.column_dimensions[col].width = 15
    build_nav_rail(ws, "HOME")

    # ---- Banner ----
    build_dashboard_banner(
        ws, "FIELD FORCE OPTIMIZER", "Plánování a řízení terénních techniků",
        col_start="C", col_end="J", title_size=26,
    )
    ws.freeze_panes = "C4"  # banner + nav rail stay visible while scrolling below

    # ==========================================================================
    # HIDDEN FORMULA PLUMBING - network-wide "latest week on record" + the
    # week before it, for the KPI card row's trend. Same AGGREGATE-based
    # combined-key technique already proven in TECHNICIAN_SCORECARD/
    # WEEK_DASHBOARD, no engine change.
    # ==========================================================================
    ws["Q1"] = "latestKey"
    ws["Q2"] = f'=IFERROR(AGGREGATE(14,6,({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)/({TP}!$A$2:$A$5000<>""),1),"")'
    ws["R1"] = "prevKey"
    ws["R2"] = (
        f'=IFERROR(AGGREGATE(14,6,({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)/'
        f'(({TP}!$A$2:$A$5000<>"")*(({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)<$Q$2)),1),"")'
    )
    prev_cond = f'({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000=$R$2)'
    ws["R3"] = (  # previous week's network-wide compliance % - computed once, referenced by the trend card below
        f'=IF($R$2="","-",IFERROR(ROUND(SUMPRODUCT({prev_cond}*{TP}!$F$2:$F$5000)/'
        f'SUMPRODUCT({prev_cond}*{TP}!$E$2:$E$5000)*100,1),"-"))'
    )
    ws.column_dimensions["Q"].hidden = True
    ws.column_dimensions["R"].hidden = True

    # ==========================================================================
    # KPI CARD ROW - network-wide totals for the most recent evaluated week
    # (docs/MANAGER_UX_ARCHITECTURE.md section 4: "same card component as
    # today's summary cards" - values are formulas over TECHNICIAN_PERFORMANCE_LOG,
    # not a new calculation).
    # ==========================================================================
    build_section_header(ws, "C5", "KOMPLIANCE - POSLEDNÍ VYHODNOCENÝ TÝDEN")
    latest_cond = f'({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000=$Q$2)'
    latest_planned = f'SUMPRODUCT({latest_cond}*{TP}!$E$2:$E$5000)'
    latest_realized = f'SUMPRODUCT({latest_cond}*{TP}!$F$2:$F$5000)'
    cards = [
        ("C", "D", "Naplánováno", f'=IF($Q$2="",0,{latest_planned})', NAVY, WHITE),
        ("E", "F", "Realizováno", f'=IF($Q$2="",0,{latest_realized})', NAVY, WHITE),
        ("G", "H", "Nesplněno", f'=IF($Q$2="",0,SUMPRODUCT({latest_cond}*{TP}!$I$2:$I$5000))', STATUS_CRITICAL, dashboard_ui.TINT_CRITICAL),
        ("I", "J", "Návštěvy navíc", f'=IF($Q$2="",0,SUMPRODUCT({latest_cond}*{TP}!$J$2:$J$5000))', STATUS_WARNING, dashboard_ui.TINT_WARNING),
        ("K", "L", "Compliance %", f'=IF($Q$2="",0,IFERROR(ROUND({latest_realized}/{latest_planned}*100,1),0))', NAVY, WHITE),
        ("M", "N", "Trend proti minulému týdnu", '=IF($Q$2="","Zatím žádná data",IF($R$2="","Zatím není s čím srovnat",""))', NAVY, WHITE),
    ]
    value_cells = build_kpi_card_row(ws, cards, label_row=6, value_row_start=7, value_row_end=9)
    home_compliance_cell = value_cells[4]  # "K7"
    home_trend_cell = value_cells[5]       # "M7"
    apply_severity_conditional_formatting(
        ws, "K7:L9", home_compliance_cell,
        thresholds=[(90, STATUS_GOOD), (70, STATUS_WARNING), (50, STATUS_SERIOUS)],
        below_color=STATUS_CRITICAL,
    )
    ws[home_trend_cell] = (
        f'=IF($Q$2="","Zatím žádná data",IF($R$2="","Zatím není s čím srovnat",'
        f'IF({home_compliance_cell}>$R$3,"▲ "&TEXT({home_compliance_cell}-$R$3,"+0.0")&" p.b.",'
        f'IF({home_compliance_cell}<$R$3,"▼ "&TEXT({home_compliance_cell}-$R$3,"+0.0;-0.0")&" p.b.","→ beze změny"))))'
    )
    ws[home_trend_cell].font = font_card_value(size=14)
    build_status_badge_conditional(ws, f"{home_trend_cell}:N9", home_trend_cell, rules=[
        ("▲", None, STATUS_GOOD),
        ("▼", None, STATUS_CRITICAL),
    ])
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 18

    # ---- Pipeline stages: each row's status (col I) is a LIVE formula that
    # reads the actual sheet the stage produces/consumes, not a static
    # instruction. Row numbers are fixed here so the "DALŠÍ KROK" callout
    # above can reference them even though it's written first. ----
    PIPE_FIRST_ROW = 32  # shifted +1 (KDO FLAKÁ) +4 (KDO JEZDÍ NEEFEKTIVNĚ) +4 (KDO JEZDÍ CIK-CAK) +5 (PLÁN AKTUÁLNOST) from 18
    stages = [
        # (num, name, description, status formula, target sheet or None, color)
        (
            "1", "IMPORT DAT", "Otevři Import Hub a vlož export(y) - lze i více najednou",
            '=IF(COUNTA(POS_MASTER!A:A)>1,"✅ Hotovo","❌ Chybí")',
            "IMPORT_HUB", "2E75B6",
        ),
        (
            "2", "PLÁN KAMPANÍ", "Nastav kampaně v ACTIVITY_PLAN",
            '=IF(COUNTA(ACTIVITY_PLAN!A:A)>1,"✅ Hotovo","❌ Chybí")',
            "ACTIVITY_PLAN", "BF8F00",
        ),
        (
            "3", "ROZPIS TECHNIKŮ", "Planning Engine vytvoří rozpis",
            '=IF(COUNTA(MANAGER_PLAN!A:A)>1,"✅ Hotovo","❌ Chybí")',
            "TECHNICIAN_PLAN", "375623",
        ),
        (
            "4", "PUBLIKACE", "Publish Engine odešle plán technikům",
            '=IF(COUNTIF(PLAN_LIFECYCLE!C:C,"Published")+COUNTIF(PLAN_LIFECYCLE!C:C,"Active")>0,"✅ Hotovo","❌ Chybí")',
            None, "BF8F00",
        ),
        (
            "5", "VYHODNOCENÍ", "Compliance + Advisor Engine porovná realitu s plánem",
            '=IF(COUNTA(COMPLIANCE_LOG!A:A)>1,"✅ Hotovo","❌ Chybí")',
            "DASHBOARD", "375623",
        ),
        (
            "6", "DASHBOARD", "Sleduj plnění, KPI a upozornění",
            '=IF(COUNTA(DASHBOARD!A5:A500)>0,"✅ Aktivní","⏳ Čeká na první běh")',
            "DASHBOARD", "375623",
        ),
    ]
    status_cells = [f"I{PIPE_FIRST_ROW + i}" for i in range(len(stages))]
    step_labels = [
        "1) Vlož export POS a SalesApp do RAW_DATA / POS_STATUS_IMPORT / SALESAPP_IMPORT a spusť Import Engine",
        "2) Vytvoř nebo prodluž kampaň v ACTIVITY_PLAN",
        "3) Spusť Planning Engine - vytvoří rozpis technikům",
        "4) Publikuj plán (Publish Engine)",
        "5) Spusť Compliance a Advisor Engine - vyhodnotí skutečné návštěvy proti plánu",
    ]

    # ==========================================================================
    # PLÁN JE AKTUÁLNÍ JEŠTĚ X DNÍ - product owner, 2026-07-11: "potřebuji
    # opravdu viditelně vidět, že ten tourplan bude aktuální ještě x dní a
    # viditelně chci dát generovat nový, který tourplan". The single most
    # prominent thing on HOME after the KPI row - a manager should see at a
    # glance whether this week's published plan is still current, without
    # opening PLAN_LIFECYCLE.
    # Deliberately NOT derived from MANAGER_PLAN_PUBLISHED's DATE column -
    # PlanningEngine.ts writes that as a locale-formatted STRING
    # (toLocaleDateString("cs-CZ")), and this session already found one real
    # bug from trusting a same-run string to behave like a Date (see
    # docs/BUSINESS_RULES.md section 20's matchedActualDate fix). Instead
    # this derives the plan's last valid day purely from PLAN_LIFECYCLE's
    # plain numeric year/week columns via ISO week arithmetic
    # (DATE(Y,1,4)-WEEKDAY(...) is always ISO week 1's Monday), so there is
    # no date-string round-trip to trust.
    # No live "click to regenerate" button is possible from an
    # openpyxl-generated file (Office Scripts button-binding is an Excel
    # Online UI action, not something storable in the .xlsx itself) - the
    # callout instead names the exact two scripts to run, and
    # NAVOD_INSTALACE.md documents how to pin them as real one-click buttons
    # via Excel Online's Automate pane ("Přidat tlačítko").
    # ==========================================================================
    ws["S1"] = "planLifecycleKey"
    ws["S2"] = (
        '=IFERROR(MAX(IF((PLAN_LIFECYCLE!$C$2:$C$2000="Published")+(PLAN_LIFECYCLE!$C$2:$C$2000="Active"),'
        'PLAN_LIFECYCLE!$A$2:$A$2000*100+PLAN_LIFECYCLE!$B$2:$B$2000)),"")'
    )
    ws["S3"] = '=IF($S$2="","",INT($S$2/100))'  # plan's year
    ws["S4"] = '=IF($S$2="","",MOD($S$2,100))'  # plan's ISO week
    ws["S5"] = '=IF($S$4="","",DATE($S$3,1,4)-WEEKDAY(DATE($S$3,1,4),3)+($S$4-1)*7)'  # Monday of that week
    ws["S6"] = '=IF($S$5="","",$S$5+4)'  # Friday - last day this plan covers
    ws["S7"] = '=IF($S$6="","",$S$6-TODAY())'  # days remaining (<=0 once expired)
    ws.column_dimensions["S"].hidden = True

    build_section_header(ws, "C10", "TOURPLAN - AKTUÁLNOST")
    r = 11
    days_left_ref = build_kpi_card(
        ws, "C", "D", r, r + 1, r + 2, "Plán aktuální ještě (dní)",
        '=IF($S$2="","–",IF($S$7>0,$S$7,0))',
        value_color=STATUS_GOOD, fill_color=dashboard_ui.TINT_GOOD,
    )
    ws.merge_cells(f"E{r}:M{r+2}")
    freshness_cell = ws.cell(r, 5)
    freshness_cell.value = (
        '=IF($S$2="","⏳ Zatím žádný publikovaný plán - spusť Planning Engine → Publish Engine, ať je co sledovat.",'
        'IF($S$7>0,"✅ Plán pro týden "&$S$4&"/"&$S$3&" je aktuální ještě "&$S$7&IF($S$7=1," den",IF($S$7<5," dny"," dní"))&".",'
        '"🔴 PLÁN PRO TÝDEN "&$S$4&"/"&$S$3&" VYPRŠEL - vygeneruj nový: spusť Planning Engine, pak Publish Engine."))'
    )
    freshness_cell.font = Font(bold=True, size=13, color=NAVY)
    freshness_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    freshness_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    build_status_badge_conditional(ws, freshness_cell.coordinate, freshness_cell.coordinate, rules=[
        ("✅", "E2EFDA", None),
        ("🔴", STATUS_CRITICAL, WHITE),
        ("⏳", "FFF2CC", None),
    ])
    build_status_badge_conditional(ws, f"{days_left_ref}:D{r+2}", days_left_ref, rules=[
        ("–", "F2F2F2", "808080"),
    ])
    ws.conditional_formatting.add(
        f"{days_left_ref}:D{r+2}",
        FormulaRule(formula=[f'AND(ISNUMBER({days_left_ref}),{days_left_ref}<3)'],
                    fill=PatternFill("solid", fgColor=dashboard_ui.TINT_WARNING), font=Font(bold=True, size=22, color=STATUS_WARNING)),
    )
    ws.conditional_formatting.add(
        f"{days_left_ref}:D{r+2}",
        FormulaRule(formula=[f'AND(ISNUMBER({days_left_ref}),{days_left_ref}=0)'],
                    fill=PatternFill("solid", fgColor=dashboard_ui.TINT_CRITICAL), font=Font(bold=True, size=22, color=STATUS_CRITICAL)),
    )
    _nav_button(ws, f"N{r}", "Rozpis →", "TECHNICIAN_PLAN", color="375623")
    for row_ in (r, r + 1, r + 2):
        ws.row_dimensions[row_].height = 18
    r += 4

    # ---- "KDO FLAKÁ" callout - product owner, 2026-07-06: "chci mit o
    # všem přehled" - a manager opening HOME should see at a glance whether
    # anyone is currently flagged (PerformanceEngine.ts's flakaRiziko, see
    # BUSINESS_RULES.md), without having to navigate to PERFORMANCE first.
    # Same single-callout-line pattern as "DALŠÍ KROK"/the pre-publish check
    # below, not a full table - this is a headline, not a report. ----
    ws.cell(r, 3, "KDO FLAKÁ").font = SECTION_FONT
    r += 1
    ws.merge_cells(f"C{r}:J{r+1}")
    flaka_summary_cell = ws.cell(r, 3)
    flaka_summary_cell.value = (
        '=IFERROR(IF(COUNTIF(TECHNICIAN_PERFORMANCE_SUMMARY!$O:$O,"Ano")=0,'
        '"✅ Žádný technik není aktuálně v riziku flákání",'
        '"⚠ "&COUNTIF(TECHNICIAN_PERFORMANCE_SUMMARY!$O:$O,"Ano")&" technik(ů) v riziku: "&'
        'TEXTJOIN(", ",TRUE,FILTER(TECHNICIAN_PERFORMANCE_SUMMARY!$A:$A,TECHNICIAN_PERFORMANCE_SUMMARY!$O:$O="Ano"))),'
        '"Zatím žádná data")'
    )
    flaka_summary_cell.font = Font(bold=True, size=13, color=NAVY)
    flaka_summary_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    flaka_summary_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    build_status_badge_conditional(ws, flaka_summary_cell.coordinate, flaka_summary_cell.coordinate, rules=[
        ("✅", "E2EFDA", None),
        ("⚠", STATUS_SERIOUS, None),
    ])
    r += 3

    # ---- "KDO JEZDÍ NEEFEKTIVNĚ" callout - found missing during a final
    # full test pass (2026-07-06): route efficiency (km/semafor,
    # PerformanceEngine.ts's maxKmDay) only ever existed per-technician on
    # TECHNICIAN_SCORECARD, with no network-wide "who had a bad day" view -
    # same gap PERFORMANCE's new "Km/den (nejhorší)" column just closed;
    # this is the HOME-level headline for the same signal, same pattern as
    # "KDO FLAKÁ" above. Reuses the same CONTROL.ROUTE_KM_CRITICAL_KM
    # threshold PERFORMANCE/TECHNICIAN_SCORECARD already use.
    ws.cell(r, 3, "KDO JEZDÍ NEEFEKTIVNĚ").font = SECTION_FONT
    r += 1
    ws.merge_cells(f"C{r}:J{r+1}")
    route_summary_cell = ws.cell(r, 3)
    route_summary_cell.value = (
        '=IFERROR(IF(COUNTIF(TECHNICIAN_PERFORMANCE_SUMMARY!$P:$P,'
        '">"&IFERROR(VLOOKUP("ROUTE_KM_CRITICAL_KM",CONTROL!$A:$B,2,FALSE),150))=0,'
        '"✅ Žádný technik nemá kritický den (km)",'
        '"⚠ "&COUNTIF(TECHNICIAN_PERFORMANCE_SUMMARY!$P:$P,'
        '">"&IFERROR(VLOOKUP("ROUTE_KM_CRITICAL_KM",CONTROL!$A:$B,2,FALSE),150))&'
        '" technik(ů) s kritickým dnem: "&TEXTJOIN(", ",TRUE,FILTER(TECHNICIAN_PERFORMANCE_SUMMARY!$A:$A,'
        'TECHNICIAN_PERFORMANCE_SUMMARY!$P:$P>IFERROR(VLOOKUP("ROUTE_KM_CRITICAL_KM",CONTROL!$A:$B,2,FALSE),150)))),'
        '"Zatím žádná data")'
    )
    route_summary_cell.font = Font(bold=True, size=13, color=NAVY)
    route_summary_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    route_summary_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    build_status_badge_conditional(ws, route_summary_cell.coordinate, route_summary_cell.coordinate, rules=[
        ("✅", "E2EFDA", None),
        ("⚠", STATUS_SERIOUS, None),
    ])
    r += 3

    # ---- "KDO ZE MĚ DĚLÁ BLBCE" callout (product owner, 2026-07-09,
    # "Monitoring efektivity" - vedoucí Field Force týmu: "chci se na
    # dashboard jen podívat a hned vidět, kdo ze mě dělá blbce") - keyed off
    # combinedRiskFlag (>= PROBLEM_SIGNAL_MIN_COUNT corroborating signals:
    # návštěvnost, hodnota/návštěva, délka návštěvy, trasa, compliance), NOT
    # off route efficiency alone - "GPS je odhad, takže to ani nemusí být na
    # vinu" (product owner, 2026-07-09). Distinct from "KDO JEZDÍ
    # NEEFEKTIVNĚ" above (that one is raw km/day, a single-day signal).
    # Links to the new EFFICIENCY heatmap, same pattern as every other HOME
    # callout.
    ws.cell(r, 3, "KDO ZE MĚ DĚLÁ BLBCE").font = SECTION_FONT
    r += 1
    ws.merge_cells(f"C{r}:H{r+1}")
    zigzag_cell = ws.cell(r, 3)
    zigzag_cell.value = (
        '=IFERROR(IF(COUNTIF(TECHNICIAN_PERFORMANCE_SUMMARY!$AC:$AC,"Ano")=0,'
        '"✅ Žádný technik nemá aktuálně víc než jeden varovný signál najednou",'
        '"🚩 "&COUNTIF(TECHNICIAN_PERFORMANCE_SUMMARY!$AC:$AC,"Ano")&'
        '" technik(ů) s více souběžnými signály (návštěvnost/hodnota/délka/trasa): "&'
        'TEXTJOIN(", ",TRUE,FILTER(TECHNICIAN_PERFORMANCE_SUMMARY!$A:$A,'
        'TECHNICIAN_PERFORMANCE_SUMMARY!$AC:$AC="Ano"))),'
        '"Zatím žádná data")'
    )
    zigzag_cell.font = Font(bold=True, size=13, color=NAVY)
    zigzag_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    zigzag_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    build_status_badge_conditional(ws, zigzag_cell.coordinate, zigzag_cell.coordinate, rules=[
        ("✅", "E2EFDA", None),
        ("🚩", STATUS_SERIOUS, None),
    ])
    _nav_button(ws, f"I{r}", "Detail →", "EFFICIENCY", color="2E75B6")
    r += 3

    # ---- "DALŠÍ KROK" callout: one live sentence, always the first
    # incomplete pipeline stage - this is the answer to "what do I do now",
    # not a checklist the user has to read themselves. ----
    ws.cell(r, 3, "DALŠÍ KROK").font = SECTION_FONT
    r += 1
    ws.merge_cells(f"C{r}:J{r+1}")
    next_step_cell = ws.cell(r, 3)
    ifs_args = []
    for cell_ref, label in zip(status_cells, step_labels):
        ifs_args.append(f'{cell_ref}="❌ Chybí"')
        ifs_args.append(f'"{label}"')
    next_step_cell.value = "=IFS(" + ",".join(ifs_args) + ',TRUE,"✅ Vše hotovo pro tento týden - sleduj plnění na DASHBOARD")'
    next_step_cell.font = Font(bold=True, size=13, color=NAVY)
    next_step_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    next_step_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    build_status_badge_conditional(ws, next_step_cell.coordinate, next_step_cell.coordinate, rules=[
        ("✅", "E2EFDA", None),
    ])
    r += 3

    # ---- Pipeline status strip ----
    ws.cell(r, 3, "STAV PROCESU").font = TITLE_FONT
    r += 1
    assert r == PIPE_FIRST_ROW, "PIPE_FIRST_ROW must match the row this loop actually starts at"
    # Each stage renders as one bordered "card row" (thin border on every
    # cell of the row, subtle alternating tint) instead of plain borderless
    # text - a small but real step toward "looks like an app, not a sheet
    # of numbers" (product owner, 2026-07-03).
    for stage_index, (num, name, desc, status_formula, target, color) in enumerate(stages):
        row_tint = "FFFFFF" if stage_index % 2 == 0 else "F7F9FB"
        for col in range(3, 11):
            cell = ws.cell(r, col)
            cell.border = CARD_BORDER
            if col != 3:  # badge (col C) sets its own fill below; conditional
                # formatting on col I overlays this when a rule matches
                cell.fill = PatternFill("solid", fgColor=row_tint)
        ws.cell(r, 3, num).font = Font(bold=True, size=16, color=WHITE)
        ws.cell(r, 3).fill = PatternFill("solid", fgColor=color)
        ws.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"D{r}:G{r}")
        ws.cell(r, 4, f"{name} — {desc}").font = Font(size=11)
        ws.cell(r, 4).alignment = Alignment(vertical="center", indent=1)
        status_cell = ws.cell(r, 9, status_formula)
        status_cell.font = Font(bold=True, size=11)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        build_status_badge_conditional(ws, status_cell.coordinate, status_cell.coordinate, rules=[
            ("✅", "E2EFDA", None),
            ("❌", STATUS_SERIOUS, None),
        ])
        if target:
            _nav_button(ws, f"J{r}", "Otevřít →", target, color=color)
        else:
            ws.cell(r, 10, "⚙ Automatizace")
            ws.cell(r, 10).font = Font(italic=True, size=10, color="808080")
            ws.cell(r, 10).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1

    # ---- Post-generation summary (product owner, 2026-07-09): "Vybráno X
    # poboček, PPT: Y" - a one-line answer to "what did the last Planning
    # Engine run actually produce", right under the pipeline stages so it's
    # the immediate next thing seen after stage 3 (ROZPIS TECHNIKŮ). Reads
    # MANAGER_PLAN directly - no new engine field needed (distinct-POS-count
    # is the same SUMPRODUCT/COUNTIF pattern as "POS pokryto plánem" below,
    # PPT is a plain SUM). ----
    ws.merge_cells(f"C{r}:J{r}")
    plan_summary_cell = ws.cell(r, 3)
    _summary_pos_range = "MANAGER_PLAN!E2:E200000"
    plan_summary_cell.value = (
        f'=IF(COUNTA(MANAGER_PLAN!A:A)<=1,"Zatím žádný vygenerovaný plán",'
        f'"✅ Vybráno "&SUMPRODUCT(({_summary_pos_range}<>"")/COUNTIF({_summary_pos_range},{_summary_pos_range}&""))'
        f'&" poboček do plánu, celkové PPT: "&TEXT(SUM(MANAGER_PLAN!M2:M200000),"#,##0"))'
    )
    plan_summary_cell.font = Font(bold=True, size=12, color=NAVY)
    plan_summary_cell.fill = PatternFill("solid", fgColor="E2EFDA")
    plan_summary_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 20
    r += 2

    # ---- Terminal-type "last used" countdown (product owner, 2026-07-09):
    # per terminal type (VELKY/SMALL/LI, from TERMINAL_RULES), how many
    # weeks on average since Active POS of that type were last visited - a
    # network-wide early-warning signal distinct from any single technician's
    # scorecard. Higher = worse (opposite direction from the compliance %
    # cards above), so severity coloring is applied manually rather than via
    # apply_severity_conditional_formatting's "higher is better" convention.
    # Reuses NEGLECTED_AFTER_WEEKS/ADVISOR_NEGLECT_WARNING_RATIO_PERCENT
    # (already the WARNING/CRITICAL cutoffs Advisor Engine itself uses) so
    # this headline agrees with whatever ADVISOR_LOG would eventually flag,
    # rather than inventing a second threshold.
    # ==========================================================================
    build_section_header(ws, f"C{r}", "TERMINÁLY - PRŮMĚRNÝ POČET TÝDNŮ OD NÁVŠTĚVY")
    r += 1
    term_status_range = "POS_MASTER!Q2:Q20000"
    term_type_range = "POS_MASTER!E2:E20000"
    term_weeks_range = "POS_MASTER!AA2:AA20000"
    neglected_after_formula = 'IFERROR(VLOOKUP("NEGLECTED_AFTER_WEEKS",CONTROL!$A:$B,2,FALSE),26)'
    warn_ratio_formula = 'IFERROR(VLOOKUP("ADVISOR_NEGLECT_WARNING_RATIO_PERCENT",CONTROL!$A:$B,2,FALSE),80)/100'
    terminal_types = [("VELKY TERMINAL", "C", "D"), ("SMALL TERMINAL", "E", "F"), ("LI", "G", "H")]
    for term_type, label_col, _value_col in terminal_types:
        ws[f"{label_col}{r}"] = term_type
        ws[f"{label_col}{r}"].font = Font(size=9, color="595959")
        value_cell = f"{label_col}{r+1}"
        ws[value_cell] = (
            f'=IFERROR(ROUND(AVERAGEIFS({term_weeks_range},{term_status_range},"Active",'
            f'{term_type_range},"{term_type}"),1)&" týdnů","Žádná data")'
        )
        ws[value_cell].font = Font(bold=True, size=16, color=NAVY)
        numeric_cell = f'AVERAGEIFS({term_weeks_range},{term_status_range},"Active",{term_type_range},"{term_type}")'
        ws.conditional_formatting.add(
            value_cell,
            FormulaRule(formula=[f"{numeric_cell}>={neglected_after_formula}"], fill=PatternFill("solid", fgColor=STATUS_CRITICAL)),
        )
        ws.conditional_formatting.add(
            value_cell,
            FormulaRule(formula=[f"{numeric_cell}>={neglected_after_formula}*{warn_ratio_formula}"], fill=PatternFill("solid", fgColor=STATUS_WARNING)),
        )
    ws.row_dimensions[r].height = 14
    ws.row_dimensions[r + 1].height = 20
    r += 3

    # ---- Status strip (live formulas - stays accurate without any manual
    # update). Operational numbers distinct from the compliance KPI cards
    # above (plan size/coverage, not plan vs. reality). Two rows of 3 tiles.
    # Distinct-POS/distinct-technician tiles reuse the same SUMPRODUCT/
    # COUNTIF distinct-count pattern already proven in
    # redesign_activity_plan()'s reference panel. ----
    ws.cell(r, 3, "PROVOZNÍ PŘEHLED").font = SECTION_FONT
    r += 1
    mp_pos_range = "MANAGER_PLAN!E2:E200000"
    mp_tech_range = "MANAGER_PLAN!D2:D200000"
    strip = [
        ("D", "Aktuální kampaň týden", '=IFERROR(VLOOKUP("CAMPAIGN_START_WEEK",CONTROL!A:B,2,FALSE),"-")'),
        ("F", "POS v systému", '=COUNTA(POS_MASTER!A:A)-1'),
        ("H", "Naplánováno návštěv", '=COUNTA(MANAGER_PLAN!A:A)-1'),
        ("D", "POS pokryto plánem", f'=SUMPRODUCT(({mp_pos_range}<>"")/COUNTIF({mp_pos_range},{mp_pos_range}&""))'),
        ("F", "Techniků naplánováno", f'=SUMPRODUCT(({mp_tech_range}<>"")/COUNTIF({mp_tech_range},{mp_tech_range}&""))'),
        ("H", "Aktivní kampaně", '=COUNTA(ACTIVITY_PLAN!A:A)-1'),
        ("D", "Otevřená upozornění", "=DASHBOARD!E3"),
    ]
    row_offsets = [0, 0, 0, 3, 3, 3, 6]
    for (col, label, formula), row_offset in zip(strip, row_offsets):
        ws[f"{col}{r + row_offset}"] = label
        ws[f"{col}{r + row_offset}"].font = Font(size=9, color="595959")
        value_font = Font(bold=True, size=20, color=NAVY) if row_offset == 0 else Font(bold=True, size=16, color=NAVY)
        ws[f"{col}{r + row_offset + 1}"] = formula
        ws[f"{col}{r + row_offset + 1}"].font = value_font
    ws.conditional_formatting.add(
        f"D{r + 7}",
        FormulaRule(formula=[f"D{r + 7}>0"], fill=PatternFill("solid", fgColor=STATUS_SERIOUS)),
    )
    r += 9

    # ---- Pre-publish sanity check: the one thing a manager actually needs
    # before hitting Publish for 10-20 technicians once a week - "did anyone
    # get silently left out of this plan". Compares distinct technicians
    # with at least one Active POS assigned (POS_MASTER) against distinct
    # technicians actually present in this week's plan (MANAGER_PLAN). A
    # mismatch is the single costliest realistic weekly mistake in this
    # workflow (a technician shows up Monday with no route) and is
    # otherwise invisible - nothing else on this workbook would surface it
    # before publish. Pure comparison of two counts already computed above,
    # no new business logic, no new engine field. ----
    tech_pm_range = f"POS_MASTER!{pos_master_tech_col}2:{pos_master_tech_col}20000"
    active_range = "POS_MASTER!Q2:Q20000"
    active_tech_count_formula = (
        f'=SUMPRODUCT(({tech_pm_range}<>"")*({active_range}="Active")'
        f'/COUNTIFS({tech_pm_range},{tech_pm_range}&"",{active_range},"Active"))'
    )
    ws.cell(r, 3, "PŘED PUBLIKACÍ ZKONTROLUJ").font = TITLE_FONT
    r += 1
    ws.cell(r, 3, "Technik. s aktivními POS").font = Font(size=9, color="595959")
    ws.cell(r, 5, "Technik. v plánu tento týden").font = Font(size=9, color="595959")
    r += 1
    ws.cell(r, 3, active_tech_count_formula).font = Font(bold=True, size=16, color=NAVY)
    ws.cell(r, 5, f'=SUMPRODUCT(({mp_tech_range}<>"")/COUNTIF({mp_tech_range},{mp_tech_range}&""))').font = Font(bold=True, size=16, color=NAVY)
    check_row = r
    r += 1
    ws.merge_cells(f"C{r}:I{r}")
    check_cell = ws.cell(
        r, 3,
        f'=IF(C{check_row}=E{check_row},"✅ Počty souhlasí",'
        f'"⚠ Nesoulad ("&C{check_row}&" vs "&E{check_row}&") - zkontroluj, jestli někdo nechybí v plánu")'
    )
    check_cell.font = Font(bold=True, size=11, color=NAVY)
    check_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    check_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 20
    build_status_badge_conditional(ws, check_cell.coordinate, check_cell.coordinate, rules=[
        ("✅", "E2EFDA", None),
        ("⚠", STATUS_SERIOUS, None),
    ])
    r += 2

    # ---- Quick navigation ----
    ws.cell(r, 3, "RYCHLÁ NAVIGACE").font = TITLE_FONT
    r += 1
    quick_links = [
        ("SCORECARD", "TECHNICIAN_SCORECARD", "2E75B6"),
        ("PERFORMANCE", "PERFORMANCE", "2E75B6"),
        ("EFEKTIVITA", "EFFICIENCY", "2E75B6"),
        ("MANUÁL", "MANUAL", "808080"),
        ("WEEK DASHBOARD", "WEEK_DASHBOARD", "2E75B6"),
        # BUG FIX (found 2026-07-06 during a full test pass): MAP was added
        # to the shared nav rail (dashboard_ui.NAV_RAIL_SHEETS) but never
        # added here, so HOME's own quick-nav row was missing it.
        ("MAPA ÚZEMÍ", "MAP", "2E75B6"),
        ("DASHBOARD", "DASHBOARD", "375623"),
        ("TECHNICIAN_PLAN", "TECHNICIAN_PLAN", "375623"),
        ("POS_MASTER", "POS_MASTER", "7030A0"),
        ("ACTIVITY_PLAN", "ACTIVITY_PLAN", "BF8F00"),
        ("TYPY TERMINÁLŮ", "TERMINAL_RULES", "BF8F00"),
        ("BLACKLIST", "BLACKLIST", "C00000"),
        ("AKTIVOVAT POS", "POS_ACTIVATE_LIST", "375623"),
    ]
    col_idx = 3
    for label, target, color in quick_links:
        col_letter = get_column_letter(col_idx)
        _nav_button(ws, f"{col_letter}{r}", label, target, color=color)
        col_idx += 2
    ws.row_dimensions[r].height = 24
    r += 2

    # ---- Legend, inline, not an appendix ----
    ws.cell(r, 3, "JAK ČÍST BARVY").font = TITLE_FONT
    r += 1
    legend_items = [
        (EDITABLE_FILL, "Editovatelné - sem zapisuješ"),
        (SYSTEM_FILL, "Systémové - počítá engine"),
        (IMPORT_FILL, "Sem vlož export"),
        (OUTPUT_FILL, "Výsledek plánování"),
    ]
    col_idx = 3
    for color, text in legend_items:
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}{r}"] = ""
        ws[f"{col_letter}{r}"].fill = PatternFill("solid", fgColor=color)
        ws.merge_cells(f"{get_column_letter(col_idx+1)}{r}:{get_column_letter(col_idx+1)}{r}")
        ws.cell(r, col_idx + 1, text).font = Font(size=9)
        ws.cell(r, col_idx + 1).alignment = Alignment(vertical="center", wrap_text=True)
        col_idx += 2
    ws.row_dimensions[r].height = 28
    r += 2

    # ---- First-time setup, kept short - detail lives in office-scripts/README.md ----
    ws.cell(r, 3, "PRVNÍ SPUŠTĚNÍ").font = TITLE_FONT
    r += 1
    for text in [
        "1) Otevři tento sešit v Excelu na webu (OneDrive/SharePoint) - Office Scripts to vyžadují.",
        "2) Záložka Automatizace → New Script → vlož obsah office-scripts/ImportEngine.ts → Spustit.",
        "3) Opakuj pro PlanningEngine.ts, PublishEngine.ts, StartTrackingEngine.ts, ComplianceEngine.ts, AdvisorEngine.ts, PerformanceEngine.ts, ReportingEngine.ts.",
    ]:
        ws.merge_cells(f"D{r}:J{r}")
        ws.cell(r, 4, text).font = Font(size=10)
        r += 1

    return ws


def build_technician_scorecard(wb):
    """The first screen of the manager UX layer (docs/MANAGER_UX_ARCHITECTURE.md
    section 4) - a technician/week-driven dashboard, not a table. Every KPI
    is a live formula over TECHNICIAN_PERFORMANCE_LOG/TECHNICIAN_TOP_ISSUES
    (PerformanceEngine.ts's output) keyed on two dropdowns; no new business
    logic here (see that engine's file header for why "TOP problematic POS"
    is computed there, not as a raw-COMPLIANCE_LOG formula).

    Built entirely out of tools/dashboard_ui.py's shared components (nav
    rail, banner, KPI cards, progress bar, severity conditional formatting,
    charts) - this function only supplies the sheet-specific formulas and
    layout positions, never re-derives styling. See that module's docstring
    for why (product owner, 2026-07-03: one UI library, dashboards only
    compose it).

    All formula plumbing (unique technician/week lists, parsed year/week,
    previous-week lookup, chart data blocks) lives in hidden columns P:W -
    per the product owner's explicit ask ("minimum mřížky Excelu... působit
    jako desktopová aplikace"), the visible area (C:N) shows only the
    finished dashboard, nothing a user would recognize as "Excel plumbing"."""
    TP = "TECHNICIAN_PERFORMANCE_LOG"
    TI = "TECHNICIAN_TOP_ISSUES"
    if "TECHNICIAN_SCORECARD" in wb.sheetnames:
        del wb["TECHNICIAN_SCORECARD"]
    ws = wb.create_sheet("TECHNICIAN_SCORECARD")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    for col in "CDEFGHIJKLMN":
        ws.column_dimensions[col].width = 12
    build_nav_rail(ws, "TECHNICIAN_SCORECARD")

    build_dashboard_banner(
        ws, "TECHNICIAN SCORECARD", "Výkon technika v čase - vyber technika a týden níže",
        col_start="C", col_end="N",
    )
    ws.freeze_panes = "C4"

    # ==========================================================================
    # HIDDEN FORMULA PLUMBING (columns P:W) - never shown to the user.
    # ==========================================================================
    ws["P1"] = "technici"
    ws["P2"] = (
        '=IFERROR(SORT(UNIQUE(FILTER('
        f'{TP}!$A$2:$A$5000,{TP}!$A$2:$A$5000<>""))),"Zatím žádná data")'
    )
    ws["Q1"] = "týdny pro vybraného technika"
    ws["Q2"] = (
        '=IFERROR(SORT(UNIQUE(FILTER('
        f'TEXT({TP}!$B$2:$B$5000,"0000")&"-W"&TEXT({TP}!$C$2:$C$5000,"00"),'
        f'{TP}!$A$2:$A$5000=$D$5))),"Vyber technika")'
    )
    ws["R1"] = '=IFERROR(VALUE(LEFT($H$5,4)),0)'   # selected year
    ws["R2"] = '=IFERROR(VALUE(MID($H$5,7,2)),0)'  # selected week
    ws["R3"] = (  # previous week on record for this technician (any gap size)
        f'=IFERROR(AGGREGATE(14,6,({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)/'
        f'(({TP}!$A$2:$A$5000=$D$5)*(({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)<($R$1*100+$R$2))),1),"")'
    )
    ws["R4"] = (  # that previous week's compliance %
        f'=IF($R$3="","-",SUMPRODUCT(({TP}!$A$2:$A$5000=$D$5)*'
        f'(({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)=$R$3)*{TP}!$K$2:$K$5000))'
    )
    # daily distribution data block (S=label, T=value)
    day_cols = [("S1", "Po", "L"), ("S2", "Út", "M"), ("S3", "St", "N"), ("S4", "Čt", "O"), ("S5", "Pá", "P")]
    for i, (cell_ref, label, tp_col) in enumerate(day_cols, start=1):
        ws[cell_ref] = label
        ws[f"T{i}"] = (
            f'=SUMPRODUCT(({TP}!$A$2:$A$5000=$D$5)*({TP}!$B$2:$B$5000=$R$1)*'
            f'({TP}!$C$2:$C$5000=$R$2)*{TP}!${tp_col}$2:${tp_col}$5000)'
        )
    # last-6-weeks trend data block (W=resolved year*100+week key, U=label, V=value)
    for i in range(1, 7):
        k = 7 - i
        ws[f"W{i}"] = (
            f'=IFERROR(AGGREGATE(14,6,({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)/'
            f'({TP}!$A$2:$A$5000=$D$5),{k}),"")'
        )
        ws[f"U{i}"] = f'=IF($W{i}="","","W"&TEXT(MOD($W{i},100),"00"))'
        ws[f"V{i}"] = (
            f'=IF($W{i}="","",SUMPRODUCT(({TP}!$A$2:$A$5000=$D$5)*'
            f'(({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)=$W{i})*{TP}!$K$2:$K$5000))'
        )
    # long-term monthly trend data block (Z=resolved monthKey, AA=label,
    # AB=avg compliance) - product owner, 2026-07-06, after the weekly/
    # 4-week views above: "je pro mě i důležitý dlouhodobý pohled" - wants
    # compliance trend across months/campaigns, not just the last 6 weeks.
    # Same AGGREGATE-LARGE-k pattern as the 6-week block above, but grouping
    # by PerformanceEngine.ts's monthKey (TP!AC, YYYYMM) instead of raw week -
    # a technician can have several weeks in the same month, so this
    # averages compliancePercent across all of that technician's rows for
    # each of their last 12 distinct months on record, oldest to newest.
    for i in range(1, 13):
        k = 13 - i
        ws[f"Z{i}"] = (
            f'=IFERROR(AGGREGATE(14,6,{TP}!$AC$2:$AC$5000/({TP}!$A$2:$A$5000=$D$5),{k}),"")'
        )
        ws[f"AA{i}"] = f'=IF($Z{i}="","",INT($Z{i}/100)&"/"&TEXT(MOD($Z{i},100),"00"))'
        ws[f"AB{i}"] = (
            f'=IF($Z{i}="","",AVERAGEIFS({TP}!$K$2:$K$5000,{TP}!$A$2:$A$5000,$D$5,'
            f'{TP}!$AC$2:$AC$5000,$Z{i}))'
        )

    # route-efficiency (km) data block - product owner, 2026-07-06: "číslo
    # kolik najel km třeba mezi těmi pos a semafor" - estimated daily
    # driving distance between visited POS, from PerformanceEngine.ts's
    # kmMon..kmFri columns (TP!R:V), with semaphore thresholds tunable via
    # CONTROL.ROUTE_KM_WARNING_KM/ROUTE_KM_CRITICAL_KM (see that sheet's
    # comment: thresholds are a proposed default, not a confirmed rule).
    ws["X1"] = '=IFERROR(VLOOKUP("ROUTE_KM_WARNING_KM",CONTROL!$A:$B,2,FALSE),80)'
    ws["X2"] = '=IFERROR(VLOOKUP("ROUTE_KM_CRITICAL_KM",CONTROL!$A:$B,2,FALSE),150)'
    km_day_cols = [("Y1", "R"), ("Y2", "S"), ("Y3", "T"), ("Y4", "U"), ("Y5", "V")]
    for cell_ref, tp_col in km_day_cols:
        ws[cell_ref] = (
            f'=SUMPRODUCT(({TP}!$A$2:$A$5000=$D$5)*({TP}!$B$2:$B$5000=$R$1)*'
            f'({TP}!$C$2:$C$5000=$R$2)*{TP}!${tp_col}$2:${tp_col}$5000)'
        )
    ws["Y6"] = "=SUM($Y$1:$Y$5)"

    for col in ["P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB"]:
        ws.column_dimensions[col].hidden = True

    # Named Ranges over the hidden spill formulas above - the standard
    # filter-dropdown pattern for every dashboard screen (product owner,
    # 2026-07-05: prefer Named Ranges + Data Validation lists over raw
    # `$P$2#`-style cell references, so the wiring is inspectable in
    # Excel's own Name Manager, not just in this Python scaffold).
    dashboard_ui.define_named_range(ws, "TechnicianList", "TECHNICIAN_SCORECARD!$P$2#")
    dashboard_ui.define_named_range(ws, "TechnicianWeekList", "TECHNICIAN_SCORECARD!$Q$2#")

    # ==========================================================================
    # FILTER BAR - the two dropdowns that drive the entire sheet.
    # ==========================================================================
    build_filter_bar_background(ws, 5, "C", "N")
    build_filter_dropdown(ws, "C5", "TECHNIK", "D5:F5", "=TechnicianList", default_formula='=IFERROR(INDEX(TechnicianList,1),"")')
    build_filter_dropdown(ws, "G5", "TÝDEN", "H5:J5", "=TechnicianWeekList", default_formula='=IFERROR(INDEX(TechnicianWeekList,1),"")')
    ws.merge_cells("K5:N5")
    # "Flaká riziko" badge appended to the region line (product owner,
    # 2026-07-06: "chci aby mi to ukazalo ktery z nich flaka a ktery ne" -
    # see PerformanceEngine.ts's "EIGHTH OUTPUT ADDITION" header comment).
    # Technician-level (not week-scoped, unlike the region lookup), so it's
    # looked up from TECHNICIAN_PERFORMANCE_SUMMARY by technician alone.
    ws["K5"] = (
        f'=IFERROR("Region: "&INDEX({TP}!$D$2:$D$5000,MATCH(1,({TP}!$A$2:$A$5000=$D$5)*'
        f'({TP}!$B$2:$B$5000=$R$1)*({TP}!$C$2:$C$5000=$R$2),0)),"-")&'
        f'IF(IFERROR(INDEX(TECHNICIAN_PERFORMANCE_SUMMARY!$O:$O,MATCH($D$5,TECHNICIAN_PERFORMANCE_SUMMARY!$A:$A,0)),"Ne")="Ano",'
        f'"  ⚠ RIZIKO FLÁKÁNÍ","")'
    )
    ws["K5"].font = Font(italic=True, size=10, color="595959")
    ws["K5"].alignment = Alignment(vertical="center", horizontal="right", indent=1)
    ws.conditional_formatting.add(
        "K5:N5",
        FormulaRule(formula=['ISNUMBER(SEARCH("FLÁKÁNÍ",K5))'], font=Font(italic=True, bold=True, size=10, color=STATUS_CRITICAL)),
    )

    # ==========================================================================
    # KPI CARD ROW - 6 cards, each a 2-column-wide tile.
    # ==========================================================================
    build_section_header(ws, "C7", "KPI PŘEHLED")
    tp_cond = f'({TP}!$A$2:$A$5000=$D$5)*({TP}!$B$2:$B$5000=$R$1)*({TP}!$C$2:$C$5000=$R$2)'
    cards = [
        ("C", "D", "Plánováno", f'=SUMPRODUCT({tp_cond}*{TP}!$E$2:$E$5000)', NAVY, WHITE),
        ("E", "F", "Realizováno", f'=SUMPRODUCT({tp_cond}*{TP}!$F$2:$F$5000)', NAVY, WHITE),
        ("G", "H", "Splněno (včas+pozdě)", f'=SUMPRODUCT({tp_cond}*({TP}!$G$2:$G$5000+{TP}!$H$2:$H$5000))', STATUS_GOOD, dashboard_ui.TINT_GOOD),
        ("I", "J", "Nesplněno", f'=SUMPRODUCT({tp_cond}*{TP}!$I$2:$I$5000)', STATUS_CRITICAL, dashboard_ui.TINT_CRITICAL),
        ("K", "L", "Návštěvy navíc", f'=SUMPRODUCT({tp_cond}*{TP}!$J$2:$J$5000)', STATUS_WARNING, dashboard_ui.TINT_WARNING),
        ("M", "N", "Compliance %", f'=SUMPRODUCT({tp_cond}*{TP}!$K$2:$K$5000)', NAVY, WHITE),
    ]
    value_cells = build_kpi_card_row(ws, cards, label_row=8, value_row_start=9, value_row_end=11)
    ws.row_dimensions[9].height = 18
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 18
    compliance_cell = value_cells[5]  # "M9"
    nesplneno_cell = value_cells[3]   # "I9"
    apply_severity_conditional_formatting(
        ws, "M9:N11", compliance_cell,
        thresholds=[(90, STATUS_GOOD), (70, STATUS_WARNING), (50, STATUS_SERIOUS)],
        below_color=STATUS_CRITICAL,
    )
    ws.conditional_formatting.add(
        "I9:J11",
        FormulaRule(formula=[f"{nesplneno_cell}=0"], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_GOOD),
                    font=Font(bold=True, size=22, color=STATUS_GOOD)),
    )

    # ---- Compliance progress bar ----
    ws.cell(12, 3, "Poměr realizace vs. plán").font = NOTE_FONT
    build_progress_bar(ws, "C13:N13", compliance_cell, max_value=100, color=dashboard_ui.ACCENT_BLUE)

    # ==========================================================================
    # LONG-RUN AVERAGE + TREND + BUSIEST DAY - 3 side-by-side cards.
    # ==========================================================================
    build_section_header(ws, "C15", "DLOUHODOBÝ VÝKON")
    build_kpi_card(
        ws, "C", "F", 16, 17, 18, "Dlouhodobý průměr (compliance %)",
        f'=IFERROR(ROUND(AVERAGEIFS({TP}!$K$2:$K$5000,{TP}!$A$2:$A$5000,$D$5),1),"-")',
        value_color=NAVY, fill_color=WHITE,
    )
    ws["C17"].font = font_card_value(size=18)
    trend_cell_ref = build_kpi_card(
        ws, "G", "J", 16, 17, 18, "Trend proti minulému týdnu",
        f'=IF($R$3="","Zatím není s čím srovnat",'
        f'IF({compliance_cell}>$R$4,"▲ "&TEXT({compliance_cell}-$R$4,"+0.0")&" p.b.",'
        f'IF({compliance_cell}<$R$4,"▼ "&TEXT({compliance_cell}-$R$4,"+0.0;-0.0")&" p.b.","→ beze změny")))',
        value_color=NAVY, fill_color=WHITE,
    )
    ws[trend_cell_ref].font = font_card_value(size=16)
    ws[trend_cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    build_status_badge_conditional(ws, "G17:J18", trend_cell_ref, rules=[
        ("▲", None, STATUS_GOOD),
        ("▼", None, STATUS_CRITICAL),
    ])
    busiest_day_ref = build_kpi_card(
        ws, "K", "N", 16, 17, 18, "Nejvytíženější den",
        '=IF(SUM($T$1:$T$5)=0,"Zatím žádná data",'
        'INDEX($S$1:$S$5,MATCH(MAX($T$1:$T$5),$T$1:$T$5,0))&" ("&MAX($T$1:$T$5)&" návštěv)")',
        value_color=NAVY, fill_color=WHITE,
    )
    ws[busiest_day_ref].font = font_card_value(size=16)
    ws[busiest_day_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ==========================================================================
    # DAILY DISTRIBUTION + 6-WEEK TREND - two small native charts.
    # ==========================================================================
    build_section_header(ws, "C20", "DENNÍ ROZLOŽENÍ NÁVŠTĚV")
    d_cats = Reference(ws, min_col=19, min_row=1, max_row=5)   # S1:S5
    d_data = Reference(ws, min_col=20, min_row=1, max_row=5)   # T1:T5
    ws.add_chart(make_bar_chart(d_cats, d_data, color=dashboard_ui.ACCENT_BLUE), "C21")

    build_section_header(ws, "I20", "VÝVOJ COMPLIANCE (posl. 6 týdnů)")
    t_cats = Reference(ws, min_col=21, min_row=1, max_row=6)  # U1:U6
    t_data = Reference(ws, min_col=22, min_row=1, max_row=6)  # V1:V6
    ws.add_chart(make_line_chart(t_cats, t_data, color=dashboard_ui.ACCENT_BLUE), "I21")

    # ==========================================================================
    # DLOUHODOBÝ TREND (12 měsíců) - product owner, 2026-07-06, after the
    # weekly/4-week views above: "je pro mě i důležitý dlouhodobý pohled" -
    # compliance trend across months/campaigns, not just the last 6 weeks.
    # ==========================================================================
    build_section_header(ws, "C31", "DLOUHODOBÝ TREND (posl. 12 měsíců)")
    m_cats = Reference(ws, min_col=27, min_row=1, max_row=12)  # AA1:AA12
    m_data = Reference(ws, min_col=28, min_row=1, max_row=12)  # AB1:AB12
    ws.add_chart(make_line_chart(m_cats, m_data, color=dashboard_ui.ACCENT_BLUE), "C32")

    # ==========================================================================
    # TRASA / EFEKTIVITA JÍZD - odhad km mezi navštívenými POS + semafor.
    # No real GPS/timestamp tracking exists in this system (see
    # PerformanceEngine.ts's routeKmForDay header comment): this is an
    # estimate based on the technician's planned visit order for that date,
    # not measured driving distance.
    # ==========================================================================
    build_section_header(ws, "C42", "TRASA / EFEKTIVITA JÍZD (odhad km mezi POS)")
    style_dashboard_table_header(ws, 43, "CDEFGHI", ["Po", "Út", "St", "Čt", "Pá", "Týden celkem", "Ostatní návštěvy"])
    km_cells = [("C44", "Y1"), ("D44", "Y2"), ("E44", "Y3"), ("F44", "Y4"), ("G44", "Y5")]
    for target, source in km_cells:
        ws[target] = f"={source}"
        ws[target].number_format = '0.0" km"'
        ws[target].font = Font(bold=True, size=12)
        ws[target].alignment = Alignment(horizontal="center", vertical="center")
        ws[target].border = CARD_BORDER
    ws["H44"] = "=$Y$6"
    ws["H44"].number_format = '0.0" km"'
    ws["H44"].font = font_card_value(size=14, color=NAVY)
    ws["H44"].fill = PatternFill("solid", fgColor=WHITE)
    ws["H44"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H44"].border = CARD_BORDER
    # Ostatní návštěvy (otherVisits, TP!W) - informational only, not part of
    # the compliance/campaign gate (see PerformanceEngine.ts's "SIXTH OUTPUT
    # ADDITION" header comment): visits at that technician's POS with a
    # non-campaign SalesApp purpose (restocking, lottery ticket downloads...).
    # Neutral gray styling on purpose - this is context, not a KPI to chase.
    ws["I44"] = f'=SUMPRODUCT({tp_cond}*{TP}!$W$2:$W$5000)'
    ws["I44"].font = font_card_value(size=14, color=NAVY)
    ws["I44"].fill = PatternFill("solid", fgColor=dashboard_ui.LIGHT_GREY)
    ws["I44"].alignment = Alignment(horizontal="center", vertical="center")
    ws["I44"].border = CARD_BORDER
    ws.row_dimensions[44].height = 20
    for col, cell in (("C", "C44"), ("D", "D44"), ("E", "E44"), ("F", "F44"), ("G", "G44")):
        rng = f"{col}44"
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f"{cell}<=$X$1"], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_GOOD),
                              font=Font(bold=True, size=12, color=STATUS_GOOD), stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f"{cell}<=$X$2"], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_WARNING),
                              font=Font(bold=True, size=12, color=STATUS_WARNING), stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f"{cell}>$X$2"], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_CRITICAL),
                              font=Font(bold=True, size=12, color=STATUS_CRITICAL), stopIfTrue=True),
        )

    # ==========================================================================
    # POS PO DNECH - which POS the technician actually visited each day, in
    # the same order the km estimate above was computed from (product owner,
    # 2026-07-06: "na tady mě to zajímá až na dny, zda jezdil efektivně,
    # kolik jich udělal a pos" - wants the concrete POS list per day, not just
    # a count/km number). Reads PerformanceEngine.ts's posListMon..posListFri
    # (TP!X..AB) - comma-separated "id - name" text, one FILTER lookup per
    # day since each day's list is a single wide text value, not a range.
    # ==========================================================================
    build_section_header(ws, "C47", "POS PO DNECH (v pořadí trasy)")
    day_pos_list_cols = [("Po", "X"), ("Út", "Y"), ("St", "Z"), ("Čt", "AA"), ("Pá", "AB")]
    row = 48
    for label, tp_col in day_pos_list_cols:
        ws.cell(row, 3, label).font = Font(bold=True, size=10, color=NAVY)
        ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=14)
        cell = ws.cell(row, 4)
        cell.value = (
            f'=IFERROR(INDEX(FILTER({TP}!${tp_col}$2:${tp_col}$5000,{tp_cond}),1),"Žádné návštěvy")'
        )
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col in range(3, 15):
            ws.cell(row, col).border = CARD_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    # ==========================================================================
    # NÁVŠTĚVY PO DNECH (kampaň / ostatní) - product owner, 2026-07-11:
    # "chybi mi tam zobrazení kolik udelal za den a podobně, neni ten
    # salesapp pořádně vytezeny". The daily counts (TP!L:P campaign,
    # TP!AD:AH ad-hoc) were already computed by PerformanceEngine.ts and
    # used elsewhere on this sheet (daily chart, Ostatní návštěvy KPI), but
    # never shown broken out by day as its own table - a pure visibility
    # gap, no new engine logic needed here.
    # ==========================================================================
    build_section_header(ws, "C53", "NÁVŠTĚVY PO DNECH (kampaň / ostatní)")
    style_dashboard_table_header(ws, 54, "CDEFGHI", ["", "Po", "Út", "St", "Čt", "Pá", "Týden celkem"])
    visit_rows = [
        (55, "Kampaň (plán)", ["L", "M", "N", "O", "P"]),
        (56, "Ostatní (SalesApp)", ["AD", "AE", "AF", "AG", "AH"]),
    ]
    for row, label, cols in visit_rows:
        ws.cell(row, 3, label).font = Font(size=10)
        for col_letter, tp_col in zip("DEFGH", cols):
            cell = ws[f"{col_letter}{row}"]
            cell.value = f'=SUMPRODUCT({tp_cond}*{TP}!${tp_col}$2:${tp_col}$5000)'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell = ws[f"I{row}"]
        total_cell.value = f"=SUM(D{row}:H{row})"
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_table_borders(ws, 55, 56, "CDEFGHI")

    # ==========================================================================
    # PRACOVNÍ DEN - REÁLNÝ ČAS (odhad z SalesApp Started at/Finished at) -
    # product owner, 2026-07-11, follow-up: "také tam nevidím ten čas".
    # Skutečná pracovní doba (první start - poslední konec toho dne) a
    # nevytížený čas (rozdíl mezi pracovní dobou a součtem trvání návštěv -
    # viz PerformanceEngine.ts recordDayTiming()/workSpanHoursByDay -
    # zahrnuje kampaňové i ostatní návštěvy). Informační zobrazení, zatím
    # bez semaforu/triggeru - uživatel žádal jen "zobrazení", ne novou
    # KPI/flag; pokud se prokáže užitečné, práh lze doplnit později stejně
    # jako u km efektivity trasy.
    # FILTER()+INDEX() (ne SUMPRODUCT) - workSpanHours/idleHours obsahují ""
    # pro dny bez záznamu Started/Finished at, což by v SUMPRODUCT
    # (aritmetika na textu) shodilo #VALUE!; FILTER/INDEX na text nesahá.
    # ==========================================================================
    build_section_header(ws, "C58", "PRACOVNÍ DEN - REÁLNÝ ČAS (SalesApp Started/Finished at)")
    style_dashboard_table_header(ws, 59, "CDEFGHI", ["", "Po", "Út", "St", "Čt", "Pá", "Týden celkem"])
    time_rows = [
        (60, "Pracovní doba (h)", ["AX", "AY", "AZ", "BA", "BB"]),
        (61, "Nevytížený čas (h)", ["BC", "BD", "BE", "BF", "BG"]),
    ]
    for row, label, cols in time_rows:
        ws.cell(row, 3, label).font = Font(size=10)
        for col_letter, tp_col in zip("DEFGH", cols):
            cell = ws[f"{col_letter}{row}"]
            cell.value = (
                f'=LET(v,IFERROR(INDEX(FILTER({TP}!${tp_col}$2:${tp_col}$5000,{tp_cond}),1),""),'
                f'IF(v="","-",v))'
            )
            cell.number_format = '0.0" h";;;@'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Week-total as its own KPI tile (product owner, 2026-07-11: "dej to
        # do kpi") - same visual treatment as the route-km week total (H44)
        # so both weekly-hours numbers a manager scans for read the same way
        # at a glance. SUM() (not SUMPRODUCT) tolerates the "-" text placeholder
        # in D:H for no-data days - it's simply skipped, no #VALUE! risk.
        total_cell = ws[f"I{row}"]
        total_cell.value = f"=SUM(D{row}:H{row})"
        total_cell.number_format = '0.0" h"'
        total_cell.font = font_card_value(size=14, color=NAVY)
        total_cell.fill = PatternFill("solid", fgColor=WHITE)
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_table_borders(ws, 60, 61, "CDEFGHI")
    ws.conditional_formatting.add(
        "D61:H61",
        FormulaRule(formula=['AND(ISNUMBER(D61),D61>=3)'], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_WARNING),
                    font=Font(color=STATUS_WARNING)),
    )

    # ==========================================================================
    # TOP PROBLÉMOVÉ POS - deduped, engine-computed (see PerformanceEngine.ts).
    # ==========================================================================
    build_section_header(ws, "C64", "TOP PROBLÉMOVÉ POS")
    style_dashboard_table_header(ws, 65, "CDEF", ["POS", "Název", "Region", "Nesplněno (celkem)"])
    ws["C66"] = f'=IFERROR(FILTER({TI}!$C$2:$F$1000,{TI}!$A$2:$A$1000=$D$5),{{"—","Žádné problémy 🎉","",0}})'
    apply_table_borders(ws, 66, 70, "CDEF")
    ws.conditional_formatting.add(
        "F66:F70", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color=STATUS_CRITICAL),
    )

    # ==========================================================================
    # POS BEZ NÁVŠTĚVY (nikdy) - product owner, 2026-07-11: "rad bych i
    # někde viděl jaké POS v kampani vůbec nejel" - distinct from "TOP
    # PROBLÉMOVÉ POS" above (which ranks POS by how many times a planned
    # visit was missed - repeated misses on a POS that HAS been visited
    # before). This is the POS that were never visited even once.
    # NOT based on POS_MASTER's lastRealVisitDate - checked against real
    # data first and that field is unusable for this: ImportEngine.ts seeds
    # it to "today" for every brand-new POS ("product owner confirmed that
    # installation counts as the first visit"), so on the real workbook
    # every single Active row already has a non-blank lastRealVisitDate -
    # that signal would always report zero results. The uncontaminated
    # "never visited" signal is a POS's total absence from
    # VISIT_HISTORY_ACTUAL, which only ever gets a row when
    # ComplianceEngine.ts processes a real SalesApp visit - untouched by
    # ImportEngine.ts's install-day default.
    # Technician match mirrors PerformanceEngine.ts's posTechnician lookup
    # (managerOverrideTechnician wins over assignedTechnician when set).
    # Sorted by businessScore descending - the most valuable un-visited POS
    # first, since that's the actionable gap a manager would want to see.
    # ==========================================================================
    build_section_header(ws, "C72", "POS BEZ NÁVŠTĚVY (nikdy)")
    pm_tech_match = 'IF(POS_MASTER!$AJ$2:$AJ$20000<>"",POS_MASTER!$AJ$2:$AJ$20000,POS_MASTER!$O$2:$O$20000)=$D$5'
    pm_active = 'POS_MASTER!$Q$2:$Q$20000="Active"'
    pm_never_visited = 'ISNA(MATCH(POS_MASTER!$A$2:$A$20000,VISIT_HISTORY_ACTUAL!$A$2:$A$200000,0))'
    pm_never_visited_cond = f'({pm_tech_match})*({pm_active})*({pm_never_visited})'
    never_visited_count_ref = build_kpi_card(
        ws, "C", "D", 73, 74, 74, "Počet POS bez návštěvy",
        f'=SUMPRODUCT({pm_never_visited_cond})',
        value_color=STATUS_WARNING, fill_color=dashboard_ui.TINT_WARNING,
    )
    ws[never_visited_count_ref].font = font_card_value(size=16, color=STATUS_WARNING)
    ws.merge_cells("E73:N74")
    ws["E73"] = "Aktivní POS přiřazené tomuto technikovi, které dosud nemají v SalesApp ani jednu zaznamenanou reálnou návštěvu."
    ws["E73"].font = Font(size=9, italic=True, color="595959")
    ws["E73"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    style_dashboard_table_header(ws, 76, "CDEFG", ["POS", "Název", "Město", "Region", "Business Score"])
    ws["C77"] = (
        f'=IFERROR(SORT(FILTER(CHOOSE({{1,2,3,4,5}},POS_MASTER!$A$2:$A$20000,POS_MASTER!$G$2:$G$20000,'
        f'POS_MASTER!$L$2:$L$20000,POS_MASTER!$H$2:$H$20000,POS_MASTER!$AC$2:$AC$20000),'
        f'{pm_never_visited_cond}),5,-1),"— Žádné, vše bylo aspoň jednou navštíveno 🎉")'
    )
    ws["C77"].font = Font(size=10)
    ws.conditional_formatting.add(
        "C77", FormulaRule(formula=['ISNUMBER(SEARCH("Žádné",C77))'], font=Font(italic=True, color=STATUS_GOOD)),
    )

    return ws


def build_performance_sheet(wb, n_rows=60):
    """PERFORMANCE: all technicians compared side by side. Built as a real,
    native Excel Table (ListObject) with AutoFilter over
    TECHNICIAN_PERFORMANCE_SUMMARY (PerformanceEngine.ts's per-technician
    snapshot) - sorting/filtering is Excel's own built-in Table behavior,
    not a custom UI (product owner, 2026-07-05: prefer native Excel Table +
    AutoFilter over a hand-rolled filter panel for a comparison grid).

    Same live-formula-view pattern as TECHNICIAN_PLAN (see that function's
    docstring): each table row is a plain, non-array formula pulling row r-8
    from TECHNICIAN_PERFORMANCE_SUMMARY's row r - a fixed n_rows cap well
    above any realistic technician count. Deliberately NOT a FILTER()/
    dynamic-array spill: Excel does not allow a spilling array formula
    inside a Table's range, so a native, sortable Table needs per-row
    formulas instead - this is why PerformanceEngine.ts pre-computes one row
    per technician (TECHNICIAN_PERFORMANCE_SUMMARY) rather than this sheet
    trying to derive that itself."""
    TS = "TECHNICIAN_PERFORMANCE_SUMMARY"
    if "PERFORMANCE" in wb.sheetnames:
        del wb["PERFORMANCE"]
    ws = wb.create_sheet("PERFORMANCE")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    for col in "CDEFGHIJKLMNOPQR":
        ws.column_dimensions[col].width = 13
    build_nav_rail(ws, "PERFORMANCE")

    build_dashboard_banner(
        ws, "PERFORMANCE", "Srovnání všech techniků - řaď a filtruj přímo v tabulce (Excel AutoFilter)",
        col_start="C", col_end="R",
    )
    ws.freeze_panes = "C9"

    # ---- Team KPI summary row ----
    build_section_header(ws, "C5", "TÝM CELKEM")
    cards = [
        ("C", "E", "Průměrná compliance (tým)",
         f'=IFERROR(ROUND(AVERAGE({TS}!$K$2:$K$5000),1)&"%","-")', NAVY, WHITE),
        ("F", "H", "Nejlepší technik",
         f'=IFERROR(INDEX({TS}!$A$2:$A$5000,MATCH(MAX({TS}!$K$2:$K$5000),{TS}!$K$2:$K$5000,0))'
         f'&" ("&TEXT(MAX({TS}!$K$2:$K$5000),"0.0")&"%)","-")', STATUS_GOOD, dashboard_ui.TINT_GOOD),
        ("I", "M", "Nejslabší technik",
         f'=IFERROR(INDEX({TS}!$A$2:$A$5000,MATCH(MIN({TS}!$K$2:$K$5000),{TS}!$K$2:$K$5000,0))'
         f'&" ("&TEXT(MIN({TS}!$K$2:$K$5000),"0.0")&"%)","-")', STATUS_CRITICAL, dashboard_ui.TINT_CRITICAL),
    ]
    build_kpi_card_row(ws, cards, label_row=6, value_row_start=7, value_row_end=7)
    for c1, c2, label, formula, color, fill in cards:
        ws[f"{c1}7"].font = font_card_value(size=13, color=color)
    ws.row_dimensions[7].height = 22

    # ==========================================================================
    # COMPARISON TABLE - native Excel Table/AutoFilter, row-anchored view of
    # TECHNICIAN_PERFORMANCE_SUMMARY (see docstring above).
    # ==========================================================================
    build_section_header(ws, "C9", "SROVNÁNÍ TECHNIKŮ")
    header_row = 10
    headers = [
        "Technik", "Region", "Naplánováno", "Realizováno", "Splněno včas", "Splněno pozdě",
        "Nesplněno", "Navíc", "Compliance %", "Dlouhodobý průměr", "Trend", "Flaká riziko",
        "Km/den (nejhorší)", "POS v kampani", "Hotovo", "Chybí",
    ]
    for col, label in zip("CDEFGHIJKLMNOPQR", headers):
        ws[f"{col}{header_row}"] = label

    # POS v kampani / Hotovo / Chybí (product owner, 2026-07-11: "dashboard
    # kde uvidím každého technika a kolik POS z kampaně už má hotovo a kolik
    # mu chybí") - not from TECHNICIAN_PERFORMANCE_SUMMARY (no engine field
    # for this), computed directly from POS_MASTER + VISIT_HISTORY_ACTUAL,
    # same definition as TECHNICIAN_SCORECARD's "POS BEZ NÁVŠTĚVY" table:
    # "Hotovo" = active assigned POS with at least one row ever in
    # VISIT_HISTORY_ACTUAL (a real campaign-purpose SalesApp visit).
    # The "is this POS ever visited" check (MATCH against 200,000 rows) is
    # computed ONCE as a hidden spilled helper column (T) instead of inside
    # each of the up to n_rows per-technician SUMPRODUCTs - re-running that
    # MATCH per technician row would multiply an already 11,605-row scan by
    # n_rows, needlessly expensive for a Table that recalculates on every
    # sort/filter.
    ws["T1"] = "posEverVisitedFlag"
    ws["T2"] = "=ISNUMBER(MATCH(POS_MASTER!$A$2:$A$20000,VISIT_HISTORY_ACTUAL!$A$2:$A$200000,0))"
    ws.column_dimensions["T"].hidden = True

    src_cols = "ABEFGHIJKLMOP"  # TECHNICIAN_PERFORMANCE_SUMMARY column per table column, in order
    first_data_row = header_row + 1
    for i in range(n_rows):
        r = first_data_row + i
        sr = i + 2  # TECHNICIAN_PERFORMANCE_SUMMARY row (header is row 1 there)
        for col, src_col in zip("CDEFGHIJKLMNO", src_cols):
            ws[f"{col}{r}"] = f'=IF({TS}!$A${sr}="","",{TS}!{src_col}{sr})'
        pm_tech_match_row = (
            f'(IF(POS_MASTER!$AJ$2:$AJ$20000<>"",POS_MASTER!$AJ$2:$AJ$20000,POS_MASTER!$O$2:$O$20000)=C{r})'
        )
        pm_active_row = 'POS_MASTER!$Q$2:$Q$20000="Active"'
        ws[f"P{r}"] = f'=IF({TS}!$A${sr}="","",SUMPRODUCT({pm_tech_match_row}*({pm_active_row})))'
        ws[f"Q{r}"] = f'=IF({TS}!$A${sr}="","",SUMPRODUCT({pm_tech_match_row}*({pm_active_row})*$T$2:$T$20001))'
        ws[f"R{r}"] = f'=IF({TS}!$A${sr}="","",P{r}-Q{r})'

    last_row = first_data_row + n_rows - 1
    table_ref = f"C{header_row}:R{last_row}"
    table = Table(displayName="PerformanceTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # Compliance % severity coloring - same 4-tier palette as every other
    # compliance figure in this workbook (STATUS_GOOD/WARNING/SERIOUS/CRITICAL).
    apply_severity_conditional_formatting(
        ws, f"K{first_data_row}:K{last_row}", f"K{first_data_row}",
        thresholds=[(90, STATUS_GOOD), (70, STATUS_WARNING), (50, STATUS_SERIOUS)],
        below_color=STATUS_CRITICAL, font_size=11, font_color=WHITE,
    )
    # Trend column: native Excel icon set (▲/flat/▼), not a custom arrow -
    # a numeric column sorts correctly in the Table, the icon set alone
    # carries direction. cfvo thresholds MUST be non-decreasing (OOXML
    # requirement) - found broken as [0, -0.001, 0.001] during a post-build
    # QA pass (2026-07-06), which is not ascending and risks Excel flagging
    # the file for repair on open. [0, 0, 0.0001] is valid and keeps the
    # same intent: negative delta -> down arrow, delta in [0, 0.0001) ->
    # flat, delta >= 0.0001 -> up arrow.
    ws.conditional_formatting.add(
        f"M{first_data_row}:M{last_row}",
        IconSetRule(icon_style="3Arrows", type="num", values=[0, 0, 0.0001], showValue=True, reverse=False),
    )

    # Flaká riziko badge (product owner, 2026-07-06: "chci aby mi to ukazalo
    # ktery z nich flaka a ktery ne" - see PerformanceEngine.ts's "EIGHTH
    # OUTPUT ADDITION" header comment for the underlying rule: 2+ of the last
    # 4 tracked weeks below 70% compliance, a repeated pattern, not one bad
    # week). "Ano" in red draws the eye when scanning the whole team; "Ne" is
    # left neutral (not green) - not-flagged is the expected default, not an
    # achievement to highlight.
    ws.conditional_formatting.add(
        f"N{first_data_row}:N{last_row}",
        FormulaRule(formula=[f'N{first_data_row}="Ano"'], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_CRITICAL),
                    font=Font(bold=True, color=STATUS_CRITICAL)),
    )

    # Km/den (nejhorší) semaphore - same CONTROL.ROUTE_KM_WARNING_KM/
    # CRITICAL_KM thresholds already used on TECHNICIAN_SCORECARD's route-
    # efficiency table, so the two screens agree on what counts as a bad
    # day. Found missing during a final full test pass (2026-07-06): route
    # efficiency had no network-wide view at all before this - only
    # per-technician on TECHNICIAN_SCORECARD.
    # Moved to column U (product owner, 2026-07-11): Q is now a visible
    # table column ("Hotovo") - these two threshold cells used to sit hidden
    # in Q1/Q2, which would otherwise now surface as stray numbers above the
    # visible "Hotovo" column.
    ws["U1"] = '=IFERROR(VLOOKUP("ROUTE_KM_WARNING_KM",CONTROL!$A:$B,2,FALSE),80)'
    ws["U2"] = '=IFERROR(VLOOKUP("ROUTE_KM_CRITICAL_KM",CONTROL!$A:$B,2,FALSE),150)'
    ws.column_dimensions["U"].hidden = True
    o_rng = f"O{first_data_row}:O{last_row}"
    ws.conditional_formatting.add(
        o_rng, FormulaRule(formula=[f"O{first_data_row}<=$U$1"], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_GOOD),
                            font=Font(color=STATUS_GOOD), stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        o_rng, FormulaRule(formula=[f"O{first_data_row}<=$U$2"], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_WARNING),
                            font=Font(color=STATUS_WARNING), stopIfTrue=True),
    )
    ws.conditional_formatting.add(
        o_rng, FormulaRule(formula=[f"O{first_data_row}>$U$2"], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_CRITICAL),
                            font=Font(bold=True, color=STATUS_CRITICAL), stopIfTrue=True),
    )

    # "Chybí" (POS never visited in the campaign) - a data bar so the worst
    # gaps are visible at a glance while scanning the whole team, same
    # component TECHNICIAN_SCORECARD's TOP PROBLÉMOVÉ POS already uses for
    # the analogous "how many is this" signal. 0 stays neutral (nothing to
    # flag) - only non-zero rows get a bar.
    r_rng = f"R{first_data_row}:R{last_row}"
    ws.conditional_formatting.add(
        r_rng, DataBarRule(start_type="num", start_value=0, end_type="max", color=STATUS_WARNING),
    )
    ws.conditional_formatting.add(
        r_rng, FormulaRule(formula=[f"R{first_data_row}=0"], font=Font(color=STATUS_GOOD)),
    )

    return ws


def build_efficiency_sheet(wb, n_rows=60):
    """EFFICIENCY: "Monitoring efektivity" (product owner, 2026-07-09,
    speaking as vedoucí Field Force týmu: "chci se na dashboard jen podívat
    a hned vidět, kdo ze mě dělá blbce") - technicians ranked automatically,
    most corroborated problem signals first, so a manager never has to sort
    anything themselves to find the outlier.

    Sorted by activeSignalCount (product owner, 2026-07-09: "GPS je odhad,
    takže to ani nemusí být na vinu" - a lone route-efficiency flag should
    never rank someone at the top; a technician with several independent
    signals corroborating each other is the real outlier), NOT by route
    efficiency alone - see combinedRiskFlag's rationale in
    PerformanceEngine.ts and docs/BUSINESS_RULES.md.

    Built as a single LET()+SORTBY()+FILTER() dynamic-array spill over
    TECHNICIAN_PERFORMANCE_SUMMARY (PerformanceEngine.ts's per-technician
    snapshot) - deliberately NOT a native Table like PERFORMANCE (that
    screen's own docstring explains why: Excel does not allow a spilling
    array formula inside a Table's range). This screen is meant to always
    already be sorted worst-first, not something a manager filters/re-sorts
    themselves - a spill is the right tool here, not a workaround.
    """
    TS = "TECHNICIAN_PERFORMANCE_SUMMARY"
    if "EFFICIENCY" in wb.sheetnames:
        del wb["EFFICIENCY"]
    ws = wb.create_sheet("EFFICIENCY")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    for col in "CDEFGHIJKLM":
        ws.column_dimensions[col].width = 14
    build_nav_rail(ws, "EFFICIENCY")

    build_dashboard_banner(
        ws, "MONITORING EFEKTIVITY",
        "Kdo jezdí cik-cak a kdo pracuje - seřazeno automaticky, nejhorší nahoře",
        col_start="C", col_end="M",
    )
    ws.freeze_panes = "C11"

    # ---- Team KPI summary row ----
    build_section_header(ws, "C5", "TÝM CELKEM")
    cards = [
        ("C", "D", "Průměrný poměr nájezd/optimum",
         f'=IFERROR(ROUND(AVERAGE({TS}!$Q$2:$Q$5000),0)&"%","-")', NAVY, WHITE),
        ("E", "F", "Nejvíc signálů najednou",
         f'=IFERROR(INDEX({TS}!$A$2:$A$5000,MATCH(MAX({TS}!$AB$2:$AB$5000),{TS}!$AB$2:$AB$5000,0))'
         f'&" ("&MAX({TS}!$AB$2:$AB$5000)&")","-")', STATUS_CRITICAL, dashboard_ui.TINT_CRITICAL),
        ("G", "H", "Technik(ů) - problémový (Ano)",
         f'=COUNTIF({TS}!$AC$2:$AC$5000,"Ano")', STATUS_CRITICAL, dashboard_ui.TINT_CRITICAL),
        ("I", "J", "Technik(ů) v KRITICKÉ trase",
         f'=COUNTIF({TS}!$T$2:$T$5000,"KRITICKÉ")', STATUS_WARNING, dashboard_ui.TINT_WARNING),
    ]
    build_kpi_card_row(ws, cards, label_row=6, value_row_start=7, value_row_end=7)
    for c1, c2, label, formula, color, fill in cards:
        ws[f"{c1}7"].font = font_card_value(size=13, color=color)
    ws.row_dimensions[7].height = 22

    # ---- Auto-surfaced "problémový technik" callout (product owner: "chci,
    # aby mi to systém sám vystrčil jako problémového technika"). Keyed off
    # combinedRiskFlag (>= PROBLEM_SIGNAL_MIN_COUNT corroborating signals),
    # NOT off route efficiency alone - "GPS je odhad, takže to ani nemusí
    # být na vinu" (product owner, 2026-07-09). ----
    ws.cell(9, 3, "PROBLÉMOVÍ TECHNICI").font = SECTION_FONT
    r = 10
    ws.merge_cells(f"C{r}:M{r+1}")
    flag_cell = ws.cell(r, 3)
    flag_cell.value = (
        f'=IFERROR(IF(COUNTIF({TS}!$AC:$AC,"Ano")=0,'
        f'"✅ Žádný technik nemá aktuálně víc než jeden varovný signál najednou",'
        f'"🚩 "&COUNTIF({TS}!$AC:$AC,"Ano")&" technik(ů) s více souběžnými signály: "&'
        f'TEXTJOIN(", ",TRUE,FILTER({TS}!$A:$A,{TS}!$AC:$AC="Ano"))),'
        f'"Zatím žádná data")'
    )
    flag_cell.font = Font(bold=True, size=13, color=NAVY)
    flag_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    flag_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 20
    build_status_badge_conditional(ws, flag_cell.coordinate, flag_cell.coordinate, rules=[
        ("✅", "E2EFDA", None),
        ("🚩", STATUS_SERIOUS, None),
    ])

    # ==========================================================================
    # HEATMAP - one spill, sorted worst-first by activeSignalCount (how many
    # independent signals corroborate each other), not by any single metric.
    # ==========================================================================
    header_row = 13
    build_section_header(ws, f"C{header_row - 1}", "TECHNICI SEŘAZENI OD NEJVÍCE PROBLÉMOVÝCH")
    headers = [
        "Technik", "Region", "Signálů", "Problémový?",
        "Návštěvnost vs kolegové", "Hodnota/návštěva vs kolegové", "Délka návštěvy vs kolegové",
        "Trasa vs optimum", "Km/den (nejhorší)", "Compliance %", "Trasa - stav",
    ]
    for col, label in zip("CDEFGHIJKLM", headers):
        ws[f"{col}{header_row}"] = label
        ws[f"{col}{header_row}"].font = HEADER_FONT
        ws[f"{col}{header_row}"].fill = PatternFill("solid", fgColor=NAVY)
        ws[f"{col}{header_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 30

    first_data_row = header_row + 1
    last_row = first_data_row + n_rows - 1
    last_src_row = 1 + n_rows * 4
    spill_formula = (
        "=IFERROR(LET("
        f"src,CHOOSE({{1,2,3,4,5,6,7,8,9,10,11}},"
        f"{TS}!$A$2:$A${last_src_row},{TS}!$B$2:$B${last_src_row},"
        f"{TS}!$AB$2:$AB${last_src_row},{TS}!$AC$2:$AC${last_src_row},"
        f"{TS}!$V$2:$V${last_src_row},{TS}!$X$2:$X${last_src_row},"
        f"{TS}!$Z$2:$Z${last_src_row},{TS}!$S$2:$S${last_src_row},"
        f"{TS}!$P$2:$P${last_src_row},{TS}!$K$2:$K${last_src_row},"
        f"{TS}!$T$2:$T${last_src_row}),"
        f"filtered,FILTER(src,{TS}!$A$2:$A${last_src_row}<>\"\"),"
        "SORTBY(filtered,INDEX(filtered,0,3),-1,INDEX(filtered,0,8),-1)"
        '),"Zatím žádná data - spusť Performance Engine")'
    )
    ws[f"C{first_data_row}"] = spill_formula
    for col in "CDEFGHIJKLM":
        for r2 in range(first_data_row, last_row + 1):
            ws[f"{col}{r2}"].border = CARD_BORDER
            ws[f"{col}{r2}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"C{first_data_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Heatmap color scales. Volume/PPT-density/duration (F,G,H here) are all
    # "% of peer average" where LOW is bad (opposite direction from route
    # efficiency) - red at/below CRITICAL (50%), green at/above 100%. Route
    # efficiency (I) keeps its own HIGH-is-bad direction.
    for col in ("F", "G", "H"):
        rng = f"{col}{first_data_row}:{col}{last_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="num", start_value=50, start_color="F8696B",
                mid_type="num", mid_value=75, mid_color="FFEB84",
                end_type="num", end_value=100, end_color="63BE7B",
            ),
        )
    ws.conditional_formatting.add(
        f"I{first_data_row}:I{last_row}",
        ColorScaleRule(
            start_type="num", start_value=100, start_color="63BE7B",
            mid_type="num", mid_value=125, mid_color="FFEB84",
            end_type="num", end_value=150, end_color="F8696B",
        ),
    )
    # "Problémový?" (col F->D here) badge.
    ws.conditional_formatting.add(
        f"D{first_data_row}:D{last_row}",
        FormulaRule(formula=[f'D{first_data_row}="Ano"'], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_CRITICAL),
                    font=Font(bold=True, color=STATUS_CRITICAL)),
    )
    # Trasa - stav (col M) - same KRITICKÉ/POZOR/OK convention as
    # PerformanceEngine.ts's efficiencyFlag - informational only here, does
    # NOT drive the row highlight (that's combinedRiskFlag's job now).
    ws.conditional_formatting.add(
        f"M{first_data_row}:M{last_row}",
        FormulaRule(formula=[f'M{first_data_row}="KRITICKÉ"'], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_CRITICAL),
                    font=Font(bold=True, color=STATUS_CRITICAL)),
    )
    ws.conditional_formatting.add(
        f"M{first_data_row}:M{last_row}",
        FormulaRule(formula=[f'M{first_data_row}="POZOR"'], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_WARNING),
                    font=Font(bold=True, color=STATUS_WARNING)),
    )
    ws.conditional_formatting.add(
        f"M{first_data_row}:M{last_row}",
        FormulaRule(formula=[f'M{first_data_row}="OK"'], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_GOOD),
                    font=Font(color=STATUS_GOOD)),
    )
    # Whole-row highlight for "Problémový? = Ano" technicians (combinedRiskFlag,
    # column D here) - the row jumps out even without reading any single
    # column (product owner: "chci, aby to začalo svítit červeně"). This is
    # the >= PROBLEM_SIGNAL_MIN_COUNT corroborated signal, not a lone metric.
    for col in "CDEFGHIJKLM":
        ws.conditional_formatting.add(
            f"{col}{first_data_row}:{col}{last_row}",
            FormulaRule(formula=[f'$D{first_data_row}="Ano"'], fill=PatternFill("solid", fgColor=dashboard_ui.TINT_CRITICAL)),
        )

    ws.row_dimensions[first_data_row].height = 20

    return ws


def build_manual_sheet(wb):
    """MANUAL: interpretation guide for the efficiency metrics (product
    owner, 2026-07-09, speaking as vedoucí Field Force týmu: "napiš, jak mám
    ty metriky interpretovat a co je už za hranou, kdy je potřeba jít a
    technika konfrontovat"). Static guidance text, not live formulas - this
    sheet explains HOW to read EFFICIENCY/PERFORMANCE, it does not duplicate
    their numbers."""
    if "MANUAL" in wb.sheetnames:
        del wb["MANUAL"]
    ws = wb.create_sheet("MANUAL")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 100
    build_dashboard_banner(
        ws, "MANUÁL - MONITORING EFEKTIVITY",
        "Jak číst metriky a kdy jít technika konfrontovat (pohled vedoucího Field Force týmu)",
        col_start="C", col_end="C", title_size=20,
    )
    ws.freeze_panes = "C4"

    sections = [
        ("1. CO TY DVĚ ČÍSLA ZNAMENAJÍ", [
            "Poměr nájezd/optimum (efficiencyRatioPercent): kolik % nad matematickým minimem "
            "technik reálně najezdil. Minimum počítá systém sám pro KAŽDÝ den zvlášť - je to "
            "nejkratší možná trasa mezi body, které ten den skutečně navštívil (ne odhad, "
            "přesný výpočet). 100 % = jel prakticky ideálně. 150 % = najezdil o polovinu víc, "
            "než musel - to je přesně ta hranice, kterou jsi zadal jako kritickou.",
            "",
            "Km na návštěvu (kmPerVisit): celkové km za týden děleno počtem realizovaných "
            "návštěv. Na rozdíl od poměru nájezd/optimum nebere v úvahu, jak blízko u sebe POS "
            "reálně jsou - proto ho čti VŽDY ve srovnání s kolegy na EFFICIENCY, ne jako "
            "izolované číslo. Technik ve městě bude mít nižší km/návštěvu než technik na "
            "venkově, i když jezdí stejně efektivně - to není chyba dat, to je geografie.",
        ]),
        ("2. JAK ČÍST BAREVNÉ PÁSMO (trasa)", [
            "🟢 OK (do 125 %): normální provoz. Sem spadá naprostá většina týdnů - to je "
            "očekávaný stav, ne úspěch, který je třeba oceňovat.",
            "🟡 POZOR (125-149 %): trasa je znatelně horší než ideál. Jeden týden v POZOR "
            "ještě nic neznamená (viz bod 4 - co NENÍ signál) - sleduj, jestli se opakuje.",
            "🔴 KRITICKÉ (150 % a víc): o polovinu nebo víc nad optimem. To je konkrétní "
            "zadání, kdy má systém začít svítit červeně.",
        ]),
        ("3. TŘI DALŠÍ TRIGGERY - a proč je NUTNÉ je číst spolu s trasou, ne samostatně", [
            "Návštěvnost vs kolegové (volumeFlag): kolik návštěv technik reálně zvládl ve "
            "srovnání s ostatními stejný týden - a zvlášť i se svým vlastním dlouhodobým "
            "průměrem (obojí najednou, systém bere tu horší z obou hodnot). Tohle chytí "
            "situaci, kterou Compliance % NEVIDÍ: technik má naplánováno málo, splní to na "
            "100 %, ale objel dvakrát méně poboček než kolega - na Compliance % to vypadá "
            "bezvadně, na Návštěvnosti ne.",
            "",
            "Hodnota/návštěva vs kolegové (pptDensityFlag): kolik obchodní hodnoty (PPT) "
            "technik v průměru přinese na jednu návštěvu, oproti kolegům. Přesně tvůj postřeh "
            "\"hodně návštěv, ale jednoúčelové\" - technik může mít skvělou trasu (nízké "
            "nájezd/optimum) a přitom objíždět jen bezcenné poblíž ležící POS. Trasa a "
            "hodnota jsou DVĚ RŮZNÉ VĚCI a tenhle sloupec je jediný, který hlídá tu druhou.",
            "",
            "Délka návštěvy vs kolegové (durationFlag): průměrná reálná délka návštěvy "
            "(\"Real duration (h)\" ze SalesApp) oproti kolegům. Na rozdíl od trasy (GPS "
            "odhad) je tohle PŘÍMO NAMĚŘENÝ údaj - technik, který je systematicky výrazně "
            "rychlejší než všichni ostatní, návštěvy pravděpodobně odbývá. Dostupné až po "
            "prvním importu, který obsahuje sloupec Real duration (h).",
        ]),
        ("4. KOMBINOVANÝ SIGNÁL (\"Problémový?\" = Ano/Ne) - tohle je to hlavní, na co se dívat", [
            "Žádný JEDNOTLIVÝ signál sám o sobě nikdy nespustí \"Problémový? = Ano\" - ani "
            "KRITICKÁ trasa, ani nízká návštěvnost, ani nic jiného samo o sobě. Musí se "
            "potkat aspoň 2 signály najednou (nastavitelné v CONTROL jako "
            "PROBLEM_SIGNAL_MIN_COUNT) - protože \"GPS je odhad, takže to ani nemusí být "
            "na vinu\" - a stejně tak jeden jediný slabý týden v jakékoli metrice.",
            "Sloupec \"Signálů\" na EFFICIENCY ukazuje PŘESNĚ kolik z pěti sledovaných "
            "signálů (compliance/flaká riziko, návštěvnost, hodnota/návštěva, délka "
            "návštěvy, trasa) je u daného technika aktuálně POZOR/KRITICKÉ zároveň - a "
            "žebříček je podle tohoto čísla seřazený, ne podle trasy.",
            "Tohle je přesně ten filtr, který odděluje \"má smůlu s GPS/terénem\" od "
            "\"má reálný problém\" - jeden signál je šum, dva a víc najednou už je vzorec.",
        ]),
        ("5. CO JE SIGNÁL A CO JEŠTĚ NENÍ (než začneš řešit)", [
            "NENÍ signál: jeden KRITICKÝ den v týdnu, kdy měl technik jen 2-3 návštěvy daleko "
            "od sebe (např. servisní výjezd mimo běžnou trasu) - u málo bodů dokáže jedna "
            "vynucená zajížďka rozhodit poměr, aniž by šlo o špatnou práci.",
            "NENÍ signál: jednorázový týden POZOR/KRITICKÉ v jedné metrice bez opakování - "
            "všechny čtyři sledované metriky (trasa, návštěvnost, hodnota/návštěva, délka) "
            "počítají svůj \"long-run\" flag z průměru za posledních několik sledovaných "
            "týdnů (stejné okno jako \"flaká riziko\"), takže jednu špatnou zajížďku nebo "
            "jeden náročný týden systém sám vyhladí.",
            "NENÍ signál: POS bez GPS souřadnic v datech - systém takové zastávky do výpočtu "
            "trasy vůbec nezahrnuje (nehádá vzdálenost), takže chybějící GPS data snižují "
            "přesnost, ne technikovo skóre.",
            "JE signál: \"Problémový? = Ano\" na EFFICIENCY/TECHNICIAN_PERFORMANCE_SUMMARY "
            "(sloupec combinedRiskFlag) - to je systém, který už sám zkombinoval aspoň dva "
            "nezávislé signály za tebe.",
        ]),
        ("6. KDY JÍT A KONFRONTOVAT TECHNIKA - konkrétní hranice", [
            "1 signál (jakýkoli, včetně KRITICKÉ trasy): nic nedělej, jen sleduj. Otevři si "
            "daný den na TECHNICIAN_PERFORMANCE_LOG (sloupce posListMon..posListFri) a "
            "mrkni, jestli tam není zjevný důvod (velký region, vynucená zajížďka, porucha "
            "vozu apod.).",
            "\"Problémový? = Ano\" (2+ signály najednou): posaď se s technikem. V tuhle "
            "chvíli už to není náhoda ani jedna metrika, která může být zkreslená - dva "
            "nezávislé zdroje dat (GPS trasa, počet návštěv, hodnota, délka návštěvy) "
            "ukazují stejným směrem zároveň.",
            "\"Problémový? = Ano\" opakovaně kampaň za kampaní / vysoký počet Signálů (4-5 "
            "najednou): to je přesně ten technik, kterého má HOME/EFFICIENCY \"sám "
            "vystrčit\" úplně nahoru žebříčku - tady už jde o celkový výkon, ne o jednu "
            "věc, kterou lze vysvětlit.",
        ]),
        ("7. POSTUP PŘED SCHŮZKOU S TECHNIKEM", [
            "1) EFFICIENCY - najdi technika v žebříčku (řazeno podle počtu Signálů, "
            "nejproblémovější nahoře), podívej se, KTERÉ konkrétní signály má aktivní.",
            "2) TECHNICIAN_PERFORMANCE_LOG - najdi jeho řádky za poslední týdny, otevři "
            "sloupce posListMon..posListFri pro dny s vysokým km - uvidíš PŘESNĚ které POS "
            "ten den navštívil a v jakém pořadí. Pokud je aktivní i signál hodnoty/návštěvy, "
            "podívej se, jaké PPT mají POS, které navštěvuje nejčastěji.",
            "3) Porovnej s plánovaným pořadím (TECHNICIAN_PLAN / MANAGER_PLAN_PUBLISHED pro "
            "stejné datum) - liší se realita od plánu? Pokud ano, zeptej se proč (mohl mít "
            "dobrý důvod, který systém nezná - objednávka od klienta, uzavírka silnice...).",
            "4) Teprve pak jdi na schůzku - s konkrétními dny, čísly a KONKRÉTNÍMI signály, "
            "ne s obecným pocitem \"něco mi tu nesedí\".",
        ]),
        ("8. OMEZENÍ, KTERÁ MUSÍŠ ZNÁT", [
            "\"Reálná trasa\" je odhad, ne GPS trasování v reálném čase: systém neví, v jakém "
            "pořadí technik body doopravdy objel - předpokládá, že to bylo v pořadí, v jakém "
            "byly naplánované (MANAGER_PLAN_PUBLISHED). Pokud technik pořadí sám změnil (a "
            "měl k tomu dobrý důvod), číslo to nepozná a může ho neprávem penalizovat.",
            "Vzdálenost je vzdušná čára s korekcí (ne silniční síť) - v členitém terénu nebo "
            "městě s řekou/dálnicí uprostřed může být reálná silniční vzdálenost delší než "
            "vzdušná čára, u všech techniků stejně, takže srovnání mezi nimi zůstává férové, "
            "ale absolutní % může být u všech mírně nadhodnocené.",
            "Bez reálných dat z SalesApp systém nic neukáže: dokud neproběhne první ostrý "
            "import, TECHNICIAN_PERFORMANCE_LOG/SUMMARY i EFFICIENCY zůstanou prázdné - to "
            "je očekávaný stav, ne chyba.",
            "Signál délky návštěvy (durationFlag) potřebuje sloupec \"Real duration (h)\" v "
            "SalesApp exportu - pokud ho konkrétní export neobsahuje, sloupec zůstane "
            "prázdný a signál se prostě nepočítá (nepočítá se jako 0 ani jako chyba).",
            "Návštěvnost a hodnota/návštěva se porovnávají PROTI OSTATNÍM TECHNIKŮM TEN "
            "SAMÝ TÝDEN (prostý průměr) - u malého týmu (do ~10 lidí) může jeden extrémní "
            "výkyv jednoho technika citelně posunout průměr, ke kterému se ostatní "
            "poměřují. S 27 techniky, jak máte teď, je to zanedbatelné.",
        ]),
    ]

    r = 5
    for title, paragraphs in sections:
        ws.cell(r, 3, title).font = TITLE_FONT
        r += 1
        for para in paragraphs:
            if para == "":
                r += 1
                continue
            ws.merge_cells(f"C{r}:C{r}")
            cell = ws.cell(r, 3, para)
            cell.font = Font(size=11)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Rough row height estimate for wrapped text at this column width
            # (100 chars wide, ~11pt font) - generous enough that text is
            # never clipped, occasionally a little tall rather than cut off.
            ws.row_dimensions[r].height = 15 * (1 + len(para) // 95)
            r += 1
        r += 1

    return ws


def _distinct_color(i, n):
    """Evenly-spaced hue around the color wheel - up to ~40 technicians on
    one scatter chart need genuinely distinguishable colors, not Excel's
    default ~10-color theme cycle repeating (which would make two different
    technicians' territories look identical on the map)."""
    import colorsys
    hue = (i / max(n, 1)) % 1.0
    r, g, b = colorsys.hsv_to_rgb(hue, 0.65, 0.82)
    return "%02X%02X%02X" % (int(r * 255), int(g * 255), int(b * 255))


def build_pos_map(wb, max_techs=40, max_rows=700):
    """MAP: territory overview - all Active POS plotted by GPS position,
    colored by assigned technician (product owner, 2026-07-06, from the
    manager-analytics review: "území techniků, barva = technik" - want to
    verify territories make geographic sense, spot overlaps/gaps, as part of
    "vhodně vybrané POS" - the whole reason for this review).

    No real basemap: this project has no online map service (architecture
    mandate - no external APIs, no online sync), so this is a flat XY
    scatter of GPS coordinates - the same flat-earth approximation already
    used by distanceKm() for route-km, not a street map. Chart X = longitude,
    chart Y = latitude (see ReportingEngine.ts's WRITE POS_MAP_DATA section
    for why POS_MASTER's own gpsX/gpsY columns need swapping to read as a
    normal north-up map).

    A real Excel chart can't bind to a variable-length range, so
    ReportingEngine.ts writes a FIXED max_techs x max_rows grid of (X, Y)
    column pairs every run (one pair per technician SLOT, not per actual
    technician) - this function pre-builds exactly that many chart series;
    slots beyond the real technician count are simply empty/unused series,
    harmless (Excel just shows nothing for them)."""
    DATA_SHEET = "POS_MAP_DATA"
    if "MAP" in wb.sheetnames:
        del wb["MAP"]
    ws = wb.create_sheet("MAP")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    build_nav_rail(ws, "MAP")

    build_dashboard_banner(
        ws, "MAPA ÚZEMÍ", "Aktivní POS podle technika - odhad z GPS souřadnic, není to skutečná mapa s ulicemi",
        col_start="C", col_end="N",
    )

    data_ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else None
    chart = ScatterChart()
    chart.style = 13
    chart.height = 26
    chart.width = 44
    chart.scatterStyle = "marker"
    chart.x_axis.title = "Zeměpisná délka (odhad)"
    chart.y_axis.title = "Zeměpisná šířka (odhad)"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend.position = "r"

    if data_ws is not None:
        for slot in range(max_techs):
            col_x = slot * 2 + 1  # 1-based column index for this slot's X column
            col_y = col_x + 1
            xvalues = Reference(data_ws, min_col=col_x, min_row=2, max_row=1 + max_rows)
            yvalues = Reference(data_ws, min_col=col_y, min_row=2, max_row=1 + max_rows)
            series = Series(yvalues, xvalues, title_from_data=False)
            series.marker.symbol = "circle"
            series.marker.size = 5
            series.marker.graphicalProperties.solidFill = _distinct_color(slot, max_techs)
            series.marker.graphicalProperties.ln.noFill = True
            series.graphicalProperties.line.noFill = True  # points only, no connecting line
            series.tx = SeriesLabel(strRef=StrRef(f"'{DATA_SHEET}'!${get_column_letter(col_x)}$1"))
            chart.series.append(series)

    ws.add_chart(chart, "C6")
    ws.cell(4, 3, "Barva = přiřazený technik (max. 40 technických slotů, viz ReportingEngine.ts).").font = NOTE_FONT
    return ws


def build_week_dashboard(wb, n_tech_rows=60):
    """WEEK_DASHBOARD: two views. The primary one (product owner, 2026-07-06:
    "je důležité vždy mít podle kampaně to vyhodnocení, nejdůležitější") is a
    CAMPAIGN WINDOW summary - totals across every week inside the current
    campaign's [MIN(ACTIVITY_PLAN.START_WEEK), MAX(ACTIVITY_PLAN.END_WEEK)]
    range for CONTROL.YEAR, network-wide (all technicians) plus best/worst
    technician over that whole window. This deliberately does NOT try to
    attribute a visit to a specific LOS/LOT product - docs/BUSINESS_RULES.md
    already documents that as blocked on input data (SalesApp doesn't
    reliably say which campaign a visit serviced), so this is a campaign
    SEASON rollup, not a per-product breakdown (confirmed with product
    owner). The secondary view is a single-week selector with a trend vs.
    the previous week and that week's best/worst technician, matching the
    architecture doc's original WEEK_DETAIL sketch.

    No engine change and no data model change (explicit product owner
    instruction, 2026-07-06) - everything here is a live formula over the
    already-existing TECHNICIAN_PERFORMANCE_LOG/ACTIVITY_PLAN/CONTROL.
    Per-technician campaign-window aggregation reuses the same bounded,
    fixed-row-count + SUMPRODUCT technique already proven in
    build_technician_scorecard() (hidden helper rows, not an untested
    dynamic-array-criteria trick) - INDEX() walks TECHNICIAN_SCORECARD's
    already-computed unique-technician spill positionally rather than
    recomputing UNIQUE/FILTER a second time."""
    TP = "TECHNICIAN_PERFORMANCE_LOG"
    if "WEEK_DASHBOARD" in wb.sheetnames:
        del wb["WEEK_DASHBOARD"]
    ws = wb.create_sheet("WEEK_DASHBOARD")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    for col in "CDEFGHIJKLM":
        ws.column_dimensions[col].width = 13
    build_nav_rail(ws, "WEEK_DASHBOARD")

    build_dashboard_banner(
        ws, "WEEK DASHBOARD", "Aktuální kampaň a jednotlivé týdny",
        col_start="C", col_end="M",
    )
    ws.freeze_panes = "C4"

    # ==========================================================================
    # HIDDEN FORMULA PLUMBING - never shown to the user.
    # ==========================================================================
    # CONTROL's own key text has real trailing whitespace on some rows in
    # the production workbook (e.g. "YEAR    ") - office-scripts engines
    # tolerate this via their norm()-based setting() lookup, but Excel's
    # VLOOKUP does an exact match, so this TRIM-based MATCH is deliberate,
    # not stylistic.
    ws["P1"] = "campaignYear"
    ws["P2"] = '=IFERROR(INDEX(CONTROL!$B:$B,MATCH("YEAR",TRIM(CONTROL!$A:$A),0)),YEAR(TODAY()))'
    ws["Q1"] = "campaignStartWeek"
    ws["Q2"] = "=IFERROR(MIN(ACTIVITY_PLAN!$C$2:$C$1000),0)"
    ws["R1"] = "campaignEndWeek"
    ws["R2"] = "=IFERROR(MAX(ACTIVITY_PLAN!$D$2:$D$1000),0)"

    # Per-technician campaign-window compliance % - fixed n_tech_rows helper
    # rows walking TECHNICIAN_SCORECARD's already-computed unique-technician
    # spill positionally (INDEX(...,ROW()-8)), each with its own plain
    # SUMPRODUCT ratio over the campaign window. Blank when that technician
    # has zero planned visits in the window (excluded from best/worst by
    # MAX/MIN, which ignore blanks/text).
    tech_first_row = 9
    for i in range(n_tech_rows):
        r = tech_first_row + i
        ws[f"S{r}"] = f'=IFERROR(INDEX(TECHNICIAN_SCORECARD!$P$2#,ROW()-{tech_first_row - 1}),"")'
        planned_expr = (
            f'SUMPRODUCT(({TP}!$A$2:$A$5000=$S{r})*({TP}!$B$2:$B$5000=$P$2)*'
            f'({TP}!$C$2:$C$5000>=$Q$2)*({TP}!$C$2:$C$5000<=$R$2)*{TP}!$E$2:$E$5000)'
        )
        realized_expr = (
            f'SUMPRODUCT(({TP}!$A$2:$A$5000=$S{r})*({TP}!$B$2:$B$5000=$P$2)*'
            f'({TP}!$C$2:$C$5000>=$Q$2)*({TP}!$C$2:$C$5000<=$R$2)*{TP}!$F$2:$F$5000)'
        )
        ws[f"T{r}"] = f'=IF($S{r}="","",IF({planned_expr}=0,"",ROUND({realized_expr}/{planned_expr}*100,1)))'
    tech_last_row = tech_first_row + n_tech_rows - 1
    for col in "PQRST":
        ws.column_dimensions[col].hidden = True

    # Unique (year,week) list across all technicians - the single-week
    # selector's source (not technician-scoped, unlike TECHNICIAN_SCORECARD's
    # own list).
    ws["U1"] = "weeks"
    ws["U2"] = (
        f'=IFERROR(SORT(UNIQUE(FILTER(TEXT({TP}!$B$2:$B$5000,"0000")&"-W"&TEXT({TP}!$C$2:$C$5000,"00"),'
        f'{TP}!$A$2:$A$5000<>""))),"Zatím žádná data")'
    )
    ws["V1"] = '=IFERROR(VALUE(LEFT($H$18,4)),0)'   # selected year
    ws["V2"] = '=IFERROR(VALUE(MID($H$18,7,2)),0)'  # selected week
    ws["W1"] = (  # previous week on record, network-wide (any gap size) - combined year*100+week key
        f'=IFERROR(AGGREGATE(14,6,({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)/'
        f'((({TP}!$B$2:$B$5000*100+{TP}!$C$2:$C$5000)<($V$1*100+$V$2))*({TP}!$A$2:$A$5000<>"")),1),"")'
    )
    prev_week_cond = f'({TP}!$B$2:$B$5000=INT($W$1/100))*({TP}!$C$2:$C$5000=MOD($W$1,100))'
    ws["W2"] = (  # that previous week's network-wide compliance % (sum realized / sum planned - NOT an average of per-technician %s)
        f'=IF($W$1="","-",IFERROR(ROUND(SUMPRODUCT({prev_week_cond}*{TP}!$F$2:$F$5000)/'
        f'SUMPRODUCT({prev_week_cond}*{TP}!$E$2:$E$5000)*100,1),"-"))'
    )
    ws.column_dimensions["U"].hidden = True
    ws.column_dimensions["V"].hidden = True
    ws.column_dimensions["W"].hidden = True

    dashboard_ui.define_named_range(ws, "WeekList", "WEEK_DASHBOARD!$U$2#")

    # ==========================================================================
    # CAMPAIGN WINDOW SUMMARY - the primary view.
    # ==========================================================================
    build_section_header(ws, "C5", "AKTUÁLNÍ KAMPAŇ")
    ws.merge_cells("F5:M5")
    ws["F5"] = '="Týdny "&$Q$2&"–"&$R$2&" / "&$P$2&" (dle ACTIVITY_PLAN, bez rozlišení LOS/LOT)"'
    ws["F5"].font = NOTE_FONT
    ws["F5"].alignment = Alignment(horizontal="right", vertical="center")

    campaign_cond = f'({TP}!$B$2:$B$5000=$P$2)*({TP}!$C$2:$C$5000>=$Q$2)*({TP}!$C$2:$C$5000<=$R$2)'
    campaign_planned = f'SUMPRODUCT({campaign_cond}*{TP}!$E$2:$E$5000)'
    campaign_realized = f'SUMPRODUCT({campaign_cond}*{TP}!$F$2:$F$5000)'
    cards = [
        ("C", "D", "Naplánováno", f'={campaign_planned}', NAVY, WHITE),
        ("E", "F", "Realizováno", f'={campaign_realized}', NAVY, WHITE),
        ("G", "H", "Splněno (včas+pozdě)", f'=SUMPRODUCT({campaign_cond}*({TP}!$G$2:$G$5000+{TP}!$H$2:$H$5000))', STATUS_GOOD, dashboard_ui.TINT_GOOD),
        ("I", "J", "Nesplněno", f'=SUMPRODUCT({campaign_cond}*{TP}!$I$2:$I$5000)', STATUS_CRITICAL, dashboard_ui.TINT_CRITICAL),
        ("K", "L", "Návštěvy navíc", f'=SUMPRODUCT({campaign_cond}*{TP}!$J$2:$J$5000)', STATUS_WARNING, dashboard_ui.TINT_WARNING),
        ("M", "M", "Compliance %", f'=IFERROR(ROUND({campaign_realized}/{campaign_planned}*100,1),0)', NAVY, WHITE),
    ]
    compliance_cell = build_kpi_card_row(ws, cards, label_row=6, value_row_start=7, value_row_end=9)[5]  # "M7"
    apply_severity_conditional_formatting(
        ws, "M7:M9", compliance_cell,
        thresholds=[(90, STATUS_GOOD), (70, STATUS_WARNING), (50, STATUS_SERIOUS)],
        below_color=STATUS_CRITICAL,
    )

    best_tech_ref = build_kpi_card(
        ws, "C", "F", 11, 12, 13, "Nejlepší technik (kampaň)",
        f'=IFERROR(INDEX($S${tech_first_row}:$S${tech_last_row},MATCH(MAX($T${tech_first_row}:$T${tech_last_row}),$T${tech_first_row}:$T${tech_last_row},0))'
        f'&" ("&TEXT(MAX($T${tech_first_row}:$T${tech_last_row}),"0.0")&"%)","Zatím žádná data")',
        value_color=STATUS_GOOD, fill_color=dashboard_ui.TINT_GOOD,
    )
    ws[best_tech_ref].font = font_card_value(size=14, color=STATUS_GOOD)
    worst_tech_ref = build_kpi_card(
        ws, "G", "M", 11, 12, 13, "Nejslabší technik (kampaň)",
        f'=IFERROR(INDEX($S${tech_first_row}:$S${tech_last_row},MATCH(MIN($T${tech_first_row}:$T${tech_last_row}),$T${tech_first_row}:$T${tech_last_row},0))'
        f'&" ("&TEXT(MIN($T${tech_first_row}:$T${tech_last_row}),"0.0")&"%)","Zatím žádná data")',
        value_color=STATUS_CRITICAL, fill_color=dashboard_ui.TINT_CRITICAL,
    )
    ws[worst_tech_ref].font = font_card_value(size=14, color=STATUS_CRITICAL)

    # ==========================================================================
    # SINGLE WEEK VIEW - secondary, matches the architecture doc's original
    # WEEK_DETAIL sketch (per-week slice + trend vs. previous week).
    # ==========================================================================
    build_section_header(ws, "C16", "JEDNOTLIVÉ TÝDNY")
    build_filter_bar_background(ws, 18, "C", "M")
    build_filter_dropdown(ws, "C18", "TÝDEN", "H18:J18", "=WeekList", default_formula='=IFERROR(INDEX(WeekList,1),"")')

    week_cond = f'({TP}!$B$2:$B$5000=$V$1)*({TP}!$C$2:$C$5000=$V$2)'
    week_planned = f'SUMPRODUCT({week_cond}*{TP}!$E$2:$E$5000)'
    week_realized = f'SUMPRODUCT({week_cond}*{TP}!$F$2:$F$5000)'
    week_cards = [
        ("C", "D", "Naplánováno", f'={week_planned}', NAVY, WHITE),
        ("E", "F", "Realizováno", f'={week_realized}', NAVY, WHITE),
        ("G", "H", "Nesplněno", f'=SUMPRODUCT({week_cond}*{TP}!$I$2:$I$5000)', STATUS_CRITICAL, dashboard_ui.TINT_CRITICAL),
        ("I", "J", "Compliance %", f'=IFERROR(ROUND({week_realized}/{week_planned}*100,1),0)', NAVY, WHITE),
        ("K", "M", "Trend proti minulému týdnu",
         f'=IF($W$1="","Zatím není s čím srovnat",'
         f'"("&TEXT($W$1,"0")&")")',  # placeholder replaced below with the real delta formula
         NAVY, WHITE),
    ]
    week_value_cells = build_kpi_card_row(ws, week_cards, label_row=19, value_row_start=20, value_row_end=21)
    week_compliance_cell = week_value_cells[3]  # "I20"
    apply_severity_conditional_formatting(
        ws, "I20:J21", week_compliance_cell,
        thresholds=[(90, STATUS_GOOD), (70, STATUS_WARNING), (50, STATUS_SERIOUS)],
        below_color=STATUS_CRITICAL, font_size=22,
    )
    # Real trend formula (needs week_compliance_cell, only known after the
    # card row above is built) - overwrite the placeholder card value. $W$2
    # is the previous week's network-wide compliance %, computed correctly
    # as sum(realized)/sum(planned) - see the W2 formula above.
    trend_cell_ref = week_value_cells[4]  # "K20"
    ws[trend_cell_ref] = (
        f'=IF($W$1="","Zatím není s čím srovnat",'
        f'IF({week_compliance_cell}>$W$2,"▲ "&TEXT({week_compliance_cell}-$W$2,"+0.0")&" p.b.",'
        f'IF({week_compliance_cell}<$W$2,"▼ "&TEXT({week_compliance_cell}-$W$2,"+0.0;-0.0")&" p.b.","→ beze změny")))'
    )
    ws[trend_cell_ref].font = font_card_value(size=13)
    build_status_badge_conditional(ws, f"{trend_cell_ref}:M21", trend_cell_ref, rules=[
        ("▲", None, STATUS_GOOD),
        ("▼", None, STATUS_CRITICAL),
    ])

    best_week_tech_ref = build_kpi_card(
        ws, "C", "F", 23, 24, 25, "Nejlepší technik (tento týden)",
        f'=IFERROR(INDEX({TP}!$A$2:$A$5000,MATCH(1,{week_cond}*'
        f'({TP}!$K$2:$K$5000=IFERROR(AGGREGATE(14,6,{TP}!$K$2:$K$5000/({week_cond}),1),-1)),0)),"Zatím žádná data")',
        value_color=STATUS_GOOD, fill_color=dashboard_ui.TINT_GOOD,
    )
    ws[best_week_tech_ref].font = font_card_value(size=14, color=STATUS_GOOD)
    worst_week_tech_ref = build_kpi_card(
        ws, "G", "M", 23, 24, 25, "Nejslabší technik (tento týden)",
        f'=IFERROR(INDEX({TP}!$A$2:$A$5000,MATCH(1,{week_cond}*'
        f'({TP}!$K$2:$K$5000=IFERROR(AGGREGATE(15,6,{TP}!$K$2:$K$5000/({week_cond}),1),-1)),0)),"Zatím žádná data")',
        value_color=STATUS_CRITICAL, fill_color=dashboard_ui.TINT_CRITICAL,
    )
    ws[worst_week_tech_ref].font = font_card_value(size=14, color=STATUS_CRITICAL)

    return ws


def redesign_activity_plan(wb, tech_column_letter):
    """Adds a live impact estimate + a Gantt-style timeline heatmap to the
    RIGHT of the existing A:F data table. The data table itself (A:F) is
    left at its original position/column order deliberately - ImportEngine.ts
    reads it positionally (row[0..5]), and moving it would be a business-
    logic-adjacent risk for a pure UX task. Everything new lives in columns
    G onward, which no engine reads."""
    ws = wb["ACTIVITY_PLAN"]
    # Real data row count from column A (TYPE) - NOT ws.max_row, which can
    # already be inflated by decorative pre-styling of empty future rows if
    # this runs after the generic per-sheet styling pass (it must run
    # before that pass - see apply_all - but this guard makes the function
    # correct regardless of call order).
    n_rows = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value not in (None, ""):
            n_rows = r
    n_rows = max(n_rows, 2)

    # ---- G: live per-campaign estimate ----
    ws["G1"] = "ODHAD_NAVSTEV_ZA_KAMPAN"
    for r in range(2, n_rows + 1):
        ws.cell(r, 7, f"=IF($C{r}=\"\",\"\",($D{r}-$C{r}+1)*$J$2*$J$5)")

    # ---- I/J: reference values panel that the estimate formulas read ----
    ws["I1"] = "REFERENČNÍ HODNOTY (pro odhad)"
    ws["I1"].font = SECTION_FONT
    pm_range = f"POS_MASTER!{tech_column_letter}2:{tech_column_letter}20000"
    ws["I2"] = "Počet techniků (distinct, z POS_MASTER)"
    ws["J2"] = f'=SUMPRODUCT(({pm_range}<>"")/COUNTIF({pm_range},{pm_range}&""))'
    ws["I3"] = "Cílový počet návštěv/den (CONTROL.TARGET_VISITS_DAY)"
    ws["J3"] = '=IFERROR(VLOOKUP("TARGET_VISITS_DAY",CONTROL!A:B,2,FALSE),8)'
    ws["I4"] = "Průměr pracovních dní/týden (odhad vč. svátků)"
    ws["J4"] = 4.8
    ws["I5"] = "→ Kapacita/technik/týden (řádky J3*J4)"
    ws["J5"] = "=J3*J4"
    for row in (1, 2, 3, 4, 5):
        ws.cell(row, 9).font = NOTE_FONT if row != 1 else SECTION_FONT
    ws["J4"].fill = PatternFill("solid", fgColor=EDITABLE_FILL)

    # ---- L onward: timeline heatmap (weeks as columns) ----
    weeks = []
    for r in range(2, n_rows + 1):
        sv, ev = ws.cell(r, 3).value, ws.cell(r, 4).value
        if isinstance(sv, (int, float)):
            weeks.append(int(sv))
        if isinstance(ev, (int, float)):
            weeks.append(int(ev))
    if weeks:
        week_start, week_end = min(weeks) - 3, max(weeks) + 3
    else:
        week_start, week_end = 1, 20
    week_start = max(1, week_start)

    ws.cell(1, 11, "").fill = PatternFill("solid", fgColor="FFFFFF")  # column K = spacer
    timeline_first_col = 12  # L
    ws.cell(1, timeline_first_col - 1, "ČASOVÁ OSA KAMPANÍ (týden)").font = SECTION_FONT
    for i, week in enumerate(range(week_start, week_end + 1)):
        col = timeline_first_col + i
        cell = ws.cell(1, col, week)
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 3.2

    los_fill = PatternFill("solid", fgColor=LOS_FILL)
    lot_fill = PatternFill("solid", fgColor=LOT_FILL)
    last_col_letter = get_column_letter(timeline_first_col + (week_end - week_start))
    for r in range(2, n_rows + 1):
        for i, week in enumerate(range(week_start, week_end + 1)):
            col = timeline_first_col + i
            col_letter = get_column_letter(col)
            week_header_ref = f"{col_letter}$1"
            in_range_formula = f"AND({week_header_ref}>=$C{r},{week_header_ref}<=$D{r})"
            ws.conditional_formatting.add(
                f"{col_letter}{r}",
                FormulaRule(formula=[f'AND({in_range_formula},$A{r}="LOS")'], fill=los_fill),
            )
            ws.conditional_formatting.add(
                f"{col_letter}{r}",
                FormulaRule(formula=[f'AND({in_range_formula},$A{r}="LOT")'], fill=lot_fill),
            )

    # "You are here" - a vertical band on whichever timeline column is the
    # current week, so orientation in a several-months-wide timeline doesn't
    # require counting columns. Safe to compare directly against
    # ISOWEEKNUM(TODAY()): PlanningEngine.ts generates real calendar dates
    # via isoMonday(CONTROL.YEAR, week) using this exact same week number
    # (see ReportingEngine.ts's PLANNING READINESS section for the same
    # reasoning applied elsewhere) - it is a real ISO week number, not an
    # unrelated campaign-relative counter, under the existing documented
    # single-year-per-run simplification (docs/BACKLOG.md).
    # Header-row only (not every data cell) so it never competes with the
    # LOS/LOT fill on the same cell - a header highlight is enough to orient
    # "we are here" in a several-months-wide timeline without hiding which
    # campaign type is active during the current week.
    today_fill = PatternFill("solid", fgColor="FFF2A6")
    for i, week in enumerate(range(week_start, week_end + 1)):
        col_letter = get_column_letter(timeline_first_col + i)
        ws.conditional_formatting.add(
            f"{col_letter}1",
            FormulaRule(formula=[f"{col_letter}1=ISOWEEKNUM(TODAY())"], fill=today_fill),
        )

    ws.cell(n_rows + 3, 12, "LOS").fill = los_fill
    ws.cell(n_rows + 3, 13, "= aktivní LOS kampaň v daném týdnu")
    ws.cell(n_rows + 4, 12, "LOT").fill = lot_fill
    ws.cell(n_rows + 4, 13, "= aktivní LOT kampaň v daném týdnu")
    ws.cell(n_rows + 5, 12, "").fill = today_fill
    ws.cell(n_rows + 5, 13, "= aktuální týden (dnes)")
    ws.cell(n_rows + 6, 12,
            "Souběh dvou kampaní ve stejném týdnu = obě barvy vidíš ve stejném sloupci u různých řádků "
            "(porovnej řádky svisle).").font = NOTE_FONT

    # ---- LOS/LOT activity as two lines - product owner, 2026-07-06: "chci
    # aby aktivity plan byl i vizualizovany jako 2 čáry a dynamicky se
    # měnil podle toho co zadam". The heatmap above shows per-row detail;
    # this collapses it to a single "is LOS/LOT running this week, or is
    # there a gap" line each, over the same week timeline - a step chart a
    # manager can read at a glance, live off the same $A/$C/$D columns the
    # heatmap already reads, so it updates the moment a row is edited, no
    # engine involved. ----
    los_row = n_rows + 9
    lot_row = n_rows + 10
    ws.cell(los_row, 11, "LOS aktivní (1/0)").font = NOTE_FONT
    ws.cell(lot_row, 11, "LOT aktivní (1/0)").font = NOTE_FONT
    for i, week in enumerate(range(week_start, week_end + 1)):
        col = timeline_first_col + i
        col_letter = get_column_letter(col)
        week_ref = f"{col_letter}$1"
        ws.cell(los_row, col).value = (
            f'=IF(SUMPRODUCT(($A$2:$A${n_rows}="LOS")*({week_ref}>=$C$2:$C${n_rows})*'
            f'({week_ref}<=$D$2:$D${n_rows}))>0,1,0)'
        )
        ws.cell(lot_row, col).value = (
            f'=IF(SUMPRODUCT(($A$2:$A${n_rows}="LOT")*({week_ref}>=$C$2:$C${n_rows})*'
            f'({week_ref}<=$D$2:$D${n_rows}))>0,1,0)'
        )
    for row in (los_row, lot_row):
        ws.row_dimensions[row].hidden = True

    chart_cats = Reference(ws, min_col=timeline_first_col, max_col=timeline_first_col + (week_end - week_start),
                            min_row=1, max_row=1)
    los_values = Reference(ws, min_col=timeline_first_col, max_col=timeline_first_col + (week_end - week_start),
                            min_row=los_row, max_row=los_row)
    lot_values = Reference(ws, min_col=timeline_first_col, max_col=timeline_first_col + (week_end - week_start),
                            min_row=lot_row, max_row=lot_row)
    activity_chart = LineChart()
    activity_chart.style = 2
    activity_chart.height = 7
    activity_chart.width = 30
    activity_chart.title = "LOS / LOT aktivita v čase"
    los_series = Series(los_values, title="LOS")
    los_series.graphicalProperties.line.solidFill = LOS_FILL
    los_series.graphicalProperties.line.width = 25000
    los_series.smooth = False  # this is a 0/1 on-off flag, not a curve - smoothing would draw misleading arcs
    lot_series = Series(lot_values, title="LOT")
    lot_series.graphicalProperties.line.solidFill = LOT_FILL
    lot_series.graphicalProperties.line.width = 25000
    lot_series.smooth = False
    activity_chart.series.append(los_series)
    activity_chart.series.append(lot_series)
    activity_chart.set_categories(chart_cats)
    activity_chart.y_axis.scaling.min = 0
    activity_chart.y_axis.scaling.max = 1.2
    activity_chart.y_axis.delete = True
    ws.add_chart(activity_chart, f"L{n_rows + 12}")

    # AutoFilter + banded rows on the editable data table (A:G, including
    # the live estimate column) - a campaign list spanning many months is
    # only actually usable if it can be filtered/sorted like the working
    # screen it is, not just displayed.
    ws.auto_filter.ref = f"A1:G{n_rows}"
    apply_banded_rows(ws, 2, n_rows, 7)

    ws.freeze_panes = "C2"


def build_technician_plan(wb, n_rows=260, pos_master_notes_col="AK", pos_master_last_visit_col="X"):
    """TOUR PLAN: the document actually sent to a technician roughly once
    per campaign (~4 weeks) - pick a technician, see their ENTIRE campaign
    route (every week currently in MANAGER_PLAN, Draft included, grouped by
    week) in one place, ready to print or export to PDF. Excel's native
    Print/Export-to-PDF always operates on the sheet's current state, and
    the current state IS already that technician's whole campaign once the
    one dropdown is set - so "select technician, then File > Print" is the
    entire delivery workflow (product owner, 2026-07-06: this replaces the
    weekly manual AutoFilter+copy-paste process documented in
    docs/EXCEL_ONLY_WORKFLOW.md step 4, and must show the full campaign a
    technician is actually sent, not a single week at a time).

    This is a VIEW over what PlanningEngine.ts already decided, not a
    second planning pass - the planning logic (PPT/cadence/scoring/GPS
    clustering) already ran and is baked into MANAGER_PLAN's rows; nothing
    here re-derives it. The one piece of decision-relevant HISTORY this
    view adds beyond MANAGER_PLAN's own columns is POSLEDNÍ NÁVŠTĚVA (POS_MASTER.lastRealVisitDate)
    - "when were you last actually here" - product owner's explicit pick
    among several candidate context fields (2026-07-06), the others
    (PPT, a human-readable REASON, per-POS compliance) deliberately not
    added this round.

    No internal category code, no POS_AREA, no system REASON tag beyond
    that (TYDEN/week number is included - product owner confirmed it's
    important for them, 2026-07-03).

    Pure live-formula view (no engine change): stays in sync automatically
    whenever Planning Engine regenerates MANAGER_PLAN, including Draft
    weeks, so a technician's plan is visible even before publish.

    DESIGN: a single hidden FILTER() spill (column R) pulls every matching
    technician row straight out of MANAGER_PLAN (A:Q, sorted by WEEK then
    DATE) in one pass; the visible columns are then simple per-row
    INDEX(spill, row, col) lookups plus the same light transforms the
    previous flat view already had (DEN Czech translation, ČÍSLO
    TERMINÁLU/POZNÁMKA/POSLEDNÍ NÁVŠTĚVA lookups into POS_MASTER, AKTIVITA
    = LOS+LOT concat) - not a second per-column FILTER(), which would have
    needed FILTER() to wrap a transform expression (SWITCH/VLOOKUP) rather
    than a plain range; this two-stage shape keeps every visible formula
    simple and independently correct-or-blank via IFERROR, rather than one
    large nested expression.

    MANAGER_PLAN column layout this reads from (fixed, see
    scaffold_workbook.py): A=WEEK, B=DATE, C=DAY, D=TECHNICIAN, E=POS,
    F=KATEGORIE, G=NAZEV_PROVOZOVNY, H=ULICE, I=CISLO, J=MESTO, K=OBLAST,
    L=POS_AREA, M=PPT, N=LOS_ACTIVITY, O=LOT_ACTIVITY, P=REASON, Q=GPS_GROUP.

    n_rows=200 note: bounded to comfortably cover one technician's WHOLE
    campaign (CAMPAIGN_LENGTH weeks, default 4, times VISITS_PER_WEEK
    default capacity plus GPS-bonus headroom - well under 200 in practice).
    Print area is sized to n_rows, not to the actual match count (openpyxl
    can't know that at build time) - a technician with fewer visits prints
    some trailing blank-but-bordered rows rather than a perfectly trimmed
    page; a known, accepted imperfection, not a silent gap."""
    TP = "MANAGER_PLAN"
    if "TECHNICIAN_PLAN" in wb.sheetnames:
        del wb["TECHNICIAN_PLAN"]
    ws = wb.create_sheet("TECHNICIAN_PLAN")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    for col in "CDEFGHIJKLMN":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["C"].width = 8
    build_nav_rail(ws, "TECHNICIAN_PLAN")

    build_dashboard_banner(
        ws, "TOUR PLAN", "Vyber technika - celá kampaň, seskupená po týdnech, připravená k tisku nebo exportu do PDF",
        col_start="C", col_end="N",
    )
    ws.freeze_panes = "C4"

    # ==========================================================================
    # HIDDEN FORMULA PLUMBING (columns R:S) - never shown to the user.
    # ==========================================================================
    HEADER_ROW = 8
    DATA_FIRST_ROW = HEADER_ROW + 1
    DATA_LAST_ROW = DATA_FIRST_ROW + n_rows - 1
    ws["S1"] = "technici"
    ws["S2"] = f'=IFERROR(SORT(UNIQUE(FILTER({TP}!$D$2:$D$3001,{TP}!$D$2:$D$3001<>""))),"Zatím žádná data")'
    dashboard_ui.define_named_range(ws, "TourTechnicianList", "TECHNICIAN_PLAN!$S$2#")
    # Raw matching MANAGER_PLAN rows (A:Q, all 17 columns) for the WHOLE
    # campaign, sorted by WEEK then DATE - the single source every visible
    # column below reads from.
    ws[f"R{DATA_FIRST_ROW}"] = (
        f'=IFERROR(SORT(FILTER({TP}!$A$2:$Q$3001,{TP}!$D$2:$D$3001=$D$5),{{1,2}},{{1,1}}),'
        f'"Zatím žádné návštěvy pro tento výběr")'
    )
    for col in "RS":
        ws.column_dimensions[col].hidden = True

    # ==========================================================================
    # FILTER BAR
    # ==========================================================================
    build_filter_bar_background(ws, 5, "C", "N")
    build_filter_dropdown(ws, "C5", "TECHNIK", "D5:F5", "=TourTechnicianList",
                           default_formula='=IFERROR(INDEX(TourTechnicianList,1),"")')
    ws.merge_cells("G5:N5")
    ws["G5"] = f'="Počet návštěv v kampani: "&IFERROR(ROWS($R${DATA_FIRST_ROW}#),0)'
    ws["G5"].font = Font(italic=True, size=10, color="595959")
    ws["G5"].alignment = Alignment(vertical="center", horizontal="right", indent=1)

    # ==========================================================================
    # VISIBLE TABLE - simple per-row lookups into the hidden spill above.
    # ==========================================================================
    headers = [
        "TYDEN", "DATUM", "DEN", "POS", "ČÍSLO TERMINÁLU", "NÁZEV PROVOZOVNY",
        "ULICE", "MĚSTO", "OBLAST", "AKTIVITA", "POSLEDNÍ NÁVŠTĚVA", "POZNÁMKA",
    ]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 3)  # starts at C
        ws[f"{col}{HEADER_ROW}"] = h
    for cell in ws[f"C{HEADER_ROW}:N{HEADER_ROW}"][0]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[HEADER_ROW].height = 26

    def staged(row, col_num):
        return f'INDEX($R${DATA_FIRST_ROW}#,{row - DATA_FIRST_ROW + 1},{col_num})'

    for r in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        pos_cell = f"F{r}"
        ws[f"C{r}"] = f'=IFERROR({staged(r, 1)},"")'   # TYDEN
        ws[f"D{r}"] = f'=IFERROR({staged(r, 2)},"")'   # DATUM
        ws[f"E{r}"] = (  # DEN - same MON->Pondělí translation as before
            f'=IFERROR(SWITCH({staged(r, 3)},'
            f'"MON","Pondělí","TUE","Úterý","WED","Středa","THU","Čtvrtek","FRI","Pátek",'
            f'{staged(r, 3)}),"")'
        )
        ws[f"F{r}"] = f'=IFERROR({staged(r, 5)},"")'   # POS
        ws[f"G{r}"] = (  # CISLO TERMINALU - see build_technician_plan docstring re: single-terminal-per-POS limitation
            f'=IF({pos_cell}="","",IFERROR(VLOOKUP({pos_cell},POS_MASTER!$A:$B,2,FALSE),""))'
        )
        ws[f"H{r}"] = f'=IFERROR({staged(r, 7)},"")'   # NAZEV PROVOZOVNY
        ws[f"I{r}"] = f'=IFERROR(TRIM({staged(r, 8)}&" "&{staged(r, 9)}),"")'  # ULICE (+ CISLO)
        ws[f"J{r}"] = f'=IFERROR({staged(r, 10)},"")'  # MESTO
        ws[f"K{r}"] = f'=IFERROR({staged(r, 11)},"")'  # OBLAST
        ws[f"L{r}"] = (  # AKTIVITA = LOS + LOT concat
            f'=IFERROR(TRIM(IF({staged(r, 14)}<>"","LOS: "&{staged(r, 14)}&" ","")'
            f'&IF({staged(r, 15)}<>"","LOT: "&{staged(r, 15)},"")),"")'
        )
        ws[f"M{r}"] = (  # POSLEDNI NAVSTEVA - POS_MASTER.lastRealVisitDate, the one history field
            f'=IF({pos_cell}="","",IFERROR(VLOOKUP({pos_cell},POS_MASTER!$A:${pos_master_last_visit_col},'
            f'{pos_master_notes_col_index(pos_master_last_visit_col)},FALSE),"–"))'
        )
        ws[f"N{r}"] = (  # POZNAMKA - manager note from POS_MASTER, not the internal REASON tag
            f'=IF({pos_cell}="","",IFERROR(VLOOKUP({pos_cell},POS_MASTER!$A:${pos_master_notes_col},'
            f'{pos_master_notes_col_index(pos_master_notes_col)},FALSE),""))'
        )
        for col in "CDEFGHIJKLMN":
            ws[f"{col}{r}"].border = CARD_BORDER

    # ==========================================================================
    # PRINT SETUP - "select technician, then File > Print / Export to PDF"
    # is the one-click deliverable this was built for.
    # ==========================================================================
    ws.print_area = f"C1:N{DATA_LAST_ROW}"
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.gridLines = False

    # Highlight today's visits - the one row (or handful) a technician
    # actually needs when they open this sheet on the day itself, so they
    # don't have to scroll through the whole campaign to find it.
    ws.conditional_formatting.add(
        f"C{DATA_FIRST_ROW}:N{DATA_LAST_ROW}",
        FormulaRule(formula=[f"$D{DATA_FIRST_ROW}=TODAY()"], fill=PatternFill("solid", fgColor="FFF2A6")),
    )
    # Week-boundary cue: a bold top border whenever TYDEN changes from the
    # row above - visually groups the whole-campaign list into per-week
    # blocks without needing actual inserted rows (which a formula view
    # can't do), per the product owner's "seskupený po týdnech" request
    # (2026-07-06). Formula is anchored to the range's top-left cell with
    # relative row references; Excel re-anchors it per row automatically.
    week_start_side = Side(style="medium", color=NAVY)
    ws.conditional_formatting.add(
        f"C{DATA_FIRST_ROW + 1}:N{DATA_LAST_ROW}",
        FormulaRule(
            formula=[f'AND($C{DATA_FIRST_ROW + 1}<>"",$C{DATA_FIRST_ROW + 1}<>$C{DATA_FIRST_ROW})'],
            border=Border(top=week_start_side),
        ),
    )
    return ws


def pos_master_notes_col_index(col_letter):
    return column_index_from_string(col_letter)


def apply_banded_rows(ws, first_data_row, last_row, n_cols, band_color="F2F2F2"):
    """Alternating row shading via conditional formatting (not direct cell
    fill) - survives engine clear(contents)/setValues() cycles because it's
    attached to the range, not to individual cell styles that clear() could
    interact with, and doesn't need the range to actually contain data to
    render correctly for a growing/shrinking table."""
    last_col_letter = get_column_letter(n_cols)
    band_range = f"A{first_data_row}:{last_col_letter}{last_row}"
    ws.conditional_formatting.add(
        band_range,
        FormulaRule(formula=[f"ISEVEN(ROW())"], fill=PatternFill("solid", fgColor=band_color)),
    )


def hide_technical_sheets(wb):
    for name in HIDDEN_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"
    # TERMINAL_RULES is deliberately excluded from HIDDEN_SHEETS (see that
    # set's comment) - forced visible here too, not just at scaffold-build
    # time, in case a workbook copy ever had it hidden from an older run.
    if "TERMINAL_RULES" in wb.sheetnames:
        wb["TERMINAL_RULES"].sheet_state = "visible"


def build_dashboard_template(wb):
    """Pre-styles a KPI-tile header band that ReportingEngine.ts writes
    numbers into (fixed cell positions, see ReportingEngine.ts). The
    detailed tables ReportingEngine already produces are kept below this
    band, just pushed down to make room - same data, better hierarchy.

    Also builds the three native Excel charts (weekly trend, technician
    workload, regional completion) bound to the FIXED chart-data ranges in
    columns H:K that ReportingEngine.ts writes on every run (see that
    file's "CHART DATA BLOCKS" comment for why fixed ranges, not the
    flowing detail sections, are what a chart can safely reference). Charts
    are openpyxl objects created once here; Office Scripts never touches
    them, only the cell values they read from - so a chart keeps rendering
    correctly across every future engine run with no further Python step."""
    ws = wb["DASHBOARD"]
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18
    for col in "HIJK":
        ws.column_dimensions[col].width = 14

    ws.merge_cells("A1:F1")
    ws["A1"] = "DASHBOARD"
    ws["A1"].font = Font(bold=True, size=20, color=WHITE)  # matches IMPORT_HUB's banner size - HOME (26) is the only intentionally-larger one, as the primary entry point
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    tile_specs = [
        ("B3", "Aktivní POS", "375623"),
        ("C3", "Splněno včas", "375623"),
        ("D3", "Nesplněno", "C00000"),
        ("E3", "Otevřené alerty", "BF8F00"),
    ]
    for cell_ref, label, color in tile_specs:
        col = cell_ref[0]
        ws[f"{col}2"] = label
        ws[f"{col}2"].font = Font(bold=True, size=9, color="595959")
        ws[f"{col}2"].alignment = Alignment(horizontal="center")
        ws[cell_ref].font = Font(bold=True, size=22, color=color)
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 34
    ws.freeze_panes = "A5"

    # Severity badges for the flowing ADVISOR ALERTS rows (ReportingEngine.ts
    # writes "TYPE (SEVERITY)" as column A of each alert row, e.g.
    # "TECHNICIAN_OVERLOAD (CRITICAL)") - a colored left-edge cue so a
    # CRITICAL alert is visually distinct from an informational one without
    # reading the text. Row-range based (A5:A2000), not tied to any fixed
    # row count, since the flowing section's length varies run to run - a
    # text-match rule keeps working regardless of exactly which row it lands
    # on, unlike the chart data blocks above which need fixed positions.
    severity_colors = [("CRITICAL", "C00000"), ("WARNING", "BF8F00"), ("INFO", "2E75B6")]
    for keyword, color in severity_colors:
        ws.conditional_formatting.add(
            "A5:A2000",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("({keyword})",$A5))'],
                fill=PatternFill("solid", fgColor=color),
                font=Font(color=WHITE, bold=True),
            ),
        )

    _build_dashboard_charts(ws)
    return ws


def _build_dashboard_charts(ws):
    # ---- WEEKLY TREND: label H1, header H2:K2, data H3:K14 (12 weeks) ----
    ws["H1"] = "📈 VÝVOJ PLNĚNÍ PO TÝDNECH"
    ws["H1"].font = SECTION_FONT
    for col, label in zip("HIJK", ["Week", "Splněno včas", "Splněno pozdě", "Nesplněno"]):
        ws[f"{col}2"] = label
        ws[f"{col}2"].font = Font(bold=True, size=9, color="595959")

    weekly_chart = LineChart()
    weekly_chart.title = "Plnění plánovaných návštěv po týdnech"
    weekly_chart.style = 2
    weekly_chart.y_axis.title = "Počet návštěv"
    weekly_chart.x_axis.title = "Týden"
    weekly_chart.height = 8
    weekly_chart.width = 22
    cats = Reference(ws, min_col=8, min_row=3, max_row=14)
    for col, name, color in [(9, "Splněno včas", "375623"), (10, "Splněno pozdě", "BF8F00"), (11, "Nesplněno", "C00000")]:
        data = Reference(ws, min_col=col, min_row=2, max_row=14)
        weekly_chart.add_data(data, titles_from_data=True)
    weekly_chart.set_categories(cats)
    for series, color in zip(weekly_chart.series, ["375623", "BF8F00", "C00000"]):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 20000
        series.smooth = False
    ws.add_chart(weekly_chart, "M1")

    # ---- TECHNICIAN WORKLOAD: label H17, header H18:K18, data H19:K32 ----
    ws["H17"] = "👥 VYTÍŽENÍ TECHNIKŮ (nejnovější týden)"
    ws["H17"].font = SECTION_FONT
    for col, label in zip("HIJK", ["Technik", "Naplánováno", "Kapacita", "Vytížení %"]):
        ws[f"{col}18"] = label
        ws[f"{col}18"].font = Font(bold=True, size=9, color="595959")

    workload_chart = BarChart()
    workload_chart.type = "col"
    workload_chart.title = "Naplánováno vs. kapacita (aktuální týden)"
    workload_chart.style = 10
    workload_chart.y_axis.title = "Počet návštěv"
    workload_chart.height = 8
    workload_chart.width = 22
    w_cats = Reference(ws, min_col=8, min_row=19, max_row=32)
    for col, color in [(9, "375623"), (10, "BFBFBF")]:
        data = Reference(ws, min_col=col, min_row=18, max_row=32)
        workload_chart.add_data(data, titles_from_data=True)
    workload_chart.set_categories(w_cats)
    for series, color in zip(workload_chart.series, ["375623", "BFBFBF"]):
        series.graphicalProperties.solidFill = color
    ws.add_chart(workload_chart, "M19")  # 2-row buffer below the weekly trend chart above (M1, ~15 rows tall) so they don't visually crowd each other

    # Progress bar on Utilization % (K19:K32) - safe here (unlike the
    # flowing detail sections, where the same column means different things
    # in different sections) since this fixed block has exactly one meaning
    # per column.
    ws.conditional_formatting.add(
        "K19:K32",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=150, color="375623"),
    )

    # ---- REGIONAL OVERVIEW: label H35, header H36:I36, data H37:I48 ----
    ws["H35"] = "🗺 REGIONÁLNÍ PŘEHLED (completion %)"
    ws["H35"].font = SECTION_FONT
    for col, label in zip("HI", ["Market", "Completion %"]):
        ws[f"{col}36"] = label
        ws[f"{col}36"].font = Font(bold=True, size=9, color="595959")

    regional_chart = BarChart()
    regional_chart.type = "bar"  # horizontal - reads better with region names
    regional_chart.title = "Splnění plánu podle regionu"
    regional_chart.style = 12
    regional_chart.y_axis.title = "Completion %"
    regional_chart.height = 8
    regional_chart.width = 22
    r_cats = Reference(ws, min_col=8, min_row=37, max_row=48)
    r_data = Reference(ws, min_col=9, min_row=36, max_row=48)
    regional_chart.add_data(r_data, titles_from_data=True)
    regional_chart.set_categories(r_cats)
    regional_chart.series[0].graphicalProperties.solidFill = "2E75B6"
    ws.add_chart(regional_chart, "M37")  # same buffer reasoning as the workload chart above

    # Progress bar on Completion % (I37:I48) - same reasoning as above.
    ws.conditional_formatting.add(
        "I37:I48",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="2E75B6"),
    )


def find_tech_column_letter(pos_master_header_row):
    for i, h in enumerate(pos_master_header_row):
        if h == "assignedTechnician":
            return get_column_letter(i + 1)
    return "O"


def enhance_pos_master(wb, max_rows=20000):
    """POS_MASTER is the planner's working registry, not a report - a
    manager scanning it needs to spot "which POS need my attention" without
    reading every row. Three visual cues, all pure presentation over fields
    engines already compute (no new business logic, no new field):
      - status badge (Active=green, Closed=grey)
      - neglected-risk highlight on weeksSinceLastVisit, using the SAME
        NEGLECTED_AFTER_WEEKS threshold AdvisorEngine.ts already uses for
        its own NEGLECT_RISK alert - one threshold, read from CONTROL, not
        a second hardcoded copy of the number
      - manual-override highlight (managerOverrideType non-blank) so an
        exception a manager set weeks ago doesn't silently get forgotten
    Plus AutoFilter, since a registry the user can't filter isn't usable as
    a working screen."""
    if "POS_MASTER" not in wb.sheetnames:
        return
    ws = wb["POS_MASTER"]
    header = [c.value for c in ws[1]]
    col = lambda name: get_column_letter(header.index(name) + 1) if name in header else None

    last_col = get_column_letter(ws.max_column or 39)
    ws.auto_filter.ref = f"A1:{last_col}{max_rows}"

    status_col = col("status")
    if status_col:
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{max_rows}",
            FormulaRule(formula=[f'{status_col}2="Active"'], font=Font(color="375623", bold=True)),
        )
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{max_rows}",
            FormulaRule(formula=[f'{status_col}2="Closed"'], font=Font(color="808080")),
        )

    weeks_col = col("weeksSinceLastVisit")
    if weeks_col:
        threshold_formula = 'IFERROR(VLOOKUP("NEGLECTED_AFTER_WEEKS",CONTROL!$A:$B,2,FALSE),26)'
        ws.conditional_formatting.add(
            f"{weeks_col}2:{weeks_col}{max_rows}",
            FormulaRule(
                formula=[f'AND({weeks_col}2<>"",{weeks_col}2>={threshold_formula})'],
                fill=PatternFill("solid", fgColor=WARNING_FILL),
            ),
        )

    override_col = col("managerOverrideType")
    if override_col:
        ws.conditional_formatting.add(
            f"{override_col}2:{override_col}{max_rows}",
            FormulaRule(
                formula=[f'{override_col}2<>""'],
                fill=PatternFill("solid", fgColor="E2D4F0"),
                font=Font(bold=True),
            ),
        )


def build_pos_activate_preview(wb):
    """Live preview block on POS_ACTIVATE_LIST, columns D onward (A:B stay
    pure paste-input, matching BLACKLIST's convention) - product owner,
    2026-07-11: "chci mit moznost... vybrat treba prvnich 500 nebo vsechny"
    needs to actually SEE what "prvních 500" would include before running
    the script, not just trust a number.

    The excluded/eligible check here is a formula approximation of
    core.ts's categoryRule() (exact match > STARTS_1 for categories
    starting with "1" > "*" wildcard > "NORMAL" fallback) - close enough for
    a preview; ActivatePOSEngine.ts/activate_pos_engine.py are the actual
    authority on what gets activated, this table never writes anything."""
    if "POS_ACTIVATE_LIST" not in wb.sheetnames:
        return
    ws = wb["POS_ACTIVATE_LIST"]
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 10

    ws["D1"] = "NÁHLED - AKTUÁLNĚ VYŘAZENÉ POS (podle CATEGORY_RULES)"
    ws["D1"].font = Font(bold=True, size=12, color=NAVY)

    # excludedFlag - a single LET spilling one array over all POS_MASTER
    # rows (elementwise VLOOKUP/IF, no BYROW/LAMBDA - keeping this to the
    # same dynamic-array function set already proven supported elsewhere in
    # this workbook: FILTER/SORT/LET/CHOOSE/IFS/AGGREGATE/SUMPRODUCT).
    ws["G1"] = "excludedFlag"
    ws["G2"] = (
        '=LET(cat,POS_MASTER!$D$2:$D$20000,'
        'status,POS_MASTER!$Q$2:$Q$20000,'
        'override,POS_MASTER!$AH$2:$AH$20000,'
        'exactMatch,IFERROR(VLOOKUP(cat,CATEGORY_RULES!$A:$B,2,FALSE),""),'
        'starsRule,IFERROR(VLOOKUP("STARTS_1",CATEGORY_RULES!$A:$B,2,FALSE),""),'
        'wildcardRule,IFERROR(VLOOKUP("*",CATEGORY_RULES!$A:$B,2,FALSE),"NORMAL"),'
        'resolvedRule,IF(exactMatch<>"",exactMatch,IF(LEFT(cat,1)="1",IF(starsRule<>"",starsRule,wildcardRule),wildcardRule)),'
        'IF((status="Active")*(override<>"FORCE_EXCLUDE")*(override<>"FORCE_INCLUDE")*(resolvedRule="EXCLUDE"),1,0))'
    )
    ws.column_dimensions["G"].hidden = True

    ws["H1"] = "=SUM($G$2:$G$20000)"
    ws.column_dimensions["H"].hidden = True

    ws["D2"] = 'Počet aktuálně vyřazených POS (lze aktivovat):'
    ws["D2"].font = Font(bold=True, size=10)
    ws["E2"] = "=$H$1"
    ws["E2"].font = Font(bold=True, size=14, color=STATUS_WARNING)
    ws["D3"] = 'CONTROL.ACTIVATE_COUNT_BY_PPT (aktuální nastavení):'
    ws["D3"].font = Font(size=10)
    ws["E3"] = '=IFERROR(VLOOKUP("ACTIVATE_COUNT_BY_PPT",CONTROL!$A:$B,2,FALSE),0)'
    ws["E3"].font = Font(bold=True, size=12, color=NAVY)

    ws["D5"] = "POS"
    ws["E5"] = "Název"
    ws["F5"] = "PPT"
    for c in "DEF":
        ws[f"{c}5"].font = HEADER_FONT
        ws[f"{c}5"].fill = HEADER_FILL
    ws["D6"] = (
        '=IFERROR(SORT(FILTER(CHOOSE({1,2,3},POS_MASTER!$A$2:$A$20000,POS_MASTER!$G$2:$G$20000,'
        'POS_MASTER!$P$2:$P$20000),$G$2:$G$20000=1),3,-1),'
        '"— Žádné, nic aktuálně vyřazeného 🎉")'
    )
    ws["D6"].font = Font(size=10)


def apply_all(wb, control_rows):
    """Single entry point called by scaffold_workbook.py after all sheets
    and data are populated."""
    control_values = {}
    for row in control_rows[1:]:
        if row and row[0]:
            control_values[str(row[0]).strip()] = row[1] if len(row) > 1 else ""

    # Build TECHNICIAN_PLAN before the generic styling pass touches
    # anything, and ACTIVITY_PLAN's timeline redesign likewise - that pass
    # decoratively pre-styles empty future rows (up to row 500), which
    # would inflate ws.max_row and confuse row-count-dependent logic.
    if "POS_MASTER" in wb.sheetnames:
        pm_header = [c.value for c in wb["POS_MASTER"][1]]
        tech_col = find_tech_column_letter(pm_header)
    else:
        tech_col = "O"
    if "ACTIVITY_PLAN" in wb.sheetnames:
        redesign_activity_plan(wb, tech_col)
    if "MANAGER_PLAN" in wb.sheetnames:
        build_technician_plan(wb)
    if "DASHBOARD" in wb.sheetnames:
        build_dashboard_template(wb)
    scorecard_built = "TECHNICIAN_PERFORMANCE_LOG" in wb.sheetnames and "TECHNICIAN_TOP_ISSUES" in wb.sheetnames
    if scorecard_built:
        build_technician_scorecard(wb)
    if "TECHNICIAN_PERFORMANCE_SUMMARY" in wb.sheetnames:
        build_performance_sheet(wb)
        build_efficiency_sheet(wb)
        build_manual_sheet(wb)
    if scorecard_built:
        # Must run after build_technician_scorecard() (just above) - reuses
        # its unique-technician spill (TECHNICIAN_SCORECARD!$P$2#) rather
        # than recomputing UNIQUE/FILTER a second time. Gated on the same
        # scorecard_built flag that gated the call above, not on
        # "does a TECHNICIAN_SCORECARD sheet happen to exist" (a stale sheet
        # from an earlier run could satisfy that check without this pass
        # having actually rebuilt it - found during a post-build QA pass,
        # 2026-07-06).
        build_week_dashboard(wb)
    if "POS_MAP_DATA" in wb.sheetnames:
        build_pos_map(wb)
    if "POS_MASTER" in wb.sheetnames:
        apply_banded_rows(wb["POS_MASTER"], 2, 500, wb["POS_MASTER"].max_column or 39)
        enhance_pos_master(wb)

    for sheet_name in list(wb.sheetnames):
        ws = wb[sheet_name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        if sheet_name == "TECHNICIAN_PLAN":
            continue  # build_technician_plan already fully styled it (TOUR PLAN)
        if sheet_name == "DASHBOARD":
            continue  # build_dashboard_template already fully styled it
        if sheet_name == "TECHNICIAN_SCORECARD":
            continue  # build_technician_scorecard already fully styled it
        if sheet_name == "PERFORMANCE":
            continue  # build_performance_sheet already fully styled it (native Table)
        if sheet_name == "WEEK_DASHBOARD":
            continue  # build_week_dashboard already fully styled it
        if sheet_name == "MAP":
            continue  # build_pos_map already fully styled it
        if sheet_name == "ACTIVITY_PLAN":
            # freeze_below=False: redesign_activity_plan() already set
            # freeze_panes="C2" (keep TYPE+ACTIVITY visible while scrolling
            # through a several-months-wide timeline) - this generic pass's
            # default freeze_below=True would silently reset it to "A2"
            # (row-only freeze), which was a real bug: found while testing
            # multi-month timeline orientation, it meant the campaign
            # name/type scrolled off-screen exactly when the timeline was
            # most useful.
            style_header_row(ws, freeze_below=False)
            color_editable_columns(ws, sheet_name)
            add_dropdowns(ws, sheet_name)
            continue
        if sheet_name == "RAW_DATA":
            # RAW_DATA's real header is row 2, not row 1 - see
            # fix_raw_data_layout()'s docstring. Handled entirely there
            # instead of the generic row-1-is-header helpers.
            fix_raw_data_layout(ws)
            continue
        style_header_row(ws)
        color_editable_columns(ws, sheet_name)
        add_dropdowns(ws, sheet_name)
        if sheet_name in IMPORT_UTILITY_SHEETS:
            # POS_STATUS_IMPORT/SALESAPP_IMPORT (RAW_DATA handled separately
            # above) - found missing during a full production review: every
            # other working screen had AutoFilter, these two didn't.
            last_col = get_column_letter(ws.max_column or 16)
            ws.auto_filter.ref = f"A1:{last_col}{max(ws.max_row, 2)}"

    build_pos_activate_preview(wb)

    for sheet_name in list(wb.sheetnames):
        protect_config_sheet(wb[sheet_name], sheet_name)

    add_sheet_purpose_notes(wb)
    build_import_hub(wb, pos_master_tech_col=tech_col)

    build_home(wb, {
        k: v for k, v in control_values.items()
        if k in ("CAMPAIGN_START_WEEK", "CAMPAIGN_LENGTH", "VISITS_PER_WEEK",
                  "TARGET_VISITS_DAY", "YEAR")
    }, pos_master_tech_col=tech_col)
    apply_sheet_order_and_colors(wb)  # re-apply so HOME lands first
    hide_technical_sheets(wb)
    wb.active = 0  # HOME is what the user sees on open
