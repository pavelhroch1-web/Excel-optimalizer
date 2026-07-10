"""One-off: generate a real 5-week tour plan (weeks 29-33) from the manager's
data using the unchanged engines, following the case study:
  week 29     = Dojezd sítě (no campaign -> fill capacity by neglect/score)
  weeks 30-33 = Kampaňový režim (campaigns active -> Smart Hold-back protects
                and covers campaigns), carrying week 29 as locked.

Writes a clean Excel: TOUR_PLAN (the plan) + SOUHRN (per-week/technician +
coverage scorecard). Runs locally (plenty of RAM); the free host's 512 MB is
the only place this doesn't fit.
"""
from __future__ import annotations

import copy
import glob
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "backend"))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from desktop_client.engines import compliance_engine, import_engine, planning_engine
from desktop_client.engines.mock_workbook import MockWorkbook
from desktop_client.engines.core_logic import POSItem, WorkDay, geo_days
from desktop_client.engines.dates_logic import work_days, to_cs_cz_date_string

import config_store
import pipeline
import snapshot_store
import brain

CAP_PER_TECH_WEEK = 40  # ~8 visits/day * 5 days

# Manager overrides: force specific POS to a specific technician + week,
# regardless of the algorithm (e.g. a technician is off, another must cover
# important POS). These POS are taken OUT of the automatic plan and placed
# exactly as instructed.
FORCE_ASSIGN = [
    {
        "technician": "Štolba Jan",
        "week": 30,
        "reason": "OVERRIDE (Štolba za Dvořáka)",
        "pos": [
            "82640301", "82703901", "82932701", "82639901", "82619503", "82611003",
            "82632101", "82640101", "82602601", "82924701", "82639902", "82636701",
            "82922801", "82641101", "82921302", "82637201", "82628001", "82640103",
            "82704101", "82640701", "82815501", "82616507", "82642001", "82642201",
            "82639701",
        ],
    },
]

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCAFFOLD = os.path.join(ROOT, "workbook", "FieldForceOptimizer_V11_scaffold.xlsx")
OUT = os.path.join(ROOT, "TOUR_PLAN_tydny_29-33.xlsx")

PLAN_COLS = ["WEEK", "DATE", "DAY", "TECHNICIAN", "POS", "KATEGORIE", "NAZEV_PROVOZOVNY",
             "ULICE", "CISLO", "MESTO", "OBLAST", "POS_AREA", "PPT", "LOS_ACTIVITY",
             "LOT_ACTIVITY", "REASON"]
DAY_ORDER = {"MON": 1, "TUE": 2, "WED": 3, "THU": 4, "FRI": 5}


def build_base():
    cfg = config_store.load_config_state(SCAFFOLD)
    snap = snapshot_store.load_snapshot(SCAFFOLD)
    for s in ("MANAGER_PLAN", "MANAGER_PLAN_PUBLISHED", "PLAN_LIFECYCLE"):
        snap[s] = [snap[s][0]]
    raw = pipeline.read_workbook_sheet(SCAFFOLD, "RAW_DATA")
    sa = [pipeline.read_workbook_sheet(SCAFFOLD, "SALESAPP_IMPORT")]
    state = pipeline.build_state(cfg, raw, pipeline.merge_salesapp(sa), snapshot=snap)
    # state already carries the accumulated POS_MASTER; run import+compliance so
    # last-visit is current, then it's ready to plan.
    pipeline.run_import_compliance(state)
    return state


def _force_excluded_posids():
    s = set()
    for o in FORCE_ASSIGN:
        s.update(str(p) for p in o["pos"])
    return s


def mark_force_exclude(state, posids):
    """Take POS out of the automatic plan (managerOverrideType=FORCE_EXCLUDE),
    so neither the week-29 cleanup nor the campaign run places them - they are
    injected manually per the override."""
    pm = state["POS_MASTER"]
    h = {n: i for i, n in enumerate(pm[0])}
    ci = h["managerOverrideType"]
    n = 0
    for r in pm[1:]:
        if str(r[h["posId"]]) in posids:
            r[ci] = "FORCE_EXCLUDE"
            n += 1
    return n


def build_forced_rows(state, posids, tech, week, reason):
    """Build MANAGER_PLAN rows for an explicit POS list assigned to `tech` in
    `week`, routed into days by the engine's geo_days. POS are placed as
    instructed regardless of score/filters (manager override)."""
    pm = state["POS_MASTER"]
    h = {n: i for i, n in enumerate(pm[0])}
    by_id = {str(r[h["posId"]]): r for r in pm[1:]}
    days = work_days(2026, week)
    wd = [WorkDay(day=d.day, dateIso=to_cs_cz_date_string(d.date)) for d in days]
    items = []
    for pid in posids:
        r = by_id.get(str(pid))
        if not r:
            print("  ! override POS nenalezen:", pid)
            continue

        def g(col):
            return r[h[col]] if col in h else ""
        it = POSItem(
            pos=str(pid), tech=tech, kategorie=g("category"), market=g("market"),
            classification=g("classification"), nazev=g("nazev"), ulice=g("street"),
            cislo=g("houseNumber"), mesto=g("city"), oblast=g("area"), posArea=g("posArea"),
            ppt=float(g("ppt") or 0), x=float(g("gpsX") or 0), y=float(g("gpsY") or 0),
            weeksSinceLastVisit=g("weeksSinceLastVisit"), forceInclude=True,
            core=False, mandatoryRuleId=None,
        )
        it.score = 1.0
        it.reason = reason + " | "
        items.append(it)
    rows = []
    for pv in geo_days(items, wd):
        p = pv.pos
        rows.append([week, pv.dateIso, pv.day, tech, p.pos, p.kategorie, p.nazev,
                     p.ulice, p.cislo, p.mesto, p.oblast, p.posArea, p.ppt, "", "", p.reason, pv.group])
    return rows


def set_capacity_override(state, tech, year, week, capacity):
    """CAPACITY_OVERRIDE row (technician|year|week -> capacity) so the engine
    plans only that many automatic visits for the technician that week -
    leaving room for forced overrides."""
    co = state.setdefault("CAPACITY_OVERRIDE", [["technician", "year", "week", "capacity"]])
    for row in co[1:]:
        if str(row[0]) == tech and str(row[1]) == str(year) and str(row[2]) == str(week):
            row[3] = capacity
            return
    co.append([tech, year, week, capacity])


def _lock_week(state, week):
    pl = state["PLAN_LIFECYCLE"]
    h = [str(x) for x in pl[0]]
    wi, si = h.index("week"), h.index("status")
    for row in pl[1:]:
        if str(row[wi]) == str(week):
            row[si] = "Published"
            return
    pl.append([2026, week, "Published", "", ""])


def build_week29_cleanup(state, week, cap_per_tech=CAP_PER_TECH_WEEK):
    """Week-29 dojezd: pick the LONGEST-neglected eligible POS per technician
    (rank by weeksSinceLastVisit), then let the engine route them into days.
    The engine decides eligibility (candidates_out = the filtered pool); the
    manager goal is neglect, not PPT."""
    import copy
    # 1) eligible pool from the engine's own filter (candidates_out).
    probe = copy.deepcopy(state)
    pipeline._set_control(probe, "CAMPAIGN_START_WEEK", week)
    pipeline._set_control(probe, "CAMPAIGN_LENGTH", 1)
    cap = []
    planning_engine.run(MockWorkbook(probe), candidates_out=cap)

    # 2) address fields (not in candidates_out) from POS_MASTER
    pm = state["POS_MASTER"]
    h = {n: i for i, n in enumerate(pm[0])}
    addr = {}
    for r in pm[1:]:
        addr[str(r[h["posId"]])] = {
            "ulice": r[h["street"]], "cislo": r[h["houseNumber"]], "mesto": r[h["city"]],
            "oblast": r[h["area"]], "posArea": r[h["posArea"]],
        }

    # 3) rank each technician's eligible POS by neglect (desc), take capacity
    by_tech = {}
    for c in cap:
        by_tech.setdefault(c["tech"], []).append(c)

    days = work_days(2026, week)
    wd = [WorkDay(day=d.day, dateIso=to_cs_cz_date_string(d.date)) for d in days]
    rows = []
    for tech, pool in by_tech.items():
        pool.sort(key=lambda c: (c.get("weeksSinceLastVisit") if c.get("weeksSinceLastVisit") is not None else -1),
                  reverse=True)
        chosen = pool[:cap_per_tech]
        items = []
        for c in chosen:
            a = addr.get(str(c["pos"]), {})
            it = POSItem(
                pos=str(c["pos"]), tech=tech, kategorie=c.get("kategorie", ""), market=c.get("market", ""),
                classification=c.get("classification", ""), nazev=c.get("nazev", ""),
                ulice=a.get("ulice", ""), cislo=a.get("cislo", ""), mesto=a.get("mesto", ""),
                oblast=a.get("oblast", ""), posArea=a.get("posArea", ""),
                ppt=c.get("ppt", 0) or 0, x=c.get("x", 0) or 0, y=c.get("y", 0) or 0,
                weeksSinceLastVisit=c.get("weeksSinceLastVisit"), forceInclude=False,
                core=bool(c.get("core")), mandatoryRuleId=c.get("mandatoryRuleId"),
            )
            wsv = c.get("weeksSinceLastVisit")
            it.score = float(wsv) if wsv is not None else 0.0  # anchor by neglect
            it.reason = (f"MANDATORY ({c['mandatoryRuleId']}) | " if c.get("mandatoryRuleId") else "") + \
                        f"DOJEZD (zanedbáno {wsv if wsv is not None else '?'} týd.) | "
            items.append(it)
        for pv in geo_days(items, wd):
            p = pv.pos
            rows.append([week, pv.dateIso, pv.day, tech, p.pos, p.kategorie, p.nazev,
                         p.ulice, p.cislo, p.mesto, p.oblast, p.posArea, p.ppt, "", "", p.reason, pv.group])
    return rows


def run_planning(state, start, length):
    pipeline._set_control(state, "CAMPAIGN_START_WEEK", start)
    pipeline._set_control(state, "CAMPAIGN_LENGTH", length)
    wb = MockWorkbook(state)
    msg = planning_engine.run(wb)
    state.update(wb.dump())
    return msg


def main():
    print("Building base state (Import + Compliance)…")
    base = build_base()

    # Manager overrides: pull the forced POS out of the automatic plan first.
    forced = _force_excluded_posids()
    if forced:
        print(f"Manager override: {mark_force_exclude(base, forced)} POS vyňato z automatiky (řízené ručně)")

    # WEEK 29 = DOJEZD: visit the LONGEST-neglected POS (rank by
    # weeksSinceLastVisit), NOT the high-PPT ones - those stay fresh for the
    # campaign weeks. The engine still decides who is eligible (filters) and
    # routes the chosen POS into days (geo_days); the manager goal here is
    # "clean up the neglected network", not "grab the strongest".
    week29_rows = build_week29_cleanup(base, 29)
    print(f"Week 29 (Dojezd, nejzanedbanější): {len(week29_rows)} návštěv")

    # Inject week 29 as a locked week so the campaign run carries it and never
    # re-plans its POS; the strong POS remain available for weeks 30-33.
    mp = base["MANAGER_PLAN"]
    for r in week29_rows:
        mp.append(r)
    _lock_week(base, 29)

    # WEEKS 30-33 = KAMPAŇ: campaigns active, but hold-back look-ahead off so
    # the campaign weeks are FULL (strong/campaign POS get visited now, not
    # deferred into emptiness). GECO/CORN cadence stays guaranteed.
    pipeline._set_control(base, "HOLDBACK_LOOKAHEAD_WEEKS", 0)
    # Leave room for forced overrides: cap the technician's automatic picks that
    # week so automatic + override stays near normal capacity.
    for o in FORCE_ASSIGN:
        auto_cap = max(CAP_PER_TECH_WEEK - len(o["pos"]), 0)
        set_capacity_override(base, o["technician"], 2026, o["week"], auto_cap)
        print(f"Kapacita: {o['technician']} t{o['week']} automaticky {auto_cap} + override {len(o['pos'])} = {auto_cap + len(o['pos'])}")
    print("Planning weeks 30-33 (Kampaň, plné týdny)…", run_planning(base, 30, 4))

    # inject manager overrides (forced POS -> specific technician + week)
    mp2 = base["MANAGER_PLAN"]
    for o in FORCE_ASSIGN:
        frows = build_forced_rows(base, o["pos"], o["technician"], o["week"], o["reason"])
        for r in frows:
            mp2.append(r)
        print(f"Override: {len(frows)} POS → {o['technician']}, týden {o['week']}")

    plan = base["MANAGER_PLAN"]
    hdr = plan[0]
    idx = {n: i for i, n in enumerate(hdr)}
    rows = [r for r in plan[1:] if r and r[idx["WEEK"]] not in (None, "")]
    rows.sort(key=lambda r: (str(r[idx["TECHNICIAN"]]), int(r[idx["WEEK"]]),
                             DAY_ORDER.get(str(r[idx["DAY"]]), 9)))
    print(f"Plan rows total: {len(rows)}")

    write_excel(rows, idx, base)
    print("Wrote", OUT, "(", round(os.path.getsize(OUT) / 1e6, 2), "MB )")


def write_excel(rows, idx, state):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TOUR_PLAN"
    head_fill = PatternFill("solid", fgColor="12807C")
    head_font = Font(bold=True, color="FFFFFF")
    ws.append(PLAN_COLS)
    for c in range(1, len(PLAN_COLS) + 1):
        ws.cell(1, c).fill = head_fill
        ws.cell(1, c).font = head_font
    for r in rows:
        ws.append([r[idx[c]] if c in idx else "" for c in PLAN_COLS])
    ws.freeze_panes = "A2"
    widths = {"NAZEV_PROVOZOVNY": 34, "ULICE": 22, "MESTO": 16, "REASON": 40,
              "TECHNICIAN": 18, "LOS_ACTIVITY": 16, "LOT_ACTIVITY": 16}
    for i, c in enumerate(PLAN_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)

    # SOUHRN sheet
    s = wb.create_sheet("SOUHRN")
    from collections import Counter
    per_week = Counter(int(r[idx["WEEK"]]) for r in rows)
    per_tech = Counter(str(r[idx["TECHNICIAN"]]) for r in rows)
    s.append(["Tour plán – týdny 29–33", ""])
    s["A1"].font = Font(bold=True, size=14)
    s.append(["Celkem naplánováno návštěv", len(rows)])
    s.append(["Počet techniků", len(per_tech)])
    s.append([])
    s.append(["Strategie", "T29 = Dojezd (nejzanedbanější POS) · T30–33 = Kampaň (silné POS, plné týdny)"])
    s.append(["", "GECO/CORN cadence garantováno, každý POS max. 1× za 5 týdnů"])
    s.append(["Návštěvy po týdnech", ""])
    s.cell(s.max_row, 1).font = Font(bold=True)
    labels = {29: "Týden 29 (Dojezd – zanedbané)", 30: "Týden 30 (Kampaň)", 31: "Týden 31 (Kampaň)",
              32: "Týden 32 (Kampaň)", 33: "Týden 33 (Kampaň)"}
    for w in (29, 30, 31, 32, 33):
        s.append([labels[w], per_week.get(w, 0)])
    s.append([])
    s.append(["Návštěvy po technicích", ""])
    s.cell(s.max_row, 1).font = Font(bold=True)
    for tech, n in sorted(per_tech.items()):
        s.append([tech, n])
    s.column_dimensions["A"].width = 34
    s.column_dimensions["B"].width = 14

    wb.save(OUT)


if __name__ == "__main__":
    main()
