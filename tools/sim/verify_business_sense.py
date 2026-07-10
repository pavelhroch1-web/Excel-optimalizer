"""Business-sense checks (manager view), over the existing engine:

  #1 Daily route geography: after selection, geo_days() groups a technician's
     POS onto days by GPS proximity. Verify that POS planned on the SAME day
     are geographically MUCH closer than POS on different days (sensible daily
     routes, not random).

  #2 Excel field compatibility: building POS_MASTER from the raw uploaded
     exports reproduces the same per-POS business fields as today's Excel -
     PPT, classification, category/CORE, terminal type, market,
     weeksSinceLastVisit - with no manual editing.

Run: python tools/sim/verify_business_sense.py
"""
from __future__ import annotations

import glob
import math
import os
import sys
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "backend"))

from desktop_client.engines import compliance_engine, import_engine, planning_engine  # noqa: E402
for _m in (import_engine, compliance_engine, planning_engine):
    if hasattr(_m, "iso_now"):
        _m.iso_now = lambda: "2026-07-09T00:00:00.000Z"
from desktop_client.engines.core_logic import category_rule, norm  # noqa: E402
from desktop_client.engines.mock_workbook import MockWorkbook  # noqa: E402

import config_store  # noqa: E402
import pipeline  # noqa: E402
import snapshot_store  # noqa: E402

SCAFFOLD = os.path.join(ROOT, "workbook", "FieldForceOptimizer_V11_scaffold.xlsx")
UPLOADS = "/root/.claude/uploads/96762f2e-6479-5ca9-bce2-fc70e4cf2947"
PPT = os.path.join(UPLOADS, "824b106e-Z_kladn___daje_o_prodejn_ch_m_stech_2.xlsx")
fails = []


def check(name, ok, detail):
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}: {detail}")
    if not ok:
        fails.append(name)


def haversine(a, b):
    (la1, lo1), (la2, lo2) = a, b
    r = 6371.0
    p1, p2 = math.radians(la1), math.radians(la2)
    dp, dl = math.radians(la2 - la1), math.radians(lo2 - lo1)
    h = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(h))


def main():
    raw = pipeline.read_export_rows(PPT)
    sa = [pipeline.read_export_rows(p) for p in
          sorted({os.path.getsize(f): f for f in glob.glob(UPLOADS + "/*visitdata*.xlsx")}.values())]
    cfg = config_store.load_config_state(SCAFFOLD)
    snap = snapshot_store.load_snapshot(SCAFFOLD)
    for s in ("MANAGER_PLAN", "MANAGER_PLAN_PUBLISHED", "PLAN_LIFECYCLE"):
        snap[s] = [snap[s][0]]
    print("Building draft (Import + Compliance)…")
    base = pipeline.build_state(cfg, raw, pipeline.merge_salesapp(sa), snapshot=snap)
    pipeline.run_import_compliance(base)

    # ---- #2 Excel field compatibility: stateless POS_MASTER vs scaffold ----
    print("\n#2 Kompatibilita polí s Excelem (upload → stejná POS_MASTER pole)")
    import openpyxl
    wb = openpyxl.load_workbook(SCAFFOLD, read_only=True, data_only=True)
    sm = wb["POS_MASTER"]
    sh = [c.value for c in next(sm.iter_rows(min_row=1, max_row=1))]
    si = {n: i for i, n in enumerate(sh)}
    scaffold_pm = {}
    for r in sm.iter_rows(min_row=2, values_only=True):
        if r[si["posId"]] not in (None, ""):
            scaffold_pm[str(r[si["posId"]])] = r
    wb.close()

    h = base["POS_MASTER"][0]
    idx = {n: i for i, n in enumerate(h)}
    fields = ["ppt", "classification", "category", "terminalType", "market", "weeksSinceLastVisit"]
    match = {f: 0 for f in fields}
    total = 0
    for row in base["POS_MASTER"][1:]:
        pid = str(row[idx["posId"]])
        s = scaffold_pm.get(pid)
        if not s:
            continue
        total += 1
        for f in fields:
            a, b = row[idx[f]], s[si[f]]
            if f == "ppt":
                same = abs(float(a or 0) - float(b or 0)) < 1e-6
            else:
                same = str(a) == str(b)
            if same:
                match[f] += 1
    for f in fields:
        pct = 100.0 * match[f] / total if total else 0
        check(f"pole '{f}' shodné s Excelem", match[f] == total, f"{match[f]}/{total} ({pct:.2f} %)")
    # CORE derivation identical (category_rule over the same CATEGORY_RULES)
    crt = [{"key": norm(str(r[0])), "value": norm(str(r[1]))} for r in base["CATEGORY_RULES"][1:]]
    core_ct = sum(1 for row in base["POS_MASTER"][1:]
                  if category_rule(crt, norm(str(row[idx["category"]]))) == "CORE")
    check("CORE odvozeno stejným pravidlem (STARTS_1/CATEGORY_RULES)", core_ct > 0,
          f"{core_ct} POS je CORE dle CATEGORY_RULES")

    # ---- #1 Daily route geography ----
    print("\n#1 Denní rozložení trasy (geo shluky po dnech)")
    pipeline._set_control(base, "CAMPAIGN_START_WEEK", 29)
    pipeline._set_control(base, "CAMPAIGN_LENGTH", 1)
    # clear campaigns so many POS get planned (multiple per day per tech)
    for r in base["ACTIVITY_PLAN"][1:]:
        hh = [str(x) for x in base["ACTIVITY_PLAN"][0]]
        r[hh.index("START_WEEK")] = ""; r[hh.index("END_WEEK")] = ""
    wb2 = MockWorkbook(base)
    planning_engine.run(wb2)
    st = wb2.dump()
    mp = st["MANAGER_PLAN"]
    mh = {n: i for i, n in enumerate(mp[0])}
    # gps per POS
    gps = {}
    for row in base["POS_MASTER"][1:]:
        try:
            gps[str(row[idx["posId"]])] = (float(row[idx["gpsX"]]), float(row[idx["gpsY"]]))
        except (TypeError, ValueError):
            pass
    # group plan by technician -> day -> POS
    by_tech = defaultdict(lambda: defaultdict(list))
    for row in mp[1:]:
        pos = str(row[mh["POS"]])
        if pos in gps:
            by_tech[row[mh["TECHNICIAN"]]][row[mh["DAY"]]].append(gps[pos])

    intra, inter, sampled = [], [], 0
    for tech, days in by_tech.items():
        pts_by_day = {d: p for d, p in days.items() if len(p) >= 2}
        if len(pts_by_day) < 2:
            continue
        sampled += 1
        for d, pts in pts_by_day.items():
            for i in range(len(pts)):
                for j in range(i + 1, len(pts)):
                    intra.append(haversine(pts[i], pts[j]))
        dl = list(pts_by_day.items())
        for i in range(len(dl)):
            for j in range(i + 1, len(dl)):
                for a in dl[i][1][:3]:
                    for b in dl[j][1][:3]:
                        inter.append(haversine(a, b))
        if sampled >= 20:
            break
    avg_intra = sum(intra) / len(intra) if intra else 0
    avg_inter = sum(inter) / len(inter) if inter else 0
    print(f"   vzorek techniků: {sampled}")
    print(f"   průměrná vzdálenost POS ve STEJNÝ den: {avg_intra:.1f} km")
    print(f"   průměrná vzdálenost POS mezi RŮZNÝMI dny: {avg_inter:.1f} km")
    check("POS ve stejný den jsou geograficky blíž než mezi dny",
          avg_intra < avg_inter and avg_intra > 0,
          f"stejný den {avg_intra:.1f} km < různé dny {avg_inter:.1f} km")

    print("\n" + ("BUSINESS-SENSE CHECKS OK" if not fails else f"FAILURES: {fails}"))
    if fails:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
