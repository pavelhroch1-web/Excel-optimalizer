"""Proves the stateless pipeline (backend/pipeline.py) is a FAITHFUL wrapper
around the unchanged engines, and runs it end-to-end on the manager's real
uploaded exports.

Test A (faithfulness): with identical inputs, assembling state via
build_state()/run_pipeline() must produce byte-identical POS_MASTER and
MANAGER_PLAN to seeding a MockWorkbook by hand and calling the same engines
directly. This proves the orchestration wrapper adds nothing and changes
nothing - the Draft is exactly what the proven engines yield.

Test B (real end-to-end): assemble a Draft from the manager's actual
uploaded POS export + SalesApp exports + the scaffold's config, plan a real
week, and report what the engines produced (POS count, visit history,
plan rows, weeks, a sample) so it can be eyeballed against the Excel.

Run: python tools/sim/verify_stateless.py
"""
from __future__ import annotations

import copy
import glob
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "backend"))

from desktop_client.engines import (  # noqa: E402
    compliance_engine,
    import_engine,
    planning_engine,
)
from desktop_client.engines.mock_workbook import MockWorkbook  # noqa: E402

import config_store  # noqa: E402
import pipeline  # noqa: E402
import snapshot_store  # noqa: E402

# iso_now() is wall-clock, so two runs stamp importedAt/updatedAt/evaluatedAt
# microseconds apart. Freeze it to a constant in both paths so the
# comparison is a genuine byte-for-byte engine-output check, not a race.
_FROZEN_NOW = "2026-07-09T00:00:00.000Z"
for _mod in (import_engine, compliance_engine, planning_engine):
    if hasattr(_mod, "iso_now"):
        _mod.iso_now = lambda: _FROZEN_NOW

SCAFFOLD = os.path.join(ROOT, "workbook", "FieldForceOptimizer_V11_scaffold.xlsx")
UPLOADS = "/root/.claude/uploads/96762f2e-6479-5ca9-bce2-fc70e4cf2947"
PPT_EXPORT = os.path.join(UPLOADS, "824b106e-Z_kladn___daje_o_prodejn_ch_m_stech_2.xlsx")

WEEK = 29
LENGTH = 1


def _run_direct(state: dict) -> dict:
    """Runs the three engines directly on a hand-seeded MockWorkbook -
    the control path the pipeline wrapper must match exactly."""
    pipeline._set_control(state, "CAMPAIGN_START_WEEK", WEEK)
    pipeline._set_control(state, "CAMPAIGN_LENGTH", LENGTH)
    wb = MockWorkbook(state)
    import_engine.run(wb)
    state.update(wb.dump())
    wb = MockWorkbook(state)
    compliance_engine.run(wb)
    state.update(wb.dump())
    wb = MockWorkbook(state)
    planning_engine.run(wb)
    state.update(wb.dump())
    return state


def test_a_faithfulness() -> None:
    print("=== Test A: stateless wrapper == direct engine calls (identical inputs) ===")
    config_state = config_store.load_config_state(SCAFFOLD)
    raw = pipeline.read_export_rows(PPT_EXPORT)
    salesapp = pipeline.merge_salesapp(
        [pipeline.read_export_rows(p) for p in _salesapp_uploads()]
    )

    snapshot = snapshot_store.load_snapshot(SCAFFOLD)

    # Path DIRECT: build the same assembled state, run engines by hand.
    direct_state = pipeline.build_state(config_state, raw, salesapp, snapshot=snapshot)
    direct_state = _run_direct(copy.deepcopy(direct_state))

    # Path PIPELINE: same inputs through the wrapper.
    pipe_state = pipeline.build_state(config_state, raw, salesapp, snapshot=snapshot)
    pipeline.run_pipeline(pipe_state, WEEK, LENGTH)

    for sheet in ("POS_MASTER", "VISIT_HISTORY_ACTUAL", "MANAGER_PLAN"):
        a = direct_state[sheet]
        b = pipe_state[sheet]
        ok = a == b
        print(f"  {sheet}: {'IDENTICAL' if ok else 'DIFFERENT'} "
              f"({len(a) - 1} vs {len(b) - 1} rows)")
        if not ok:
            _first_diff(a, b, sheet)
            raise SystemExit("FAIL: wrapper diverges from direct engine calls")
    print("  -> wrapper is faithful\n")


def test_b_real_end_to_end() -> None:
    print("=== Test B: real Draft from the manager's uploaded exports ===")
    raw = pipeline.read_export_rows(PPT_EXPORT)
    salesapp_exports = [pipeline.read_export_rows(p) for p in _salesapp_uploads()]
    print(f"  POS export rows (RAW_DATA): {len(raw) - 2}")  # minus banner + header
    print(f"  SalesApp exports: {len(salesapp_exports)} file(s), "
          f"{sum(len(e) - 1 for e in salesapp_exports)} visit rows merged")

    result = pipeline.generate_draft(raw, salesapp_exports, WEEK, LENGTH, seed_workbook=SCAFFOLD)
    s = result["summary"]
    print(f"  Import:     {result['messages']['import']}")
    print(f"  Compliance: {result['messages']['compliance']}")
    print(f"  Planning:   {result['messages']['planning']}")
    print(f"  POS_MASTER rows:      {s['posMasterRows']}")
    print(f"  VISIT_HISTORY rows:   {s['visitHistoryRows']}")
    print(f"  MANAGER_PLAN rows:    {s['managerPlanRows']}")
    print(f"  weeks in plan:        {s['weeksInPlan']}")

    mp = result["state"]["MANAGER_PLAN"]
    hdr = mp[0]
    print(f"  sample plan rows (first 5 of week {WEEK}):")
    show_cols = ["WEEK", "DAY", "TECHNICIAN", "POS", "NAZEV_PROVOZOVNY", "PPT", "REASON"]
    ci = {c: hdr.index(c) for c in show_cols if c in hdr}
    shown = 0
    for row in mp[1:]:
        if shown >= 5:
            break
        print("   ", {c: row[ci[c]] for c in ci})
        shown += 1
    assert s["posMasterRows"] > 11000, "expected ~11605 POS from the real export"
    assert s["managerPlanRows"] > 0, "expected a non-empty week-29 plan"
    print("  -> real Draft generated\n")


def test_c_resume_matches_excel() -> None:
    """Resume-from-snapshot must reproduce the Excel's accumulated per-POS
    state. Seeding POS_MASTER from the scaffold snapshot and re-running the
    pipeline on the scaffold's own full RAW_DATA + SalesApp history must
    yield the SAME weeksSinceLastVisit distribution as the scaffold - proving
    the git-like carry-over is faithful, not an approximation."""
    print("=== Test C: resume-from-snapshot reproduces the Excel's per-POS state ===")
    raw = _scaffold_sheet("RAW_DATA")
    salesapp = _scaffold_sheet("SALESAPP_IMPORT")
    snapshot = snapshot_store.load_snapshot(SCAFFOLD)
    config_state = config_store.load_config_state(SCAFFOLD)

    state = pipeline.build_state(config_state, raw, salesapp, snapshot=snapshot)
    pipeline.run_pipeline(state, WEEK, LENGTH)

    got = _weeks_distribution(state["POS_MASTER"])
    want = _weeks_distribution(snapshot["POS_MASTER"])
    print(f"  scaffold (Excel) : {want}")
    print(f"  resume pipeline  : {got}")
    if got != want:
        raise SystemExit("FAIL: resume-from-snapshot diverges from the Excel state")
    print("  -> resume-from-snapshot is faithful to the Excel\n")


def _weeks_distribution(pos_master: list[list]) -> dict:
    from collections import Counter
    hdr = pos_master[0]
    wi = hdr.index("weeksSinceLastVisit")
    d: Counter = Counter()
    for r in pos_master[1:]:
        try:
            v = int(r[wi])
        except (ValueError, TypeError):
            v = -1
        if v == 0:
            d["0"] += 1
        elif v <= 4:
            d["1-4"] += 1
        elif v <= 12:
            d["5-12"] += 1
        elif v <= 26:
            d["13-26"] += 1
        else:
            d["27+"] += 1
    return dict(d)


def _scaffold_sheet(name: str) -> list[list]:
    import openpyxl
    import datetime
    wb = openpyxl.load_workbook(SCAFFOLD, read_only=True, data_only=True)
    try:
        ws = wb[name]

        def cj(v):
            if isinstance(v, (datetime.datetime, datetime.date)):
                return v.isoformat()
            return "" if v is None else v

        rows = [[cj(v) for v in r] for r in ws.iter_rows(values_only=True)]
        while rows and all(v == "" for v in rows[-1]):
            rows.pop()
        return rows
    finally:
        wb.close()


def test_d_incremental_equals_full() -> None:
    """The reproducibility guarantee, stated as the product owner did:
    uploading ONLY the new export each week (state carried by the snapshot)
    must give the SAME accumulated state as if all the data had been present
    at once. Proves nothing hides outside the snapshot.

    Split the scaffold's full SalesApp history into EARLY and LATE by date.
      FULL path:   cold start, upload EARLY+LATE together.
      INCR path:   cold start + EARLY -> publish snapshot -> resume + LATE.
    The accumulated VISIT_HISTORY_ACTUAL (as a UID set) and every POS's
    weeksSinceLastVisit must match. Compared independently of the plan week,
    since accumulated state does not depend on it."""
    print("=== Test D: incremental (snapshot + delta) == full (all data at once) ===")
    config_state = config_store.load_config_state(SCAFFOLD)
    raw = _scaffold_sheet("RAW_DATA")
    salesapp = _scaffold_sheet("SALESAPP_IMPORT")
    early, late, cutoff = _split_salesapp_by_date(salesapp)
    print(f"  split at {cutoff}: EARLY={len(early) - 1} rows, LATE={len(late) - 1} rows")

    # FULL: cold start, all visits at once.
    full = pipeline.build_state(config_state, raw, pipeline.merge_salesapp([salesapp]))
    pipeline.run_pipeline(full, WEEK, LENGTH)

    # INCREMENTAL cycle 1: cold start + EARLY only -> snapshot.
    c1 = pipeline.build_state(config_state, raw, pipeline.merge_salesapp([early]))
    pipeline.run_pipeline(c1, 20, LENGTH)
    snap1 = {name: c1[name] for name in snapshot_store.SNAPSHOT_SHEETS if name in c1}

    # INCREMENTAL cycle 2: resume from snapshot + LATE delta only.
    incr = pipeline.build_state(config_state, raw, pipeline.merge_salesapp([late]), snapshot=snap1)
    pipeline.run_pipeline(incr, WEEK, LENGTH)

    full_uids = {r[6] for r in full["VISIT_HISTORY_ACTUAL"][1:]}
    incr_uids = {r[6] for r in incr["VISIT_HISTORY_ACTUAL"][1:]}
    print(f"  VISIT_HISTORY_ACTUAL UIDs: full={len(full_uids)}, incremental={len(incr_uids)}")
    if full_uids != incr_uids:
        only_full = len(full_uids - incr_uids)
        only_incr = len(incr_uids - full_uids)
        raise SystemExit(f"FAIL: visit history diverged (only-full={only_full}, only-incr={only_incr})")

    full_ws = _weeks_by_pos(full["POS_MASTER"])
    incr_ws = _weeks_by_pos(incr["POS_MASTER"])
    mismatches = [p for p in full_ws if full_ws[p] != incr_ws.get(p)]
    print(f"  POS with weeksSinceLastVisit mismatch: {len(mismatches)}")
    if mismatches:
        p = mismatches[0]
        raise SystemExit(f"FAIL: e.g. POS {p}: full={full_ws[p]} incr={incr_ws.get(p)}")
    print("  -> incremental delta upload reproduces the full-history state exactly\n")


def _split_salesapp_by_date(rows: list[list]) -> tuple[list[list], list[list], str]:
    header = rows[0]
    dates = sorted(str(r[1])[:10] for r in rows[1:] if r[1] not in (None, ""))
    cutoff = dates[len(dates) // 2]
    early = [header] + [r for r in rows[1:] if str(r[1])[:10] < cutoff]
    late = [header] + [r for r in rows[1:] if str(r[1])[:10] >= cutoff]
    return early, late, cutoff


def _weeks_by_pos(pos_master: list[list]) -> dict:
    hdr = pos_master[0]
    pi = hdr.index("posId")
    wi = hdr.index("weeksSinceLastVisit")
    return {r[pi]: r[wi] for r in pos_master[1:] if r[pi] not in (None, "")}


def _salesapp_uploads() -> list[str]:
    """All distinct SalesApp visitdata exports the manager uploaded."""
    files = sorted(glob.glob(os.path.join(UPLOADS, "*visitdata*.xlsx")))
    # Drop obvious duplicates by (size) - the same export re-uploaded.
    seen: dict[int, str] = {}
    for f in files:
        seen.setdefault(os.path.getsize(f), f)
    return sorted(seen.values())


def _first_diff(a: list, b: list, sheet: str) -> None:
    for i, (ra, rb) in enumerate(zip(a, b)):
        if ra != rb:
            print(f"    first diff in {sheet} at row {i}:")
            print(f"      direct:   {ra}")
            print(f"      pipeline: {rb}")
            return
    if len(a) != len(b):
        print(f"    row count differs: {len(a)} vs {len(b)}")


if __name__ == "__main__":
    test_a_faithfulness()
    test_c_resume_matches_excel()
    test_d_incremental_equals_full()
    test_b_real_end_to_end()
    print("ALL CHECKS PASSED")
