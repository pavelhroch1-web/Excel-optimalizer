"""End-to-end check of the production live workflow WITHOUT GitHub: exercises
the exact code the endpoints call (pipeline + state_xlsx + plan_io +
candidates + run_publish) against local temp files, so the full
upload -> generate -> candidates -> edit -> publish -> resume chain is
proven before deploying. Only the thin gh.py/store.py I/O is not exercised
here (it just moves these same files to/from GitHub).

Run: python tools/sim/verify_live_workflow.py
"""
from __future__ import annotations

import os
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "backend"))

from desktop_client.engines import compliance_engine, import_engine, planning_engine  # noqa: E402
for _m in (import_engine, compliance_engine, planning_engine):
    if hasattr(_m, "iso_now"):
        _m.iso_now = lambda: "2026-07-09T00:00:00.000Z"

import candidates as candidates_mod  # noqa: E402
import pipeline  # noqa: E402
import plan_io  # noqa: E402
import state_xlsx  # noqa: E402

SCAFFOLD = os.path.join(ROOT, "workbook", "FieldForceOptimizer_V11_scaffold.xlsx")
UPLOADS = "/root/.claude/uploads/96762f2e-6479-5ca9-bce2-fc70e4cf2947"
PPT = os.path.join(UPLOADS, "824b106e-Z_kladn___daje_o_prodejn_ch_m_stech_2.xlsx")
WEEK = 33


def _tmp():
    fd, p = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    return p


def _salesapp():
    import glob
    return sorted({os.path.getsize(f): f for f in glob.glob(UPLOADS + "/*visitdata*.xlsx")}.values())


def main() -> None:
    draft = _tmp()
    snap = _tmp()
    try:
        # 1) UPLOAD: resume from scaffold snapshot + fresh exports -> draft.xlsx
        print("1) upload")
        raw = pipeline.read_export_rows(PPT)
        sa = [pipeline.read_export_rows(p) for p in _salesapp()]
        res = pipeline.build_upload_draft(raw, sa, seed_workbook=SCAFFOLD)
        # The shipped scaffold already has a large catch-up plan published for
        # weeks 29-32, so a normal week right after it legitimately has no
        # eligible candidates (min-gap not elapsed) - correct engine behaviour,
        # but it makes this PLUMBING test non-deterministic. Simulate a planner
        # with no prior published plan so Generate yields a real fresh week and
        # the upload->generate->edit->publish->resume chain is exercised
        # deterministically. (Engine fidelity itself is proven separately in
        # verify_stateless.py against the real scaffold state.)
        for s in ("MANAGER_PLAN", "MANAGER_PLAN_PUBLISHED", "PLAN_LIFECYCLE"):
            res["state"][s] = [res["state"][s][0]]
        state_xlsx.save_state(res["state"], draft)
        print(f"   {res['messages']['import']}")
        print(f"   {res['messages']['compliance']}")
        assert os.path.getsize(draft) > 0

        # 3) CANDIDATES (read-only) on the draft file
        print("3) candidates")
        cand = candidates_mod.list_candidates(draft, WEEK)
        print(f"   week {WEEK}: {cand['total']} candidates, {cand['selected']} selected")
        assert cand["total"] > 0

        # 2) GENERATE: run Planning on the draft, save back
        print("2) generate")
        st = state_xlsx.load_state(draft)
        msg = pipeline.run_planning(st, WEEK, 1)
        state_xlsx.save_state(st, draft)
        print(f"   {msg['planning']}")
        draft_rows = plan_io.read_enriched_draft(draft)
        wk_rows = [r for r in draft_rows if r["WEEK"] == WEEK]
        print(f"   MANAGER_PLAN week {WEEK}: {len(wk_rows)} rows")
        assert len(wk_rows) > 0, "expected a non-empty new week"

        # 4) EDIT: remove one POS from the new week, confirm it's gone
        print("4) edit (remove one POS)")
        victim = wk_rows[0]
        removed = plan_io.remove_pos(draft, WEEK, str(victim["POS"]), victim["TECHNICIAN"])
        assert removed == 1, f"expected to remove 1, removed {removed}"
        after = [r for r in plan_io.read_enriched_draft(draft) if r["WEEK"] == WEEK]
        assert len(after) == len(wk_rows) - 1
        print(f"   removed POS {victim['POS']} -> week now {len(after)} rows")

        # 5) PUBLISH: freeze the lowest draft week into a snapshot
        print("5) publish")
        st = state_xlsx.load_state(draft)
        pub_before = len(st.get("MANAGER_PLAN_PUBLISHED", [])) - 1
        pres = pipeline.run_publish(st)
        state_xlsx.save_state(st, snap)
        state_xlsx.save_state(st, draft)  # draft now reflects the locked state

        # locked-week guard: the just-published week must now refuse edits
        try:
            plan_io.remove_pos(draft, WEEK, str(after[0]["POS"]), after[0]["TECHNICIAN"])
            raise SystemExit("FAIL: editing the just-published week was allowed")
        except ValueError:
            print(f"   locked-week guard OK (published week {WEEK} refused)")
        print(f"   {pres['message']}")
        print(f"   publishedWeeks={pres['publishedWeeks']}")
        assert pres["publishedWeeks"], "expected a week to be published"
        pub_after = len(st.get("MANAGER_PLAN_PUBLISHED", [])) - 1
        assert pub_after > pub_before, "MANAGER_PLAN_PUBLISHED should grow"
        print(f"   MANAGER_PLAN_PUBLISHED: {pub_before} -> {pub_after} rows")

        # 7) DOWNLOAD: the published plan sheet is extractable
        import openpyxl
        wb = openpyxl.load_workbook(snap, read_only=True, data_only=True)
        assert "MANAGER_PLAN_PUBLISHED" in wb.sheetnames
        pub_week_rows = sum(
            1 for r in wb["MANAGER_PLAN_PUBLISHED"].iter_rows(min_row=2, values_only=True)
            if r and r[0] == WEEK
        )
        wb.close()
        print(f"7) download: MANAGER_PLAN_PUBLISHED has {pub_week_rows} rows for week {WEEK}")
        assert pub_week_rows > 0

        # RESUME: next upload from the new snapshot must treat WEEK as locked
        print("resume: new upload from published snapshot locks week", WEEK)
        res2 = pipeline.build_upload_draft(raw, sa, seed_workbook=snap)
        st2 = res2["state"]
        msg2 = pipeline.run_planning(st2, WEEK, 1)
        print(f"   {msg2['planning']}")
        assert "0 new planned visits" in msg2["planning"] or "carried over" in msg2["planning"], \
            "published week should not be regenerated"
        print("   published week is immutable on resume OK")

        print("\nLIVE WORKFLOW OK")
    finally:
        for p in (draft, snap):
            if os.path.exists(p):
                os.remove(p)


if __name__ == "__main__":
    main()
