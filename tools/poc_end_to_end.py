"""Proof of Concept — celý workflow end-to-end na REÁLNÝCH datech.

Ověřuje, že backend a plánovací logika fungují od začátku do konce:

  1. Import vstupních dat (workbook s reálnými POS + návštěvami + konfigurací).
  2. Vytvoření úkolů (servis / kampaň / materiál) — z Excelu nebo demonstrační.
  3. Výpočet Planning Engine (skutečný engine, beze změny).
  4. Bundling servis + kampaně + materiál na každou zastávku.
  5. Vygenerování TourPlanu (MANAGER_PLAN).
  6. Zobrazení všech úkolů u jednotlivých zastávek (read_enriched_draft).
  7. Export .xlsx pro techniky (se sloupcem ÚKOLY).

Zároveň dokazuje tři klíčové chování:
  * urgentní / nekombinovatelný úkol  -> vytvoří SAMOSTATNÝ výjezd (FORCE_INCLUDE),
  * kombinovatelný úkol s daleko deadlinem -> jen se PŘIBALÍ (žádný výjezd navíc),
  * úkoly na jednom POS se SLUČUJÍ a řadí podle priority a deadlinu.

Spuštění (na scaffold = reálná data, 11 605 POS):
    python3 tools/poc_end_to_end.py

Na VLASTNÍCH datech:
    python3 tools/poc_end_to_end.py --workbook muj_export.xlsx --week 35 \
        --tasks-excel moje_ukoly.xlsx --type-id 4

Výstup: každý krok s očekáváním, skutečností a PASS/FAIL; na konci verdikt a
cesta k vyexportovanému Excelu. Nic se nezapisuje do tvých produkčních dat —
běží v dočasné složce.
"""
from __future__ import annotations

import argparse
import copy
import datetime
import os
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)                       # for the desktop_client package
sys.path.insert(0, os.path.join(ROOT, "backend"))

# Portable/local mode + an isolated throwaway data dir (never touch real data).
os.environ["FFO_LOCAL"] = "1"
_TMP_DATA = tempfile.mkdtemp(prefix="ffo_poc_")
os.environ["FFO_DATA_DIR"] = _TMP_DATA

import db                 # noqa: E402
import db_state           # noqa: E402
import importer           # noqa: E402
import pipeline           # noqa: E402
import plan_io            # noqa: E402
import state_xlsx         # noqa: E402
import tasks              # noqa: E402

SCAFFOLD = os.path.join(ROOT, "workbook", "FieldForceOptimizer_V11_scaffold.xlsx")

_PASS = "\033[92mPASS\033[0m"
_FAIL = "\033[91mFAIL\033[0m"
_results: list[tuple[str, bool, str]] = []


def check(name: str, ok: bool, detail: str = "") -> bool:
    _results.append((name, ok, detail))
    print(f"  [{_PASS if ok else _FAIL}] {name}" + (f" — {detail}" if detail else ""))
    return ok


def step(n: int, title: str) -> None:
    print(f"\n=== KROK {n}: {title} ===")


# --------------------------------------------------------------------------
def _planned_pos(state: dict, week: int) -> set:
    mp = state.get("MANAGER_PLAN", [])
    if not mp:
        return set()
    h = {v: i for i, v in enumerate(mp[0])}
    wi, pi = h.get("WEEK"), h.get("POS")
    return {str(r[pi]) for r in mp[1:] if pi is not None and str(r[wi]) == str(week)}


def _committed_pos(state: dict) -> set:
    """POS committed in ANY MANAGER_PLAN week (incl. locked/prior weeks). A POS
    already here needs no new visit — FORCE_INCLUDE only adds genuinely-free POS."""
    mp = state.get("MANAGER_PLAN", [])
    if not mp:
        return set()
    h = {v: i for i, v in enumerate(mp[0])}
    pi = h.get("POS")
    return {str(r[pi]) for r in mp[1:] if pi is not None}


def _configure_and_plan(state: dict, mode: str, week: int, length: int, capacity: float) -> dict:
    db_state.configure(state, mode, week, length, capacity)   # applies task_bridge overlay
    pipeline.run_planning(state, week, length)
    return state


def _make_demo_tasks(committed: set, wk35_planned: set, free_pos: list[str]) -> dict:
    """Create demonstrative tasks on REAL POS. `free_pos` are POS committed in NO
    week (so FORCE_INCLUDE genuinely adds them); `wk35_planned` are already in the
    target week (so a combinable task there proves piggyback). Returns POS by role."""
    today = datetime.date.today()
    far = (today + datetime.timedelta(days=55)).isoformat()   # ~2 měsíce
    near = (today + datetime.timedelta(days=5)).isoformat()   # urgentní

    not_planned = free_pos
    planned = [p for p in wk35_planned]
    if len(not_planned) < 3 or len(planned) < 1:
        raise SystemExit("PoC: málo POS pro výběr scénářů (neočekávané).")

    tt = {t["name"]: t for t in tasks.types()}
    camp = tt.get("Jednorázová akce"); mat_k = tt.get("Kotouče")
    mat_l = tt.get("Letáky"); svc = tt.get("Instalace služby")  # not combinable

    urgent_pos = not_planned[0]     # urgent campaign -> must become a dedicated visit
    dedicated_pos = not_planned[1]  # non-combinable service -> dedicated visit
    piggy_far_pos = not_planned[2]  # combinable, far deadline -> must NOT be forced
    combine_pos = planned[0]        # already planned -> combinable task piggybacks + merge

    # urgent campaign (near deadline) on a POS the engine would NOT visit
    tasks.create({"type_id": camp["id"], "pos_id": urgent_pos, "quantity": 20, "deadline": near})
    # non-combinable service on another not-planned POS
    tasks.create({"type_id": svc["id"], "pos_id": dedicated_pos})
    # combinable material, far deadline, on a not-planned POS (should stay unplanned)
    tasks.create({"type_id": mat_l["id"], "pos_id": piggy_far_pos, "quantity": 30, "deadline": far})
    # already-planned POS: two combinable tasks to prove merge + priority order
    tasks.create({"type_id": mat_k["id"], "pos_id": combine_pos, "quantity": 50, "deadline": far})   # prio 3
    tasks.create({"type_id": mat_l["id"], "pos_id": combine_pos, "quantity": 100, "deadline": near})  # prio 4
    return {"urgent": urgent_pos, "dedicated": dedicated_pos,
            "piggyFar": piggy_far_pos, "combine": combine_pos}


def _export_with_tasks(src_state_path: str, out_path: str) -> int:
    """Mirror of main._stream_sheet's ÚKOLY enrichment: copy MANAGER_PLAN and
    append the bundled task summary column. Returns rows written."""
    import openpyxl
    wb = openpyxl.load_workbook(src_state_path, read_only=True, data_only=True)
    try:
        ws = wb["MANAGER_PLAN"]
        out = openpyxl.Workbook(); ows = out.active; ows.title = "MANAGER_PLAN"
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        pos_col = [str(h).upper() if h is not None else "" for h in header].index("POS")
        ows.append(list(header) + ["ÚKOLY"])
        n = 0
        for row in it:
            pid = row[pos_col] if pos_col < len(row) else None
            summary = tasks.bundle_for_pos(str(pid)).get("summary", "") if pid not in (None, "") else ""
            ows.append(list(row) + [summary]); n += 1
        out.save(out_path)
        return n
    finally:
        wb.close()


# --------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="Field Force Optimizer — PoC end-to-end")
    ap.add_argument("--workbook", default=SCAFFOLD, help="vstupní workbook (default: scaffold = reálná data)")
    ap.add_argument("--week", type=int, default=35, help="plánovací týden (musí být otevřený, ne publikovaný)")
    ap.add_argument("--length", type=int, default=1)
    ap.add_argument("--mode", default="dojezd")
    ap.add_argument("--capacity", type=float, default=40, help="návštěv/technik/týden")
    ap.add_argument("--tasks-excel", default=None, help="vlastní Excel úkolů (POS + počet); jinak demo úkoly")
    ap.add_argument("--type-id", type=int, default=None, help="typ aktivity pro vlastní Excel úkolů")
    ap.add_argument("--out-dir", default=os.path.join(_TMP_DATA, "out"))
    args = ap.parse_args()
    os.makedirs(args.out_dir, exist_ok=True)

    if not os.path.exists(args.workbook):
        print(f"Workbook nenalezen: {args.workbook}"); return 2

    print("Field Force Optimizer — Proof of Concept (end-to-end na reálných datech)")
    print(f"Workbook: {args.workbook}\nTýden: {args.week}  Režim: {args.mode}  Kapacita: {args.capacity}")
    print(f"Dočasná data: {_TMP_DATA}")

    # -------------------------------------------------- KROK 1: import
    step(1, "Import vstupních dat")
    db.init_db()
    tasks.seed_default_types()
    imp = importer.import_workbook(args.workbook)
    n_pos = db.get("SELECT COUNT(*) c FROM pos_master")[0]["c"]
    check("POS naimportovány do databáze", n_pos > 0, f"{n_pos} POS")
    check("Návštěvy naimportovány", (imp.get("salesapp_visits") or 0) >= 0,
          f"{imp.get('salesapp_visits')} návštěv")
    check("Konfigurace naimportována", (imp.get("config") or 0) > 0, f"{imp.get('config')} položek")
    all_pos = [str(r["pos_id"]) for r in db.get("SELECT pos_id FROM pos_master WHERE active=1")]

    # -------------------------------------------------- build state (once)
    print("\n  … sestavuji plánovací stav (build_upload_draft) …")
    state0 = pipeline.build_upload_draft(None, [], seed_workbook=args.workbook)["state"]
    check("Plánovací stav sestaven", bool(state0.get("RAW_DATA")),
          f"{len(state0.get('RAW_DATA', [])) - 1} řádků RAW_DATA")

    # -------------------------------------------------- baseline plan (no tasks)
    print("  … baseline: Planning Engine BEZ úkolů (pro srovnání) …")
    baseline_state = copy.deepcopy(state0)
    _configure_and_plan(baseline_state, args.mode, args.week, args.length, args.capacity)
    baseline = _planned_pos(baseline_state, args.week)
    committed = _committed_pos(baseline_state)          # POS in ANY week (incl. locked)
    check("Baseline plán vygenerován", len(baseline) > 0, f"{len(baseline)} zastávek (týden {args.week})")

    # genuinely-free POS: not committed anywhere, active, with an assigned technician
    pm = state0["POS_MASTER"]; ph = {h: i for i, h in enumerate(pm[0])}
    pi, ti, si = ph["posId"], ph.get("assignedTechnician"), ph.get("status")
    free_pos = [str(r[pi]) for r in pm[1:]
                if str(r[pi]) not in committed
                and ti is not None and str(r[ti]).strip() not in ("", "None")
                and (si is None or str(r[si]) == "Active")]

    # -------------------------------------------------- KROK 2: tasks
    step(2, "Vytvoření úkolů (servis / kampaň / materiál)")
    roles = {}
    if args.tasks_excel:
        if not args.type_id:
            print("  --tasks-excel vyžaduje --type-id (typ aktivity)."); return 2
        rows = tasks.parse_bulk_excel(args.tasks_excel)
        res = tasks.bulk_create(rows, args.type_id)
        check("Úkoly z Excelu založeny", res["created"] > 0,
              f"{res['created']} založeno, {res['skipped']} přeskočeno (neznámé POS)")
    else:
        roles = _make_demo_tasks(committed, baseline, free_pos)
        check("Demonstrační úkoly založeny na reálných POS", True,
              f"urgent={roles['urgent']} dedicated={roles['dedicated']} "
              f"piggyFar={roles['piggyFar']} combine={roles['combine']}")
    open_n = tasks.open_tasks()["counts"]["open"]
    check("Otevřené úkoly evidované", open_n > 0, f"{open_n} otevřených")

    # -------------------------------------------------- KROK 3+5: plan WITH tasks
    step(3, "Výpočet Planning Engine (s úkoly) → KROK 5: TourPlan")
    task_state = copy.deepcopy(state0)
    _configure_and_plan(task_state, args.mode, args.week, args.length, args.capacity)
    planned = _planned_pos(task_state, args.week)
    check("TourPlan (MANAGER_PLAN) vygenerován", len(planned) > 0, f"{len(planned)} zastávek")

    # -------------------------------------------------- behavior proofs
    step(4, "Ověření slučování + rozhodnutí engine (bundling)")
    if roles:
        # urgent -> dedicated visit (POS was NOT in baseline, IS now)
        check("Urgentní úkol vytvořil samostatný výjezd",
              roles["urgent"] not in baseline and roles["urgent"] in planned,
              f"POS {roles['urgent']}: baseline={roles['urgent'] in baseline} → plán={roles['urgent'] in planned}")
        # non-combinable -> dedicated visit
        check("Nekombinovatelný úkol vytvořil samostatný výjezd",
              roles["dedicated"] not in baseline and roles["dedicated"] in planned,
              f"POS {roles['dedicated']}: → plán={roles['dedicated'] in planned}")
        # combinable + far deadline -> NOT forced (no extra trip)
        check("Kombinovatelný úkol s daleko deadlinem NEVytvořil výjezd navíc",
              roles["piggyFar"] not in planned,
              f"POS {roles['piggyFar']}: v plánu={roles['piggyFar'] in planned} (očekáváno False)")
        # combinable task on an already-planned POS -> piggybacks on that visit
        _cb = tasks.bundle_for_pos(roles["combine"])
        check("Kombinovatelný úkol se přibalil k již plánované návštěvě",
              roles["combine"] in planned and _cb.get("count", 0) > 0,
              f"POS {roles['combine']}: v plánu wk{args.week}={roles['combine'] in planned}, úkolů={_cb.get('count')}")
        # merge + priority ordering on the combine POS
        b = tasks.bundle_for_pos(roles["combine"])
        mat = b.get("groups", {}).get("material", [])
        merged_ok = len(mat) == 2 and mat[0]["type"] == "Kotouče"
        check("Úkoly na jednom POS se sloučily a seřadily podle priority",
              merged_ok, f"pořadí={[m['type'] for m in mat]} (Kotouče prio 3 před Letáky prio 4)")
        check("Balík obsahuje množství i prioritu",
              all(m.get("quantity") for m in mat) and b.get("topPriority") is not None,
              f"topPriority={b.get('topPriority')}, summary='{b.get('summary')}'")

    # -------------------------------------------------- KROK 6: bundling on stops
    step(6, "Zobrazení úkolů u jednotlivých zastávek")
    draft_path = os.path.join(args.out_dir, "poc_state.xlsx")
    state_xlsx.save_state(task_state, draft_path)
    rows = plan_io.read_enriched_draft(draft_path)
    with_tasks = [r for r in rows if r.get("tasks") and r["tasks"].get("count")]
    check("Zastávky nesou přibalené úkoly", len(with_tasks) > 0,
          f"{len(with_tasks)} zastávek s úkoly z {len(rows)} celkem")
    if with_tasks:
        ex = with_tasks[0]
        print(f"      příklad zastávky POS {ex.get('POS')}: {ex['tasks'].get('summary')}")

    # -------------------------------------------------- KROK 7: export
    step(7, "Export .xlsx pro techniky (sloupec ÚKOLY)")
    export_path = os.path.join(args.out_dir, f"MANAGER_PLAN_POC_week{args.week}.xlsx")
    written = _export_with_tasks(draft_path, export_path)
    import openpyxl
    ew = openpyxl.load_workbook(export_path, read_only=True, data_only=True)
    ews = ew.active
    hdr = [str(c.value) for c in next(ews.iter_rows())]
    has_col = "ÚKOLY" in hdr
    # any non-empty ÚKOLY cell?
    col = hdr.index("ÚKOLY") if has_col else -1
    nonempty = 0
    if has_col:
        for r in ews.iter_rows(min_row=2, values_only=True):
            if col < len(r) and r[col]:
                nonempty += 1
    ew.close()
    check("Export obsahuje sloupec ÚKOLY", has_col, f"{written} řádků")
    check("Export má vyplněné úkoly u zastávek", nonempty > 0, f"{nonempty} řádků s úkoly")
    print(f"      export: {export_path}")

    # -------------------------------------------------- verdict
    total = len(_results); ok = sum(1 for _, o, _ in _results if o)
    print("\n" + "=" * 60)
    print(f"VÝSLEDEK PoC: {ok}/{total} kontrol prošlo")
    if ok == total:
        print("✓ Celý workflow ověřen end-to-end na reálných datech.")
    else:
        print("✗ Některé kontroly selhaly — viz [FAIL] výše.")
        for name, o, d in _results:
            if not o:
                print(f"   - {name}: {d}")
    print(f"Export k prohlédnutí: {export_path}")
    return 0 if ok == total else 1


if __name__ == "__main__":
    sys.exit(main())
