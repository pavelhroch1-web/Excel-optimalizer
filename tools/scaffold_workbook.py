"""
Builds the V11 scaffold workbook.

Copies the production-data sheets from the reference V10.5.5 workbook unchanged
(RAW_DATA, CONTROL, ACTIVITY_PLAN, CATEGORY_RULES, TERMINAL_RULES, VISIT_HISTORY),
adds an explicit default row to CATEGORY_RULES, and adds the new V11 sheets
(POS_MASTER, config tables, import staging) as headers-only or seeded with values
that are already confirmed in docs/BUSINESS_RULES.md. Anything not yet confirmed
is seeded as an inactive TODO row, never silently activated.

Usage: python3 tools/scaffold_workbook.py <path-to-reference-workbook.xlsx> <output-path.xlsx>
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from ux_style import apply_all

REFERENCE_SHEETS_KEEP_AS_IS = [
    "RAW_DATA", "CONTROL", "ACTIVITY_PLAN", "TERMINAL_RULES", "VISIT_HISTORY",
]


def copy_sheet(src_wb, dst_wb, name):
    src = src_wb[name]
    dst = dst_wb.create_sheet(name)
    for row in src.iter_rows(values_only=False):
        for cell in row:
            dst.cell(row=cell.row, column=cell.column, value=cell.value)
    return dst


def write_table(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(h)) + 2)
    return ws


def main(ref_path, out_path):
    src_wb = openpyxl.load_workbook(ref_path, data_only=True)
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)  # drop default empty sheet
    # A brand-new openpyxl.Workbook() has no embedded theme (loaded_theme is
    # None) - borrow the reference workbook's, so apply_all()'s
    # set_modern_theme_fonts() has an actual theme to modify instead of
    # silently no-op-ing on a from-scratch build.
    dst_wb.loaded_theme = src_wb.loaded_theme

    for name in REFERENCE_SHEETS_KEEP_AS_IS:
        copy_sheet(src_wb, dst_wb, name)
    # TERMINAL_RULES is copied verbatim above, including whatever
    # sheet_state the reference workbook happened to have - explicitly force
    # it visible (product owner, 2026-07-08: wants the VELKY TERMINAL/SMALL
    # TERMINAL/LI Ano/Ne toggles reachable, not hidden) rather than relying
    # on ux_style.HIDDEN_SHEETS alone, since that set only ever HIDES sheets
    # named in it - it never un-hides a sheet that arrived already hidden.
    dst_wb["TERMINAL_RULES"].sheet_state = "visible"

    # Add GPS bonus config keys to CONTROL (docs/BUSINESS_RULES.md 6a - corrected
    # spec: bounded, capacity-aware overflow, not a hard cap at capacity).
    control_ws = dst_wb["CONTROL"]
    control_ws.append(["GPS_EXTRA_ENABLED", 1, "1=on. Allows a small overflow beyond capacity for POS very close to an already-selected visit."])
    control_ws.append(["GPS_EXTRA_RADIUS_METERS", 300, "Radius for the GPS bonus overflow rule."])
    control_ws.append(["GPS_EXTRA_MAX_VISITS", 5, "Max extra visits per technician/week from the GPS bonus rule."])
    control_ws.append(["COMPLIANCE_LATE_CUTOFF_WEEKS", 1, "Weeks past the planned week before a still-unrealized visit becomes Nesplneno instead of Pending. Proposed default per docs/BUSINESS_RULES.md section 12, not yet formally reconfirmed."])
    control_ws.append(["ADVISOR_NEGLECT_WARNING_RATIO_PERCENT", 80, "Proposed default (not a confirmed business rule): WARNING neglect alert fires at this % of NEGLECTED_AFTER_WEEKS, CRITICAL fires at 100%. Tune on real data."])
    control_ws.append(["ADVISOR_TREND_WINDOW_WEEKS", 4, "Proposed default: how many recent weeks of COMPLIANCE_LOG feed the technician/region overload alerts. Tune on real data."])
    control_ws.append(["ADVISOR_OVERLOAD_WARNING_RATE_PERCENT", 20, "Proposed default (not a confirmed business rule): technician/region Nesplneno rate over the trend window that triggers a WARNING. Tune on real data."])
    control_ws.append(["ADVISOR_OVERLOAD_CRITICAL_RATE_PERCENT", 35, "Proposed default (not a confirmed business rule): technician/region Nesplneno rate over the trend window that triggers a CRITICAL alert. Tune on real data."])
    control_ws.append(["ADVISOR_VOLUME_TRAILING_WEEKS", 8, "Planning Cycle Advisor v1 (deterministic, informational only - docs/ARCHITECTURE.md section 19): how many of the most recent weeks in VISIT_HISTORY_ACTUAL count as the 'trailing' period for the volume trend signal."])
    control_ws.append(["ADVISOR_VOLUME_BASELINE_WEEKS", 8, "Planning Cycle Advisor v1: how many weeks immediately before the trailing period form the 'baseline' the trailing period is compared against. Needs TRAILING+BASELINE weeks of history before this signal can fire at all - with under a year of SalesApp history, expect it to stay silent (correct, not a bug)."])
    control_ws.append(["ADVISOR_VOLUME_THRESHOLD_PERCENT", 25, "Planning Cycle Advisor v1: trailing-vs-baseline deviation (%) that triggers a VOLUME_TREND_SIGNAL alert. Proposed default, not a confirmed business rule - tune once a real season of history exists."])
    control_ws.append(["ROUTE_KM_WARNING_KM", 80, "Proposed default (not a confirmed business rule): daily route-efficiency estimate (PerformanceEngine.ts kmMon..kmFri) above this shows the WARNING semafor color on TECHNICIAN_SCORECARD. Tune on real data - product owner requested the metric 2026-07-06, thresholds are a starting guess."])
    control_ws.append(["ROUTE_KM_CRITICAL_KM", 150, "Proposed default (not a confirmed business rule): daily route-efficiency estimate above this shows the CRITICAL semafor color. Tune on real data."])
    control_ws.append(["FLAKANI_WINDOW_WEEKS", 4, "Confirmed (product owner, 2026-07-06): how many of a technician's most recent tracked weeks PerformanceEngine.ts looks at for the 'flaka riziko' (persistent-underperformance) flag on TECHNICIAN_PERFORMANCE_SUMMARY/PERFORMANCE."])
    control_ws.append(["FLAKANI_BAD_WEEK_THRESHOLD_PERCENT", 70, "Confirmed (product owner, 2026-07-06): a week counts as 'bad' for the flaka-riziko flag when compliancePercent falls below this. Same cutoff as the existing WARNING semafor elsewhere in the workbook."])
    control_ws.append(["FLAKANI_BAD_WEEKS_COUNT", 2, "Confirmed (product owner, 2026-07-06): a technician is flagged 'flaka riziko' when at least this many of their last FLAKANI_WINDOW_WEEKS tracked weeks are 'bad' (see FLAKANI_BAD_WEEK_THRESHOLD_PERCENT) - requires a repeated pattern, not a single bad week."])
    control_ws.append(["GEO_CLUSTER_RADIUS_KM", 3, "Confirmed (product owner, 2026-07-06): PlanningEngine.ts's geo cluster bonus - a candidate POS within this radius of another valuable candidate for the same technician gets a small score bonus (see GEO_CLUSTER_BONUS_FACTOR/GEO_CLUSTER_MAX_BONUS), nudging selection toward tighter daily routes without overriding value as the primary driver."])
    control_ws.append(["GEO_CLUSTER_BONUS_FACTOR", 0.01, "Confirmed (product owner, 2026-07-06): fraction of a nearby neighbor's own score added as this candidate's geo cluster bonus (1% by default)."])
    control_ws.append(["GEO_CLUSTER_MAX_BONUS", 5000, "Confirmed (product owner, 2026-07-06): cap on the total geo cluster bonus a single candidate can accumulate - kept well below the smallest meaningful score tier (NEGLECTED_BONUS=50000) so this can only break near-ties, never outweigh being CORE/classification A/neglected."])
    control_ws.append(["HOLDBACK_LOOKAHEAD_WEEKS", 3, "Confirmed (product owner, 2026-07-09, 'Kriticke'): Smart Hold-back's widest possible elastic lookahead - a non-mandatory POS is only ever considered for deferral if an ACTIVITY_PLAN campaign starts within this many weeks."])
    control_ws.append(["HOLDBACK_TOLERANCE_A_WEEKS", 1, "Confirmed (product owner, 2026-07-09): classification A POS can only be held back for a campaign starting within 1 week (capped further by HOLDBACK_LOOKAHEAD_WEEKS)."])
    control_ws.append(["HOLDBACK_TOLERANCE_OTHER_WEEKS", 3, "Confirmed (product owner, 2026-07-09): classification B/C POS can be held back for a campaign starting up to this many weeks out (capped by HOLDBACK_LOOKAHEAD_WEEKS). Smart Hold-back never defers past a POS's own hard deadline (its matched CADENCE_RULES maxIntervalWeeks, else NEGLECTED_AFTER_WEEKS) regardless of this tolerance - see shouldHoldBack() in office-scripts/shared/core.ts."])
    control_ws.append(["URGENCY_BOOST_MAX", 20000, "Confirmed (product owner, 2026-07-09): max score boost computeUrgencyBoost() adds to a POS as it approaches its own deadline (deadlineWeeks), so it isn't starved out of selection by a campaign-driven surge. Kept well below NEGLECTED_BONUS (50000) and classification A (10000000) - only nudges among non-neglected candidates, never overrides the existing hard priority tiers."])
    control_ws.append(["URGENCY_BOOST_RAMP_START_RATIO", 0.5, "Confirmed (product owner, 2026-07-09): computeUrgencyBoost() starts ramping once weeksSinceLastVisit reaches this fraction of the POS's own deadline (0.5 = halfway), reaching URGENCY_BOOST_MAX exactly at the deadline. A smooth ramp, not a step function - see computeUrgencyBoost()'s own comment."])
    control_ws.append(["ROUTE_EFFICIENCY_WARNING_PERCENT", 125, "Confirmed (product owner, 2026-07-09, 'Monitoring efektivity - kdo jezdi cik-cak'): a technician's weekly actual-vs-optimal ('matematicke minimum') route km ratio at or above this % triggers the POZOR efficiencyFlag on TECHNICIAN_PERFORMANCE_LOG/SUMMARY/PERFORMANCE."])
    control_ws.append(["ROUTE_EFFICIENCY_CRITICAL_PERCENT", 150, "Confirmed (product owner, 2026-07-09): 'o 50 %+ vyssi nez optimum' - the explicit CRITICAL bar for efficiencyRatioPercent (KRITICKE efficiencyFlag)."])
    control_ws.append(["VOLUME_WARNING_PERCENT", 70, "Confirmed (product owner, 2026-07-09, 'manazerske triggery' - vyrazne mene navstevnosti nez ostatni): realizedVisits below this % of the network peer average (or the technician's own recent average, whichever is worse) triggers the POZOR volumeFlag."])
    control_ws.append(["VOLUME_CRITICAL_PERCENT", 50, "Confirmed (product owner, 2026-07-09): the KRITICKE bar for volumeFlag."])
    control_ws.append(["PPT_DENSITY_WARNING_PERCENT", 70, "Confirmed (product owner, 2026-07-09): 'hodne navstev, ale jednoucelove' - PPT captured per realized visit below this % of the network peer average triggers the POZOR pptDensityFlag. Independent of route km efficiency - a technician can have a perfect route and still be visiting low-value POS."])
    control_ws.append(["PPT_DENSITY_CRITICAL_PERCENT", 50, "Confirmed (product owner, 2026-07-09): the KRITICKE bar for pptDensityFlag."])
    control_ws.append(["DURATION_WARNING_PERCENT", 70, "Confirmed (product owner, 2026-07-09): average realized-visit duration (Real duration (h) from SalesApp) below this % of the network peer average triggers the POZOR durationFlag - a directly-measured signal, not a GPS estimate."])
    control_ws.append(["DURATION_CRITICAL_PERCENT", 50, "Confirmed (product owner, 2026-07-09): the KRITICKE bar for durationFlag."])
    control_ws.append(["PROBLEM_SIGNAL_MIN_COUNT", 2, "Confirmed (product owner, 2026-07-09: 'GPS je odhad, takze to ani nemusi byt na vinu'): how many of {flaka riziko, volumeFlag, pptDensityFlag, durationFlag, efficiencyFlag} must be simultaneously POZOR/KRITICKE before combinedRiskFlag='Ano' - the gate for the automatic 'problemovy technik' callouts on HOME/EFFICIENCY. No single signal alone (including route efficiency) triggers it."])
    control_ws.append(["ACTIVATE_COUNT_BY_PPT", 0, "Confirmed (product owner, 2026-07-11: 'chci mit moznost je pridat... prvnich 500 nehlede kolik techniku to bude'): ActivatePOSEngine.ts's count-based mode - when POS_ACTIVATE_LIST is empty and this is > 0, activates (managerOverrideType=FORCE_INCLUDE) this many currently CATEGORY_RULES-EXCLUDE-d Active POS, highest PPT first. 0 = disabled (count mode does nothing; only POS_ACTIVATE_LIST's explicit list is processed). Ignored whenever POS_ACTIVATE_LIST has any rows - explicit list always wins over count."])

    # CATEGORY_RULES: copy reference rows + add explicit confirmed default row
    cat_ws = src_wb["CATEGORY_RULES"]
    cat_rows = [list(r) for r in cat_ws.iter_rows(values_only=True)]
    header, body = cat_rows[0], cat_rows[1:]
    if not any(str(r[0]).strip() == "*" for r in body):
        body.append(["*", "NORMAL"])  # confirmed V10.5.5 default fallback, made explicit
    write_table(dst_wb, "CATEGORY_RULES", header, body)

    # MARKET_RULES (new filter layer)
    write_table(
        dst_wb, "MARKET_RULES",
        ["MARKET", "ACTIVE"],
        [
            ["IDT", "YES"],
            ["ČESKÁ POŠTA", "YES"],
            ["KA PARTNERS", "YES"],
            ["PETROL", "YES"],
            ["CORN", "YES"],
        ],
    )

    # BLACKLIST (product owner, 2026-07-09): manual paste-list of POS IDs
    # Planning Engine must ignore completely, regardless of any other data -
    # a dedicated quick-scan list, distinct from POS_MASTER's per-row
    # managerOverrideType=FORCE_EXCLUDE dropdown.
    write_table(
        dst_wb, "BLACKLIST",
        ["POS", "NOTES"],
        [],
    )

    # POS_ACTIVATE_LIST (product owner, 2026-07-11: "chci mit moznost je
    # pridat jako mame treba ted vyrazene 1CD... urcit bud jake, nebo
    # prvnich 500 nehlede kolik techniku to bude") - same minimal
    # paste-list pattern as BLACKLIST above, but the opposite direction:
    # POS IDs pasted here get ActivatePOSEngine.ts's managerOverrideType=
    # FORCE_INCLUDE, overriding a CATEGORY_RULES EXCLUDE rule (e.g. "1CD",
    # "1POSTA") for just those POS - the existing PlanningEngine.ts
    # FORCE_INCLUDE bypass already does the rest (each POS keeps its own
    # assignedTechnician, so this naturally spreads across however many
    # technicians those POS already belong to - no manual per-technician
    # assignment needed). Leave empty and set CONTROL.ACTIVATE_COUNT_BY_PPT
    # instead to activate the top N by PPT from the excluded pool.
    write_table(
        dst_wb, "POS_ACTIVATE_LIST",
        ["POS", "NOTES"],
        [],
    )

    # CADENCE_RULES (unifies CORE / Mandatory / GECO / CORN)
    write_table(
        dst_wb, "CADENCE_RULES",
        ["ruleId", "scope", "matchValue", "minGapWeeks", "maxIntervalWeeks",
         "intervalType", "guaranteeType", "dedupBy", "campaignChangeOverride",
         "priority", "active", "validFrom", "validTo", "notes"],
        [
            ["CORE", "categoryPrefix", "1", 2, "", "RECURRING", "SOFT_HIGH_WEIGHT",
             "NONE", "YES", 90, "YES", "", "",
             "Evolution of V10.5.5 score constant (CORE+=1e8). minGapWeeks=2 preserves "
             "PREMIUM_GAP with campaign-change override."],
            ["MANDATORY_9PODNIK", "category", "9PODNIKC;9PODNIKFC", "", "",
             "ONCE_PER_CAMPAIGN", "HARD", "ADDRESS", "NO", 100, "YES", "", "",
             "Preserves V10.5.5 mandatoryPodnik(): one guaranteed slot per campaign run, "
             "best-PTT POS per street+city."],
            ["GECO", "category", "1GECO", "", 5, "RECURRING", "HARD", "ADDRESS", "NO",
             80, "YES", "", "",
             "Scope confirmed as 1GECO only (not the broader KA PARTNERS market), "
             "guaranteeType=HARD confirmed - product owner, 2026-07-03. dedupBy=ADDRESS "
             "added 2026-07-08 (product owner: two same-address terminals under the same "
             "cadence rule must dedupe to the higher-PPT one, same mechanism already used "
             "by MANDATORY_9PODNIK - see pickMandatory() in core.ts, no code change needed)."],
            ["CORN", "market", "CORN", "", 4, "RECURRING", "HARD", "ADDRESS", "NO",
             80, "YES", "", "",
             "HARD/4 weeks confirmed (16 POS = negligible capacity impact). dedupBy=ADDRESS "
             "added 2026-07-08 - see GECO's note above, same rationale."],
        ],
    )

    # PLANNING_HORIZON_RULES (data-model placeholder, NOT yet read by any
    # engine - product owner asked the data model to be ready for a future
    # "when should the next plan be prepared" advisor before the decision
    # logic itself is designed/approved. Today PlanningEngine.ts still uses
    # CONTROL.CAMPAIGN_LENGTH as a flat constant, unchanged. This table lets
    # a future PlanningCycleAdvisor read a seasonal override (e.g. "start
    # planning 8 weeks ahead in the weeks around Christmas") as config
    # instead of a code change, mirroring how CADENCE_RULES/PARETO_GROUPS
    # were seeded inactive before their logic existed. See
    # docs/ARCHITECTURE.md section 18 for the full design and the open
    # business question (how "vhodna chvile" gets defined) still pending
    # product-owner confirmation before any engine reads this table.
    write_table(
        dst_wb, "PLANNING_HORIZON_RULES",
        ["ruleId", "appliesFromWeek", "appliesToWeek", "horizonWeeks", "reason", "active", "notes"],
        [
            ["DEFAULT", "", "", 4, "Standard rok", "NO",
             "Placeholder mirroring today's implicit CONTROL.CAMPAIGN_LENGTH behaviour - "
             "not yet read by any engine. Not activating changes nothing."],
            ["SEZONA_VANOCE", 47, 52, 8, "Vanoce a konec roku - delsi priprava", "NO",
             "TODO: exact week range and horizon value are illustrative, not confirmed - "
             "needs product-owner sign-off before this table is ever read by an engine."],
        ],
    )

    # PARETO_GROUPS
    write_table(
        dst_wb, "PARETO_GROUPS",
        ["tierId", "name", "scope", "boundaryType", "boundaryValue", "active", "notes"],
        [
            ["PREMIUM_TOP20", "Premium (top 20% per portfolio)", "PER_TECHNICIAN",
             "PERCENTILE", 20, "YES",
             "Preserves V10.5.5 behaviour exactly (relative ranking within each "
             "technician's own portfolio). Scope field kept switchable to GLOBAL / "
             "PER_REGION / PER_MARKET for a later decision, not activated now."],
            ["KA", "KA (business-significant outlets)", "", "", "", "NO",
             "TODO: scope + threshold not yet confirmed (BUSINESS_RULES.md)."],
            ["IDT_ABOVE_THRESHOLD", "IDT above PPT threshold", "", "", "", "NO",
             "TODO: boundaryType/scope not yet confirmed (BUSINESS_RULES.md)."],
        ],
    )

    # SCORE_PROFILES
    write_table(
        dst_wb, "SCORE_PROFILES",
        ["profileId", "component", "weight", "notes"],
        [
            ["DEFAULT", "CORE", 100000000, "Preserves V10.5.5 magnitude as-is; normalize later."],
            ["DEFAULT", "KATEGORIZACE_A", 10000000, "Preserves V10.5.5 magnitude as-is."],
            ["DEFAULT", "PPT", 1, "Raw PTT value, unweighted (as in V10.5.5)."],
            ["DEFAULT", "NEGLECTED_BONUS", 50000, "Applied when weeksSinceLastVisit >= NEGLECTED_AFTER_WEEKS."],
        ],
    )

    # ADVISOR_RULES (reserved for a future fully config-driven rule table -
    # AdvisorEngine.ts v1 reads its three alert types' thresholds directly
    # from CONTROL instead, see docs/BACKLOG.md for the planned generalization)
    write_table(
        dst_wb, "ADVISOR_RULES",
        ["ruleId", "type", "condition", "threshold", "severity", "messageTemplate", "active"],
        [],
    )

    # ADVISOR_LOG (append-only output of AdvisorEngine.ts)
    write_table(
        dst_wb, "ADVISOR_LOG",
        ["type", "severity", "subjectType", "subjectId", "message", "evaluatedAt"],
        [],
    )

    # CAPACITY_OVERRIDE
    write_table(
        dst_wb, "CAPACITY_OVERRIDE",
        ["technician", "year", "week", "capacity"],
        [],
    )

    # POS_STATUS_IMPORT (staging import sheet)
    write_table(
        dst_wb, "POS_STATUS_IMPORT",
        ["POS", "ACTIVE"],
        [],
    )

    # POS_MASTER (headers only — populated by Import Engine)
    write_table(
        dst_wb, "POS_MASTER",
        [
            "posId", "terminalId",
            "market", "category", "terminalType", "classification", "nazev", "area", "posArea",
            "street", "houseNumber", "city", "gpsX", "gpsY", "assignedTechnician", "ppt",
            "status", "closedSinceWeek", "closedSinceYear",
            "currentLosActivity", "currentLotActivity", "targetLosActivity", "targetLotActivity",
            "lastRealVisitDate", "lastRealVisitWeek", "lastPlannedVisitDate",
            "weeksSinceLastVisit", "visitCountThisCampaign",
            "businessScore",
            "plannerStatus", "assignedWeek", "assignedDay", "gpsGroup",
            "managerOverrideType", "managerOverridePriority", "managerOverrideTechnician",
            "plannerNotes",
            "importedAt", "updatedAt",
        ],
        [],
    )

    # MANAGER_PLAN (Planning Engine output - headers match legacy OUTPUT_PLAN
    # column order so the shape is familiar, populated by PlanningEngine.ts)
    write_table(
        dst_wb, "MANAGER_PLAN",
        ["WEEK", "DATE", "DAY", "TECHNICIAN", "POS", "KATEGORIE", "NAZEV_PROVOZOVNY",
         "ULICE", "CISLO", "MESTO", "OBLAST", "POS_AREA", "PPT", "LOS_ACTIVITY",
         "LOT_ACTIVITY", "REASON", "GPS_GROUP"],
        [],
    )

    # PLAN_LIFECYCLE (Draft -> Published -> Active -> Closed, one row per
    # (year, week). Draft->Published only via PublishEngine.ts (explicit
    # manager action); Published->Active->Closed recomputed by
    # ComplianceEngine.ts on every run - see docs/BUSINESS_RULES.md section 11.
    # trackingStartedAt: a SEPARATE explicit manager action
    # (StartTrackingEngine.ts) - blank until the manager runs it, even after
    # the week is Published/Active. PerformanceEngine.ts only includes a
    # week's numbers in the manager dashboards once this is set (product
    # owner, 2026-07-06: "abych ho začal sledovat až řeknu já"). Appended as
    # a 6th column - ComplianceEngine.ts/PublishEngine.ts only ever write
    # columns A-E by index, so this is safe to add without touching them.)
    write_table(
        dst_wb, "PLAN_LIFECYCLE",
        ["year", "week", "status", "publishedAt", "closedAt", "trackingStartedAt"],
        [],
    )

    # MANAGER_PLAN_PUBLISHED (immutable snapshot, append-only - what was
    # actually sent to technicians. ComplianceEngine.ts compares against
    # this, never against the freely-regenerated MANAGER_PLAN, per product
    # owner: "Compliance vzdy porovnava pouze Published snapshot")
    write_table(
        dst_wb, "MANAGER_PLAN_PUBLISHED",
        ["WEEK", "DATE", "DAY", "TECHNICIAN", "POS", "KATEGORIE", "NAZEV_PROVOZOVNY",
         "ULICE", "CISLO", "MESTO", "OBLAST", "POS_AREA", "PPT", "LOS_ACTIVITY",
         "LOT_ACTIVITY", "REASON", "GPS_GROUP", "publishedAt"],
        [],
    )

    # SALESAPP_IMPORT (staging - paste the weekly SalesApp export here; header
    # row matches the real export format so column-name lookup in
    # ComplianceEngine.ts works regardless of export column order/extras.
    # "Ucel navstevy - Technik - MCHD - Nabeh kampane" (Ano/blank) is included
    # deliberately, not just the identity columns: ComplianceEngine.ts only
    # counts a Completed/Finalized row as a realized CAMPAIGN visit when this
    # column is Ano - confirmed by product owner. Column order/extras beyond
    # what's listed here don't matter (paste the full real export as-is), but
    # this specific column must be present with this name for compliance
    # matching to work at all.)
    write_table(
        dst_wb, "SALESAPP_IMPORT",
        ["UID", "Date", "State", "Started at", "Finished at", "Real duration (h)",
         "Chain UID", "Chain", "Store UID", "Store", "Store address",
         "Agency region", "Executor UID", "Executor",
         "Účel návštevy -  Technik - MCHD - Náběh kampaně"],
        [],
    )

    # VISIT_HISTORY_ACTUAL (real, append-only visit log from SalesApp -
    # distinct from the legacy VISIT_HISTORY sheet carried over from V10.5.5,
    # which recorded the script's own planned output, not reality; see
    # docs/BUSINESS_RULES.md 15c and ARCHITECTURE.md Compliance Engine entry)
    write_table(
        dst_wb, "VISIT_HISTORY_ACTUAL",
        ["posId", "date", "week", "year", "executor", "state", "salesAppUid", "durationHours", "startedAt", "finishedAt"],
        [],
    )

    # OTHER_VISIT_LOG (append-only, dedup by uid - see ComplianceEngine.ts):
    # Completed/Finalized SalesApp visits whose purpose is NOT the campaign
    # ("MCHD - Nabeh kampane") signal - real visits (restocking, lottery
    # ticket downloads, etc.) that don't count toward compliance, logged here
    # purely so PerformanceEngine.ts can surface an informational "Ostatní
    # návštěvy" count on TECHNICIAN_SCORECARD (product owner, 2026-07-06).
    write_table(
        dst_wb, "OTHER_VISIT_LOG",
        ["posId", "date", "week", "year", "executor", "salesAppUid", "durationHours", "startedAt", "finishedAt"],
        [],
    )

    # COMPLIANCE_LOG (append-only, one row per planned-visit evaluation)
    write_table(
        dst_wb, "COMPLIANCE_LOG",
        ["posId", "technician", "plannedWeek", "plannedYear", "status",
         "matchedActualDate", "matchedActualWeek", "evaluatedAt", "matchedActualDurationHours",
         "matchedActualStartedAt", "matchedActualFinishedAt"],
        [],
    )

    # DASHBOARD (Reporting Engine output - written fresh on every run)
    write_table(dst_wb, "DASHBOARD", ["", "", "", "", "", ""], [])

    # POS_MAP_DATA (Reporting Engine output - written fresh on every run):
    # fixed-size (MAX_MAP_TECHS=40 slots x 2 columns) X/Y coordinate pairs
    # per technician, feeding the MAP sheet's territory scatter chart. Empty
    # headers here are just placeholders - the engine writes each slot's
    # technician name into row 1 on first run.
    write_table(dst_wb, "POS_MAP_DATA", [""] * 80, [])

    # TECHNICIAN_PERFORMANCE_LOG (Performance Engine output - full rebuild
    # every run, one row per technician/ISO-week - see
    # docs/MANAGER_UX_ARCHITECTURE.md section 1)
    write_table(
        dst_wb, "TECHNICIAN_PERFORMANCE_LOG",
        ["technician", "year", "week", "region",
         "plannedVisits", "realizedVisits",
         "splnenoVcas", "splnenoPozde", "nesplneno", "navicEvidovano",
         "compliancePercent",
         "visitsMon", "visitsTue", "visitsWed", "visitsThu", "visitsFri",
         "updatedAt",
         "kmMon", "kmTue", "kmWed", "kmThu", "kmFri",
         "otherVisits",
         "posListMon", "posListTue", "posListWed", "posListThu", "posListFri",
         "monthKey",
         "otherVisitsMon", "otherVisitsTue", "otherVisitsWed", "otherVisitsThu", "otherVisitsFri",
         "totalActualKmWeek", "totalOptimalKmWeek", "efficiencyRatioPercent", "kmPerVisit", "efficiencyFlag",
         "pptPerVisit", "avgVisitDurationHours",
         "volumeVsPeerPercent", "pptDensityVsPeerPercent", "durationVsPeerPercent",
         "volumeFlag", "pptDensityFlag", "durationFlag", "activeSignalCount", "combinedRiskFlag",
         "workSpanHoursMon", "workSpanHoursTue", "workSpanHoursWed", "workSpanHoursThu", "workSpanHoursFri",
         "idleHoursMon", "idleHoursTue", "idleHoursWed", "idleHoursThu", "idleHoursFri"],
        [],
    )

    # TECHNICIAN_PERFORMANCE_SUMMARY (Performance Engine's second output -
    # one row per technician, most recent week + long-run avg + trend,
    # feeds PERFORMANCE)
    write_table(
        dst_wb, "TECHNICIAN_PERFORMANCE_SUMMARY",
        ["technician", "region", "latestYear", "latestWeek",
         "plannedVisits", "realizedVisits", "splnenoVcas", "splnenoPozde", "nesplneno", "navicEvidovano",
         "compliancePercent", "longRunAvgCompliance", "trendDelta",
         "badWeeksInWindow", "flakaRiziko", "maxKmDay",
         "efficiencyRatioPercent", "kmPerVisit", "longRunAvgEfficiencyRatio", "efficiencyFlag",
         "volumeVsOwnAvgPercent", "longRunAvgVolumeVsPeerPercent", "volumeFlag",
         "longRunAvgPptDensityVsPeerPercent", "pptDensityFlag",
         "longRunAvgDurationVsPeerPercent", "durationFlag",
         "activeSignalCount", "combinedRiskFlag"],
        [],
    )

    # TECHNICIAN_TOP_ISSUES (Performance Engine's third output - top 5
    # all-time Nesplneno POS per technician, feeds TECHNICIAN_SCORECARD)
    write_table(
        dst_wb, "TECHNICIAN_TOP_ISSUES",
        ["technician", "rank", "posId", "posName", "region", "nesplnenoCount"],
        [],
    )

    # UX pass: sheet organization, color coding, dropdowns, legend,
    # START_HERE, ACTIVITY_PLAN timeline - pure presentation, see
    # tools/ux_style.py. Runs last so it sees the final sheet content.
    control_rows_for_ux = list(dst_wb["CONTROL"].iter_rows(values_only=True))
    apply_all(dst_wb, control_rows_for_ux)

    dst_wb.save(out_path)
    print(f"Scaffold written to {out_path}")
    print("Sheets:", dst_wb.sheetnames)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
