"""
Shared UI component library for the manager dashboard layer (HOME,
TECHNICIAN_SCORECARD, and the PERFORMANCE/WEEK_DASHBOARD screens still to
come - docs/MANAGER_UX_ARCHITECTURE.md). Every dashboard screen composes
these functions instead of hand-rolling its own banner/card/nav-rail code
(product owner, 2026-07-03: "Nechci kopírovaný kód mezi dashboardy. Cílem
je mít jednu UI knihovnu a jednotlivé dashboardy budou pouze skládat
komponenty.").

This is the ONE place the dashboard color palette and typography are
defined - ux_style.py imports its NAVY/WHITE/STATUS_*/TITLE_FONT/etc names
from here rather than redefining them, so there is exactly one accessibility-
validated palette in the codebase (see the dataviz-skill-derived values
below), not two that could quietly drift apart.

Deliberately NOT for the data-entry/config sheets (POS_MASTER, CONTROL,
RAW_DATA, CADENCE_RULES, etc.) - those have a different visual language
(editable-cell color coding, dropdowns, sheet protection) that stays in
ux_style.py's sheet-styling functions. This file only knows about dashboard
screens: banners, KPI cards, nav rails, charts, filters.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import column_index_from_string
from openpyxl.chart import BarChart, LineChart
from openpyxl.workbook.defined_name import DefinedName

# ============================================================================
# COLOR PALETTE
# ============================================================================
NAVY = "1F4E78"
WHITE = "FFFFFF"
RAIL_DARK = "2B2B2B"
MUTED_GREY = "595959"
NOTE_GREY = "808080"
LIGHT_GREY = "F2F2F2"
BORDER_GREY = "D9D9D9"
ACCENT_BLUE = "2E75B6"

# Status colors - accessibility-validated (see tools/ux_style.py's original
# note: checked with the project's color-accessibility validator; WARNING/
# SERIOUS fall under 3:1 contrast against white alone, an accepted tradeoff
# because every use pairs the color with an icon/label, never color-alone).
STATUS_GOOD = "0CA30C"
STATUS_WARNING = "FAB219"
STATUS_SERIOUS = "EC835A"
STATUS_CRITICAL = "D03B3B"

# Light tints of the status colors, for KPI card backgrounds (readable body
# text on top, unlike the saturated status colors above which are for small
# badges/borders only).
TINT_GOOD = "E2EFDA"
TINT_WARNING = "FFF2CC"
TINT_SERIOUS = "FCE4D6"
TINT_CRITICAL = "FCE4D6"
TINT_NEUTRAL = WHITE

ICONS = {
    "home": "🏠", "scorecard": "📊", "performance": "📋", "dashboard": "📈", "plan": "🗺",
    "week": "🗓", "map": "📍", "efficiency": "🚗",
    "good": "✅", "bad": "❌", "warning": "⚠", "celebrate": "🎉",
    "up": "▲", "down": "▼", "flat": "→",
}

# ============================================================================
# TYPOGRAPHY
# ============================================================================
FONT_DASHBOARD_TITLE = Font(bold=True, size=24, color=WHITE)
FONT_DASHBOARD_SUBTITLE = Font(italic=True, size=11, color=WHITE)
FONT_HEADER = Font(color=WHITE, bold=True, size=11)          # data-table header row
FONT_TITLE = Font(bold=True, size=14, color=NAVY)             # section/screen title
FONT_SECTION = Font(bold=True, size=11, color=NAVY)            # sub-section label
FONT_NOTE = Font(italic=True, size=9, color=NOTE_GREY)
FONT_CARD_LABEL = Font(bold=True, size=9, color=MUTED_GREY)


def font_card_value(size=22, color=NAVY, bold=True):
    return Font(bold=bold, size=size, color=color)


CARD_BORDER = Border(*(Side(style="thin", color=BORDER_GREY),) * 4)

# ============================================================================
# NAVIGATION RAIL
# ============================================================================
# The persistent side menu every dashboard screen shares. Add a screen here
# once it exists (PERFORMANCE, WEEK_DASHBOARD) and every existing rail picks
# it up automatically on the next build - this list is the single source of
# truth for "what screens does the app have", not per-sheet duplication.
NAV_RAIL_SHEETS = [
    ("HOME", f"{ICONS['home']} Domů", "404040"),
    ("TECHNICIAN_SCORECARD", f"{ICONS['scorecard']} Scorecard", ACCENT_BLUE),
    ("PERFORMANCE", f"{ICONS['performance']} Performance", ACCENT_BLUE),
    ("EFFICIENCY", f"{ICONS['efficiency']} Efektivita", ACCENT_BLUE),
    ("WEEK_DASHBOARD", f"{ICONS['week']} Week Dashboard", ACCENT_BLUE),
    ("MAP", f"{ICONS['map']} Mapa území", ACCENT_BLUE),
    ("DASHBOARD", f"{ICONS['dashboard']} Dashboard", "375623"),
    ("TECHNICIAN_PLAN", f"{ICONS['plan']} Plán týdne", "375623"),
]


def build_nav_rail(ws, current_sheet, rail_items=None, first_row=1, nav_col="A", spacer_col="B",
                    nav_width=20, spacer_width=3):
    """A persistent vertical stack of nav buttons in nav_col - the "left
    menu" a plain Excel tab strip can't give you (see
    docs/MANAGER_UX_ARCHITECTURE.md section 3). Pair with a frozen-pane call
    that keeps column A/B in view (e.g. ws.freeze_panes = "C4") so this rail
    never scrolls away."""
    rail_items = rail_items or NAV_RAIL_SHEETS
    ws.column_dimensions[nav_col].width = nav_width
    ws.column_dimensions[spacer_col].width = spacer_width
    r = first_row
    for sheet_name, label, color in rail_items:
        is_current = sheet_name == current_sheet
        cell = ws[f"{nav_col}{r}"]
        cell.value = ("▶ " if is_current else "   ") + label
        if not is_current:
            cell.hyperlink = f"#{sheet_name}!A1"
        cell.font = Font(bold=is_current, size=11, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color if is_current else RAIL_DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 26
        r += 1
    # fill the rest of the rail down to a generous depth so it reads as one
    # continuous dark sidebar, not a stack of buttons floating on white
    for rr in range(r, r + 40):
        ws[f"{nav_col}{rr}"].fill = PatternFill("solid", fgColor=RAIL_DARK)
    return r


def build_nav_button(ws, cell_ref, label, target_sheet, color=NAVY):
    """A single hyperlink button (as opposed to the vertical rail above) -
    used for one-off "jump to X" links, e.g. HOME's quick-nav row or a
    pipeline stage's "Otevřít →" action.

    A real cell-level hyperlink (openpyxl Cell.hyperlink), NOT a
    =HYPERLINK() formula: openpyxl never calculates formulas, so a
    formula-based button is stored with an empty cached value - some Excel
    contexts don't render it until a forced recalc, which was a real bug
    (product owner confirmed nav buttons did not work). A native hyperlink
    has no calculation dependency: it works the instant the file opens."""
    cell = ws[cell_ref]
    cell.value = label
    cell.hyperlink = f"#{target_sheet}!A1"
    cell.font = Font(bold=True, size=12, color=WHITE, underline=None)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = CARD_BORDER


# ============================================================================
# DASHBOARD TITLE / BANNER
# ============================================================================
def build_dashboard_banner(ws, title, subtitle, col_start, col_end, title_row=1, subtitle_row=3, title_size=24):
    """The big navy banner every dashboard screen opens with. Returns the
    first free row below the banner so the caller can keep laying out
    without hardcoding row numbers twice."""
    ws.merge_cells(f"{col_start}{title_row}:{col_end}{title_row + 1}")
    cell = ws[f"{col_start}{title_row}"]
    cell.value = title
    cell.font = Font(bold=True, size=title_size, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[title_row].height = 30
    ws.row_dimensions[title_row + 1].height = 22
    ws.merge_cells(f"{col_start}{subtitle_row}:{col_end}{subtitle_row}")
    sub = ws[f"{col_start}{subtitle_row}"]
    sub.value = subtitle
    sub.font = FONT_DASHBOARD_SUBTITLE
    sub.fill = PatternFill("solid", fgColor=NAVY)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    sub.border = Border(bottom=Side(style="medium", color=ACCENT_BLUE))
    ws.row_dimensions[subtitle_row].height = 20
    return subtitle_row + 1


# ============================================================================
# SECTION HEADER
# ============================================================================
def build_section_header(ws, cell_ref, text):
    """A section title with a short colored underline accent (product
    owner, 2026-07-11: "aby to vypadalo opravdu jako systém") - a border on
    a single unmerged cell, not a new row, so this never shifts any of the
    many hardcoded row numbers elsewhere in this codebase that assume a
    section header consumes exactly one row."""
    ws[cell_ref] = text
    ws[cell_ref].font = FONT_TITLE
    ws[cell_ref].border = Border(bottom=Side(style="medium", color=ACCENT_BLUE))


# ============================================================================
# NAMED RANGES
# ============================================================================
def define_named_range(ws, name, formula):
    """Registers a sheet-scoped Named Range (Excel's native "give this
    formula/range a readable name" feature) - preferred over raw
    `=$P$2#`-style cell references baked into DataValidation/formulas
    (product owner, 2026-07-05: prefer Named Ranges + Data Validation
    lists as the standard filter-dropdown pattern, so the wiring is
    readable and editable directly in Excel without touching this Python
    scaffold). Scoped to the sheet (not workbook-global) so each dashboard
    screen can define its own "TechnicianList"-style name without
    colliding with another screen's. formula should be a full sheet-
    qualified reference/formula, e.g. "TECHNICIAN_SCORECARD!$P$2#"."""
    ws.defined_names[name] = DefinedName(name, attr_text=formula)


# ============================================================================
# FILTER PANEL
# ============================================================================
def build_filter_bar_background(ws, row, col_start, col_end, fill=LIGHT_GREY, height=26):
    """The light-grey toolbar strip a row of filter controls sits on."""
    for col_idx in range(column_index_from_string(col_start), column_index_from_string(col_end) + 1):
        ws.cell(row, col_idx).fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[row].height = height


def build_filter_dropdown(ws, label_cell, label_text, input_range, source_formula, default_formula=None):
    """One dropdown filter control: a label cell plus a (possibly merged)
    input cell with a list-type DataValidation sourced from source_formula.
    Pass a Named Range reference (e.g. "=TechnicianList", defined via
    define_named_range() above) rather than a raw cell reference - the
    standard pattern for every dashboard filter in this workbook. Returns
    the input cell's coordinate so other formulas on the sheet can
    reference the selection."""
    ws[label_cell] = label_text
    ws[label_cell].font = FONT_CARD_LABEL
    ws[label_cell].alignment = Alignment(vertical="center", indent=1)
    first_input_cell = input_range.split(":")[0]
    if ":" in input_range:
        ws.merge_cells(input_range)
    cell = ws[first_input_cell]
    if default_formula:
        cell.value = default_formula
    cell.font = font_card_value(size=13)
    cell.fill = PatternFill("solid", fgColor=WHITE)
    cell.alignment = Alignment(vertical="center", indent=1)
    cell.border = CARD_BORDER
    dv = DataValidation(type="list", formula1=source_formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell)
    return first_input_cell


# ============================================================================
# KPI CARD
# ============================================================================
def build_kpi_card(ws, col_start, col_end, label_row, value_row_start, value_row_end,
                    label, value_formula, value_color=NAVY, fill_color=WHITE, number_format=None):
    """One KPI tile: a merged label cell above a merged, bordered value
    block. Returns the value cell's coordinate (top-left of the merge) so
    the caller can wire conditional formatting or reference it in other
    formulas on the sheet.

    Label row gets a colored top accent (value_color) plus a light grey
    fill, so label+value read as one cohesive card instead of the label
    floating unstyled on the sheet's background above a separate bordered
    box (product owner, 2026-07-11: "aby to vypadalo opravdu jako
    systém")."""
    ws.merge_cells(f"{col_start}{label_row}:{col_end}{label_row}")
    lbl = ws[f"{col_start}{label_row}"]
    lbl.value = label
    lbl.font = FONT_CARD_LABEL
    lbl.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    accent_top = Border(top=Side(style="medium", color=value_color), left=Side(style="thin", color=BORDER_GREY),
                         right=Side(style="thin", color=BORDER_GREY))
    for col in (col_start, col_end):
        ws[f"{col}{label_row}"].border = accent_top
    ws.merge_cells(f"{col_start}{value_row_start}:{col_end}{value_row_end}")
    val = ws[f"{col_start}{value_row_start}"]
    val.value = value_formula
    val.font = font_card_value(color=value_color)
    val.fill = PatternFill("solid", fgColor=fill_color)
    val.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        val.number_format = number_format
    for row in range(value_row_start, value_row_end + 1):
        for col in (col_start, col_end):
            ws[f"{col}{row}"].border = CARD_BORDER
    return f"{col_start}{value_row_start}"


def build_kpi_card_row(ws, cards, label_row, value_row_start, value_row_end):
    """cards: list of (col_start, col_end, label, value_formula, value_color,
    fill_color) tuples. Convenience wrapper for laying out a full row of
    tiles in one call. Returns the list of value-cell coordinates, in order."""
    return [
        build_kpi_card(ws, c1, c2, label_row, value_row_start, value_row_end, label, formula, color, fill)
        for (c1, c2, label, formula, color, fill) in cards
    ]


# ============================================================================
# PROGRESS BAR
# ============================================================================
def build_progress_bar(ws, cell_range, value_ref_or_formula, max_value=100, color=ACCENT_BLUE,
                        number_format='0.0"%"', row_height=20):
    """A single (optionally merged) cell showing value_ref_or_formula, with
    a DataBarRule fill proportional to max_value - the closest Excel-native
    equivalent of a progress bar. value_ref_or_formula may be a bare cell
    reference (e.g. "M9") or a full formula string; a bare reference is
    auto-prefixed with "=" ."""
    anchor = cell_range.split(":")[0]
    if ":" in cell_range:
        ws.merge_cells(cell_range)
    cell = ws[anchor]
    formula = str(value_ref_or_formula)
    cell.value = formula if formula.startswith("=") else f"={formula}"
    cell.number_format = number_format
    cell.font = Font(bold=True, size=11, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row = int("".join(ch for ch in anchor if ch.isdigit()))
    ws.row_dimensions[row].height = row_height
    ws.conditional_formatting.add(
        anchor, DataBarRule(start_type="num", start_value=0, end_type="num", end_value=max_value, color=color),
    )
    return anchor


# ============================================================================
# STATUS BADGE / CONDITIONAL TEXT FORMATTING
# ============================================================================
def build_status_badge_conditional(ws, cell_range, anchor_cell, rules):
    """rules: list of (prefix, fill_color_or_None, font_color_or_None).
    Wraps the `LEFT(anchor,len(prefix))="X"` pattern already used across
    HOME's pipeline status strip, pre-publish check, and (via
    apply_severity_conditional_formatting below) trend indicators, into one
    reusable call instead of each screen hand-writing FormulaRule calls."""
    for prefix, fill_color, font_color in rules:
        kwargs = {}
        if fill_color:
            kwargs["fill"] = PatternFill("solid", fgColor=fill_color)
        if font_color:
            kwargs["font"] = Font(bold=True, color=font_color)
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'LEFT({anchor_cell},{len(prefix)})="{prefix}"'], **kwargs),
        )


def apply_severity_conditional_formatting(ws, cell_range, anchor_cell, thresholds,
                                           below_color=STATUS_CRITICAL, font_size=22, font_color=WHITE):
    """thresholds: [(min_value, color), ...] sorted descending (e.g.
    [(90, STATUS_GOOD), (70, STATUS_WARNING), (50, STATUS_SERIOUS)]).
    Applies stopIfTrue fill+font rules top to bottom for "value >= min",
    then one final rule for "below the lowest threshold" using below_color -
    the 4-tier compliance-severity pattern used on every KPI card that
    carries a percentage."""
    for min_value, color in thresholds:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{anchor_cell}>={min_value}"], fill=PatternFill("solid", fgColor=color),
                        font=Font(bold=True, size=font_size, color=font_color), stopIfTrue=True),
        )
    if thresholds:
        last_threshold = thresholds[-1][0]
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{anchor_cell}<{last_threshold}"], fill=PatternFill("solid", fgColor=below_color),
                        font=Font(bold=True, size=font_size, color=font_color), stopIfTrue=True),
        )


# ============================================================================
# CHART CONTAINER
# ============================================================================
def make_bar_chart(cats_ref, data_ref, color=ACCENT_BLUE, chart_type="col", height=6.5, width=14,
                    legend=False, style=10):
    """A small, undecorated bar chart (no title, no legend by default) -
    the project's consistent "compact chart inside a dashboard tile" look,
    as opposed to DASHBOARD's larger analytical charts (ux_style.py's
    _build_dashboard_charts, which keep titles/legends - a different,
    denser screen)."""
    chart = BarChart()
    chart.type = chart_type
    chart.style = style
    chart.height = height
    chart.width = width
    if not legend:
        chart.legend = None
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = color
    return chart


def make_line_chart(cats_ref, data_ref, color=ACCENT_BLUE, height=6.5, width=14, legend=False, style=2,
                     smooth=False):
    """A small, undecorated line chart - the sparkline substitute (openpyxl
    has no native Excel Sparkline support - see PerformanceEngine.ts's
    consumer, tools/ux_style.py's build_technician_scorecard, for the
    "why")."""
    chart = LineChart()
    chart.style = style
    chart.height = height
    chart.width = width
    if not legend:
        chart.legend = None
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = color
    chart.series[0].graphicalProperties.line.width = 20000
    chart.series[0].smooth = smooth
    return chart


# ============================================================================
# TABLE STYLE (small in-dashboard tables, e.g. "TOP problematic POS" -
# distinct from the full data-entry sheet headers, which stay in
# ux_style.style_header_row/apply_banded_rows since those also carry
# editable-column coloring and dropdown wiring that dashboards don't need)
# ============================================================================
def style_dashboard_table_header(ws, row, columns, labels, fill=LIGHT_GREY):
    for col, label in zip(columns, labels):
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = FONT_CARD_LABEL
        cell.fill = PatternFill("solid", fgColor=fill)


def apply_table_borders(ws, row_start, row_end, columns, border=CARD_BORDER):
    for row in range(row_start, row_end + 1):
        for col in columns:
            ws[f"{col}{row}"].border = border
