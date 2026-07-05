"""
Parses the two raw weekly export files (POS/PPT report, SalesApp report) as
the user actually receives them, so they never have to manually copy-paste
into RAW_DATA/SALESAPP_IMPORT themselves. This is presentation/import-format
convenience only - it does not decide or reinterpret any business rule; it
just finds the real header row (the export may or may not have an
instruction row above it, and column order/exact text varies release to
release) and hands the rows to the same header-name matching that
ImportEngine.ts/ComplianceEngine.ts already do (see desktop_client/engines/
import_engine.py, compliance_engine.py) - so a file that would work if
pasted into the workbook by hand also works dropped in here.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from desktop_client.engines.core_logic import norm  # noqa: E402

Row = list


class ReportParseError(Exception):
    """A raw export file doesn't look like what it was expected to be -
    e.g. neither of the two known shapes (POS report / SalesApp report)
    could find their required columns. Message is user-facing (shown in
    the GUI), so it names what was actually being looked for."""


def _read_all_sheets(path: str) -> list[list[Row]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        while rows and all(v in (None, "") for v in rows[-1]):
            rows.pop()
        if rows:
            sheets.append(rows)
    wb.close()
    return sheets


def _find_header_row(rows: list[Row], required_tokens: list[str], scan_rows: int = 10) -> int:
    """Returns the 0-based index of the first row (within the first
    `scan_rows`) whose normalized cells contain every one of
    `required_tokens` as an exact match or substring - the same
    norm()-based fuzzy matching ImportEngine.ts/ComplianceEngine.ts use for
    column lookup, applied here to first find WHICH row is the header."""
    required = [norm(t) for t in required_tokens]
    for i, row in enumerate(rows[:scan_rows]):
        cells = [norm(str(v)) for v in row if v is not None]
        if all(any(tok in c for c in cells) for tok in required):
            return i
    raise ReportParseError(
        f"Nenašel jsem hlavičkový řádek se sloupci: {', '.join(required_tokens)}. "
        f"Zkontroluj, že jde o správný soubor exportu."
    )


def parse_pos_report(path: str) -> Row:
    """Returns rows shaped like RAW_DATA expects (header row first, data
    rows after) - the POS/PPT weekly export, whichever sheet/row it starts
    on. Required columns mirror ImportEngine.ts's exactCol/col lookups."""
    for rows in _read_all_sheets(path):
        try:
            header_idx = _find_header_row(rows, ["POS", "MARKET", "KATEGORIE", "PTT"])
        except ReportParseError:
            continue
        return rows[header_idx:]
    raise ReportParseError(
        "Tenhle soubor nevypadá jako report POS/PPT (chybí sloupce POS/MARKET/KATEGORIE/PTT)."
    )


def parse_salesapp_report(path: str) -> Row:
    """Returns rows shaped like SALESAPP_IMPORT expects (header row first,
    data rows after) - required columns mirror ComplianceEngine.ts's
    UID/STATE/STORE UID lookups."""
    for rows in _read_all_sheets(path):
        try:
            header_idx = _find_header_row(rows, ["UID", "STATE", "STORE UID"])
        except ReportParseError:
            continue
        return rows[header_idx:]
    raise ReportParseError(
        "Tenhle soubor nevypadá jako report ze SalesApp (chybí sloupce UID/State/Store UID)."
    )


SHEET_HEADER_OFFSET = {
    # RAW_DATA: ImportEngine.ts reads `raw[1]` (0-indexed) as the header row
    # - `raw[0]` is a placeholder/instruction row, not decoration; this is a
    # structural expectation of the engine itself, not just the styled
    # workbook's UI hint text, so a blank row 1 must genuinely be written.
    "RAW_DATA": 1,
    # SALESAPP_IMPORT: ComplianceEngine.ts reads `salesApp[0]` as the header
    # row directly - no placeholder row.
    "SALESAPP_IMPORT": 0,
}


def write_sheet_rows(path: str, sheet_name: str, rows: Row) -> None:
    """Fully overwrites `sheet_name` (must already exist in the workbook,
    e.g. RAW_DATA/SALESAPP_IMPORT) with `rows` (rows[0] is the header),
    respecting each sheet's own header-row offset (see
    SHEET_HEADER_OFFSET) - RAW_DATA needs one blank row above its header
    because ImportEngine.ts reads that fixed position, SALESAPP_IMPORT
    does not."""
    offset = SHEET_HEADER_OFFSET.get(sheet_name, 0)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        ws = wb[sheet_name]
        max_col = max((len(r) for r in rows), default=0)
        max_row = max(ws.max_row, len(rows) + offset)
        for r in range(1, max_row + 1):
            src_idx = r - 1 - offset
            row_values = rows[src_idx] if 0 <= src_idx < len(rows) else []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                v = row_values[c - 1] if c - 1 < len(row_values) else None
                cell.value = v
        wb.save(path)
    finally:
        wb.close()
