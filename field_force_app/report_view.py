"""
Renders a single, self-contained HTML report (tour plan + KPI overview +
"kdo flaká" + long-term trend) straight from the numbers the engines already
computed and wrote into the workbook - no new calculation happens here, this
is presentation only, reading TECHNICIAN_PERFORMANCE_LOG/SUMMARY,
MANAGER_PLAN_PUBLISHED/MANAGER_PLAN and DASHBOARD's already-written cells via
openpyxl. No external assets (no CDN, no network) - opens directly from disk
in any browser, matching the "no external API" constraint the rest of this
project follows.
"""
from __future__ import annotations

import html
from pathlib import Path

import openpyxl


def _sheet_rows(wb, name: str) -> list[list]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    while rows and all(v in (None, "") for v in rows[-1]):
        rows.pop()
    return rows


def _col_index(header: list, name: str) -> int:
    return header.index(name) if name in header else -1


def _read_tour_plan(wb) -> tuple[dict[str, list[dict]], str]:
    """Prefers the published/locked plan (MANAGER_PLAN_PUBLISHED); falls
    back to the Draft (MANAGER_PLAN) if nothing has been published yet, so
    the report is useful right after the Planning stage too."""
    rows = _sheet_rows(wb, "MANAGER_PLAN_PUBLISHED")
    source = "published"
    if len(rows) < 2:
        rows = _sheet_rows(wb, "MANAGER_PLAN")
        source = "draft"
    if len(rows) < 2:
        return {}, source
    header = rows[0]
    c_tech = _col_index(header, "TECHNICIAN")
    c_day = _col_index(header, "DAY")
    c_date = _col_index(header, "DATE")
    c_pos = _col_index(header, "POS")
    c_nazev = _col_index(header, "NAZEV_PROVOZOVNY")
    c_ulice = _col_index(header, "ULICE")
    c_mesto = _col_index(header, "MESTO")
    by_tech: dict[str, list[dict]] = {}
    for row in rows[1:]:
        tech = row[c_tech] if c_tech >= 0 else ""
        if not tech:
            continue
        by_tech.setdefault(str(tech), []).append({
            "day": row[c_day] if c_day >= 0 else "",
            "date": row[c_date] if c_date >= 0 else "",
            "pos": row[c_pos] if c_pos >= 0 else "",
            "nazev": row[c_nazev] if c_nazev >= 0 else "",
            "ulice": row[c_ulice] if c_ulice >= 0 else "",
            "mesto": row[c_mesto] if c_mesto >= 0 else "",
        })
    return by_tech, source


def _read_kpis(wb) -> dict[str, object]:
    if "DASHBOARD" not in wb.sheetnames:
        return {}
    ws = wb["DASHBOARD"]
    return {
        "active_pos": ws["B3"].value or 0,
        "splneno_vcas": ws["C3"].value or 0,
        "nesplneno": ws["D3"].value or 0,
        "open_alerts": ws["E3"].value or 0,
    }


def _read_summary(wb) -> list[dict]:
    rows = _sheet_rows(wb, "TECHNICIAN_PERFORMANCE_SUMMARY")
    if len(rows) < 2:
        return []
    header = rows[0]
    idx = {name: _col_index(header, name) for name in [
        "technician", "region", "compliancePercent", "longRunAvgCompliance",
        "trendDelta", "flakaRiziko", "maxKmDay", "latestYear", "latestWeek",
    ]}
    result = []
    for row in rows[1:]:
        result.append({k: (row[i] if i >= 0 else "") for k, i in idx.items()})
    return result


def _read_long_term_trend(wb) -> dict[str, list[tuple[int, float]]]:
    """Per technician, (monthKey, average compliancePercent that month)
    sorted by monthKey - feeds the long-term trend mini-chart."""
    rows = _sheet_rows(wb, "TECHNICIAN_PERFORMANCE_LOG")
    if len(rows) < 2:
        return {}
    header = rows[0]
    c_tech = _col_index(header, "technician")
    c_month = _col_index(header, "monthKey")
    c_compl = _col_index(header, "compliancePercent")
    by_tech_month: dict[str, dict[int, list[float]]] = {}
    for row in rows[1:]:
        tech = row[c_tech]
        month = row[c_month]
        compl = row[c_compl]
        if not tech or not month:
            continue
        by_tech_month.setdefault(str(tech), {}).setdefault(int(month), []).append(float(compl or 0))
    result: dict[str, list[tuple[int, float]]] = {}
    for tech, months in by_tech_month.items():
        avg_by_month = sorted((m, sum(vs) / len(vs)) for m, vs in months.items())
        result[tech] = avg_by_month
    return result


def _svg_line_chart(points: list[tuple[int, float]], width: int = 320, height: int = 80) -> str:
    """Tiny dependency-free inline SVG line chart (0-100% compliance axis) -
    no matplotlib/JS/CDN needed, keeps this app's only hard dependency
    openpyxl (+ tkinter for the picker window)."""
    if len(points) < 2:
        return "<em>Zatím málo historie na graf.</em>"
    pad = 6
    xs = [p[0] for p in points]
    x_min, x_max = min(xs), max(xs)
    x_span = max(x_max - x_min, 1)

    def sx(x):
        return pad + (width - 2 * pad) * (x - x_min) / x_span

    def sy(y):
        return height - pad - (height - 2 * pad) * max(0, min(100, y)) / 100

    coords = " ".join(f"{sx(x):.1f},{sy(y):.1f}" for x, y in points)
    return (
        f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
        f'xmlns="http://www.w3.org/2000/svg">'
        f'<polyline fill="none" stroke="#1F4E78" stroke-width="2" points="{coords}"/>'
        f"</svg>"
    )


def generate_html_report(workbook_path: str, out_path: str) -> str:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    try:
        kpis = _read_kpis(wb)
        tour_plan, tour_plan_source = _read_tour_plan(wb)
        summary = _read_summary(wb)
        trend = _read_long_term_trend(wb)
    finally:
        wb.close()

    e = html.escape

    def d(v, suffix: str = "") -> str:
        """Displays a cell value, or an em-dash for blank/None (a week with
        no prior data point, e.g. trendDelta on a technician's first
        tracked week) instead of the literal string "None"."""
        return "–" if v in (None, "") else f"{v}{suffix}"

    flaka_rows = "".join(
        f"<tr><td>{e(str(s['technician']))}</td><td>{e(d(s['region']))}</td>"
        f"<td>{d(s['compliancePercent'], '%')}</td><td>{d(s['maxKmDay'], ' km')}</td></tr>"
        for s in summary if str(s.get("flakaRiziko")) == "Ano"
    ) or '<tr><td colspan="4"><em>Nikdo aktuálně nefláká.</em></td></tr>'

    summary_rows = "".join(
        f"<tr><td>{e(str(s['technician']))}</td><td>{e(d(s['region']))}</td>"
        f"<td>{d(s['compliancePercent'], '%')}</td><td>{d(s['longRunAvgCompliance'], '%')}</td>"
        f"<td>{d(s['trendDelta'], '%')}</td><td>{d(s['maxKmDay'], ' km')}</td>"
        f"<td>{'⚠️ Ano' if str(s.get('flakaRiziko')) == 'Ano' else 'Ne'}</td></tr>"
        for s in summary
    ) or '<tr><td colspan="7"><em>Zatím žádná data - spusť nejdřív vyhodnocení.</em></td></tr>'

    trend_blocks = "".join(
        f'<div class="trend-card"><h4>{e(tech)}</h4>{_svg_line_chart(points)}</div>'
        for tech, points in sorted(trend.items())
    ) or "<em>Zatím žádná historie.</em>"

    tour_source_label = "publikovaný plán" if tour_plan_source == "published" else "Draft (ještě nepublikováno)"
    tour_blocks = "".join(
        f'<div class="tour-card"><h4>{e(tech)}</h4><table><tr><th>Den</th><th>Datum</th>'
        f"<th>POS</th><th>Název</th><th>Adresa</th></tr>"
        + "".join(
            f"<tr><td>{e(str(v['day']))}</td><td>{e(str(v['date']))}</td><td>{e(str(v['pos']))}</td>"
            f"<td>{e(str(v['nazev']))}</td><td>{e(str(v['ulice']))}, {e(str(v['mesto']))}</td></tr>"
            for v in visits
        )
        + "</table></div>"
        for tech, visits in sorted(tour_plan.items())
    ) or "<em>Zatím žádný plán - spusť nejdřív zpracování POS/PPT reportu.</em>"

    doc = f"""<!doctype html>
<html lang="cs"><head><meta charset="utf-8">
<title>Field Force Optimizer — report</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Arial, sans-serif; margin: 0; padding: 24px;
          background: #f4f6f9; color: #1a1a1a; }}
  h1 {{ color: #1F4E78; }}
  h2 {{ color: #1F4E78; border-bottom: 2px solid #1F4E78; padding-bottom: 4px; margin-top: 36px; }}
  .kpi-row {{ display: flex; gap: 16px; margin: 16px 0; }}
  .kpi {{ background: white; border-radius: 8px; padding: 16px 20px; box-shadow: 0 1px 3px rgba(0,0,0,.1); flex: 1; }}
  .kpi .label {{ font-size: 12px; color: #6c757d; font-weight: bold; }}
  .kpi .value {{ font-size: 28px; font-weight: bold; color: #1F4E78; }}
  table {{ border-collapse: collapse; width: 100%; background: white; }}
  th, td {{ padding: 6px 10px; border-bottom: 1px solid #e5e5e5; text-align: left; font-size: 13px; }}
  th {{ background: #1F4E78; color: white; }}
  .tour-card, .trend-card {{ background: white; border-radius: 8px; padding: 12px 16px; margin-bottom: 16px;
                              box-shadow: 0 1px 3px rgba(0,0,0,.1); }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 16px; }}
  .note {{ color: #6c757d; font-size: 12px; }}
</style></head>
<body>
<h1>Field Force Optimizer — report</h1>
<p class="note">Vygenerováno appkou z reálných dat workbooku, žádné ruční počítání.</p>

<h2>Přehled</h2>
<div class="kpi-row">
  <div class="kpi"><div class="label">AKTIVNÍ POS</div><div class="value">{kpis.get('active_pos', 0)}</div></div>
  <div class="kpi"><div class="label">SPLNĚNO VČAS</div><div class="value">{kpis.get('splneno_vcas', 0)}</div></div>
  <div class="kpi"><div class="label">NESPLNĚNO</div><div class="value">{kpis.get('nesplneno', 0)}</div></div>
  <div class="kpi"><div class="label">OTEVŘENÁ UPOZORNĚNÍ</div><div class="value">{kpis.get('open_alerts', 0)}</div></div>
</div>

<h2>Kdo fláká</h2>
<table><tr><th>Technik</th><th>Region</th><th>Compliance (posl. týden)</th><th>Nejhorší den (km)</th></tr>
{flaka_rows}
</table>

<h2>Technici — souhrn</h2>
<table><tr><th>Technik</th><th>Region</th><th>Compliance</th><th>Dlouhodobý průměr</th>
<th>Trend</th><th>Nejhorší den (km)</th><th>Fláká?</th></tr>
{summary_rows}
</table>

<h2>Dlouhodobý trend (podle měsíce)</h2>
<div class="grid">{trend_blocks}</div>

<h2>Tour plan ({e(tour_source_label)})</h2>
<div class="grid">{tour_blocks}</div>

</body></html>"""

    Path(out_path).write_text(doc, encoding="utf-8")
    return out_path
