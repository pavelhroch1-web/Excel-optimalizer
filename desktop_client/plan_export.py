"""
Pure logic for the Distribution Client - no GUI dependency, so it can be
unit tested and reused without a display. See distribution_client.py's
module docstring for the full scope/boundary of this app.

FieldForceOptimizer remains the sole source of business logic. Everything
here is file I/O and formatting: read the already-published
TECHNICIAN_PLAN sheet, group it by technician, write one plain (non-
formula) .xlsx per technician. No planning, no compliance, no recompute,
no write-back to the source workbook.
"""

import os
import re
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_NAME = "TECHNICIAN_PLAN"
TECH_COLUMN = "TECHNIK"

# Characters not safe in Windows/macOS/Linux filenames.
_UNSAFE_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')


def sanitize_filename_part(text: str) -> str:
    """Makes a technician name safe to use as a filename component. Pure
    string cleanup - not a business rule, just filesystem safety."""
    cleaned = _UNSAFE_FILENAME_CHARS.sub("_", text).strip()
    return cleaned or "technik"


def read_technician_plan(workbook_path: str):
    """Reads TECHNICIAN_PLAN and groups rows by technician.

    TECHNICIAN_PLAN's cells are live formulas (see
    tools/ux_style.py:build_technician_plan) referencing MANAGER_PLAN, so
    this reads with data_only=True to get the last value Excel calculated
    and cached - present as long as the workbook has been opened (and
    calculated/saved) in real Excel at least once since the last change,
    which is always true in the real workflow (publish happens in Excel,
    this app runs afterward on the saved file).

    Returns (headers, {technician_name: [row_dict, ...]}).
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"List '{SHEET_NAME}' v tomto souboru neexistuje.")
    ws = wb[SHEET_NAME]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) for h in header_row if h not in (None, "")]
    if TECH_COLUMN not in headers:
        raise ValueError(f"Sloupec '{TECH_COLUMN}' v listu {SHEET_NAME} nebyl nalezen.")
    tech_idx = headers.index(TECH_COLUMN)

    by_technician = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = row[: len(headers)]
        if not row or not row[tech_idx]:
            continue
        tech = str(row[tech_idx]).strip()
        if not tech:
            continue
        row_dict = dict(zip(headers, row))
        by_technician.setdefault(tech, []).append(row_dict)

    return headers, by_technician


def week_year_label(rows: list) -> str:
    """Derives the "<Rok>_W<Tyden>" part of the filename from the earliest
    DATUM among a technician's rows, using ISO-8601 week numbering
    (Python's date.isocalendar() - the same ISO week definition
    office-scripts/shared/core.ts's isoWeekNumber() uses, so a filename
    generated here means the same week as anywhere else in the system).
    Falls back to today's week if no valid date is found (should not
    normally happen for a real published plan)."""
    dates = []
    for row in rows:
        v = row.get("DATUM")
        if isinstance(v, datetime):
            dates.append(v.date())
        elif isinstance(v, date):
            dates.append(v)
    reference = min(dates) if dates else date.today()
    iso_year, iso_week, _ = reference.isocalendar()
    return f"{iso_year}_W{iso_week:02d}"


def export_technician_file(headers: list, technician: str, rows: list, output_dir: str) -> str:
    """Writes a new, standalone .xlsx containing only this technician's
    rows (same columns as TECHNICIAN_PLAN, plain values - not formulas, not
    linked to the source workbook in any way). Returns the path written."""
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "ROZPIS"
    out_ws.append(headers)
    for cell in out_ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    date_col_indexes = {i for i, h in enumerate(headers) if h == "DATUM"}
    for row in rows:
        out_ws.append([row.get(h, "") for h in headers])
    if date_col_indexes:
        for row_cells in out_ws.iter_rows(min_row=2, max_row=out_ws.max_row):
            for i in date_col_indexes:
                row_cells[i].number_format = "DD.MM.YYYY"
    for i, h in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        candidate_widths = [len(h) + 2] + [len(str(r.get(h, ""))) + 2 for r in rows]
        out_ws.column_dimensions[col_letter].width = min(max(candidate_widths, default=12), 40)
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = out_ws.dimensions

    filename = f"{sanitize_filename_part(technician)}_{week_year_label(rows)}.xlsx"
    path = os.path.join(output_dir, filename)
    out_wb.save(path)
    return path
