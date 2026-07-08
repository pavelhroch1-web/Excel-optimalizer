"""
Bridges a real .xlsx file (via openpyxl) to the {sheetName: [[cell,...],...]}
dict shape desktop_client/engines/ (and tools/sim/mockWorkbook.ts) operate
on, so the SAME engine port that was verified against the real TypeScript
engines (tools/sim/compare_engines.py) can run directly against a workbook
file on disk.

Scope, deliberately narrow:
  - Reads every sheet's cell values (data_only=False, so any live formula
    elsewhere in the workbook - e.g. TECHNICIAN_PLAN - is left completely
    untouched; the sheets these engines actually consume are plain data
    tables with no formulas, so this has no effect on what they read).
  - After running the engine(s), writes back ONLY the sheets an engine can
    legally touch (see ENGINE_OUTPUT_SHEETS below) - every other sheet in
    the workbook, including any formulas, chart, or styling on it, is
    never opened for writing at all.
  - Always makes a timestamped backup copy of the file before writing,
    since this - unlike the rest of the Distribution Client - does write
    to the source workbook. See docs/ARCHITECTURE.md "Desktop Client local
    engine execution" for why this exists and its risk tradeoffs.
"""
from __future__ import annotations

import datetime
import shutil
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

# Sheets that Import/Planning/Publish/StartTracking/Compliance/Advisor/
# Performance/Reporting Engine may write to, in the real office-scripts/*.ts.
# Nothing outside this set is ever opened for writing - see each engine's
# own file header for its authoritative scope.
ENGINE_OUTPUT_SHEETS = {
    "POS_MASTER", "MANAGER_PLAN", "MANAGER_PLAN_PUBLISHED", "PLAN_LIFECYCLE",
    "VISIT_HISTORY_ACTUAL", "OTHER_VISIT_LOG", "COMPLIANCE_LOG", "ADVISOR_LOG",
    "TECHNICIAN_PERFORMANCE_LOG", "TECHNICIAN_PERFORMANCE_SUMMARY", "TECHNICIAN_TOP_ISSUES",
    "DASHBOARD", "POS_MAP_DATA",
}

# Sheets the engines read as input (beyond the ones they also write).
ENGINE_INPUT_SHEETS = ENGINE_OUTPUT_SHEETS | {
    "RAW_DATA", "POS_STATUS_IMPORT", "ACTIVITY_PLAN", "CONTROL", "TERMINAL_RULES",
    "MARKET_RULES", "CATEGORY_RULES", "CADENCE_RULES", "PARETO_GROUPS", "SCORE_PROFILES",
    "CAPACITY_OVERRIDE", "SALESAPP_IMPORT", "BLACKLIST",
}


def _cell_to_json(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return "" if v is None else v


def read_state(path: str) -> dict[str, list[list]]:
    """Reads only the sheets the engines actually need, as plain values -
    same shape as tools/sim/xlsx_to_json.py produces."""
    wb = openpyxl.load_workbook(path, data_only=False)
    state: dict[str, list[list]] = {}
    for name in ENGINE_INPUT_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_cell_to_json(v) for v in row])
        while rows and all(v == "" for v in rows[-1]):
            rows.pop()
        state[name] = rows
    wb.close()
    return state


def backup_workbook(path: str) -> str:
    p = Path(path)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = p.with_name(f"{p.stem}.backup_{stamp}{p.suffix}")
    shutil.copy2(p, backup_path)
    return str(backup_path)


def write_state(path: str, state: dict[str, list[list]], sheets: set[str]) -> None:
    """Overwrites ONLY `sheets` (must be a subset of ENGINE_OUTPUT_SHEETS)
    in the real workbook file with the engine's output rows. Only cell
    VALUES are touched - existing formatting (header style, banded rows
    from tools/ux_style.py) is left as-is, same as Office Scripts'
    clear(ClearApplyTo.contents) + setValues() pattern."""
    assert sheets <= ENGINE_OUTPUT_SHEETS, f"refusing to write non-engine-output sheet(s): {sheets - ENGINE_OUTPUT_SHEETS}"
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        for name in sheets:
            if name not in state:
                continue
            ws = wb[name]
            new_rows = state[name]
            max_col = max((len(r) for r in new_rows), default=0)
            max_row = max(ws.max_row, len(new_rows))

            for r in range(1, max_row + 1):
                row_values = new_rows[r - 1] if r - 1 < len(new_rows) else []
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        # A non-top-left cell of a merged range (e.g. a
                        # dashboard title banner or KPI tile label styled by
                        # tools/ux_style.py) - openpyxl only allows writing
                        # the top-left cell of a merge; every other cell in
                        # the range carries no independent value to begin
                        # with, so there is nothing to preserve or write
                        # here. Skipping it (rather than erroring, as a bare
                        # `.value = v` assignment does) is the correct
                        # behaviour, not a fallback - real Excel's own
                        # setValues() has the same constraint on merged
                        # ranges.
                        continue
                    v = row_values[c - 1] if c - 1 < len(row_values) else None
                    if v == "":
                        v = None
                    cell.value = v
        wb.save(path)
    finally:
        wb.close()
