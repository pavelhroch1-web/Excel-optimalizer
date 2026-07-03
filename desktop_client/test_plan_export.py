"""
Unit tests for plan_export.py (no GUI dependency - see distribution_client.py
for the tkinter app that uses this).

Run with: python3 desktop_client/test_plan_export.py
No pytest dependency - plain asserts, matching this project's existing
lightweight testing style (tests/core.test.ts).
"""

import os
import shutil
import tempfile
from datetime import date, datetime

import openpyxl

import plan_export as pe

passed = 0
failed = 0


def test(name, fn):
    global passed, failed
    try:
        fn()
        passed += 1
        print(f"  PASS  {name}")
    except Exception as e:
        failed += 1
        print(f"  FAIL  {name}\n        {e}")


# ==============================================================================
print("sanitize_filename_part()")
# ==============================================================================

def _sanitize_removes_unsafe_chars():
    assert pe.sanitize_filename_part("Novak/Jan") == "Novak_Jan"


def _sanitize_strips_whitespace():
    assert pe.sanitize_filename_part("  Novak  ") == "Novak"


def _sanitize_blank_name_falls_back():
    assert pe.sanitize_filename_part("   ") == "technik"


def _sanitize_strips_diacritics():
    assert pe.sanitize_filename_part("Novák Dvořák Malý") == "Novak Dvorak Maly"


def _sanitize_preserves_case():
    assert pe.sanitize_filename_part("ŘEHOŘ") == "REHOR"


test("removes filesystem-unsafe characters", _sanitize_removes_unsafe_chars)
test("strips surrounding whitespace", _sanitize_strips_whitespace)
test("falls back to a placeholder for an empty/blank name", _sanitize_blank_name_falls_back)
test("strips diacritics", _sanitize_strips_diacritics)
test("diacritic stripping preserves original case", _sanitize_preserves_case)


# ==============================================================================
print("week_year_label()")
# ==============================================================================

def _week_year_ordinary():
    label = pe.week_year_label([{"DATUM": datetime(2026, 7, 27)}])  # known ISO week 31
    assert label == "2026_W31", label


def _week_year_multiple_dates_same_week():
    label = pe.week_year_label([{"DATUM": datetime(2026, 7, 28)}, {"DATUM": datetime(2026, 7, 27)}])
    assert label == "2026_W31", label


def _week_year_iso53_boundary():
    label = pe.week_year_label([{"DATUM": date(2027, 1, 1)}])  # Jan 1 2027 = ISO week 53 of 2026
    assert label == "2026_W53", label


def _week_year_plain_date_object():
    label = pe.week_year_label([{"DATUM": date(2026, 7, 27)}])
    assert label == "2026_W31", label


def _week_year_contiguous_range():
    # weeks 31 (27.7.) through 34 (17.8.2026), no gaps - a technician's
    # Draft+Published rows can legitimately span the whole campaign, not
    # just one week.
    label = pe.week_year_label([
        {"DATUM": date(2026, 8, 17)},  # week 34
        {"DATUM": date(2026, 7, 27)},  # week 31
        {"DATUM": date(2026, 8, 3)},   # week 32
        {"DATUM": date(2026, 8, 10)},  # week 33
    ])
    assert label == "2026_W31-W34", label


def _week_year_noncontiguous_weeks():
    label = pe.week_year_label([{"DATUM": date(2026, 7, 27)}, {"DATUM": date(2026, 8, 17)}])  # W31, W34, gap
    assert label == "2026W31+2026W34", label


def _week_year_crosses_iso_year_boundary():
    label = pe.week_year_label([{"DATUM": date(2026, 12, 28)}, {"DATUM": date(2027, 1, 4)}])  # W53/2026, W01/2027
    assert label == "2026W53+2027W01", label


test("ordinary date -> correct ISO week/year", _week_year_ordinary)
test("multiple dates within the same week collapse to one week label", _week_year_multiple_dates_same_week)
test("ISO week 53 boundary matches office-scripts/shared/core.ts's isoWeekNumber()", _week_year_iso53_boundary)
test("accepts a plain date object, not just datetime", _week_year_plain_date_object)
test("contiguous multi-week span -> W31-W34 range", _week_year_contiguous_range)
test("non-contiguous weeks -> explicit list, not a misleading range", _week_year_noncontiguous_weeks)
test("weeks crossing an ISO year boundary -> explicit per-week year", _week_year_crosses_iso_year_boundary)


# ==============================================================================
print("read_technician_plan() / export_technician_file()")
# ==============================================================================

def _make_seed_workbook(path, rows):
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet(pe.SHEET_NAME)
    headers = ["DATUM", "DEN", "TECHNIK", "POS", "NÁZEV PROVOZOVNY", "ULICE", "MĚSTO", "OBLAST", "AKTIVITA", "POZNÁMKA"]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)
    return headers


def _read_groups_by_technician():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "seed.xlsx")
        _make_seed_workbook(path, [
            [date(2026, 7, 27), "Pondělí", "Novak", "POS1", "P1", "U1", "M1", "O1", "LOS: Gems", ""],
            [date(2026, 7, 28), "Úterý", "Novak", "POS2", "P2", "U2", "M2", "O2", "LOS: Gems", ""],
            [date(2026, 7, 27), "Pondělí", "Svoboda", "POS3", "P3", "U3", "M3", "O3", "LOT: Sportka", ""],
        ])
        headers, by_tech = pe.read_technician_plan(path)
        assert headers[0] == "DATUM" and "TECHNIK" in headers
        assert set(by_tech.keys()) == {"Novak", "Svoboda"}
        assert len(by_tech["Novak"]) == 2
        assert len(by_tech["Svoboda"]) == 1


def _blank_technician_rows_are_skipped():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "seed.xlsx")
        _make_seed_workbook(path, [
            [date(2026, 7, 27), "Pondělí", "", "", "", "", "", "", "", ""],  # blank TECHNIK/POS row (unpublished cap rows in the real sheet)
            [date(2026, 7, 27), "Pondělí", "Novak", "POS1", "P1", "U1", "M1", "O1", "", ""],
        ])
        headers, by_tech = pe.read_technician_plan(path)
        assert list(by_tech.keys()) == ["Novak"]
        assert len(by_tech["Novak"]) == 1


def _missing_sheet_raises_clear_error():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "empty.xlsx")
        openpyxl.Workbook().save(path)
        try:
            pe.read_technician_plan(path)
            raise AssertionError("expected a ValueError for a missing TECHNICIAN_PLAN sheet")
        except ValueError as e:
            assert pe.SHEET_NAME in str(e)


def _export_writes_only_that_technicians_rows():
    with tempfile.TemporaryDirectory() as tmp:
        headers = ["DATUM", "DEN", "TECHNIK", "POS", "NÁZEV PROVOZOVNY", "ULICE", "MĚSTO", "OBLAST", "AKTIVITA", "POZNÁMKA"]
        rows = [
            {"DATUM": date(2026, 7, 27), "DEN": "Pondělí", "TECHNIK": "Novak", "POS": "POS1",
             "NÁZEV PROVOZOVNY": "P1", "ULICE": "U1", "MĚSTO": "M1", "OBLAST": "O1", "AKTIVITA": "", "POZNÁMKA": ""},
        ]
        out_path = pe.export_technician_file(headers, "Novak", rows, tmp)
        assert os.path.basename(out_path) == "Novak_2026_W31.xlsx", out_path
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        written_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(written_rows) == 1
        assert written_rows[0][2] == "Novak"  # TECHNIK column


def _export_does_not_touch_source_workbook():
    with tempfile.TemporaryDirectory() as tmp:
        src_path = os.path.join(tmp, "seed.xlsx")
        _make_seed_workbook(src_path, [
            [date(2026, 7, 27), "Pondělí", "Novak", "POS1", "P1", "U1", "M1", "O1", "", ""],
        ])
        before_mtime = os.path.getmtime(src_path)
        headers, by_tech = pe.read_technician_plan(src_path)
        out_dir = os.path.join(tmp, "out")
        os.makedirs(out_dir)
        pe.export_technician_file(headers, "Novak", by_tech["Novak"], out_dir)
        after_mtime = os.path.getmtime(src_path)
        assert before_mtime == after_mtime, "export must never modify the source workbook"


test("groups rows by technician, preserving headers", _read_groups_by_technician)
test("rows with a blank TECHNIK/POS (unused formula-view cap rows) are skipped", _blank_technician_rows_are_skipped)
test("a workbook without TECHNICIAN_PLAN raises a clear, catchable error", _missing_sheet_raises_clear_error)
test("exported file contains only the requested technician's rows, correctly named", _export_writes_only_that_technicians_rows)
test("exporting never modifies the source workbook file", _export_does_not_touch_source_workbook)


print(f"\n{passed} passed, {failed} failed")
if failed:
    raise SystemExit(1)
