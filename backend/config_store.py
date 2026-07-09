"""Persisted planning configuration - the "config, separated from data"
half of the stateless architecture (governing principle #7).

Config is everything the manager decides ONCE and reuses every week: the
rule tables (terminal/market/category/cadence/scoring/capacity), the
CONTROL settings, the translated Activity Plan, and any per-POS manager
overrides. It is deliberately kept apart from the weekly *data* (the fresh
SalesApp / POS / Activity-Plan exports the manager uploads each run), so
that uploading data never silently changes the rules and editing the rules
never needs a fresh upload.

For this increment the store is *seeded from the proven scaffold workbook*
- the exact config already validated in the Excel - rather than re-typed by
hand (governing principle #12: reuse proven logic, don't reinvent it). The
seam is intentionally narrow (`load_config_state()` returns a plain
{sheetName: rows} dict) so the backing store can later become a standalone
JSON/DB file without any engine or pipeline change.

This module owns NO business logic. It only reads config cells.
"""
from __future__ import annotations

import datetime
import os

import openpyxl

# Sheets that are configuration (persisted, reused every week), as opposed
# to the weekly data uploads (RAW_DATA, SALESAPP_IMPORT). ACTIVITY_PLAN is
# config here because the engine consumes the *translated* Activity Plan;
# translating a fresh raw Activity-Plan export into this shape is a separate,
# already-existing step (see BACKLOG) and is deliberately out of scope for
# this increment.
CONFIG_SHEETS = (
    "CONTROL",
    "ACTIVITY_PLAN",
    "TERMINAL_RULES",
    "MARKET_RULES",
    "CATEGORY_RULES",
    "CADENCE_RULES",
    "PARETO_GROUPS",
    "SCORE_PROFILES",
    "CAPACITY_OVERRIDE",
    "BLACKLIST",
)

# Per-POS manager decisions that must survive a from-scratch rebuild of
# POS_MASTER. They live keyed by posId in the config store, not in the
# uploaded data, and are re-applied to the fresh POS_MASTER on every run.
OVERRIDE_COLUMNS = (
    "managerOverrideType",
    "managerOverridePriority",
    "managerOverrideTechnician",
    "plannerNotes",
)

# Where the seed workbook lives. Defaults to the repo's scaffold; the
# backend overrides this with the workbook it downloads from GitHub.
DEFAULT_SEED_WORKBOOK = os.environ.get(
    "CONFIG_SEED_WORKBOOK",
    os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "workbook",
        "FieldForceOptimizer_V11_scaffold.xlsx",
    ),
)


def _cell_to_json(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return "" if v is None else v


def _read_sheet(wb, name: str) -> list[list]:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_to_json(v) for v in row])
    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    return rows


def load_config_state(seed_workbook: str | None = None) -> dict[str, list[list]]:
    """Returns the config sheets as a {sheetName: rows} dict - the config
    half of the engine state the stateless pipeline assembles."""
    path = seed_workbook or DEFAULT_SEED_WORKBOOK
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        state: dict[str, list[list]] = {}
        for name in CONFIG_SHEETS:
            if name in wb.sheetnames:
                state[name] = _read_sheet(wb, name)
        return state
    finally:
        wb.close()


def load_pos_overrides(seed_workbook: str | None = None) -> dict[str, dict]:
    """Returns {posId: {overrideColumn: value}} for every POS that carries a
    non-empty manager override, so the pipeline can re-apply them to a
    freshly imported POS_MASTER. Empty today (no overrides set yet), but the
    seam exists so the manager's future overrides are honoured without
    re-touching the engine."""
    path = seed_workbook or DEFAULT_SEED_WORKBOOK
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        pm = wb["POS_MASTER"]
        header = [c.value for c in next(pm.iter_rows(min_row=1, max_row=1))]
        idx = {n: i for i, n in enumerate(header)}
        if "posId" not in idx:
            return {}
        overrides: dict[str, dict] = {}
        for row in pm.iter_rows(min_row=2, values_only=True):
            pos_id = row[idx["posId"]]
            if pos_id in (None, ""):
                continue
            present = {
                col: row[idx[col]]
                for col in OVERRIDE_COLUMNS
                if col in idx and row[idx[col]] not in (None, "")
            }
            if present:
                overrides[str(pos_id)] = present
        return overrides
    finally:
        wb.close()
