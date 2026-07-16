"""Excel -> SQLite import (the ONLY way real data enters the datastore).

Reads an authoritative workbook (the current real Excel, or fresh exports)
and loads it into the normalised SQLite tables. Faithful, repeatable, and
idempotent where it matters:
  - pos_master     : upsert by pos_id (+ audit changed fields to history)
  - salesapp_visits: insert, deduped by UID (re-importing overlapping exports
    is harmless - matches the engine's UID dedup)
  - campaigns      : replaced from ACTIVITY_PLAN
  - config         : CONTROL settings + rule tables (as JSON)
  - technicians    : derived from POS assignments + SalesApp executors
  - closed_pos     : derived from POS status/closedSince*

Linking a SalesApp visit to a POS is done by the engine (authoritative); here
we store the raw SalesApp fields faithfully (incl. store_uid + start/finish
times for later route ordering). Plan-vs-reality matching is a later layer.
"""
from __future__ import annotations

import json

import openpyxl

import db


# Raw client exports vary: some have a blank leading row (header on row 2) and
# localized column names. _rows finds the real header row (first row with enough
# non-empty cells) and, given an alias map, translates raw column names to the
# normalized keys the importers use — so files import with no manual edits.
def _norm_header(h) -> str:
    return " ".join(str(h).strip().upper().split())


def _rows(ws, aliases: dict | None = None, min_cells: int = 2):
    it = ws.iter_rows(values_only=True)
    header = []
    for row in it:
        if sum(1 for c in row if c not in (None, "")) >= min_cells:
            header = list(row)
            break  # `it` is now positioned on the first data row
    hidx: dict = {}
    for i, h in enumerate(header):
        if h in (None, ""):
            continue
        raw = str(h).strip()
        hidx.setdefault(raw, i)  # exact raw name always resolvable
        if aliases:
            key = _norm_header(h)
            alias = aliases.get(key)
            if alias is None:  # tolerate truncated/variant headers via prefix
                alias = next((v for k, v in aliases.items() if key.startswith(k) or k.startswith(key)), None)
            if alias:
                hidx.setdefault(alias, i)
    return hidx, it, header


def _g(row, hidx, name):
    i = hidx.get(name)
    if i is None or i >= len(row):
        return None
    v = row[i]
    return v if v not in ("", None) else None


# ---------------------------------------------------------------------------

# Raw client POS export ("Základní údaje o prodejních místech") column names ->
# the normalized keys _POS_MAP consumes. Keys are upper-cased/space-collapsed
# (see _norm_header). The app's own template/scaffold already uses the
# normalized names, so those pass through unchanged.
_POS_HEADER_ALIASES = {
    "ČÍSLO TERMINÁLU": "terminalId",
    "POS": "posId",
    "TYP TERMINÁLU": "terminalType",
    "MARKET": "market",
    "KATEGORIE": "category",
    "KATEGORIZACE": "classification",
    "NAZEV PROVOZOVNY": "nazev", "NÁZEV PROVOZOVNY": "nazev",
    "ULICE": "street",
    "ČÍSLO POPISNÉ/ORIENTAČNÍ": "houseNumber",
    "MĚSTO": "city",
    "OBLAST": "area",
    "POS AREA": "posArea",
    "TECHNIK": "assignedTechnician",
    "X": "gpsX", "Y": "gpsY",
    "PTT": "ppt", "PPT": "ppt",
    "STATUS": "status",
}

_POS_MAP = {
    "posId": "pos_id", "terminalId": "terminal_id",
    "nazev": "name", "street": "street", "houseNumber": "house_number",
    "city": "city", "area": "area", "posArea": "pos_area", "category": "category",
    "market": "market", "classification": "classification", "terminalType": "terminal_type",
    "ppt": "ppt", "gpsX": "gps_x", "gpsY": "gps_y", "assignedTechnician": "technician",
    "managerOverrideType": "manager_override_type",
}


def import_pos_master(conn, ws) -> dict:
    """Import POS master with historical memory: every changed field (esp. PPT,
    status) is recorded to pos_master_history, and POS absent from this import
    are marked inactive (not deleted). Returns a diff summary for the UI."""
    import history
    hidx, it, _ = _rows(ws, _POS_HEADER_ALIASES)
    # Preload the tracked fields of every existing POS once, so we can diff
    # without a SELECT per row (11k+ POS).
    existing = {str(r["pos_id"]): dict(r) for r in db.get(
        "SELECT pos_id, ppt, active, technician, terminal_type, classification, "
        "market, category, name FROM pos_master")}
    seen: set[str] = set()
    n = new_pos = changed_pos = ppt_changed = 0
    for row in it:
        pid = _g(row, hidx, "posId")
        if pid is None:
            continue
        pid = str(pid)
        seen.add(pid)
        vals = {dst: _g(row, hidx, src) for src, dst in _POS_MAP.items()}
        vals["pos_id"] = pid
        status = _g(row, hidx, "status")
        vals["active"] = 0 if (status and str(status).upper() in ("CLOSED", "ZAVRENO", "ZAVŘENO")) else 1

        old = existing.get(pid)
        changes = history.record_pos_changes(conn, pid, old, vals, source="import")
        if old is None:
            new_pos += 1
        elif changes:
            changed_pos += 1
            if history._norm(old.get("ppt")) != history._norm(vals.get("ppt")):
                ppt_changed += 1

        fields = ", ".join(vals.keys())
        marks = ", ".join("?" for _ in vals)
        updates = ", ".join(f"{k}=excluded.{k}" for k in vals if k != "pos_id")
        conn.execute(
            f"INSERT INTO pos_master ({fields}) VALUES ({marks}) "
            f"ON CONFLICT(pos_id) DO UPDATE SET {updates}, updated_at=datetime('now'), last_seen=datetime('now')",
            tuple(vals.values()))
        # closed POS
        csw = _g(row, hidx, "closedSinceWeek")
        if csw is not None or vals["active"] == 0:
            conn.execute(
                "INSERT INTO closed_pos (pos_id, closed_on, reason, source) VALUES (?, ?, ?, 'import') "
                "ON CONFLICT(pos_id) DO NOTHING",
                (pid, str(csw) if csw is not None else None, "status/closedSince"))
        n += 1

    # POS present in the DB but missing from this import -> Inactive (kept).
    inactivated = history.mark_missing_inactive(conn, seen, source="import")
    summary = {"total": n, "new": new_pos, "changed": changed_pos,
               "pptChanged": ppt_changed, "inactivated": inactivated}
    history.log_event("import", "pos_master", None, summary, conn=conn)
    try:
        import clustering
        clustering.rebuild()  # GPS may have changed -> refresh micro-clusters
    except Exception:  # noqa: BLE001
        pass
    return summary


def import_salesapp(conn, ws, filename: str | None = None) -> int:
    hidx, it, header = _rows(ws)
    purpose_cols = [(str(h), i) for i, h in enumerate(header) if str(h).startswith("Účel")]

    # Link SalesApp visit -> POS the same way the engine does:
    # SalesApp "Store UID" == pos_master.terminal_id -> posId.
    term_to_pos = {str(r["terminal_id"]): r["pos_id"]
                   for r in conn.execute(
                       "SELECT terminal_id, pos_id FROM pos_master WHERE terminal_id IS NOT NULL")}

    cur = conn.execute(
        "INSERT INTO salesapp_imports (filename, row_count) VALUES (?, 0)", (filename,))
    import_id = cur.lastrowid

    def purpose_and_role(row):
        hit, is_oz, is_tech = [], False, False
        for h, i in purpose_cols:
            if i < len(row) and row[i] not in (None, "", 0):
                hit.append(h.replace("Účel návštevy", "").strip(" -"))
                # header form: "Účel návštevy - OZ - ..." / "- Technik - ..."
                seg = h.replace("Účel návštevy", "")
                if "OZ" in seg:
                    is_oz = True
                if "Technik" in seg:
                    is_tech = True
        role = "OZ" if (is_oz and not is_tech) else ("TECHNIK" if is_tech and not is_oz else None)
        return ("; ".join(hit) if hit else None), role

    batch, n = [], 0
    for row in it:
        uid = _g(row, hidx, "UID")
        if uid is None:
            continue
        purpose, role = purpose_and_role(row)
        store_uid = _g(row, hidx, "Store UID")
        pos_id = term_to_pos.get(str(store_uid)) if store_uid is not None else None
        batch.append((
            str(uid), pos_id, store_uid,
            _g(row, hidx, "Store"), _g(row, hidx, "Store address"), _g(row, hidx, "Agency region"),
            _g(row, hidx, "Executor"), _g(row, hidx, "Executor UID"), role,
            _g(row, hidx, "Date"), _g(row, hidx, "Started at"), _g(row, hidx, "Finished at"),
            _g(row, hidx, "Real duration (h)"), purpose, import_id))
        n += 1
        if len(batch) >= 1000:
            _flush_visits(conn, batch); batch = []
    if batch:
        _flush_visits(conn, batch)
    conn.execute("UPDATE salesapp_imports SET row_count=? WHERE id=?", (n, import_id))
    return n


def _flush_visits(conn, batch):
    conn.executemany(
        "INSERT INTO salesapp_visits (uid, pos_id, store_uid, store_name, store_address, region, "
        "technician, executor_uid, visitor_role, visit_date, started_at, finished_at, real_duration, purpose, import_id) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(uid) DO NOTHING",
        batch)


def import_activity_plan(conn, ws) -> int:
    hidx, it, _ = _rows(ws)
    conn.execute("DELETE FROM campaigns")
    n = 0
    for row in it:
        name = _g(row, hidx, "ACTIVITY")
        if name is None:
            continue
        odhad_raw = _g(row, hidx, "ODHAD_NAVSTEV_ZA_KAMPAN")
        try:
            target = int(float(odhad_raw)) if odhad_raw not in (None, "") else None
        except (ValueError, TypeError):
            target = None
        conn.execute(
            "INSERT INTO campaigns (kind, name, year, start_week, end_week, priority, override_gap, estimate, target_visits) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (_g(row, hidx, "TYPE"), str(name), 2026, _g(row, hidx, "START_WEEK"),
             _g(row, hidx, "END_WEEK"), _g(row, hidx, "PRIORITY"), _g(row, hidx, "OVERRIDE_GAP"),
             str(odhad_raw or ""), target))
        n += 1
    return n


def import_config(conn, wb) -> int:
    n = 0
    if "CONTROL" in wb.sheetnames:
        hidx, it, _ = _rows(wb["CONTROL"])
        for row in it:
            key = _g(row, hidx, "SETTING")
            if key is None:
                continue
            conn.execute("INSERT INTO config (key, value) VALUES (?, ?) "
                         "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                         (str(key), str(_g(row, hidx, "VALUE") or "")))
            n += 1
    # rule tables stored as JSON under config (engine-ready, extensible)
    for sheet, key, cols in (
        ("TERMINAL_RULES", "terminal_rules", ("TYP TERMINALU", "ACTIVE")),
        ("MARKET_RULES", "market_rules", ("MARKET", "ACTIVE")),
        ("CATEGORY_RULES", "category_rules", ("CATEGORY", "RULE")),
    ):
        if sheet not in wb.sheetnames:
            continue
        hidx, it, _ = _rows(wb[sheet])
        data = []
        for row in it:
            k = _g(row, hidx, cols[0])
            if k is None:
                continue
            data.append({cols[0]: k, cols[1]: _g(row, hidx, cols[1])})
        conn.execute("INSERT INTO config (key, value) VALUES (?, ?) "
                     "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                     (key, json.dumps(data, ensure_ascii=False)))
    return n


def derive_technicians(conn) -> int:
    conn.execute(
        "INSERT OR IGNORE INTO technicians (name) "
        "SELECT DISTINCT technician FROM pos_master WHERE technician IS NOT NULL AND technician<>'' "
        "UNION SELECT DISTINCT technician FROM salesapp_visits WHERE technician IS NOT NULL AND technician<>''")
    # Auto-rule (only where NOT manually set): service number 3xx => OZ, or the
    # person's OZ visits outnumber TECHNIK visits. Manual roles (Admin/Manager,
    # or hand overrides) are never touched.
    conn.execute(
        "UPDATE technicians SET role='OZ' WHERE manual_role=0 AND ("
        "  name GLOB '3[0-9][0-9]*' "
        "  OR name IN (SELECT technician FROM salesapp_visits WHERE technician IS NOT NULL "
        "              GROUP BY technician HAVING SUM(visitor_role='OZ') > SUM(visitor_role='TECHNIK')))")
    # anyone auto-classified and not OZ defaults back to TECHNIK
    conn.execute("UPDATE technicians SET role='TECHNIK' WHERE manual_role=0 AND role NOT IN ('OZ')")
    return conn.execute("SELECT COUNT(*) c FROM technicians").fetchone()[0]


# TourPlan column -> published_plans column. Header matching is case/diacritics
# tolerant, so the manager's own exported TourPlan loads as-is.
_TOURPLAN_MAP = {
    "WEEK": "week", "TYDEN": "week", "TÝDEN": "week", "TOURPLAN": "week",
    "DATE": "plan_date", "DATUM": "plan_date",
    "DAY": "day", "DEN": "day",
    "TECHNICIAN": "technician", "TECHNIK": "technician",
    "POS": "pos_id", "POSID": "pos_id",
    "KATEGORIE": "category", "CATEGORY": "category",
    "NAZEV": "name", "NÁZEV": "name", "NAME": "name",
    "ULICE": "street", "CISLO": "house_number", "ČÍSLO": "house_number",
    "MESTO": "city", "MĚSTO": "city", "CITY": "city",
    "OBLAST": "area", "AREA": "area", "POSAREA": "pos_area",
    "PPT": "ppt", "REASON": "reason", "DUVOD": "reason", "DŮVOD": "reason",
    "GROUP": "day_group", "SKUPINA": "day_group", "SEQ": "day_seq",
}


def _tourplan_header_row(ws, max_scan: int = 8):
    """The header can sit below blank rows; find the row that has POS + a
    technician + a week/plan column. Returns (row_index, [UPPER headers])."""
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).strip().upper() if c is not None else "" for c in row]
        if "POS" in cells and ("TECHNICIAN" in cells or "TECHNIK" in cells) and \
           ("WEEK" in cells or "TYDEN" in cells or "TÝDEN" in cells or "TOURPLAN" in cells):
            return ri, cells
    return None, None


def _find_tourplan_sheet(wb):
    for name in wb.sheetnames:
        ri, _ = _tourplan_header_row(wb[name])
        if ri:
            return wb[name]
    return None


def _norm_name_tokens(s):
    """A name as a set of alphabetic tokens (diacritics/order/number-prefix
    ignored), so 'Vlk Pavel', ' Pavel Vlk' and '604 Pavel Vlk' all match."""
    import re
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    return frozenset(t for t in re.split(r"[^a-z]+", s) if len(t) >= 2)


def _canonical_tech_map():
    """token-set -> canonical technician name AS IT APPEARS IN SALESAPP (reality),
    so an uploaded plan's 'Surname Firstname' names line up with the reality
    'Firstname Surname' names — plan-vs-reality per technician then matches. The
    reality name (most frequent) wins over any other spelling."""
    m = {}
    for r in db.get("SELECT technician, COUNT(*) n FROM salesapp_visits "
                    "WHERE technician IS NOT NULL AND technician<>'' "
                    "GROUP BY technician ORDER BY n DESC"):
        toks = _norm_name_tokens(r["technician"])
        if toks and toks not in m:   # most frequent (reality) name wins
            m[toks] = r["technician"]
    return m


def import_tourplan(path: str, filename: str | None = None) -> dict:
    """Ingest a manager-exported TourPlan as a PUBLISHED plan, so plan-vs-reality
    (TourPlan fulfilment, missed planned POS) can be computed against SalesApp.
    The plan rows come straight from the uploaded file - nothing is hardcoded.
    Re-importing a week replaces which snapshot is 'Published' for that week."""
    import datetime
    db.init_db()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _find_tourplan_sheet(wb)
        if ws is None:
            raise ValueError("TourPlan: nenašel jsem list se sloupci WEEK/Tourplan + TECHNICIAN/Technik + POS.")
        header_row, header = _tourplan_header_row(ws)
        col = {}  # published_plans column -> source index
        for i, h in enumerate(header):
            dst = _TOURPLAN_MAP.get(h.replace(" ", ""))
            if dst and dst not in col:
                col[dst] = i
        if "week" not in col or "technician" not in col or "pos_id" not in col:
            raise ValueError("TourPlan: chybí sloupec s týdnem, technikem nebo POS.")
        tech_map = _canonical_tech_map()

        def year_of(datev):
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d. %m. %Y"):
                try:
                    return datetime.datetime.strptime(str(datev).strip()[:10], fmt).year
                except (ValueError, TypeError):
                    pass
            if isinstance(datev, datetime.datetime):
                return datev.year
            return None

        sid = "imp-" + datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        with open(path, "rb") as fh:
            blob = fh.read()
        rows = []
        weeks = {}   # (year, week) seen
        unmatched = set()
        this_year = datetime.date.today().year
        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            wk = r[col["week"]] if col["week"] < len(r) else None
            pos = r[col["pos_id"]] if col["pos_id"] < len(r) else None
            if wk in (None, "") or pos in (None, ""):
                continue
            try:
                wk = int(wk)
            except (ValueError, TypeError):
                continue
            datev = r[col["plan_date"]] if "plan_date" in col and col["plan_date"] < len(r) else None
            yr = year_of(datev) or this_year
            vals = {"snapshot_id": sid, "year": yr, "week": wk}
            for dst, i in col.items():
                if dst in ("week",):
                    continue
                v = r[i] if i < len(r) else None
                vals[dst] = str(v).strip() if (dst not in ("ppt", "day_group", "day_seq") and v is not None) else v
            vals["pos_id"] = str(pos).strip()
            # Map the plan's technician name to the datastore's canonical name so
            # plan-vs-reality per technician lines up.
            raw_tech = vals.get("technician")
            if raw_tech:
                canon = tech_map.get(_norm_name_tokens(raw_tech))
                if canon:
                    vals["technician"] = canon
                else:
                    unmatched.add(raw_tech)
            rows.append(vals)
            weeks[(yr, wk)] = True

        if not rows:
            raise ValueError("TourPlan: žádné platné řádky (WEEK + POS).")

        conn = db.connect()
        try:
            conn.execute(
                "INSERT INTO snapshots (id, message, published_week, source_files, kind, state_blob) "
                "VALUES (?,?,?,?, 'imported_plan', ?)",
                (sid, f"Nahraný TourPlan {filename or ''}".strip(),
                 min(w for _, w in weeks), json.dumps({"tourplan": filename}), blob))
            cols = ["snapshot_id", "year", "week", "plan_date", "day", "technician", "pos_id",
                    "category", "name", "street", "house_number", "city", "area", "pos_area",
                    "ppt", "reason", "day_group", "day_seq"]
            conn.executemany(
                f"INSERT INTO published_plans ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})",
                [tuple(rw.get(c) for c in cols) for rw in rows])
            # This snapshot becomes the Published plan for each of its weeks
            # (PK year,week -> only one snapshot is 'Published' per week).
            for (yr, wk) in weeks:
                conn.execute(
                    "INSERT INTO plan_lifecycle (year, week, status, snapshot_id) VALUES (?,?, 'Published', ?) "
                    "ON CONFLICT(year, week) DO UPDATE SET status='Published', snapshot_id=excluded.snapshot_id, "
                    "updated_at=datetime('now')", (yr, wk, sid))
            conn.commit()
        finally:
            conn.close()
        try:
            import history
            history.log_event("import", "tourplan", sid,
                              {"rows": len(rows), "weeks": sorted(w for _, w in weeks)})
        except Exception:  # noqa: BLE001
            pass
        return {"snapshot": sid, "rows": len(rows),
                "weeks": sorted(w for _, w in weeks),
                "technicians": len({rw["technician"] for rw in rows}),
                "unmatchedTechnicians": sorted(unmatched)}
    finally:
        wb.close()


def import_workbook(path: str, filename: str | None = None) -> dict:
    """Import everything from one workbook into SQLite. Returns per-table counts."""
    db.init_db()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    conn = db.connect()
    try:
        pos_summary = import_pos_master(conn, wb["POS_MASTER"]) if "POS_MASTER" in wb.sheetnames else {"total": 0}
        result = {
            "pos_master": pos_summary["total"],
            "pos_diff": pos_summary,
            "salesapp_visits": import_salesapp(conn, wb["SALESAPP_IMPORT"], filename) if "SALESAPP_IMPORT" in wb.sheetnames else 0,
            "campaigns": import_activity_plan(conn, wb["ACTIVITY_PLAN"]) if "ACTIVITY_PLAN" in wb.sheetnames else 0,
            "config": import_config(conn, wb),
        }
        result["technicians"] = derive_technicians(conn)
        result["closed_pos"] = conn.execute("SELECT COUNT(*) c FROM closed_pos").fetchone()[0]
        conn.commit()
    finally:
        conn.close()
        wb.close()
    sync_rules_from_config()  # keep business_rules/settings in step with imported CONTROL
    # Long-term memory: snapshot network + per-technician KPIs after each import,
    # tagged with this import as provenance. Non-blocking.
    try:
        import history
        ev = db.get("SELECT id FROM events WHERE kind='import' ORDER BY id DESC LIMIT 1")
        result["metrics_week"] = history.capture_metrics("import", ev[0]["id"] if ev else None)
    except Exception:  # noqa: BLE001 - never fail an import on a metrics snapshot
        pass
    try:
        import diagnostics
        diagnostics.invalidate_cache()  # fresh data -> recompute peer baselines
    except Exception:  # noqa: BLE001
        pass
    try:
        import duration
        duration.rebuild()  # refresh the collective visit-duration model
    except Exception:  # noqa: BLE001
        pass
    try:
        import capacity
        capacity.rebuild()  # refresh the learned daily-capacity standard
    except Exception:  # noqa: BLE001
        pass
    return result


# CONTROL key -> where the DB config that maps to it should be synced, so the
# DB (source of truth) reflects the real imported CONTROL and db_state's overlay
# reproduces the baseline plan. (Only knobs the engine actually reads.)
_CONTROL_TO_RULE = {
    "STANDARD_VISIT_GAP":          ("MIN_GAP", "weeks", float),
    "NEGLECTED_AFTER_WEEKS":       ("NEGLECTED_AFTER", "weeks", float),
    "HOLDBACK_LOOKAHEAD_WEEKS":    ("HOLDBACK", "lookahead_weeks", float),
    "HOLDBACK_TOLERANCE_A_WEEKS":  ("HOLDBACK", "tolerance_a", float),
    "HOLDBACK_TOLERANCE_OTHER_WEEKS": ("HOLDBACK", "tolerance_other", float),
    "TARGET_VISITS_WEEK":          ("MAX_VISITS_WEEK", "per_week", float),
    "TARGET_VISITS_DAY":           ("MAX_VISITS_WEEK", "per_day", float),
    "GPS_EXTRA_MAX_VISITS":        ("GPS_EXTRA", "max_extra_visits", float),
}


def sync_rules_from_config() -> None:
    """Sync business_rules params + planner settings from the imported CONTROL
    (config table), so the DB is the source of truth and the engine, driven by
    db_state, reproduces the current behaviour. Idempotent."""
    import business_rules
    import settings as settings_mod

    ctrl = {r["key"]: r["value"] for r in db.get("SELECT key, value FROM config")}
    eff = business_rules.effective()
    for ckey, (rule, pkey, cast) in _CONTROL_TO_RULE.items():
        if ckey not in ctrl or ctrl[ckey] in (None, ""):
            continue
        try:
            val = cast(ctrl[ckey])
        except (ValueError, TypeError):
            continue
        val = int(val) if isinstance(val, float) and val.is_integer() else val
        params = dict(eff.get(rule, {}).get("params", {}))
        params[pkey] = val
        business_rules.set_params(rule, params)
        eff.setdefault(rule, {"params": {}})["params"] = params
    if "TARGET_VISITS_DAY" in ctrl and ctrl["TARGET_VISITS_DAY"] not in (None, ""):
        try:
            settings_mod.set_value("planner", "max_visits_per_day", int(float(ctrl["TARGET_VISITS_DAY"])))
        except (ValueError, TypeError):
            pass
    # GPS extra visits on/off follows the imported CONTROL flag.
    if "GPS_EXTRA_ENABLED" in ctrl and ctrl["GPS_EXTRA_ENABLED"] not in (None, ""):
        try:
            business_rules.set_enabled("GPS_EXTRA", float(ctrl["GPS_EXTRA_ENABLED"]) == 1)
        except (ValueError, TypeError):
            pass
