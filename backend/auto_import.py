"""Automatic import - drop an Excel and the system figures out what it is.

Detects the file type from the sheet headers (no manual mapping), imports it
into SQLite, and returns what it did. Supports single-sheet real exports
(SalesApp / POS Master / Activity Plan) and the full scaffold workbook.
After a SalesApp import it recomputes reality metrics + alerts so the platform
stays current on its own.
"""
from __future__ import annotations

import openpyxl

import db
import import_validate
import importer


def _headers(ws) -> set[str]:
    it = ws.iter_rows(values_only=True)
    for row in it:
        if row and any(v not in (None, "") for v in row):
            return {str(v).strip() for v in row if v not in (None, "")}
    return set()


def _classify(ws) -> str | None:
    h = _headers(ws)
    if "UID" in h and ("Store UID" in h or "Store" in h) and "Executor" in h:
        return "salesapp"
    if "posId" in h and "terminalId" in h:
        return "pos_master"
    if "TYPE" in h and "ACTIVITY" in h and "START_WEEK" in h:
        return "activity_plan"
    hu = {str(x).strip().upper() for x in h}
    # raw client Activity Plan calendar matrix
    if "KALENDÁŘ" in hu and "MĚSÍC" in hu:
        return "activity_plan"
    # Tourplan and the raw POS master export share most columns; a Tourplan file
    # additionally carries a week column ("TOURPLAN"/"WEEK"), so check it first.
    if ("WEEK" in hu or "TÝDEN" in hu or "TYDEN" in hu or "TOURPLAN" in hu) and \
       ("TECHNICIAN" in hu or "TECHNIK" in hu) and "POS" in hu:
        return "tourplan"
    # Raw client POS master ("Základní údaje o prodejních místech"): Czech
    # headers, no week column.
    if "ČÍSLO TERMINÁLU" in hu and "POS" in hu and \
       ("PTT" in hu or "PPT" in hu or "NAZEV PROVOZOVNY" in hu or "POS AREA" in hu):
        return "pos_master"
    return None


def detect(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = set(wb.sheetnames)
        if {"POS_MASTER", "SALESAPP_IMPORT"} & names and "CONTROL" in names:
            return {"type": "workbook", "sheet": None}
        for name in wb.sheetnames:
            t = _classify(wb[name])
            if t:
                return {"type": t, "sheet": name}
        return {"type": "unknown", "sheet": None}
    finally:
        wb.close()


# The primary table whose row-count decides "did anything actually import".
# If this is 0 after a run, we report ok:false — never a zero-row success.
_PRIMARY = {"pos_master": "pos_master", "salesapp": "salesapp_visits",
            "activity_plan": "campaigns", "tourplan": "tourplan"}


def _primary_total(kind: str, counts: dict) -> int:
    if kind == "workbook":
        return sum(int(counts.get(k) or 0) for k in ("pos_master", "salesapp_visits", "campaigns"))
    key = _PRIMARY.get(kind)
    v = counts.get(key)
    if isinstance(v, dict):
        return int(v.get("rows") or v.get("total") or 0)
    return int(v or 0)


def _result(ok, kind, counts=None, warnings=None, error=None, filename=None,
            recomputed=None, sheet=None) -> dict:
    """The single import-response DTO every import endpoint returns. See
    web/contracts.js / docs/API_CONTRACT.md for the mirror on the frontend."""
    counts = counts or {}
    return {
        "ok": bool(ok),
        "kind": kind,
        "detected": kind,                       # back-compat alias
        "kindLabel": import_validate._KIND_LABEL.get(kind, kind),
        "imported": counts,
        "counts": counts,                       # back-compat alias
        "total": _primary_total(kind, counts) if ok else 0,
        "warnings": warnings or [],
        "error": error,
        "file": filename,
        "sheet": sheet,
        "recomputed": recomputed or [],
    }


def import_file(path: str, filename: str | None = None, force_kind: str | None = None) -> dict:
    """Detect (or force), VALIDATE, then import one file. Always returns the
    unified ImportResult DTO (see _result). Validation runs before any write, so
    a structurally wrong file is refused with a precise reason — never a silent
    zero-row success.

    force_kind (pos_master | salesapp | activity_plan | tourplan | workbook)
    skips auto-detection — the explicit, predictable path."""
    db.init_db()
    if force_kind:
        det = {"type": force_kind, "sheet": None if force_kind == "workbook" else _first_sheet(path)}
    else:
        det = detect(path)
    t = det["type"]

    if t == "unknown":
        return _result(False, "unknown", error=(
            "Nepodařilo se rozpoznat typ souboru. Čekám POS Master (číslo POS + PPT), "
            "SalesApp (UID + Executor), Activity Plan nebo kompletní workbook."),
            filename=filename)

    # --- structural pre-check: refuse before writing if columns/rows are wrong
    v = import_validate.validate(path, t)
    if not v["ok"]:
        return _result(False, t, error=v["reason"], warnings=v["warnings"], filename=filename)

    counts: dict = {}
    if t == "workbook":
        counts = importer.import_workbook(path, filename)
    elif t == "tourplan":
        counts["tourplan"] = importer.import_tourplan(path, filename)
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        conn = db.connect()
        try:
            ws = wb[det["sheet"] or wb.sheetnames[0]]
            if t == "salesapp":
                counts["salesapp_visits"] = importer.import_salesapp(conn, ws, filename)
            elif t == "pos_master":
                pos_summary = importer.import_pos_master(conn, ws)
                counts["pos_master"] = pos_summary["total"]
                counts["pos_diff"] = pos_summary
                counts["technicians"] = importer.derive_technicians(conn)
            elif t == "activity_plan":
                counts["campaigns"] = importer.import_activity_plan(conn, ws)
            conn.commit()
        finally:
            conn.close()
            wb.close()
        if t == "salesapp":
            # Register visit executors as technicians (+refresh roles) so
            # team/health analytics can match visits. Separate committed conn.
            conn2 = db.connect()
            try:
                counts["technicians"] = importer.derive_technicians(conn2)
                conn2.commit()
            finally:
                conn2.close()
        importer.sync_rules_from_config()

    total = _primary_total(t, counts)
    if total == 0:
        return _result(False, t, counts=counts, warnings=v["warnings"], filename=filename,
                       error=("Soubor prošel kontrolou, ale nenaimportoval se žádný řádek. "
                              "Zkontroluj, že list obsahuje data (ne jen hlavičku)."))

    rec = _recompute_after(counts)
    return _result(True, t, counts=counts, warnings=v["warnings"], filename=filename,
                   recomputed=rec, sheet=det.get("sheet"))


def _first_sheet(path: str) -> str:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return wb.sheetnames[0]
    finally:
        wb.close()


def _recompute_after(counts: dict) -> list[str]:
    """After a sync, auto-compute what depends on the new data (reality metrics
    + alerts). Central place so every import path stays current."""
    done = []
    if counts.get("salesapp_visits"):
        try:
            import alerts
            n = alerts.recompute()
            done.append(f"alerts:{n}")
        except Exception as e:  # noqa: BLE001
            done.append(f"alerts_failed:{e}")
    return done
