"""Stateless planning pipeline.

Takes this week's raw exports - the same files the manager pastes into the
Excel: the POS/PPT export, one or more SalesApp visit-data exports, and
(optionally) an Activity Plan - plus the persisted planning config
(config_store.py), assembles a fresh in-memory engine state, and runs the
UNCHANGED engines Import -> Compliance -> Planning to produce a Draft plan.

This module contains NO business logic. It only ASSEMBLES the sheet state
the existing desktop_client/engines/ already expect and calls them, exactly
as the Excel workbook feeds them - so the Draft is byte-for-byte what those
proven engines produce (equivalence proof: tools/sim/verify_stateless.py;
TS<->Py equivalence: tools/sim/compare_engines.py).

Every call is fully isolated (governing principle #3: one upload = one
isolated run). The state is built from scratch, lives only in memory, and
is discarded unless the manager later publishes it. Uploading never mutates
any published plan (governing principle #4 - publish is a separate step).

The uploaded POS export maps 1:1 onto RAW_DATA and each SalesApp export
maps 1:1 onto SALESAPP_IMPORT (verified: identical headers, identical
column order) - so assembly here is a sheet copy, never a re-mapping that
could diverge from what Excel does.
"""
from __future__ import annotations

import datetime
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from desktop_client.engines import (  # noqa: E402
    compliance_engine,
    import_engine,
    planning_engine,
)
from desktop_client.engines.core_logic import norm  # noqa: E402
from desktop_client.engines.mock_workbook import MockWorkbook  # noqa: E402

import config_store  # noqa: E402
import snapshot_store  # noqa: E402

# Header rows for the ledger / output sheets that start EMPTY on a fresh
# stateless run (captured from the scaffold). The engines fill them in.
# POS_MASTER's header is rewritten wholesale by the Import Engine, but the
# sheet must still exist (MockWorkbook refuses unknown sheets), so we seed
# its header too.
EMPTY_LEDGER_HEADERS: dict[str, list[str]] = {
    "POS_MASTER": [
        "posId", "terminalId", "market", "category", "terminalType", "classification",
        "nazev", "area", "posArea", "street", "houseNumber", "city", "gpsX", "gpsY",
        "assignedTechnician", "ppt", "status", "closedSinceWeek", "closedSinceYear",
        "currentLosActivity", "currentLotActivity", "targetLosActivity", "targetLotActivity",
        "lastRealVisitDate", "lastRealVisitWeek", "lastPlannedVisitDate",
        "weeksSinceLastVisit", "visitCountThisCampaign", "businessScore",
        "plannerStatus", "assignedWeek", "assignedDay", "gpsGroup",
        "managerOverrideType", "managerOverridePriority", "managerOverrideTechnician",
        "plannerNotes", "importedAt", "updatedAt",
    ],
    "MANAGER_PLAN": [
        "WEEK", "DATE", "DAY", "TECHNICIAN", "POS", "KATEGORIE", "NAZEV_PROVOZOVNY",
        "ULICE", "CISLO", "MESTO", "OBLAST", "POS_AREA", "PPT", "LOS_ACTIVITY",
        "LOT_ACTIVITY", "REASON", "GPS_GROUP",
    ],
    "MANAGER_PLAN_PUBLISHED": [
        "WEEK", "DATE", "DAY", "TECHNICIAN", "POS", "KATEGORIE", "NAZEV_PROVOZOVNY",
        "ULICE", "CISLO", "MESTO", "OBLAST", "POS_AREA", "PPT", "LOS_ACTIVITY",
        "LOT_ACTIVITY", "REASON", "GPS_GROUP", "publishedAt",
    ],
    "PLAN_LIFECYCLE": ["year", "week", "status", "publishedAt", "closedAt", "trackingStartedAt"],
    "VISIT_HISTORY_ACTUAL": [
        "posId", "date", "week", "year", "executor", "state", "salesAppUid",
        "durationHours", "startedAt", "finishedAt",
    ],
    "OTHER_VISIT_LOG": [
        "posId", "date", "week", "year", "executor", "salesAppUid",
        "durationHours", "startedAt", "finishedAt",
    ],
    "OZ_VISIT_LOG": [
        "posId", "date", "week", "year", "executor", "salesAppUid",
        "durationHours", "startedAt", "finishedAt",
    ],
    "COMPLIANCE_LOG": [
        "posId", "technician", "plannedWeek", "plannedYear", "status",
        "matchedActualDate", "matchedActualWeek", "evaluatedAt",
        "matchedActualDurationHours", "matchedActualStartedAt", "matchedActualFinishedAt",
    ],
}


def _cell_to_json(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return "" if v is None else v


def read_export_rows(path: str, sheet: str | None = None) -> list[list]:
    """Reads one uploaded export's first (or named) worksheet as plain
    rows, exactly the shape read_state() produces for an in-workbook sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = [[_cell_to_json(v) for v in row] for row in ws.iter_rows(values_only=True)]
        while rows and all(v == "" for v in rows[-1]):
            rows.pop()
        return rows
    finally:
        wb.close()


def merge_salesapp(exports: list[list[list]]) -> list[list]:
    """Concatenates several SalesApp exports into one SALESAPP_IMPORT table:
    the header of the first, then every data row of every export. The
    Compliance Engine dedups visits by UID, so overlapping exports are
    harmless - this only assembles, it does not decide anything."""
    merged: list[list] = []
    for rows in exports:
        if not rows:
            continue
        if not merged:
            merged.append(list(rows[0]))
        merged.extend(list(r) for r in rows[1:])
    return merged


def _seed_pos_master(header: list[str], overrides: dict[str, dict]) -> list[list]:
    """Builds the initial POS_MASTER: just its header, plus (if the manager
    has set any per-POS overrides in the config store) one thin row per
    overridden POS carrying only posId + override columns. The Import Engine
    already knows how to preserve those columns for an existing POS, so this
    re-applies persisted manager decisions through the engine's own,
    unchanged preservation path - no new logic."""
    rows: list[list] = [list(header)]
    if not overrides:
        return rows
    idx = {name: i for i, name in enumerate(header)}
    for pos_id, cols in overrides.items():
        row = [""] * len(header)
        row[idx["posId"]] = pos_id
        for col, value in cols.items():
            if col in idx:
                row[idx[col]] = value
        rows.append(row)
    return rows


def build_state(
    config_state: dict[str, list[list]],
    raw_data_rows: list[list],
    salesapp_rows: list[list],
    snapshot: dict[str, list[list]] | None = None,
    pos_overrides: dict[str, dict] | None = None,
) -> dict[str, list[list]]:
    """Assembles the full engine state a MockWorkbook seeds from:

      config sheets (from the config store)
    + weekly data sheets (from this run's uploads: RAW_DATA, SALESAPP_IMPORT)
    + ALL accumulated state sheets (from the last published snapshot)

    The snapshot is the complete carry-over (snapshot_store.SNAPSHOT_SHEETS):
    POS_MASTER *and* the visit logs, compliance log, published plan and
    lifecycle - not just POS_MASTER. Seeding every one of them is what makes
    "upload only the new export" work: Compliance dedups the fresh visits
    against the snapshot's VISIT_HISTORY_ACTUAL (by UID) and recomputes
    weeksSinceLastVisit from the accumulated history, so nothing has to be
    re-derived from years of exports.

      - snapshot given -> resume from it (the normal, git-like path).
      - snapshot None -> cold from-scratch state (empty state sheets, header
        only; POS_MASTER carries any per-POS overrides). Only for the very
        first bootstrap, before any snapshot exists.

    Any state sheet absent from the snapshot falls back to its empty header
    (EMPTY_LEDGER_HEADERS), so a partial snapshot still assembles cleanly.
    """
    state: dict[str, list[list]] = {name: [list(r) for r in rows] for name, rows in config_state.items()}
    state["RAW_DATA"] = [list(r) for r in raw_data_rows]
    state["SALESAPP_IMPORT"] = [list(r) for r in salesapp_rows]

    snapshot = snapshot or {}
    for name, header in EMPTY_LEDGER_HEADERS.items():
        if name in snapshot and snapshot[name]:
            state[name] = [list(r) for r in snapshot[name]]
        elif name == "POS_MASTER":
            state[name] = _seed_pos_master(header, pos_overrides or {})
        else:
            state[name] = [list(header)]

    # Derived report sheets (ADVISOR_LOG, TECHNICIAN_PERFORMANCE_*, DASHBOARD,
    # POS_MAP_DATA): carry them through if the snapshot has them, so a
    # snapshot round-trips as a complete restore point. Import/Compliance/
    # Planning never read them, so they do not affect the Draft.
    for name, rows in snapshot.items():
        if name not in state:
            state[name] = [list(r) for r in rows]
    return state


def _set_control(state: dict, key: str, value) -> None:
    control = state.setdefault("CONTROL", [["KEY", "VALUE", "NOTE"]])
    key_norm = norm(key)
    for row in control[1:]:
        if norm(str(row[0])) == key_norm:
            row[1] = value
            return
    control.append([key, value, ""])


def run_pipeline(state: dict[str, list[list]], start_week: int, length: int) -> dict:
    """Runs Import -> Compliance -> Planning against the assembled state,
    in place, and returns the per-engine messages. The Planning window is
    set exactly the way the Excel does it (CONTROL!CAMPAIGN_START_WEEK /
    CAMPAIGN_LENGTH)."""
    _set_control(state, "CAMPAIGN_START_WEEK", start_week)
    _set_control(state, "CAMPAIGN_LENGTH", length)

    wb = MockWorkbook(state)
    import_msg = import_engine.run(wb)
    # Import rewrote POS_MASTER in the MockWorkbook; pull it back into `state`
    # so the next engine (which seeds its own MockWorkbook view from the same
    # sheet objects) sees it. MockWorkbook shares the row lists by reference,
    # so Compliance/Planning already see Import's output - re-dump to be safe
    # and to return the final state to the caller.
    state.update(wb.dump())

    wb = MockWorkbook(state)
    compliance_msg = compliance_engine.run(wb)
    state.update(wb.dump())

    wb = MockWorkbook(state)
    planning_msg = planning_engine.run(wb)
    state.update(wb.dump())

    return {
        "import": import_msg,
        "compliance": compliance_msg,
        "planning": planning_msg,
    }


def generate_draft(
    raw_data_rows: list[list],
    salesapp_exports: list[list[list]],
    start_week: int,
    length: int,
    seed_workbook: str | None = None,
    resume_from_snapshot: bool = True,
) -> dict:
    """Full stateless run from raw uploads to a Draft plan.

    - raw_data_rows: the POS/PPT export's rows (maps 1:1 to RAW_DATA)
    - salesapp_exports: a list of SalesApp exports' rows (merged, deduped by
      the engine via UID)
    - start_week / length: the Planning window
    - seed_workbook: where config + the POS_MASTER baseline come from
    - resume_from_snapshot: seed POS_MASTER from the last published snapshot
      (the normal, Excel-faithful path). False = cold from-scratch bootstrap.

    Returns {state, messages, summary}. `state` is the whole Draft; nothing
    is persisted here (publish is a separate step).
    """
    config_state = config_store.load_config_state(seed_workbook)
    salesapp_rows = merge_salesapp(salesapp_exports)

    if resume_from_snapshot:
        snapshot = snapshot_store.load_snapshot(seed_workbook)
        state = build_state(config_state, raw_data_rows, salesapp_rows, snapshot=snapshot)
    else:
        overrides = config_store.load_pos_overrides(seed_workbook)
        state = build_state(config_state, raw_data_rows, salesapp_rows, pos_overrides=overrides)
    messages = run_pipeline(state, start_week, length)

    return {
        "state": state,
        "messages": messages,
        "summary": _summarize(state, start_week, length),
    }


def _summarize(state: dict, start_week: int, length: int) -> dict:
    pos_master = state.get("POS_MASTER", [])
    manager_plan = state.get("MANAGER_PLAN", [])
    visit_history = state.get("VISIT_HISTORY_ACTUAL", [])
    weeks_in_plan = sorted({
        row[0] for row in manager_plan[1:] if row and row[0] not in (None, "")
    }) if len(manager_plan) > 1 else []
    return {
        "startWeek": start_week,
        "length": length,
        "posMasterRows": max(len(pos_master) - 1, 0),
        "managerPlanRows": max(len(manager_plan) - 1, 0),
        "visitHistoryRows": max(len(visit_history) - 1, 0),
        "weeksInPlan": weeks_in_plan,
    }
