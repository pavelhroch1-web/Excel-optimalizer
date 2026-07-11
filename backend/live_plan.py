"""Living published TourPlan - the main daily working screen.

The published plan is the immutable reference (published_plans). This module
keeps it "alive":
  * board(): the current published plan (latest published version) with a LIVE
    status overlaid from reality (SalesApp) per stop - done / due this week /
    overdue / upcoming - plus a countdown (days) to each planned visit. It stays
    the reference until a NEW version is published.
  * next_due(): per-POS countdown to the next visit by GECO/CORN cadence (the
    only hard cadence per the product owner). last visit = any role (an OZ visit
    still "covers" a POS, so a technician need not re-drive it); the technician's
    own last visit is surfaced separately for context.

Read-only over SQLite + the current snapshot's CADENCE_RULES. No engine change,
no planning here - this only reports the published plan and cadence state.
"""
from __future__ import annotations

import datetime
import os

import db
import store
from desktop_client.engines.core_logic import CadenceRule, matches_cadence_rule_scope


def _norm(s) -> str:
    return str(s if s is not None else "").strip().upper()


def _iso_week(date_str) -> int | None:
    if not date_str:
        return None
    try:
        y, m, dd = (int(x) for x in str(date_str)[:10].split("-"))
        return datetime.date(y, m, dd).isocalendar()[1]
    except (ValueError, TypeError):
        return None


def _date(date_str):
    """Parse ISO (YYYY-MM-DD) or the engine's Czech plan_date (e.g. '20. 7. 2026')."""
    if not date_str:
        return None
    s = str(date_str).strip()
    try:
        y, m, dd = (int(x) for x in s[:10].split("-"))
        return datetime.date(y, m, dd)
    except (ValueError, TypeError):
        pass
    try:
        dd, m, y = (int(p.strip()) for p in s.split(".") if p.strip())
        return datetime.date(y, m, dd)
    except (ValueError, TypeError):
        return None


# ---- cadence rules from the current snapshot (cached per snapshot id) -------

_cache: dict = {"snap": object(), "rules": None, "neglected": 12}


def _cadence_rules():
    """Active RECURRING+HARD cadence rules (GECO/CORN) from the current snapshot,
    parsed exactly as the engine parses them. Cached per snapshot id."""
    rows = db.get("SELECT id FROM snapshots ORDER BY created_at DESC LIMIT 1")
    sid = rows[0]["id"] if rows else None
    if _cache["snap"] == sid and _cache["rules"] is not None:
        return _cache["rules"], _cache["neglected"]

    # Read ONLY the two sheets we need (CADENCE_RULES + CONTROL) straight from the
    # snapshot xlsx - far faster than state_xlsx.load_state() which parses all ~40
    # sheets incl. the 11k-row POS_MASTER.
    import openpyxl
    path = store.snapshot_temp()
    cad_raw: list = []
    control_raw: list = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if "CADENCE_RULES" in wb.sheetnames:
                cad_raw = [list(r) for r in wb["CADENCE_RULES"].iter_rows(values_only=True)]
            if "CONTROL" in wb.sheetnames:
                control_raw = [list(r) for r in wb["CONTROL"].iter_rows(values_only=True)]
        finally:
            wb.close()
    finally:
        try:
            os.remove(path)
        except OSError:
            pass

    raw = cad_raw or []
    headers = [str(h).strip() for h in raw[0]] if raw else []

    def ci(name):
        return headers.index(name) if name in headers else -1

    def at(row, i):
        return row[i] if 0 <= i < len(row) else ""

    def num(v):
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    rules: list[CadenceRule] = []
    for i in range(1, len(raw)):
        row = raw[i]
        if _norm(at(row, ci("active"))) != "YES":
            continue
        if _norm(at(row, ci("intervalType"))) != "RECURRING" or _norm(at(row, ci("guaranteeType"))) != "HARD":
            continue
        mv = str(at(row, ci("matchValue")) or "")
        rules.append(CadenceRule(
            ruleId=str(at(row, ci("ruleId")) or ""),
            scope=_norm(at(row, ci("scope"))),
            matchValue=[_norm(p) for p in mv.split(";") if _norm(p)],
            minGapWeeks=None,
            maxIntervalWeeks=num(at(row, ci("maxIntervalWeeks"))),
            intervalType="RECURRING",
            guaranteeType="HARD",
            dedupBy=_norm(at(row, ci("dedupBy"))),
            campaignChangeOverride=False,
            priority=num(at(row, ci("priority"))) or 0,
        ))

    # NEGLECTED_AFTER from CONTROL (soft due for non-cadence POS).
    neglected = 12
    for r in (control_raw or [])[1:]:
        if r and str(r[0]).strip() == "NEGLECTED_AFTER_WEEKS":
            neglected = num(at(r, 1)) or 12
            break

    _cache.update(snap=sid, rules=rules, neglected=neglected)
    return rules, neglected


def _match_rule(rules, category, market):
    cat, mkt = _norm(category), _norm(market)
    for r in rules:
        if matches_cadence_rule_scope(r, cat, mkt):
            return r
    return None


# ---- last real visit per POS (any role, + technician-only) -----------------

def _last_visits():
    any_role: dict = {}
    tech_only: dict = {}
    for r in db.get("SELECT pos_id, visit_date, visitor_role FROM salesapp_visits "
                    "WHERE pos_id IS NOT NULL AND visit_date IS NOT NULL"):
        pid = str(r["pos_id"]); d = str(r["visit_date"])[:10]
        if pid not in any_role or d > any_role[pid]:
            any_role[pid] = d
        if _norm(r["visitor_role"]) == "TECHNIK" and (pid not in tech_only or d > tech_only[pid]):
            tech_only[pid] = d
    return any_role, tech_only


# ---- next-due countdown by cadence -----------------------------------------

def next_due(technician: str | None = None, status: str | None = None,
             limit: int | None = 1000) -> dict:
    """Per-POS countdown to the next visit by GECO/CORN cadence."""
    rules, _neglected = _cadence_rules()
    any_role, tech_only = _last_visits()
    today = datetime.date.today()

    q = ("SELECT pos_id, name, city, category, market, classification, technician "
         "FROM pos_master WHERE active=1")
    params: tuple = ()
    if technician:
        q += " AND technician=?"; params = (technician,)

    out = []
    counts = {"overdue": 0, "dueSoon": 0, "ok": 0, "neverVisited": 0}
    for p in db.get(q, params):
        rule = _match_rule(rules, p["category"], p["market"])
        if not rule or rule.maxIntervalWeeks is None:
            continue  # only POS under a hard GECO/CORN cadence have a countdown
        pid = str(p["pos_id"])
        lv = any_role.get(pid)
        lv_date = _date(lv)
        cadence_days = int(round(rule.maxIntervalWeeks * 7))
        if lv_date:
            next_due_date = lv_date + datetime.timedelta(days=cadence_days)
            days_remaining = (next_due_date - today).days
        else:
            next_due_date = None
            days_remaining = None

        if days_remaining is None:
            st = "overdue"; counts["neverVisited"] += 1
        elif days_remaining < 0:
            st = "overdue"
        elif days_remaining <= 14:
            st = "dueSoon"
        else:
            st = "ok"
        counts[st] = counts.get(st, 0) + 1

        out.append({
            "pos": pid, "name": p["name"], "city": p["city"],
            "technician": p["technician"], "cadence": rule.ruleId,
            "cadenceWeeks": rule.maxIntervalWeeks,
            "lastVisit": lv, "lastTechnicianVisit": tech_only.get(pid),
            "nextDue": next_due_date.isoformat() if next_due_date else None,
            "daysRemaining": days_remaining, "status": st,
        })

    # never-visited & most-overdue first (None sorts before any negative)
    out.sort(key=lambda x: (x["daysRemaining"] is not None, x["daysRemaining"]))
    if status:
        out = [x for x in out if x["status"] == status]
    total = len(out)
    if limit:
        out = out[:limit]
    return {"today": today.isoformat(), "counts": counts,
            "total": total, "shown": len(out), "posList": out}


# ---- the living published plan ---------------------------------------------

def board(technician: str | None = None) -> dict:
    """The current published plan with a live per-stop status from reality.

    The current plan per week is whatever snapshot plan_lifecycle has locked for
    that week (publish freezes one week at a time), so this shows the whole
    living forward plan across however many publishes produced it."""
    locked = db.get("SELECT COUNT(*) AS c FROM plan_lifecycle WHERE status='Published'")
    if not locked or not locked[0]["c"]:
        return {"published": False, "stops": [], "weeks": [], "perWeek": [],
                "message": "Zatím nebyl publikován žádný plán."}

    latest = db.get("SELECT id, created_at, message, published_by FROM snapshots "
                    "ORDER BY created_at DESC LIMIT 1")
    meta = dict(latest[0]) if latest else {}
    version_count = db.get("SELECT COUNT(*) AS c FROM snapshots")[0]["c"]

    q = ("SELECT pp.week, pp.plan_date, pp.day, pp.technician, pp.pos_id, pp.name, "
         "pp.city, pp.ppt, pp.day_seq "
         "FROM published_plans pp JOIN plan_lifecycle pl "
         "ON pl.week=pp.week AND pl.snapshot_id=pp.snapshot_id AND pl.status='Published'")
    params: list = []
    if technician:
        q += " AND pp.technician=?"; params.append(technician)
    q += " ORDER BY pp.week, pp.plan_date, pp.day_seq"
    rows = [dict(r) for r in db.get(q, tuple(params))]

    # reality: {(pos, week)} actually visited (±1 week tolerance folded in later)
    visited: dict = {}
    for r in db.get("SELECT pos_id, visit_date FROM salesapp_visits "
                    "WHERE pos_id IS NOT NULL AND visit_date IS NOT NULL"):
        wk = _iso_week(r["visit_date"])
        if wk is not None:
            visited.setdefault(str(r["pos_id"]), set()).add(wk)

    today = datetime.date.today()
    cur_week = today.isocalendar()[1]
    stops = []
    rollup = {"total": 0, "done": 0, "overdue": 0, "due": 0, "upcoming": 0}
    for r in rows:
        pid = str(r["pos_id"]) if r["pos_id"] else ""
        wk = r["week"]
        pd = _date(r["plan_date"])
        seen = visited.get(pid, set())
        if wk in seen or (wk is not None and ((wk - 1) in seen or (wk + 1) in seen)):
            st = "done"
        elif wk is not None and wk < cur_week:
            st = "overdue"
        elif wk == cur_week:
            st = "due"
        else:
            st = "upcoming"
        rollup["total"] += 1
        rollup[st] = rollup.get(st, 0) + 1
        stops.append({
            "week": wk, "planDate": r["plan_date"], "day": r["day"],
            "technician": r["technician"], "pos": pid, "name": r["name"],
            "city": r["city"], "ppt": r["ppt"], "daySeq": r["day_seq"],
            "status": st,
            "daysUntil": (pd - today).days if pd else None,
        })

    weeks = sorted({s["week"] for s in stops if s["week"] is not None})
    per_week = []
    for w in weeks:
        wk_stops = [s for s in stops if s["week"] == w]
        per_week.append({
            "week": w, "stops": len(wk_stops),
            "done": sum(1 for s in wk_stops if s["status"] == "done"),
            "overdue": sum(1 for s in wk_stops if s["status"] == "overdue"),
        })
    pct = round(100 * rollup["done"] / rollup["total"], 1) if rollup["total"] else None
    return {
        "published": True,
        "version": meta.get("id"), "versionCount": version_count,
        "publishedAt": meta.get("created_at"), "publishedBy": meta.get("published_by"),
        "message": meta.get("message"),
        "weeks": weeks, "weekRange": (f"{weeks[0]}–{weeks[-1]}" if weeks else None),
        "currentWeek": cur_week,
        "rollup": {**rollup, "fulfilmentPct": pct},
        "perWeek": per_week,
        "stops": stops,
    }
