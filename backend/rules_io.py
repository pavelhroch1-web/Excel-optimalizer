"""Read/write for the existing manager-editable rule tables (TERMINAL_RULES,
MARKET_RULES, CATEGORY_RULES, ACTIVITY_PLAN) - the exact same tables a
manager edits directly in Excel today. No new rules or switches are
introduced here; this only lets the same editable cells be set from the web
UI instead. Editable columns per sheet match tools/ux_style.py's
EDITABLE_COLUMNS exactly - every other column (including ACTIVITY_PLAN's
live formulas) is never opened for writing.
"""
from __future__ import annotations

import openpyxl

# sheet -> (key_column, editable_columns). key_column is used to match a
# request row back to its sheet row (TYPE+ACTIVITY for ACTIVITY_PLAN, since
# neither alone is guaranteed unique).
RULE_SHEETS = {
    "TERMINAL_RULES": (["TYP TERMINALU"], ["ACTIVE"]),
    "MARKET_RULES": (["MARKET"], ["ACTIVE"]),
    "CATEGORY_RULES": (["CATEGORY"], ["RULE"]),
    "ACTIVITY_PLAN": (
        ["TYPE", "ACTIVITY"],
        ["TYPE", "ACTIVITY", "START_WEEK", "END_WEEK", "PRIORITY", "OVERRIDE_GAP"],
    ),
}


def _header_index(ws) -> dict[str, int]:
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value not in (None, "")}


def read_rule_sheet(path: str, sheet_name: str) -> list[dict]:
    key_cols, editable_cols = RULE_SHEETS[sheet_name]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    idx = _header_index(ws)
    cols = list(dict.fromkeys(key_cols + editable_cols))  # de-dup, keep order
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        blank = True
        for c in cols:
            if c not in idx:
                continue
            v = row[idx[c] - 1]
            row_dict[c] = v
            if v not in (None, ""):
                blank = False
        if not blank:
            rows.append(row_dict)
    wb.close()
    return rows


def write_rule_sheet(path: str, sheet_name: str, rows: list[dict]) -> None:
    """Overwrites only the editable-column cells of existing rows, matched
    by key_cols. Never adds/removes rows, never touches non-editable
    columns (e.g. ACTIVITY_PLAN's ODHAD_NAVSTEV_ZA_KAMPAN formula)."""
    key_cols, editable_cols = RULE_SHEETS[sheet_name]
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb[sheet_name]
    idx = _header_index(ws)

    def sheet_key(r: int) -> tuple:
        return tuple(ws.cell(r, idx[c]).value for c in key_cols if c in idx)

    existing_rows = {}
    for r in range(2, ws.max_row + 1):
        k = sheet_key(r)
        if any(v not in (None, "") for v in k):
            existing_rows[k] = r

    for incoming in rows:
        k = tuple(incoming.get(c) for c in key_cols)
        r = existing_rows.get(k)
        if r is None:
            continue  # never adds new rows - only edits what already exists
        for c in editable_cols:
            if c in incoming and c in idx:
                ws.cell(r, idx[c], incoming[c])

    wb.save(path)
    wb.close()
