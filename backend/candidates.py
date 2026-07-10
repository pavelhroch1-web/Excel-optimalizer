"""'Kandidáti POS' read-out: runs the REAL Planning Engine
(desktop_client/engines/planning_engine.py) for one week with its
observability hook enabled, and returns every candidate POS with the score
and component breakdown the engine itself computed - plus each POS's real
last-visit date and recent SalesApp visit history.

This module contains NO scoring logic. It calls planning_engine.run() with
candidates_out=[] - the identical function, identical algorithm, that
writes MANAGER_PLAN. The engine's own _assert_breakdown() guarantees the
component numbers shown here sum to the score it actually used. Confirmed
byte-equivalent to PlanningEngine.ts via tools/sim/compare_engines.py.
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from desktop_client import xlsx_engine_io
from desktop_client.engines import planning_engine
from desktop_client.engines.mock_workbook import MockWorkbook

import decision


def _set_control(state: dict, key: str, value) -> None:
    control = state["CONTROL"]
    key_norm = key.strip().upper()
    for row in control[1:]:
        if str(row[0]).strip().upper() == key_norm:
            row[1] = value
            return
    control.append([key, value, ""])


def _pos_master_extras(path: str) -> tuple[dict, dict]:
    """Returns (lastRealVisitDate by posId, sorted recent visit-date list by
    posId from VISIT_HISTORY_ACTUAL - the real SalesApp visit history)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        pm = wb["POS_MASTER"]
        h = [c.value for c in next(pm.iter_rows(min_row=1, max_row=1))]
        idx = {n: i for i, n in enumerate(h)}
        last_visit: dict[str, str] = {}
        for row in pm.iter_rows(min_row=2, values_only=True):
            pid = row[idx["posId"]]
            if pid:
                lv = row[idx["lastRealVisitDate"]]
                last_visit[str(pid)] = "" if lv in (None, "") else str(lv)[:10]

        history: dict[str, list] = {}
        vh = wb["VISIT_HISTORY_ACTUAL"]
        vh_h = [c.value for c in next(vh.iter_rows(min_row=1, max_row=1))]
        vidx = {n: i for i, n in enumerate(vh_h)}
        for row in vh.iter_rows(min_row=2, values_only=True):
            pid = row[vidx["posId"]]
            if not pid:
                continue
            d = row[vidx["date"]]
            ex = row[vidx["executor"]]
            if d:
                history.setdefault(str(pid), []).append({"date": str(d)[:10], "executor": ex})
        for pid in history:
            history[pid].sort(key=lambda v: v["date"], reverse=True)
        return last_visit, history
    finally:
        wb.close()


def _explain(c: dict) -> str:
    """Human-readable 'why selected / why not' derived purely from the score
    components the engine already produced - no new logic, just a translation
    of the numbers into a sentence the manager can read at a glance."""
    if c["status"] == "Vybráno":
        parts = []
        if c["core"]:
            parts.append("CORE (garantováno)")
        if c["mandatoryRuleId"]:
            parts.append(f"povinné pravidlo {c['mandatoryRuleId']}")
        if c["classification"] == "A":
            parts.append("klasifikace A")
        if c["neglectedBonus"]:
            parts.append("dlouho nenavštíveno")
        if c["urgencyBoost"]:
            parts.append("blíží se termín (urgence)")
        if c["premium"]:
            parts.append("vysoké PPT (top 20 %)")
        if c["gpsBonus"]:
            parts.append("výhodná trasa (GPS shluk)")
        return "Vybráno: " + (", ".join(parts) if parts else "vysoké skóre")
    if c["status"].startswith("Odloženo"):
        return "Odloženo: blíží se kampaň, POS ještě není urgentní (Smart Hold-back)"
    # Not selected
    if c["gapPenalty"] < 0:
        return "Nevybráno: navštíveno příliš nedávno (penalizace min. rozestupu)"
    return "Nevybráno: nižší skóre než vybrané POS / kapacita technika naplněna"


def list_candidates(path: str, week: int, technician: str | None = None) -> dict:
    """Runs the engine for `week` and returns its scored candidate pool. Does
    NOT persist anything - the engine writes MANAGER_PLAN into the in-memory
    MockWorkbook only, which is discarded."""
    state = xlsx_engine_io.read_state(path)
    _set_control(state, "CAMPAIGN_START_WEEK", week)
    _set_control(state, "CAMPAIGN_LENGTH", 1)

    wb = MockWorkbook(state)
    captured: list[dict] = []
    planning_engine.run(wb, candidates_out=captured)

    last_visit, history = _pos_master_extras(path)
    for c in captured:
        c["lastRealVisitDate"] = last_visit.get(str(c["pos"]), "")
        c["visitHistory"] = history.get(str(c["pos"]), [])
        c["explanation"] = _explain(c)

    if technician:
        captured = [c for c in captured if c["tech"] == technician]

    captured.sort(key=lambda c: (0 if c["status"] == "Vybráno" else 1, -c["score"]))
    techs = sorted({c["tech"] for c in captured})
    return {
        "week": week,
        "technicians": techs,
        "total": len(captured),
        "selected": sum(1 for c in captured if c["status"] == "Vybráno"),
        "candidates": captured,
    }


def _sheet_index(state: dict, sheet: str) -> tuple[list, dict]:
    rows = state.get(sheet, [])
    header = [str(h) for h in rows[0]] if rows else []
    return rows, {n: i for i, n in enumerate(header)}


def _last_compliance(state: dict, pos_id: str) -> dict | None:
    """Latest COMPLIANCE_LOG row for this POS - the same compliance the
    engine's inputs carry, no recomputation."""
    rows, idx = _sheet_index(state, "COMPLIANCE_LOG")
    if not idx:
        return None
    best = None
    for row in rows[1:]:
        if str(row[idx["posId"]]) != str(pos_id):
            continue
        if best is None or str(row[idx["evaluatedAt"]]) >= str(best[idx["evaluatedAt"]]):
            best = row
    if best is None:
        return None
    return {
        "status": best[idx["status"]],
        "plannedWeek": best[idx["plannedWeek"]],
        "plannedYear": best[idx["plannedYear"]],
        "matchedActualDate": best[idx.get("matchedActualDate", -1)] if "matchedActualDate" in idx else "",
    }


def _active_campaigns(state: dict, week: int) -> list[dict]:
    """ACTIVITY_PLAN campaigns whose window covers `week` - the same rows the
    Planning Engine reshapes into its hold-back windows."""
    rows, idx = _sheet_index(state, "ACTIVITY_PLAN")
    if not idx or "START_WEEK" not in idx or "END_WEEK" not in idx:
        return []
    out = []
    for row in rows[1:]:
        sw, ew = row[idx["START_WEEK"]], row[idx["END_WEEK"]]
        try:
            if sw in ("", None) or ew in ("", None):
                continue
            if int(float(sw)) <= week <= int(float(ew)):
                out.append({
                    "type": row[idx["TYPE"]] if "TYPE" in idx else "",
                    "activity": row[idx["ACTIVITY"]] if "ACTIVITY" in idx else "",
                    "startWeek": int(float(sw)), "endWeek": int(float(ew)),
                    "priority": row[idx["PRIORITY"]] if "PRIORITY" in idx else "",
                })
        except (TypeError, ValueError):
            continue
    return out


def pos_detail(path: str, pos_id: str, week: int) -> dict:
    """Full read-only diagnostic for one POS: exactly the data and score the
    Planning Engine used for `week`, plus why it is (or is not) a candidate.
    Runs the SAME engine with both observability hooks; adds no new logic."""
    pos_id = str(pos_id)
    state = xlsx_engine_io.read_state(path)
    _set_control(state, "CAMPAIGN_START_WEEK", week)
    _set_control(state, "CAMPAIGN_LENGTH", 1)

    wb = MockWorkbook(state)
    captured: list[dict] = []
    rejected: list[dict] = []
    planning_engine.run(wb, candidates_out=captured, rejected_out=rejected)

    scored = next((c for c in captured if str(c["pos"]) == pos_id), None)
    reject = next((c for c in rejected if str(c["pos"]) == pos_id), None)

    # POS_MASTER identity (also covers a POS that is neither scored nor
    # rejected - e.g. not present at all).
    pm_rows, pm_idx = _sheet_index(state, "POS_MASTER")
    master = None
    for row in pm_rows[1:]:
        if str(row[pm_idx["posId"]]) == pos_id:
            master = row
            break
    if master is None and scored is None and reject is None:
        return {"found": False, "pos": pos_id}

    def m(col):
        return master[pm_idx[col]] if master is not None and col in pm_idx else None

    if scored is not None:
        detail = dict(scored)
        detail["isCandidate"] = True
        detail["explanation"] = _explain(scored)
    elif reject is not None:
        detail = dict(reject)
        detail["isCandidate"] = False
        detail["score"] = None
        detail["explanation"] = "Není kandidát: " + reject.get("rejectReason", "")
    else:
        detail = {
            "pos": pos_id, "nazev": m("nazev"), "market": m("market"),
            "terminalType": m("terminalType"), "kategorie": m("category"),
            "classification": m("classification"), "tech": m("assignedTechnician"),
            "ppt": m("ppt"), "weeksSinceLastVisit": m("weeksSinceLastVisit"),
            "status": "Neznámý", "isCandidate": False, "score": None,
            "explanation": "POS existuje, ale nebyl v tomto běhu vyhodnocen.",
        }

    # Enrichment straight from the engine's own input sheets:
    detail["lastRealVisitDate"] = m("lastRealVisitDate") or detail.get("lastRealVisitDate", "")
    detail["assignedTechnician"] = m("assignedTechnician")
    detail["managerOverrideTechnician"] = m("managerOverrideTechnician")
    detail["managerOverrideType"] = m("managerOverrideType")
    detail["posStatus"] = m("status")
    detail["street"] = m("street")
    detail["city"] = m("city")
    detail["lastCompliance"] = _last_compliance(state, pos_id)
    detail["activeCampaigns"] = _active_campaigns(state, week)

    _, history = _pos_master_extras(path)
    detail["visitHistory"] = history.get(pos_id, [])
    detail["week"] = week
    detail["found"] = True

    # Decision Support layer (interpretation only - no new logic):
    detail["recommendation"] = decision.recommend(detail)
    detail["includeLever"] = decision.include_lever(detail)
    return detail
