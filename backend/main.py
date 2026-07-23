"""Field Force Optimizer - production live-planner backend.

A thin HTTP layer over the UNCHANGED desktop_client/engines/. It owns no
business logic; every endpoint assembles state and calls an already-verified
engine (Import / Compliance / Planning / Publish), then persists via the
versioned store (store.py + gh.py).

The git-like model (confirmed with the product owner):
  - The latest published SNAPSHOT is the single source of truth.
  - Uploading fresh exports only ever builds a new DRAFT (snapshot + this
    run's exports, Import+Compliance). It never touches a published plan.
  - Generate runs Planning on the draft; the manager can edit it manually.
  - Publish freezes the draft's lowest draft-week into a NEW immutable
    snapshot with an audit record; all later runs resume from it.

The seven live features, one endpoint group each:
  1 upload      POST /api/draft/upload         (multipart exports)
  2 generate    POST /api/draft/generate       {start_week, length}
  3 candidates  GET  /api/draft/candidates     ?week=&technician=
  4 edit        POST /api/draft/remove-pos | add-pos | change-technician
  5 publish     POST /api/publish              {message}
  6 history     GET  /api/versions
  7 download    GET  /api/versions/{id}/manager-plan  (+ /api/draft/download)
"""
from __future__ import annotations

import hashlib
import io
import os
import sys
import tempfile

import openpyxl
from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import auth  # noqa: E402
from auth import issue_token, require_auth  # noqa: E402
import brain as brain_mod  # noqa: E402
import gh  # noqa: E402
import candidates as candidates_mod  # noqa: E402
import decision as decision_mod  # noqa: E402
import pipeline  # noqa: E402
import plan_io  # noqa: E402
import state_xlsx  # noqa: E402
import store  # noqa: E402

ENGINE_VERSION = os.environ.get("ENGINE_VERSION", "FFO-V11")
LOCAL_MODE = os.environ.get("FFO_LOCAL") == "1"

app = FastAPI(title="Field Force Optimizer API")

if LOCAL_MODE:
    import db  # noqa: E402

    # First-run bootstrap: if there is no runtime DB yet and the build bundled a
    # default one, copy it into the data dir so the app is usable immediately
    # after download. Never overwrites an existing DB (a later user import wins).
    db.bootstrap_db()
    # Create the schema EAGERLY at import time: some routes seed config at
    # module load (e.g. task_types via seed_default_types below), which runs
    # BEFORE the startup event — on a fresh DB (first .exe launch) the tables
    # would not exist yet and the import would crash. init_db() is idempotent,
    # so it also safely migrates a bundled seed DB.
    db.init_db()

    @app.on_event("startup")
    def _init_local_db() -> None:
        db.bootstrap_db()
        db.init_db()

_allowed_origins = os.environ.get("ALLOWED_ORIGIN", "*")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[_allowed_origins] if _allowed_origins != "*" else ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(Exception)
async def _unhandled_exc(request, exc):  # noqa: ANN001
    """Any uncaught endpoint error → a real, visible message instead of a bare
    "Chyba 500". In the local desktop app the traceback is written next to the
    app (FieldForceData/api_errors.log) AND returned in the response so the UI
    can show *what* broke — a windowed .exe has no console to read otherwise.
    Re-raise HTTPException so intended 4xx/detail responses are untouched."""
    import traceback
    from fastapi import HTTPException as _HTTPExc
    from fastapi.responses import JSONResponse
    if isinstance(exc, _HTTPExc):
        return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
    tb = traceback.format_exc()
    try:
        import db as _db
        with open(os.path.join(_db.data_dir(), "api_errors.log"), "a", encoding="utf-8") as f:
            f.write(f"\n===== {request.method} {request.url.path} =====\n{tb}\n")
    except Exception:  # noqa: BLE001
        pass
    detail = f"{type(exc).__name__}: {exc}"
    body = {"detail": detail}
    if LOCAL_MODE:
        body["trace"] = tb.splitlines()[-12:]  # last frames, enough to locate it
    return JSONResponse(status_code=500, content=body)


class LoginRequest(BaseModel):
    password: str


class GenerateRequest(BaseModel):
    start_week: int
    length: int = 1
    mode: str = "vyvazeny"
    visits_per_tech_week: float | None = None


class PreflightRequest(BaseModel):
    start_week: int
    length: int = 5
    mode: str = "vyvazeny"
    visits_per_tech_week: float | None = None
    tech_count_override: int | None = None


class PublishRequest(BaseModel):
    message: str = ""


class RemovePosRequest(BaseModel):
    week: int
    pos_id: str
    technician: str


class ChangeTechnicianRequest(BaseModel):
    week: int
    pos_id: str
    old_technician: str
    new_technician: str


class AddPosRequest(BaseModel):
    week: int
    day: str
    technician: str
    pos_id: str


class CloudGenerateRequest(BaseModel):
    start_week: int
    length: int = 5
    visits_per_tech: int = 40


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------

def _tmp(suffix=".xlsx") -> str:
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    return path


def _now_iso() -> str:
    import datetime
    return datetime.datetime.now(datetime.timezone.utc).isoformat()


def _require_draft_path() -> str:
    """Downloads the current draft to a temp file, or 409 if none exists."""
    if not store.draft_exists():
        raise HTTPException(status_code=409, detail="Zatím není žádný Draft. Nejdřív nahraj exporty.")
    path = _tmp()
    store.download_draft(path)
    return path


async def _save_upload(upload: UploadFile) -> tuple[str, dict]:
    """Persists an UploadFile to a temp path; returns (path, provenance)."""
    data = await upload.read()
    path = _tmp(suffix=os.path.splitext(upload.filename or "")[1] or ".xlsx")
    with open(path, "wb") as f:
        f.write(data)
    return path, {
        "filename": upload.filename,
        "bytes": len(data),
        "sha256": hashlib.sha256(data).hexdigest(),
    }


def _stream_sheet(workbook_path: str, sheet_name: str, download_name: str) -> StreamingResponse:
    src_wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in src_wb.sheetnames:
            raise HTTPException(status_code=404, detail=f"List {sheet_name} chybí.")
        src_ws = src_wb[sheet_name]
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = sheet_name[:31]
        rows_iter = src_ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        # Bundling in the export: if the sheet lists POS, append an "ÚKOLY" column
        # so the technician sees service/campaign/material to do at each stop.
        pos_col = None
        if header:
            try:
                pos_col = [str(h).upper() if h is not None else "" for h in header].index("POS")
            except ValueError:
                pos_col = None
            out_ws.append(list(header) + (["ÚKOLY"] if pos_col is not None else []))
        for row in rows_iter:
            if pos_col is not None:
                pid = row[pos_col] if pos_col < len(row) else None
                summary = ""
                if pid not in (None, ""):
                    try:
                        import tasks as _tasks
                        summary = _tasks.bundle_for_pos(str(pid)).get("summary", "")
                    except Exception:  # noqa: BLE001 - export must not fail on task overlay
                        summary = ""
                out_ws.append(list(row) + [summary])
            else:
                out_ws.append(row)
    finally:
        src_wb.close()
    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={download_name}"},
    )


# --------------------------------------------------------------------------
# health / auth / status
# --------------------------------------------------------------------------

@app.get("/api/health")
def health():
    """Unauthenticated deployment self-check: confirms the latest snapshot
    (source of truth) downloads and parses. No business data exposed."""
    path = None
    try:
        path = store.snapshot_temp()
        size = os.path.getsize(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        # The config-only scaffold has no POS_MASTER; the real count lives in
        # the DB (built solely by import). Prefer the DB, fall back to 0.
        try:
            import db as _db
            pos_rows = _db.get("SELECT COUNT(*) AS c FROM pos_master")[0]["c"]
        except Exception:  # noqa: BLE001
            pos_rows = (wb["POS_MASTER"].max_row - 1) if "POS_MASTER" in wb.sheetnames else 0
        wb.close()
        return {
            "ok": True,
            "workbookBytes": size,
            "posMasterRows": pos_rows,
            "publishedVersions": len(store.read_index()),
            "hasDraft": store.draft_exists(),
        }
    except Exception as e:  # noqa: BLE001 - surface reason for ops
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}
    finally:
        if path and os.path.exists(path):
            os.remove(path)


@app.post("/api/login")
def login(body: LoginRequest):
    # Local desktop app: single user on localhost, any/no password is fine.
    if not LOCAL_MODE and body.password != auth.APP_PASSWORD:
        raise HTTPException(status_code=401, detail="Nesprávné heslo.")
    return {"token": issue_token()}


@app.get("/api/status", dependencies=[Depends(require_auth)])
def status():
    """Current planner state: last published week, published-version count,
    and whether a draft is waiting."""
    index = store.read_index()
    published_weeks = [w for rec in index for w in rec.get("publishedWeeks", [])]
    return {
        "lastPublishedWeek": max(published_weeks) if published_weeks else None,
        "publishedVersions": len(index),
        "hasDraft": store.draft_exists(),
        "draftMeta": store.read_draft_meta() if store.draft_exists() else None,
    }


@app.get("/api/planner/next-week", dependencies=[Depends(require_auth)])
def planner_next_week():
    """The first week not yet covered by a plan = the sensible default so the
    operator does not have to think about where to start. Covered = published
    weeks (immutable) or the current draft. Falls back to the current ISO week
    when nothing is planned yet."""
    import datetime
    weeks: list[int] = []
    try:
        for rec in store.read_index():
            weeks += [int(w) for w in rec.get("publishedWeeks", []) if w is not None]
    except Exception:  # noqa: BLE001
        pass
    try:
        r = db.get("SELECT MAX(week) m FROM draft_plans")
        if r and r[0]["m"] is not None:
            weeks.append(int(r[0]["m"]))
    except Exception:  # noqa: BLE001
        pass
    cur = datetime.date.today().isocalendar()[1]
    if weeks:
        last = max(weeks)
        nxt = last + 1
        if nxt > 53:
            nxt = 1
        return {"suggestedWeek": nxt, "lastPlannedWeek": last,
                "source": "následující nepokrytý týden", "currentWeek": cur}
    return {"suggestedWeek": cur, "lastPlannedWeek": None,
            "source": "aktuální týden (zatím nic naplánováno)", "currentWeek": cur}


# --------------------------------------------------------------------------
# 1) upload -> build Draft (resume from latest snapshot + fresh exports)
# --------------------------------------------------------------------------

@app.post("/api/draft/upload", dependencies=[Depends(require_auth)])
async def draft_upload(
    pos_export: UploadFile | None = File(default=None),
    salesapp: list[UploadFile] = File(default=[]),
):
    """Builds a fresh Draft: resume from the latest published snapshot, fold
    in this run's exports, run Import + Compliance. Does NOT plan yet and does
    NOT touch any published plan.

    The POS export is OPTIONAL: a normal weekly run only brings a fresh
    SalesApp export (where technicians actually went). When no POS export is
    uploaded, the network (RAW_DATA) is taken from the latest snapshot - the
    POS structure changes rarely."""
    seed = None
    pos_path = None
    sa_paths: list[str] = []
    draft_path = None
    try:
        if pos_export is not None and (pos_export.filename or ""):
            pos_path, pos_meta = await _save_upload(pos_export)
            raw = pipeline.read_export_rows(pos_path)
        else:
            pos_meta = None
            raw = None  # pipeline falls back to the snapshot's RAW_DATA

        sa_exports = []
        sa_meta = []
        for f in salesapp:
            p, m = await _save_upload(f)
            sa_paths.append(p)
            sa_exports.append(pipeline.read_export_rows(p))
            sa_meta.append(m)

        if raw is None and not sa_exports:
            raise HTTPException(status_code=400, detail="Nahraj aspoň jeden soubor (SalesApp a/nebo POS export).")

        seed = store.snapshot_temp()  # latest snapshot (or bootstrap)
        result = pipeline.build_upload_draft(raw, sa_exports, seed_workbook=seed)

        draft_path = _tmp()
        state_xlsx.save_state(result["state"], draft_path)
        meta = {
            "uploadedAt": _now_iso(),
            "posExport": pos_meta,
            "salesAppExports": sa_meta,
            "resumedFrom": store.latest_snapshot_repo_path(),
            "engineVersion": ENGINE_VERSION,
        }
        store.save_draft(draft_path, "Upload: novy Draft z cerstvych exportu", meta=meta)

        return {"messages": result["messages"], "summary": pipeline._summarize(result["state"], 0, 0)}
    finally:
        for p in [seed, pos_path, draft_path, *sa_paths]:
            if p and os.path.exists(p):
                os.remove(p)


# --------------------------------------------------------------------------
# 2) generate -> run Planning on the Draft
# --------------------------------------------------------------------------

def _generate_from_runtime(mode: str, start_week: int, length: int,
                           visits_per_tech_week: float | None, source: str = "generate") -> dict:
    """The single generation path: build the engine state from the SQLite
    runtime state, run Planning under mode/horizon, save it as the current draft
    (so Review / edits / Publish work), mirror it into draft_plans, and record
    the run. No workbook upload, no snapshot — one source of truth."""
    import db_state
    import route_planner
    import runtime_state
    # Smart Fill: if the caller didn't pass an explicit per-tech/week target,
    # fall back to the persisted planner default so "set it once in Nastavení
    # Planneru, generate uses it" holds for every run. GPS extra (nearby-POS
    # top-up) + its radius are already read from config inside db_state.
    if visits_per_tech_week is None:
        import settings
        pv = settings.get("planner", "visits_per_tech_week")
        if pv:
            try:
                visits_per_tech_week = float(pv)
            except (TypeError, ValueError):
                pass
    state = runtime_state.build()
    db_state.configure(state, mode, start_week, length, visits_per_tech_week)
    cands_out: list = []
    rej_out: list = []
    messages = pipeline.run_planning(state, start_week, length,
                                     candidates_out=cands_out, rejected_out=rej_out)
    path = _tmp()
    try:
        state_xlsx.save_state(state, path)
        store.save_draft(path, f"Generovat ({source}): tyden {start_week}, delka {length}, rezim {mode}",
                         meta={"source": "runtime_state", "generatedAt": _now_iso(),
                               "engineVersion": ENGINE_VERSION})
    finally:
        os.remove(path)
    route_planner.materialize_draft_plans(state)
    summary = pipeline._summarize(state, start_week, length)
    try:
        import history
        assessment = history.run_assessment_from_candidates(cands_out, rej_out)
        assessment.update({k: summary.get(k) for k in summary if k not in assessment})
        summary["run_id"] = history.record_planner_run(
            source, mode, start_week, length, visits_per_tech_week, result=assessment)
        summary["assessment"] = assessment
    except Exception:  # noqa: BLE001 - never block planning on memory write
        pass
    return {"messages": messages, "summary": summary}


@app.post("/api/draft/generate", dependencies=[Depends(require_auth)])
def draft_generate(body: GenerateRequest):
    # Local desktop: generate straight from the SQLite runtime state — the
    # unified path (no uploaded snapshot draft).
    if LOCAL_MODE:
        return _generate_from_runtime(body.mode, body.start_week, body.length,
                                      body.visits_per_tech_week, source="generate")
    # Hosted/cloud path: runs on the uploaded draft workbook (Field Brain sets
    # goals/weights via config; the Planning Engine algorithm is unchanged).
    path = _require_draft_path()
    try:
        state = state_xlsx.load_state(path)
        brain_mod.apply_mode(state, body.mode)
        brain_mod.apply_capacity(state, body.visits_per_tech_week)
        messages = pipeline.run_planning(state, body.start_week, body.length)
        state_xlsx.save_state(state, path)
        store.save_draft(path, f"Generovat tour plan: tyden {body.start_week}, delka {body.length}, rezim {body.mode}")
        summary = pipeline._summarize(state, body.start_week, body.length)
        return {"messages": messages, "summary": summary}
    finally:
        os.remove(path)


@app.post("/api/draft/preflight", dependencies=[Depends(require_auth)])
def draft_preflight(body: PreflightRequest):
    """Field Brain pre-flight: simulate the horizon under a strategy mode +
    capacity and COMPUTE the business scorecard (CORE / cadence / neglect /
    campaign coverage / capacity) + a managerial recommendation BEFORE
    generating. Read-only."""
    path = _require_draft_path()
    try:
        return brain_mod.preflight(
            path, body.start_week, body.length, body.mode,
            body.visits_per_tech_week, body.tech_count_override,
        )
    finally:
        os.remove(path)


@app.get("/api/strategy-modes", dependencies=[Depends(require_auth)])
def strategy_modes():
    modes = [{"id": k, "label": v["label"], "desc": v["desc"]}
             for k, v in brain_mod.STRATEGY_MODES.items()]
    if LOCAL_MODE:
        modes.append({"id": "cela_sit", "label": "Celá síť",
                      "desc": "Sweep celé sítě podle zanedbanosti a skóre (kampaně vypnuté)."})
    return {"modes": modes}


# --------------------------------------------------------------------------
# 3) candidates (read-only; real Planning Engine with observability)
# --------------------------------------------------------------------------

@app.get("/api/draft/candidates", dependencies=[Depends(require_auth)])
def draft_candidates(week: int, technician: str | None = None):
    path = _require_draft_path()
    try:
        return candidates_mod.list_candidates(path, week, technician)
    finally:
        os.remove(path)


@app.get("/api/draft/pos/{pos_id}", dependencies=[Depends(require_auth)])
def draft_pos_detail(pos_id: str, week: int):
    """Full read-only diagnostic for one POS for the given week - the same
    data and score the Planning Engine used, plus why it is / is not a
    candidate, plus the Decision Support recommendation. No new logic; runs
    the same engine."""
    path = _require_draft_path()
    try:
        return candidates_mod.pos_detail(path, pos_id, week)
    finally:
        os.remove(path)


@app.get("/api/draft/what-if", dependencies=[Depends(require_auth)])
def draft_what_if(week: int):
    """Decision Support 'Co kdyby...': impact of manager levers on the
    candidate pool for `week`, derived from ONE engine run's own capture -
    interpretation/simulation only, the Planning Engine is unchanged."""
    path = _require_draft_path()
    try:
        return decision_mod.what_if(path, week)
    finally:
        os.remove(path)


# --------------------------------------------------------------------------
# 4) view + manual edits (Draft weeks only; locked weeks are protected)
# --------------------------------------------------------------------------

@app.get("/api/draft", dependencies=[Depends(require_auth)])
def draft_view():
    path = _require_draft_path()
    try:
        return {"rows": plan_io.read_enriched_draft(path)}
    finally:
        os.remove(path)


@app.get("/api/draft/geo", dependencies=[Depends(require_auth)])
def draft_geo():
    """Read-only map coordinates of the current draft plan (from draft_plans).
    Plumbing for the Review map view — no logic, just the coordinates the plan
    already has, so the frontend can plot what the engine produced."""
    rows = db.get(
        "SELECT pos_id pos, technician, day, name nazev, gps_x x, gps_y y "
        "FROM draft_plans WHERE gps_x IS NOT NULL AND gps_y IS NOT NULL "
        "AND gps_x<>0 AND gps_y<>0")
    return {"points": [dict(r) for r in rows]}


def _resync_draft_plans(path: str) -> int:
    """Keep SQLite draft_plans consistent with the edited draft workbook so the
    Review cockpit (feasibility / map / unserved) reflects manual edits. Reads
    only the MANAGER_PLAN sheet and re-runs the same persistence the generator
    uses. Pure sync — no planning logic, no engine run."""
    try:
        import openpyxl
        import route_planner
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if "MANAGER_PLAN" not in wb.sheetnames:
                return 0
            rows = [list(r) for r in wb["MANAGER_PLAN"].iter_rows(values_only=True)]
        finally:
            wb.close()
        return route_planner.materialize_draft_plans({"MANAGER_PLAN": rows}) if rows else 0
    except Exception:  # noqa: BLE001 — sync is best-effort; never break the edit
        return 0


@app.post("/api/draft/remove-pos", dependencies=[Depends(require_auth)])
def draft_remove_pos(body: RemovePosRequest):
    path = _require_draft_path()
    try:
        try:
            removed = plan_io.remove_pos(path, body.week, body.pos_id, body.technician)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        if removed == 0:
            raise HTTPException(status_code=404, detail="POS v navrhu nenalezen.")
        store.save_draft(path, f"Odebrat POS {body.pos_id} z navrhu")
        _resync_draft_plans(path)
        return {"removed": removed}
    finally:
        os.remove(path)


@app.post("/api/draft/change-technician", dependencies=[Depends(require_auth)])
def draft_change_technician(body: ChangeTechnicianRequest):
    path = _require_draft_path()
    try:
        try:
            changed = plan_io.change_technician(
                path, body.week, body.pos_id, body.old_technician, body.new_technician
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        if changed == 0:
            raise HTTPException(status_code=404, detail="POS v navrhu nenalezen.")
        store.save_draft(path, f"Presunout POS {body.pos_id} na {body.new_technician}")
        _resync_draft_plans(path)
        return {"changed": changed}
    finally:
        os.remove(path)


@app.post("/api/draft/add-pos", dependencies=[Depends(require_auth)])
def draft_add_pos(body: AddPosRequest):
    path = _require_draft_path()
    try:
        try:
            new_row = plan_io.add_pos(path, body.week, body.day, body.technician, body.pos_id)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        store.save_draft(path, f"Pridat POS {body.pos_id} do navrhu")
        _resync_draft_plans(path)
        return {"row": new_row}
    finally:
        os.remove(path)


# --------------------------------------------------------------------------
# 5) publish -> freeze the Draft's lowest week into an immutable snapshot
# --------------------------------------------------------------------------

@app.post("/api/publish", dependencies=[Depends(require_auth)])
def publish(body: PublishRequest):
    path = _require_draft_path()
    snap_path = None
    try:
        state = state_xlsx.load_state(path)
        result = pipeline.run_publish(state)
        if not result["publishedWeeks"]:
            raise HTTPException(status_code=400, detail=result["message"])

        snap_path = _tmp()
        state_xlsx.save_state(state, snap_path)

        draft_meta = store.read_draft_meta()
        meta = {
            "publishedAt": _now_iso(),
            "publishedWeeks": result["publishedWeeks"],
            "message": body.message,
            "engineVersion": ENGINE_VERSION,
            "sourceExports": {
                "posExport": draft_meta.get("posExport"),
                "salesAppExports": draft_meta.get("salesAppExports"),
            },
            "resumedFrom": draft_meta.get("resumedFrom"),
        }
        record = store.publish_snapshot(snap_path, meta)

        # The draft now reflects the published (locked) state, so the manager
        # sees the freshly-locked week; the next upload resumes from the new
        # snapshot regardless.
        store.save_draft(path, f"Draft po publikaci {record['id']}")
        if LOCAL_MODE:
            # Loop: a publish is a memory event + a KPI snapshot (incl. plan
            # fulfilment vs reality), tagged with this publish as provenance.
            try:
                import history
                eid = history.log_event("publish", "snapshot", str(record.get("id")),
                                        {"publishedWeeks": result["publishedWeeks"]})
                history.capture_metrics("publish", eid)
            except Exception:  # noqa: BLE001 - never fail a publish on a memory write
                pass
        return {"published": record, "engineMessage": result["message"]}
    finally:
        for p in (path, snap_path):
            if p and os.path.exists(p):
                os.remove(p)


# --------------------------------------------------------------------------
# 6) history of published versions
# --------------------------------------------------------------------------

@app.get("/api/versions", dependencies=[Depends(require_auth)])
def versions():
    return {"versions": list(reversed(store.read_index()))}


# --------------------------------------------------------------------------
# 7) download a published (or the draft) MANAGER_PLAN
# --------------------------------------------------------------------------

@app.get("/api/versions/{version_id}/manager-plan", dependencies=[Depends(require_auth)])
def download_published_plan(version_id: str):
    path = _tmp()
    try:
        try:
            store.download_snapshot(version_id, path)
        except Exception:
            raise HTTPException(status_code=404, detail=f"Verze {version_id} nenalezena.")
        return _stream_sheet(path, "MANAGER_PLAN_PUBLISHED", f"MANAGER_PLAN_{version_id}.xlsx")
    finally:
        if os.path.exists(path):
            os.remove(path)


@app.get("/api/draft/download", dependencies=[Depends(require_auth)])
def download_draft_plan():
    path = _require_draft_path()
    try:
        return _stream_sheet(path, "MANAGER_PLAN", "MANAGER_PLAN_draft.xlsx")
    finally:
        os.remove(path)


# --------------------------------------------------------------------------
# Cloud generate: run the heavy multi-week plan on a GitHub Actions runner
# (~7 GB RAM) instead of this 512 MB host. The backend only orchestrates:
# it dispatches the workflow, reports the run status, and streams the Excel
# the runner committed into output/. One button on the web, no OOM here.
# --------------------------------------------------------------------------

WORKFLOW_FILE = "generate-tourplan.yml"
ARTIFACT_NAME = "tour-plan"


@app.post("/api/cloud/generate", dependencies=[Depends(require_auth)])
def cloud_generate(body: CloudGenerateRequest):
    """Trigger the GitHub Actions workflow that builds the plan and uploads
    the Excel as an artifact. Returns immediately; poll /status."""
    try:
        gh.dispatch_workflow(WORKFLOW_FILE, {
            "start_week": body.start_week,
            "length": body.length,
            "visits_per_tech": body.visits_per_tech,
        })
    except Exception as e:
        raise HTTPException(status_code=502,
                            detail=f"Nepodařilo se spustit GitHub workflow: {e}")
    return {"ok": True, "start_week": body.start_week}


@app.get("/api/cloud/status", dependencies=[Depends(require_auth)])
def cloud_status():
    """Status of the newest workflow run + whether its Excel artifact is ready.
    Cloud generation is optional (GitHub Actions); when it isn't configured —
    the normal case for the local desktop app — report it as unavailable rather
    than erroring, since the local Planner generates without it."""
    try:
        run = gh.latest_run(WORKFLOW_FILE)
    except Exception as e:  # noqa: BLE001 - GH not configured / offline
        return {"run": None, "ready": False, "available": False, "reason": str(e)}
    ready = False
    if run and run["status"] == "completed" and run["conclusion"] == "success":
        ready = gh.run_artifact(run["id"], ARTIFACT_NAME) is not None
    return {"run": run, "ready": ready, "available": True}


@app.get("/api/cloud/download", dependencies=[Depends(require_auth)])
def cloud_download(start_week: int = 0):
    """Stream the Excel from the newest successful run's artifact."""
    try:
        run = gh.latest_run(WORKFLOW_FILE)
    except Exception as e:  # noqa: BLE001 - cloud generation not configured
        raise HTTPException(status_code=404, detail=f"Cloudové generování není dostupné: {e}")
    art = gh.run_artifact(run["id"], ARTIFACT_NAME) if run else None
    if not art:
        raise HTTPException(status_code=404,
                            detail="Plán ještě není hotový nebo artifact vypršel.")
    try:
        data = gh.download_artifact_xlsx(art["id"])
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Stažení artifactu selhalo: {e}")
    name = f"TOUR_PLAN_tydny_{start_week}.xlsx" if start_week else "TOUR_PLAN.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={name}"},
    )


# --------------------------------------------------------------------------
# The old GET/POST /api/rules endpoints (terminal / market / category /
# activity toggles written back to the draft workbook) have been removed.
# That whole planning model is now edited through the SQLite-backed
# configurator — GET/PUT /api/model — and applied to the runtime state by
# db_state.configure() -> model_config.apply_to_state() right before the
# engine runs. One source of truth (SQLite), no draft-workbook write path.
# --------------------------------------------------------------------------


# --------------------------------------------------------------------------
# Local desktop app: serve the existing frontend from THIS server so the UI
# and API share one origin (127.0.0.1). config.js is generated to point the
# frontend at the same origin; everything else in web/ is served as-is. The
# GitHub Pages copy of web/ is untouched - this only runs when FFO_LOCAL=1.
# --------------------------------------------------------------------------

if LOCAL_MODE:
    # ----------------------------------------------------------------------
    # Import (Excel -> SQLite): the only way real data enters the datastore.
    # POS Master, SalesApp history, Activity Plan, config in one workbook.
    # ----------------------------------------------------------------------
    import importer  # noqa: E402

    _DATA_TABLES = ["pos_master", "salesapp_visits", "campaigns", "technicians",
                    "closed_pos", "snapshots", "published_plans", "draft_plans",
                    "route_metrics"]

    @app.post("/api/import/workbook", dependencies=[Depends(require_auth)])
    async def import_workbook_ep(workbook: UploadFile = File(...)):
        path, prov = await _save_upload(workbook)
        try:
            counts = importer.import_workbook(path, prov["filename"])
            return {"ok": True, "counts": counts, "file": prov}
        finally:
            os.remove(path)

    @app.get("/api/data/summary", dependencies=[Depends(require_auth)])
    def data_summary():
        return {t: db.get(f"SELECT COUNT(*) AS c FROM {t}")[0]["c"] for t in _DATA_TABLES}

    # Technician configuration: role (TECHNIK/OZ/ADMIN/MANAGER) + active. All
    # technician metrics count only active TECHNIK. Manual edits stick across
    # imports (manual_role=1); auto-rule 3xx=OZ otherwise.
    class TechnicianUpdate(BaseModel):
        role: str | None = None
        active: bool | None = None
        excluded: bool | None = None

    @app.get("/api/technicians", dependencies=[Depends(require_auth)])
    def list_technicians():
        rows = db.get("SELECT name, role, manual_role, active, excluded, region, capacity_per_week "
                      "FROM technicians ORDER BY role, name")
        return {"technicians": [dict(r) for r in rows]}

    @app.put("/api/technicians/{name}", dependencies=[Depends(require_auth)])
    def update_technician(name: str, body: TechnicianUpdate):
        sets, params = [], []
        if body.role is not None:
            sets += ["role=?", "manual_role=1"]; params.append(body.role)
        if body.active is not None:
            sets.append("active=?"); params.append(1 if body.active else 0)
        if body.excluded is not None:
            sets.append("excluded=?"); params.append(1 if body.excluded else 0)
        if sets:
            params.append(name)
            db.run(f"UPDATE technicians SET {', '.join(sets)}, updated_at=datetime('now') WHERE name=?", tuple(params))
            # role/active/exclude change the peer baselines -> drop cached profiles
            # so insights and the technician detail reflect it without a restart.
            import diagnostics  # noqa: E402
            diagnostics.invalidate_cache()
        return {"ok": True}

    # Automatic import: drop a file, the system detects its type and processes it.
    import auto_import  # noqa: E402

    @app.post("/api/import/auto", dependencies=[Depends(require_auth)])
    async def import_auto(file: UploadFile = File(...)):
        path, prov = await _save_upload(file)
        try:
            return auto_import.import_file(path, prov["filename"])
        finally:
            os.remove(path)

    # Explicit, template-based import: the user picks the type, no guessing.
    import import_templates  # noqa: E402
    _IMPORT_KINDS = {"pos_master", "salesapp", "activity_plan", "workbook"}

    @app.get("/api/import/template/{kind}", dependencies=[Depends(require_auth)])
    def import_template(kind: str):
        try:
            data = import_templates.build(kind)
        except KeyError:
            raise HTTPException(status_code=404, detail=f"Neznámá šablona: {kind}")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=sablona_{kind}.xlsx"})

    @app.post("/api/import/sample", dependencies=[Depends(require_auth)])
    def import_sample():
        """Load SYNTHETIC demo data (sample_data/) on demand, so a fresh install
        has something to explore. This is fabricated data (fake POS/technicians),
        never real client data — the scaffold is config-only. Defined BEFORE
        /api/import/{kind} so it isn't captured as kind='sample'."""
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        base = os.environ.get("SAMPLE_DATA_DIR", os.path.join(base, "sample_data"))
        files = [("POS_master.xlsx", "pos_master"), ("SalesApp_export.xlsx", "salesapp")]
        out: dict = {"counts": {}}
        loaded = 0
        for fname, kind in files:
            p = os.path.join(base, fname)
            if os.path.exists(p):
                r = auto_import.import_file(p, f"demo_{fname}", force_kind=kind)
                out["counts"].update(r.get("counts", {}))
                loaded += 1
        if not loaded:
            raise HTTPException(status_code=404, detail="Ukázková (syntetická) data nejsou k dispozici.")
        out["detected"] = "sample_synthetic"
        return out

    @app.post("/api/import/{kind}", dependencies=[Depends(require_auth)])
    async def import_explicit(kind: str, file: UploadFile = File(...)):
        if kind not in _IMPORT_KINDS:
            raise HTTPException(status_code=400, detail=f"Neznámý typ importu: {kind}")
        path, prov = await _save_upload(file)
        try:
            return auto_import.import_file(path, prov["filename"], force_kind=kind)
        finally:
            os.remove(path)

    # Automatic anomaly alerts.
    import alerts as _alerts  # noqa: E402

    @app.get("/api/alerts", dependencies=[Depends(require_auth)])
    def get_alerts():
        return {"alerts": _alerts.list_alerts()}

    @app.post("/api/alerts/recompute", dependencies=[Depends(require_auth)])
    def recompute_alerts():
        return {"ok": True, "count": _alerts.recompute()}

    # Hard-exclude POS from planning (manager blacklist). Paste POS IDs; they
    # are injected into the engine BLACKLIST by db_state and never planned.
    import re as _re

    class ExclusionsRequest(BaseModel):
        pos_ids: str | list = ""
        reason: str | None = None

    @app.get("/api/exclusions", dependencies=[Depends(require_auth)])
    def list_exclusions():
        rows = db.get("SELECT pos_id, reason, created_at FROM pos_exclusions ORDER BY pos_id")
        return {"exclusions": [dict(r) for r in rows], "count": len(rows)}

    @app.post("/api/exclusions", dependencies=[Depends(require_auth)])
    def add_exclusions(body: ExclusionsRequest):
        raw = body.pos_ids if isinstance(body.pos_ids, list) else _re.split(r"[\s,;]+", str(body.pos_ids))
        ids = [str(x).strip() for x in raw if str(x).strip()]
        for pid in ids:
            db.run("INSERT INTO pos_exclusions (pos_id, reason) VALUES (?, ?) "
                   "ON CONFLICT(pos_id) DO UPDATE SET reason=excluded.reason", (pid, body.reason))
        return {"ok": True, "added": len(ids),
                "count": db.get("SELECT COUNT(*) AS c FROM pos_exclusions")[0]["c"]}

    @app.delete("/api/exclusions/{pos_id}", dependencies=[Depends(require_auth)])
    def delete_exclusion(pos_id: str):
        if pos_id == "_all":
            db.run("DELETE FROM pos_exclusions")
        else:
            db.run("DELETE FROM pos_exclusions WHERE pos_id=?", (pos_id,))
        return {"ok": True, "count": db.get("SELECT COUNT(*) AS c FROM pos_exclusions")[0]["c"]}

    # Temporary POS reassignment (dovolená/nemoc/výpověď) + manual per-POS
    # override. Whole-technician cover (from_technician) or specific POS list.
    # valid_from/valid_to give it a window; after valid_to it auto-returns
    # (the engine simply stops applying it). No SalesForce data is changed.
    class ReassignRequest(BaseModel):
        from_technician: str | None = None
        pos_ids: str | list = ""
        to_technician: str = ""
        reason: str | None = None
        valid_from: str | None = None
        valid_to: str | None = None

    @app.get("/api/reassignments", dependencies=[Depends(require_auth)])
    def list_reassignments():
        rows = db.get("SELECT id, from_technician, pos_id, to_technician, reason, "
                      "valid_from, valid_to, active, created_at FROM pos_reassignments "
                      "WHERE active=1 ORDER BY created_at DESC")
        today = __import__("datetime").date.today().isoformat()
        out = []
        for r in rows:
            d = dict(r)
            vf, vt = d.get("valid_from"), d.get("valid_to")
            d["current"] = (not vf or vf <= today) and (not vt or vt >= today)
            out.append(d)
        return {"reassignments": out, "count": len(out)}

    @app.post("/api/reassignments", dependencies=[Depends(require_auth)])
    def add_reassignment(body: ReassignRequest):
        if not body.to_technician:
            return {"ok": False, "error": "to_technician required"}
        raw = body.pos_ids if isinstance(body.pos_ids, list) else _re.split(r"[\s,;]+", str(body.pos_ids))
        ids = [str(x).strip() for x in raw if str(x).strip()]
        n = 0
        if body.from_technician and not ids:
            db.run("INSERT INTO pos_reassignments (from_technician, to_technician, reason, "
                   "valid_from, valid_to) VALUES (?, ?, ?, ?, ?)",
                   (body.from_technician, body.to_technician, body.reason,
                    body.valid_from, body.valid_to))
            n = 1
        else:
            for pid in ids:
                db.run("INSERT INTO pos_reassignments (from_technician, pos_id, to_technician, "
                       "reason, valid_from, valid_to) VALUES (?, ?, ?, ?, ?, ?)",
                       (body.from_technician, pid, body.to_technician, body.reason,
                        body.valid_from, body.valid_to))
                n += 1
        return {"ok": True, "added": n}

    @app.delete("/api/reassignments/{rid}", dependencies=[Depends(require_auth)])
    def delete_reassignment(rid: str):
        if rid == "_all":
            db.run("UPDATE pos_reassignments SET active=0")
        else:
            db.run("UPDATE pos_reassignments SET active=0 WHERE id=?", (rid,))
        return {"ok": True}

    # OZ campaign prep list: upload POS numbers to be prepared for an upcoming OZ
    # campaign -> planned with top priority (FORCE_INCLUDE).
    class PriorityRequest(BaseModel):
        pos_ids: str | list = ""
        campaign: str | None = None
        reason: str | None = None

    @app.get("/api/priority", dependencies=[Depends(require_auth)])
    def list_priority():
        rows = db.get("SELECT pos_id, campaign, reason, created_at FROM pos_priority "
                      "WHERE active=1 ORDER BY campaign, pos_id")
        return {"priority": [dict(r) for r in rows], "count": len(rows)}

    @app.post("/api/priority", dependencies=[Depends(require_auth)])
    def add_priority(body: PriorityRequest):
        raw = body.pos_ids if isinstance(body.pos_ids, list) else _re.split(r"[\s,;]+", str(body.pos_ids))
        ids = [str(x).strip() for x in raw if str(x).strip()]
        for pid in ids:
            db.run("INSERT INTO pos_priority (pos_id, campaign, reason, active) VALUES (?, ?, ?, 1) "
                   "ON CONFLICT(pos_id) DO UPDATE SET campaign=excluded.campaign, "
                   "reason=excluded.reason, active=1", (pid, body.campaign, body.reason))
        return {"ok": True, "added": len(ids),
                "count": db.get("SELECT COUNT(*) AS c FROM pos_priority WHERE active=1")[0]["c"]}

    @app.delete("/api/priority/{pos_id}", dependencies=[Depends(require_auth)])
    def delete_priority(pos_id: str):
        if pos_id == "_all":
            db.run("DELETE FROM pos_priority")
        else:
            db.run("DELETE FROM pos_priority WHERE pos_id=?", (pos_id,))
        return {"ok": True, "count": db.get("SELECT COUNT(*) AS c FROM pos_priority WHERE active=1")[0]["c"]}

    # Campaigns: editable in-app (target_visits = campaign goal; Excel ODHAD is
    # often empty, and the app is the source of truth).
    class CampaignUpdate(BaseModel):
        target_visits: int | None = None
        priority: int | None = None
        objective_id: int | None = None

    @app.get("/api/campaigns", dependencies=[Depends(require_auth)])
    def list_campaigns():
        rows = db.get("SELECT id, kind, name, year, start_week, end_week, priority, "
                      "override_gap, estimate, target_visits, objective_id, active "
                      "FROM campaigns ORDER BY start_week, name")
        return {"campaigns": [dict(r) for r in rows]}

    @app.put("/api/campaigns/{campaign_id}", dependencies=[Depends(require_auth)])
    def update_campaign(campaign_id: int, body: CampaignUpdate):
        sets, params = [], []
        for f in ("target_visits", "priority", "objective_id"):
            v = getattr(body, f)
            if v is not None:
                sets.append(f"{f}=?"); params.append(v)
        if sets:
            params.append(campaign_id)
            db.run(f"UPDATE campaigns SET {', '.join(sets)}, updated_at=datetime('now') WHERE id=?", tuple(params))
        return {"ok": True}

    import pos_insights  # noqa: E402

    @app.get("/api/pos/search", dependencies=[Depends(require_auth)])
    def pos_search(q: str = ""):
        """Search POS by number / name / city (command bar)."""
        return pos_insights.search(q)

    @app.get("/api/pos/list", dependencies=[Depends(require_auth)])
    def pos_list_ep(q: str | None = None, area: str | None = None, market: str | None = None,
                    technician: str | None = None, status: str = "all",
                    limit: int = 200, offset: int = 0):
        """All POS with last visit + weeks-since + cadence risk (the POS table)."""
        return pos_insights.pos_list(q, area, market, technician, status, limit, offset)

    @app.get("/api/pos/list/filters", dependencies=[Depends(require_auth)])
    def pos_list_filters():
        return pos_insights.list_filters()

    @app.get("/api/pos/{pos_id}/visits", dependencies=[Depends(require_auth)])
    def pos_visits(pos_id: str):
        """Informational: who visited this POS (technician vs OZ), when, what."""
        return pos_insights.pos_visit_summary(pos_id)

    @app.get("/api/pos/{pos_id}/card", dependencies=[Depends(require_auth)])
    def pos_card(pos_id: str):
        """Full POS card for the TourPlan controller: attributes, tech/OZ
        frequency, recommended vs actual cadence, deviation, trend, next-due,
        recommendation."""
        return pos_insights.pos_card(pos_id)

    # Living published TourPlan (main working screen) + cadence countdown.
    import live_plan  # noqa: E402

    @app.get("/api/live/board", dependencies=[Depends(require_auth)])
    def live_board(technician: str | None = None):
        return live_plan.board(technician)

    @app.get("/api/live/next-due", dependencies=[Depends(require_auth)])
    def live_next_due(technician: str | None = None, status: str | None = None):
        return live_plan.next_due(technician, status)

    # Route Planner: long-term per-technician visit plan (read model over the
    # draft the engine produced; km are supportive info only).
    import route_planner  # noqa: E402

    @app.get("/api/planner/technicians", dependencies=[Depends(require_auth)])
    def planner_technicians():
        return {"technicians": route_planner.planned_technicians()}

    @app.get("/api/planner/route", dependencies=[Depends(require_auth)])
    def planner_route(technician: str, week_from: int | None = None, week_to: int | None = None):
        return route_planner.technician_route(technician, week_from, week_to)

    # Planner simulation / decision-support: run the engine under a scenario
    # (mode + capacity) and measure workload / region load / totals.
    import planner_sim  # noqa: E402

    class SimRequest(BaseModel):
        mode: str = "vyvazeny"
        start_week: int
        length: int = 5
        visits_per_tech_week: float | None = None
        tech_count: int | None = None

    class WhatIfRequest(BaseModel):
        base: SimRequest
        scenario: SimRequest

    @app.post("/api/planner/simulate", dependencies=[Depends(require_auth)])
    def planner_simulate(body: SimRequest):
        return planner_sim.simulate(body.mode, body.start_week, body.length,
                                    body.visits_per_tech_week, body.tech_count)

    @app.post("/api/planner/assess", dependencies=[Depends(require_auth)])
    def planner_assess(body: SimRequest):
        return planner_sim.assess(body.mode, body.start_week, body.length,
                                  body.visits_per_tech_week, body.tech_count)

    @app.post("/api/planner/generate-runtime", dependencies=[Depends(require_auth)])
    def planner_generate_runtime(body: SimRequest):
        """Generate a draft straight from the SQLite runtime state — no upload.
        Same single generation path as /api/draft/generate in local mode."""
        return _generate_from_runtime(body.mode, body.start_week, body.length,
                                      body.visits_per_tech_week, source="generate-runtime")

    import planner_advisor  # noqa: E402

    class AdviseRequest(SimRequest):
        clear_neglect_weeks: int | None = None

    @app.post("/api/planner/advise", dependencies=[Depends(require_auth)])
    def planner_advise(body: AdviseRequest):
        """Decision-support: assessment + verdict, weakest link, binding
        constraint, what-to-change recommendations, and goal-seek."""
        return planner_advisor.advise(body.mode, body.start_week, body.length,
                                      body.visits_per_tech_week, body.tech_count,
                                      body.clear_neglect_weeks)

    import planner_sweep  # noqa: E402

    class SweepRequest(BaseModel):
        mode: str = "vyvazeny"
        start_week: int
        length: int = 5
        capacities: list[int] | None = None
        tech_count: int | None = None

    @app.post("/api/planner/sweep", dependencies=[Depends(require_auth)])
    def planner_sweep_ep(body: SweepRequest):
        """Predictions: POS served, network coverage and weeks-to-cover across
        capacities (e.g. 35/40/45 per technician-week)."""
        return planner_sweep.sweep(body.mode, body.start_week, body.length,
                                   body.capacities, body.tech_count)

    # Plan vs. reality (SalesApp) - the tracking / evaluation half.
    import plan_reality  # noqa: E402

    @app.get("/api/reality/fulfillment", dependencies=[Depends(require_auth)])
    def reality_fulfillment(week_from: int, week_to: int):
        return plan_reality.fulfillment(week_from, week_to)

    @app.get("/api/reality/technicians", dependencies=[Depends(require_auth)])
    def reality_technicians(week_from: int | None = None, week_to: int | None = None):
        return plan_reality.reality(week_from, week_to)

    # Advisory time-feasibility of the generated plan (duration + travel model)
    # vs available hours. Read-only over draft_plans; never changes the plan.
    import plan_feasibility  # noqa: E402

    @app.get("/api/plan/feasibility", dependencies=[Depends(require_auth)])
    def plan_feasibility_ep(week_from: int | None = None, week_to: int | None = None):
        return plan_feasibility.feasibility(week_from, week_to)

    # Actual driven route (order known from SalesApp times) + km + travel time.
    import route_actual  # noqa: E402

    @app.get("/api/route/days", dependencies=[Depends(require_auth)])
    def route_days(technician: str):
        return {"days": route_actual.technician_days(technician)}

    @app.get("/api/route/actual", dependencies=[Depends(require_auth)])
    def route_actual_ep(technician: str, date_from: str | None = None, date_to: str | None = None):
        return route_actual.technician_route(technician, date_from, date_to)

    # Route analytics: metrics + map layers + efficiency findings + trends.
    import route_analytics  # noqa: E402

    @app.get("/api/analytics/day", dependencies=[Depends(require_auth)])
    def analytics_day(technician: str, date: str, radius_km: float = 2.0):
        return route_analytics.day(technician, date, radius_km)

    @app.get("/api/analytics/trends", dependencies=[Depends(require_auth)])
    def analytics_trends(technician: str, days_back: int = 90):
        return route_analytics.trends(technician, days_back)

    import team_analytics  # noqa: E402

    @app.get("/api/analytics/team", dependencies=[Depends(require_auth)])
    def analytics_team(days_back: int = 21):
        return team_analytics.overview(days_back)

    # All-technicians filterable time series (dashboard graphs). Fast SQL — no
    # per-day route reconstruction. Filter by grain/date/region/chain/campaign.
    import tech_trends  # noqa: E402

    @app.get("/api/analytics/technicians/series", dependencies=[Depends(require_auth)])
    def analytics_tech_series(grain: str = "week", date_from: str | None = None,
                              date_to: str | None = None, region: str | None = None,
                              market: str | None = None, campaign: str | None = None,
                              role: str = "TECHNIK"):
        return tech_trends.all_series(grain, date_from, date_to, region, market, campaign, role)

    @app.get("/api/analytics/technicians/filters", dependencies=[Depends(require_auth)])
    def analytics_tech_filters():
        return tech_trends.filter_options()

    import planner_unserved  # noqa: E402

    @app.post("/api/planner/unserved", dependencies=[Depends(require_auth)])
    def planner_unserved_ep(body: SimRequest):
        """Which important POS did NOT get planned, grouped by the engine's own
        reason (capacity / hold-back / min-gap / filtered)."""
        return planner_unserved.unserved(body.mode, body.start_week, body.length,
                                         body.visits_per_tech_week, body.tech_count)

    @app.post("/api/planner/whatif", dependencies=[Depends(require_auth)])
    def planner_whatif(body: WhatIfRequest):
        return planner_sim.what_if(body.base.model_dump(), body.scenario.model_dump())

    # Business Rules: planning logic as data (toggle / edit params, no code).
    import business_rules as _rules  # noqa: E402

    class RuleUpdate(BaseModel):
        enabled: bool | None = None
        params: dict | None = None
        scope: str = "global"
        scope_value: str | None = None

    @app.get("/api/rules/business", dependencies=[Depends(require_auth)])
    def get_business_rules():
        return {"rules": _rules.list_rules(), "effective": _rules.effective()}

    @app.put("/api/rules/business/{code}", dependencies=[Depends(require_auth)])
    def update_business_rule(code: str, body: RuleUpdate):
        if body.enabled is not None:
            _rules.set_enabled(code, body.enabled, body.scope, body.scope_value)
        if body.params is not None:
            _rules.set_params(code, body.params, body.scope, body.scope_value)
        _log_config("business_rule", code, body.model_dump(exclude_none=True))
        return {"ok": True, "effective": _rules.effective()}

    # Config-change audit: every planning-model / rule / setting edit becomes an
    # event, so "historie změn konfigurace" is queryable next to imports/publishes.
    import history as _history  # noqa: E402

    def _log_config(area: str, key: str, detail: dict) -> None:
        try:
            _history.log_event("config_change", area, key, detail)
        except Exception:  # noqa: BLE001 - never block a config edit on logging
            pass

    # Business cadence rules (CORN/CORE/GECO/segment) - editable + effective.
    import cadence_config  # noqa: E402

    class CadenceUpdate(BaseModel):
        min_gap_weeks: float | None = None
        max_interval_weeks: float | None = None
        active: bool | None = None
        priority: int | None = None

    @app.get("/api/cadence", dependencies=[Depends(require_auth)])
    def cadence_list():
        return {"rules": cadence_config.list_rules()}

    @app.put("/api/cadence/{rule_id}", dependencies=[Depends(require_auth)])
    def cadence_update(rule_id: str, body: CadenceUpdate):
        cadence_config.set_override(rule_id, body.min_gap_weeks, body.max_interval_weeks,
                                    body.active, body.priority)
        _log_config("cadence", rule_id, body.model_dump(exclude_none=True))
        return {"ok": True}

    @app.delete("/api/cadence/{rule_id}", dependencies=[Depends(require_auth)])
    def cadence_reset(rule_id: str):
        cadence_config.reset(rule_id)
        return {"ok": True}

    class CustomCadenceRequest(BaseModel):
        scope: str = "category"
        match_value: str
        min_gap_weeks: float | None = None
        max_interval_weeks: float | None = None
        guarantee_type: str = "SOFT"
        interval_type: str = "RECURRING"
        priority: int = 100

    @app.post("/api/cadence/custom", dependencies=[Depends(require_auth)])
    def cadence_add_custom(body: CustomCadenceRequest):
        try:
            r = cadence_config.add_custom_rule(
                body.scope, body.match_value, body.min_gap_weeks, body.max_interval_weeks,
                body.guarantee_type, body.interval_type, body.priority)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        _log_config("cadence_custom", r["ruleId"], body.model_dump())
        return {"ok": True, "rule": r}

    @app.put("/api/cadence/custom/{rule_id}/active", dependencies=[Depends(require_auth)])
    def cadence_custom_active(rule_id: str, active: bool = True):
        cadence_config.set_custom_active(rule_id, active)
        return {"ok": True}

    @app.delete("/api/cadence/custom/{rule_id}", dependencies=[Depends(require_auth)])
    def cadence_custom_delete(rule_id: str):
        cadence_config.delete_custom_rule(rule_id)
        return {"ok": True}

    @app.get("/api/planner/customer-types", dependencies=[Depends(require_auth)])
    def planner_customer_types():
        cats = db.get("SELECT category v, COUNT(*) c FROM pos_master WHERE category IS NOT NULL "
                      "AND category<>'' GROUP BY category ORDER BY c DESC")
        mkts = db.get("SELECT market v, COUNT(*) c FROM pos_master WHERE market IS NOT NULL "
                      "AND market<>'' GROUP BY market ORDER BY c DESC")
        return {"categories": [{"value": r["v"], "count": r["c"]} for r in cats],
                "markets": [{"value": r["v"], "count": r["c"]} for r in mkts]}

    # Planning-model configurator: terminals / partners / categories /
    # activities as editable sections (checkboxes / choices), overlaid onto
    # the engine's config sheets before planning. Same pattern as cadence.
    import model_config  # noqa: E402

    class ModelUpdate(BaseModel):
        col: str
        value: object

    @app.get("/api/model", dependencies=[Depends(require_auth)])
    def model_list():
        return {"sections": model_config.sections()}

    @app.put("/api/model/{section}/{match_key:path}", dependencies=[Depends(require_auth)])
    def model_update(section: str, match_key: str, body: ModelUpdate):
        try:
            model_config.set_override(section, match_key, body.col, body.value)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        _log_config("model", f"{section}/{match_key}", {"col": body.col, "value": body.value})
        return {"ok": True}

    @app.delete("/api/model/{section}/{match_key:path}", dependencies=[Depends(require_auth)])
    def model_reset(section: str, match_key: str, col: str | None = None):
        model_config.reset(section, match_key, col)
        return {"ok": True}

    # Engine business-parameter inventory: every business constant / score
    # weight the engine uses when deciding, its default, effective value, and
    # what part of the algorithm it drives. Read-out for the "Inventura
    # parametrů" screen so nothing the planner uses is invisible or hardcoded.
    import engine_config  # noqa: E402

    @app.get("/api/engine/inventory", dependencies=[Depends(require_auth)])
    def engine_inventory():
        return {"parameters": engine_config.inventory()}

    # Historical memory: unified activity timeline (import / publish / config
    # change), per-POS change history (esp. PPT), and KPI trend series. Reads
    # the substrate that the importer / config setters now populate.
    import history  # noqa: E402

    @app.get("/api/history/events", dependencies=[Depends(require_auth)])
    def history_events(kind: str | None = None, limit: int = 200):
        return {"events": history.events(kind, limit)}

    @app.get("/api/pos/{pos_id}/history", dependencies=[Depends(require_auth)])
    def pos_change_history(pos_id: str, limit: int = 100):
        return {"pos": pos_id, "history": history.pos_history(pos_id, limit)}

    @app.get("/api/history/metrics", dependencies=[Depends(require_auth)])
    def history_metrics(entity_type: str, metric_key: str, entity_id: str | None = None):
        return {"series": history.metric_series(entity_type, metric_key, entity_id)}

    @app.get("/api/history/planner-runs", dependencies=[Depends(require_auth)])
    def history_planner_runs(limit: int = 100):
        return {"runs": history.planner_runs(limit)}

    # Memory read layer: the stable query contract the cockpit and every future
    # layer (AI / alerts / prediction / benchmarking) attaches to. Pure reads.
    import memory  # noqa: E402

    @app.get("/api/memory/catalog", dependencies=[Depends(require_auth)])
    def memory_catalog():
        return {"metrics": memory.catalog()}

    @app.get("/api/memory/trend", dependencies=[Depends(require_auth)])
    def memory_trend(entity_type: str, metric_key: str, entity_id: str | None = None,
                     grain: str = "week"):
        return memory.trend(entity_type, metric_key, entity_id, grain)

    @app.get("/api/memory/pos/{pos_id}/evolution", dependencies=[Depends(require_auth)])
    def memory_pos_evolution(pos_id: str):
        return memory.pos_evolution(pos_id)

    @app.get("/api/memory/planner-run/{run_id}", dependencies=[Depends(require_auth)])
    def memory_run_explain(run_id: int):
        r = memory.planner_run_explain(run_id)
        if r is None:
            raise HTTPException(status_code=404, detail="Běh planneru nenalezen")
        return r

    @app.get("/api/memory/config-diff", dependencies=[Depends(require_auth)])
    def memory_config_diff(run_a: int, run_b: int):
        return memory.config_diff(run_a, run_b)

    # Insight layer: anomaly / inefficiency / opportunity discovery over the
    # SalesApp truth. Shows what the manager wouldn't notice, with a "why".
    import insights  # noqa: E402

    @app.get("/api/insights", dependencies=[Depends(require_auth)])
    def insights_list(days_back: int = 90):
        return insights.insights(days_back)

    # Cause analysis for one technician: WHY they are inefficient + the biggest
    # improvement opportunity (scattered area / bad ordering / few POS/day /
    # single-purpose / isolated visits / short field time), each vs peers.
    import diagnostics  # noqa: E402

    @app.get("/api/insights/diagnose", dependencies=[Depends(require_auth)])
    def insights_diagnose(technician: str, days_back: int = 90):
        d = diagnostics.diagnose(technician, days_back)
        if d is None:
            raise HTTPException(status_code=404, detail="Pro technika nejsou data trasy")
        return d

    # Company-wide view in the language of time: total lost capacity, where the
    # reserves are (regions), and the biggest opportunities.
    @app.get("/api/insights/company", dependencies=[Depends(require_auth)])
    def insights_company(days_back: int = 90):
        return diagnostics.company_overview(days_back)

    # Health Score: composite per-technician efficiency (100 = healthy, low =
    # critical). Powers the "Kritické případy" section.
    @app.get("/api/insights/health", dependencies=[Depends(require_auth)])
    def insights_health(days_back: int = 90, role: str = "TECHNIK"):
        return diagnostics.health_scores(days_back, role)

    # Deep technician profile: KPIs + Health breakdown + cause diagnosis +
    # TourPlan fulfilment + missed planned POS driven past + the day list.
    import tech_detail  # noqa: E402

    @app.get("/api/technician/{name}", dependencies=[Depends(require_auth)])
    def technician_profile(name: str, days_back: int = 120):
        return tech_detail.profile(name, days_back)

    @app.get("/api/technician/{name}/day/{date}", dependencies=[Depends(require_auth)])
    def technician_day(name: str, date: str):
        return tech_detail.day(name, date)

    @app.get("/api/technician/{name}/hotspots", dependencies=[Depends(require_auth)])
    def technician_hotspots(name: str, days_back: int = 90):
        import tech_hotspots  # noqa: E402
        return tech_hotspots.hotspots(name, days_back)

    # Time-series trends for a technician or a region (středisko), with flexible
    # time filtering (week/month grain, any date range).
    import trends as _trends  # noqa: E402

    @app.get("/api/trends/regions", dependencies=[Depends(require_auth)])
    def trends_regions():
        return {"regions": _trends.regions()}

    @app.get("/api/trends", dependencies=[Depends(require_auth)])
    def trends_series(entity: str, type: str = "technician", grain: str = "week",
                      days_back: int = 180, date_from: str | None = None,
                      date_to: str | None = None):
        return _trends.series(type, entity, grain, days_back, date_from, date_to)

    # Monthly Summary: management overview for a chosen period, with filters and
    # drill-down. Heavy (route reconstruction across the scope) — the frontend
    # calls it once per filter change.
    import summary as _summary  # noqa: E402

    @app.get("/api/summary/dimensions", dependencies=[Depends(require_auth)])
    def summary_dimensions():
        return _summary.dimensions()

    # GIS layers: big network map (summary) + road-routed technician day.
    import gis as _gis  # noqa: E402

    @app.get("/api/gis/network", dependencies=[Depends(require_auth)])
    def gis_network(period: str = "month", year: int | None = None, month: int | None = None,
                    quarter: int | None = None, date_from: str | None = None, date_to: str | None = None,
                    role: str = "TECHNIK", region: str | None = None, technician: str | None = None,
                    chain: str | None = None, visit_type: str | None = None, active: str | None = "active",
                    include_optimal: bool = False):
        return _gis.network(period, year, month, quarter, date_from, date_to,
                            role, region, technician, chain, visit_type, active, include_optimal)

    @app.get("/api/gis/technician/{name}/day/{date}", dependencies=[Depends(require_auth)])
    def gis_day(name: str, date: str, radius_m: int = 250):
        return _gis.technician_day(name, date, radius_m)

    @app.get("/api/gis/pos/{pos_id}", dependencies=[Depends(require_auth)])
    def gis_pos(pos_id: str):
        return _gis.pos_detail(pos_id)

    # Planner Phase 1: predictive visit duration (collective, nationwide).
    import duration as _duration  # noqa: E402

    @app.get("/api/planner/duration/overview", dependencies=[Depends(require_auth)])
    def duration_overview():
        ov = _duration.overview()
        if not ov.get("national"):
            _duration.rebuild(); ov = _duration.overview()
        return ov

    @app.post("/api/planner/duration/rebuild", dependencies=[Depends(require_auth)])
    def duration_rebuild():
        return _duration.rebuild()

    @app.get("/api/planner/duration/pos/{pos_id}", dependencies=[Depends(require_auth)])
    def duration_pos(pos_id: str):
        return _duration.predict(pos_id)

    # Planner: learned TRANSITION model (real cost of moving between two stops,
    # objective predictors only, ambitious-but-achievable). Replaces the constant
    # crow-flight travel model in the day-time budget.
    import transition_model as _transition  # noqa: E402

    @app.get("/api/planner/transition/overview", dependencies=[Depends(require_auth)])
    def transition_overview():
        ov = _transition.overview()
        if not any(b.get("n") for b in ov.get("byBand", [])):
            _transition.rebuild(); ov = _transition.overview()
        return ov

    @app.post("/api/planner/transition/rebuild", dependencies=[Depends(require_auth)])
    def transition_rebuild():
        return _transition.rebuild()

    # The assembled REFERENCE DAY (budget = learned productive − reserve − on-top;
    # stop cost = learned duration + learned transition). One payload for the UI.
    import reference_day as _refday  # noqa: E402

    @app.get("/api/planner/reference-day", dependencies=[Depends(require_auth)])
    def reference_day_overview(role: str = "TECHNIK"):
        return _refday.calibration(role)

    # Planner v2 (parallel, feasibility-by-construction) — runs beside v1 for A/B.
    # Read-only: builds a v2 plan in memory and compares, never touches the draft.
    import planner_v2 as _pv2  # noqa: E402

    class V2Request(BaseModel):
        start_week: int
        length: int = 1
        mode: str = "vyvazeny"
        visits_per_tech_week: float | None = None

    @app.post("/api/planner/v2/simulate", dependencies=[Depends(require_auth)])
    def planner_v2_simulate(body: V2Request):
        return _pv2.simulate(body.start_week, body.length, body.mode,
                             body.visits_per_tech_week)

    @app.get("/api/planner/v2/history", dependencies=[Depends(require_auth)])
    def planner_v2_history(limit: int = 50):
        return {"runs": _pv2.ab_history(limit)}

    # Planner Phase 2: micro-clustering of nearby POS.
    import clustering as _clustering  # noqa: E402

    @app.get("/api/planner/clusters/overview", dependencies=[Depends(require_auth)])
    def clusters_overview():
        ov = _clustering.overview()
        if not ov.get("clusters"):
            _clustering.rebuild(); ov = _clustering.overview()
        return ov

    @app.post("/api/planner/clusters/rebuild", dependencies=[Depends(require_auth)])
    def clusters_rebuild():
        return _clustering.rebuild()

    @app.get("/api/planner/clusters/pos/{pos_id}", dependencies=[Depends(require_auth)])
    def clusters_pos(pos_id: str):
        return _clustering.cluster_of(pos_id)

    # Planner: learned daily productive-capacity standard (per role).
    import capacity as _capacity  # noqa: E402

    @app.get("/api/planner/capacity", dependencies=[Depends(require_auth)])
    def capacity_overview():
        ov = _capacity.overview()
        if not ov.get("roles"):
            _capacity.rebuild(); ov = _capacity.overview()
        return ov

    @app.post("/api/planner/capacity/rebuild", dependencies=[Depends(require_auth)])
    def capacity_rebuild():
        return _capacity.rebuild()

    # Planner [S] Coverage & Campaign: configurable segments + coverage state.
    import segments as _segments  # noqa: E402

    @app.get("/api/planner/coverage", dependencies=[Depends(require_auth)])
    def planner_coverage():
        return _segments.coverage()

    @app.get("/api/planner/segments", dependencies=[Depends(require_auth)])
    def planner_segments():
        return {"segments": _segments.definitions(), "fieldsMeta": _segments.fields_meta()}

    @app.post("/api/planner/segments/seed", dependencies=[Depends(require_auth)])
    def planner_segments_seed():
        """Create the default segment set (Velké/Malé/LI terminály, klasifikace…)
        if none exist. Exposes the existing seed — no new logic."""
        return _segments.seed_defaults()

    class SegmentBody(BaseModel):
        id: int | None = None
        name: str
        rule: dict
        target_cadence_weeks: float | None = None
        priority: int = 3
        business_weight: float = 1.0
        include_in_campaign: bool = True
        min_coverage_pct: float = 80.0
        active: bool = True
        sort_order: int = 100

    @app.post("/api/planner/segments", dependencies=[Depends(require_auth)])
    def planner_segment_upsert(body: SegmentBody):
        return _segments.upsert(body.model_dump())

    @app.delete("/api/planner/segments/{seg_id}", dependencies=[Depends(require_auth)])
    def planner_segment_delete(seg_id: int):
        return _segments.delete(seg_id)

    # Planner [T] Task Engine: generic tasks over POS.
    import tasks as _tasks  # noqa: E402
    _tasks.seed_default_types()

    @app.get("/api/planner/tasks", dependencies=[Depends(require_auth)])
    def planner_tasks():
        return _tasks.open_tasks()

    @app.get("/api/planner/task-types", dependencies=[Depends(require_auth)])
    def planner_task_types():
        return {"types": _tasks.types()}

    @app.post("/api/planner/task-types", dependencies=[Depends(require_auth)])
    def planner_task_type_upsert(body: dict):
        return _tasks.upsert_type(body)

    @app.post("/api/planner/tasks", dependencies=[Depends(require_auth)])
    def planner_task_create(body: dict):
        return _tasks.create(body)

    @app.post("/api/planner/tasks/bulk", dependencies=[Depends(require_auth)])
    def planner_tasks_bulk(body: dict):
        """Create tasks for many POS at once from a parsed list (rows =
        [{pos, quantity?, note?}]) with one shared type/deadline/priority."""
        return _tasks.bulk_create(body.get("rows", []), body.get("type_id"),
                                  body.get("deadline"), body.get("priority"),
                                  body.get("est_minutes"), body.get("combinable"))

    @app.post("/api/planner/tasks/bulk-upload", dependencies=[Depends(require_auth)])
    async def planner_tasks_bulk_upload(file: UploadFile = File(...), type_id: int = Form(...),
                                        deadline: str | None = Form(None), priority: int | None = Form(None),
                                        est_minutes: float | None = Form(None)):
        """Upload an Excel of POS (+ optional quantity/note) → one task per POS."""
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(await file.read()); path = tmp.name
        try:
            rows = _tasks.parse_bulk_excel(path)
        finally:
            os.unlink(path)
        res = _tasks.bulk_create(rows, type_id, deadline, priority, est_minutes)
        res["parsed"] = len(rows)
        return res

    @app.put("/api/planner/tasks/{task_id}/status", dependencies=[Depends(require_auth)])
    def planner_task_status(task_id: int, body: dict):
        return _tasks.set_status(task_id, body.get("status", "done"))

    @app.get("/api/planner/tasks/pos/{pos_id}", dependencies=[Depends(require_auth)])
    def planner_tasks_pos(pos_id: str):
        return {"tasks": _tasks.for_pos(pos_id)}

    @app.get("/api/summary", dependencies=[Depends(require_auth)])
    def summary_overview(period: str = "month", year: int | None = None,
                         month: int | None = None, quarter: int | None = None,
                         date_from: str | None = None, date_to: str | None = None,
                         role: str = "TECHNIK", region: str | None = None,
                         technician: str | None = None, chain: str | None = None,
                         visit_type: str | None = None, active: str | None = "active",
                         grain: str = "week"):
        return _summary.summary(period, year, month, quarter, date_from, date_to,
                                role, region, technician, chain, visit_type, active, grain)

    # Settings platform: configure planner/optimization/dashboard/report/map/
    # scoring from the app. Definitions drive a generic admin UI; values override.
    import settings as _settings  # noqa: E402

    class SettingUpdate(BaseModel):
        value: object
        scope: str = "global"
        scope_value: str | None = None

    class ViewUpsert(BaseModel):
        definition: object
        is_default: bool = False

    @app.get("/api/settings/definitions", dependencies=[Depends(require_auth)])
    def settings_definitions(namespace: str | None = None):
        return {"definitions": _settings.definitions(namespace)}

    @app.get("/api/settings/{namespace}", dependencies=[Depends(require_auth)])
    def settings_get(namespace: str):
        return {"namespace": namespace, "values": _settings.effective(namespace),
                "definitions": _settings.definitions(namespace)}

    @app.put("/api/settings/{namespace}/{key}", dependencies=[Depends(require_auth)])
    def settings_put(namespace: str, key: str, body: SettingUpdate):
        _settings.set_value(namespace, key, body.value, body.scope, body.scope_value)
        _log_config("setting", f"{namespace}.{key}", {"value": body.value})
        return {"ok": True, "values": _settings.effective(namespace)}

    @app.get("/api/views/{namespace}", dependencies=[Depends(require_auth)])
    def views_list(namespace: str):
        return {"views": _settings.list_views(namespace)}

    @app.put("/api/views/{namespace}/{name}", dependencies=[Depends(require_auth)])
    def views_put(namespace: str, name: str, body: ViewUpsert):
        _settings.save_view(namespace, name, body.definition, body.is_default)
        return {"ok": True, "views": _settings.list_views(namespace)}

    @app.delete("/api/views/{namespace}/{name}", dependencies=[Depends(require_auth)])
    def views_delete(namespace: str, name: str):
        _settings.delete_view(namespace, name)
        return {"ok": True}

    import sys as _sys

    from fastapi.responses import FileResponse, HTMLResponse, Response
    from fastapi.staticfiles import StaticFiles

    def _web_dir() -> str:
        for base in (getattr(_sys, "_MEIPASS", None),
                     os.path.dirname(os.path.dirname(os.path.abspath(__file__)))):
            if base:
                d = os.path.join(base, "web")
                if os.path.isdir(d):
                    return d
        return os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "web")

    _WEB = _web_dir()

    @app.get("/config.js")
    def _local_config():
        # Same-origin API + a marker so the frontend can auto-login locally.
        return Response('window.FFO_API_BASE="";window.FFO_LOCAL=true;',
                        media_type="application/javascript")

    @app.get("/", response_class=HTMLResponse)
    def _index():
        with open(os.path.join(_WEB, "index.html"), encoding="utf-8") as f:
            return f.read()

    @app.get("/favicon.ico")
    def _favicon():
        ico = os.path.join(_WEB, "favicon.ico")
        if os.path.exists(ico):
            return FileResponse(ico)
        return Response(status_code=204)  # no icon shipped: silence the 404

    # Everything else (app.js, styles.css, ...) straight from web/.
    app.mount("/", StaticFiles(directory=_WEB), name="web")
