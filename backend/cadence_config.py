"""Editable business cadence rules (CORN=4 w, CORE=2 w, GECO=5 w, ...).

The base cadence rules are structured config in the CADENCE_RULES sheet
(ruleId, scope, matchValue, minGapWeeks, maxIntervalWeeks, intervalType,
guaranteeType, active). This module reads them and merges the user's edits from
the cadence_overrides table, and db_state applies those edits onto the engine's
CADENCE_RULES before planning - so the whole cadence model is editable from the
UI and takes effect, with no code change and no new scoring logic.
"""
from __future__ import annotations

import os

import db
import store

# columns we surface / allow editing
_COLS = ("ruleId", "scope", "matchValue", "minGapWeeks", "maxIntervalWeeks",
         "intervalType", "guaranteeType", "active", "priority", "notes")


def _base_rules() -> list[dict]:
    """Read the CADENCE_RULES sheet from the current snapshot (light: one sheet)."""
    import openpyxl
    path = store.snapshot_temp()
    out: list[dict] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if "CADENCE_RULES" not in wb.sheetnames:
                return []
            rows = list(wb["CADENCE_RULES"].iter_rows(values_only=True))
            if not rows:
                return []
            header = [str(h) for h in rows[0]]
            hi = {n: i for i, n in enumerate(header)}
            for r in rows[1:]:
                rid = r[hi["ruleId"]] if "ruleId" in hi else None
                if rid in (None, ""):
                    continue
                out.append({c: (r[hi[c]] if c in hi and hi[c] < len(r) else None) for c in _COLS})
        finally:
            wb.close()
    finally:
        try:
            os.remove(path)
        except OSError:
            pass
    return out


def _overrides() -> dict:
    return {r["rule_id"]: dict(r) for r in db.get("SELECT * FROM cadence_overrides")}


def _custom_rules() -> list[dict]:
    return [dict(r) for r in db.get("SELECT * FROM cadence_custom_rules ORDER BY match_value")]


def list_rules() -> list[dict]:
    """Base cadence rules merged with the user's overrides (effective values),
    plus any UI-added custom rules (marked custom=True)."""
    ov = _overrides()
    out = []
    for base in _base_rules():
        rid = str(base["ruleId"])
        o = ov.get(rid)
        eff = dict(base)
        eff["overridden"] = bool(o)
        eff["custom"] = False
        if o:
            if o["min_gap_weeks"] is not None:
                eff["minGapWeeks"] = o["min_gap_weeks"]
            if o["max_interval_weeks"] is not None:
                eff["maxIntervalWeeks"] = o["max_interval_weeks"]
            if o["active"] is not None:
                eff["active"] = "YES" if o["active"] else "NO"
            if o["priority"] is not None:
                eff["priority"] = o["priority"]
        out.append(eff)
    for c in _custom_rules():
        out.append({
            "ruleId": c["rule_id"], "scope": c["scope"], "matchValue": c["match_value"],
            "minGapWeeks": c["min_gap_weeks"], "maxIntervalWeeks": c["max_interval_weeks"],
            "intervalType": c["interval_type"], "guaranteeType": c["guarantee_type"],
            "active": "YES" if c["active"] else "NO", "priority": c["priority"],
            "notes": c["notes"], "overridden": False, "custom": True,
        })
    return out


def add_custom_rule(scope: str, match_value: str, min_gap_weeks=None,
                    max_interval_weeks=None, guarantee_type: str = "SOFT",
                    interval_type: str = "RECURRING", priority: int = 100,
                    notes: str | None = None) -> dict:
    """Add a new cadence rule for a customer type from the UI — no code, no
    re-import. It is appended to the engine's CADENCE_RULES at plan time."""
    scope = (scope or "category").strip().lower()
    if scope not in ("category", "market"):
        scope = "category"
    mv = (match_value or "").strip()
    if not mv:
        raise ValueError("Vyplň typ zákazníka (kategorie / market).")
    rid = f"UI_{scope.upper()}_{mv}".replace(" ", "_")
    db.run(
        "INSERT INTO cadence_custom_rules (rule_id, scope, match_value, min_gap_weeks, "
        "max_interval_weeks, interval_type, guarantee_type, priority, active, notes) "
        "VALUES (?,?,?,?,?,?,?,?,1,?) "
        "ON CONFLICT(rule_id) DO UPDATE SET min_gap_weeks=excluded.min_gap_weeks, "
        "max_interval_weeks=excluded.max_interval_weeks, guarantee_type=excluded.guarantee_type, "
        "interval_type=excluded.interval_type, priority=excluded.priority, notes=excluded.notes, "
        "active=1",
        (rid, scope, mv, min_gap_weeks, max_interval_weeks,
         (interval_type or "RECURRING").upper(), (guarantee_type or "SOFT").upper(),
         priority, notes))
    return {"ruleId": rid, "matchValue": mv, "scope": scope}


def delete_custom_rule(rule_id: str) -> None:
    db.run("DELETE FROM cadence_custom_rules WHERE rule_id=?", (rule_id,))


def set_override(rule_id: str, min_gap_weeks=None, max_interval_weeks=None,
                 active=None, priority=None) -> None:
    db.run(
        "INSERT INTO cadence_overrides (rule_id, min_gap_weeks, max_interval_weeks, active, priority, updated_at) "
        "VALUES (?, ?, ?, ?, ?, datetime('now')) "
        "ON CONFLICT(rule_id) DO UPDATE SET "
        "min_gap_weeks=COALESCE(excluded.min_gap_weeks, cadence_overrides.min_gap_weeks), "
        "max_interval_weeks=COALESCE(excluded.max_interval_weeks, cadence_overrides.max_interval_weeks), "
        "active=COALESCE(excluded.active, cadence_overrides.active), "
        "priority=COALESCE(excluded.priority, cadence_overrides.priority), updated_at=datetime('now')",
        (rule_id, min_gap_weeks, max_interval_weeks,
         (1 if active else 0) if active is not None else None, priority))


def reset(rule_id: str) -> None:
    db.run("DELETE FROM cadence_overrides WHERE rule_id=?", (rule_id,))


def set_custom_active(rule_id: str, active: bool) -> None:
    db.run("UPDATE cadence_custom_rules SET active=? WHERE rule_id=?",
           (1 if active else 0, rule_id))


def apply_to_state(state: dict) -> int:
    """Overlay cadence_overrides onto the engine's CADENCE_RULES sheet AND append
    the UI-added custom rules as new rows. Called by db_state.configure before the
    engine runs. Returns rows changed/added."""
    ov = _overrides()
    custom = [c for c in _custom_rules() if c["active"]]
    sheet = state.get("CADENCE_RULES")
    if not sheet:
        return 0
    h = {str(n): i for i, n in enumerate(sheet[0])}
    ri, mg, mx, ac = h.get("ruleId"), h.get("minGapWeeks"), h.get("maxIntervalWeeks"), h.get("active")
    pr, sc, mvv = h.get("priority"), h.get("scope"), h.get("matchValue")
    it, gt = h.get("intervalType"), h.get("guaranteeType")
    if ri is None:
        return 0
    n = 0
    # 1) overrides onto existing rows
    for row in sheet[1:]:
        rid = str(row[ri]) if ri < len(row) else ""
        o = ov.get(rid)
        if not o:
            continue
        if mg is not None and o["min_gap_weeks"] is not None:
            row[mg] = o["min_gap_weeks"]
        if mx is not None and o["max_interval_weeks"] is not None:
            row[mx] = o["max_interval_weeks"]
        if ac is not None and o["active"] is not None:
            row[ac] = "YES" if o["active"] else "NO"
        if pr is not None and o["priority"] is not None:
            row[pr] = o["priority"]
        n += 1
    # 2) append custom rules as brand-new CADENCE_RULES rows
    width = len(sheet[0])
    existing_ids = {str(row[ri]) for row in sheet[1:] if ri < len(row)}
    for c in custom:
        if c["rule_id"] in existing_ids:
            continue
        row = [""] * width
        def _put(idx, val):
            if idx is not None and idx < width:
                row[idx] = val
        _put(ri, c["rule_id"]); _put(sc, c["scope"]); _put(mvv, c["match_value"])
        _put(mg, c["min_gap_weeks"]); _put(mx, c["max_interval_weeks"])
        _put(it, c["interval_type"]); _put(gt, c["guarantee_type"])
        _put(ac, "YES"); _put(pr, c["priority"])
        sheet.append(row)
        n += 1
    return n
