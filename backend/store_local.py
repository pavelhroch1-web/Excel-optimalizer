"""The versioned production store, now on LOCAL SQLite (desktop app).

Model (unchanged, git-like):
  - A SNAPSHOT is the complete, immutable planner state at publish time
    (full engine state). The latest snapshot is the single source of truth
    every new run resumes from. Snapshots can never be modified (DB triggers).
  - The DRAFT is the one mutable working file. Uploading rebuilds it from the
    latest snapshot + this run's fresh exports; edits mutate it; Publish
    freezes it into a new snapshot. Uploading never touches any snapshot.

Storage:
  - snapshots.state_blob / drafts.state_blob hold the whole engine state as
    xlsx bytes, so the engine resumes byte-identically (config_store /
    snapshot_store still read a workbook path -> we materialise the blob to a
    temp file on demand).
  - On publish we ALSO write normalised rows into published_plans and lock the
    week in plan_lifecycle, so history/reporting can query without opening xlsx.

Bootstrap: before the first publish, the "latest snapshot" is the proven
scaffold workbook bundled with the app - a one-time seed from the Excel.
"""
from __future__ import annotations

import json
import os
import sys
import tempfile

import openpyxl

import db


def _bootstrap_workbook() -> str:
    """Path to the seed scaffold (bundled in the .exe, or in the repo)."""
    override = os.environ.get("WORKBOOK_PATH")
    if override and os.path.exists(override):
        return override
    rel = os.path.join("workbook", "FieldForceOptimizer_V11_scaffold.xlsx")
    for base in (getattr(sys, "_MEIPASS", None),
                 os.path.dirname(os.path.dirname(os.path.abspath(__file__)))):
        if base:
            p = os.path.join(base, rel)
            if os.path.exists(p):
                return p
    return rel


def _tmp_xlsx() -> str:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    return path


# ---- version index (published history) ------------------------------------

def read_index() -> list[dict]:
    """Published-version history, newest last. [] before the first publish."""
    rows = db.get("SELECT id, created_at, message, published_week, published_by, "
                  "engine_version, source_files FROM snapshots WHERE kind='state' ORDER BY created_at ASC")
    out = []
    for r in rows:
        rec = dict(r)
        if rec.get("source_files"):
            try:
                rec["sourceFiles"] = json.loads(rec["source_files"])
            except Exception:
                pass
        out.append(rec)
    return out


def _next_version_id() -> str:
    n = db.get("SELECT COUNT(*) AS c FROM snapshots WHERE kind='state'")[0]["c"]
    return f"v{n + 1:04d}"


# ---- latest snapshot (source of truth) ------------------------------------

def _latest_snapshot_id() -> str | None:
    rows = db.get("SELECT id FROM snapshots WHERE kind='state' ORDER BY created_at DESC LIMIT 1")
    return rows[0]["id"] if rows else None


def latest_snapshot_repo_path() -> str:
    """Display label of what a new run resumes from."""
    sid = _latest_snapshot_id()
    return f"snapshot {sid}" if sid else "bootstrap (scaffold)"


def _write_snapshot_blob(version_id: str, dest_path: str) -> None:
    rows = db.get("SELECT state_blob FROM snapshots WHERE id = ?", (version_id,))
    if not rows:
        raise FileNotFoundError(f"Snapshot {version_id} nenalezen.")
    with open(dest_path, "wb") as f:
        f.write(rows[0]["state_blob"])


def download_latest_snapshot(dest_path: str) -> None:
    sid = _latest_snapshot_id()
    if sid:
        _write_snapshot_blob(sid, dest_path)
    else:
        import shutil
        shutil.copyfile(_bootstrap_workbook(), dest_path)


def download_snapshot(version_id: str, dest_path: str) -> None:
    _write_snapshot_blob(version_id, dest_path)


def snapshot_temp(version_id: str | None = None) -> str:
    """Materialise a snapshot (or the latest/bootstrap) to a temp xlsx and
    return its path. Caller deletes it."""
    path = _tmp_xlsx()
    if version_id:
        download_snapshot(version_id, path)
    else:
        download_latest_snapshot(path)
    return path


# ---- draft ----------------------------------------------------------------

def draft_exists() -> bool:
    return bool(db.get("SELECT 1 FROM drafts WHERE id = 'current'"))


def download_draft(dest_path: str) -> None:
    rows = db.get("SELECT state_blob FROM drafts WHERE id = 'current'")
    if not rows:
        raise FileNotFoundError("Žádný draft.")
    with open(dest_path, "wb") as f:
        f.write(rows[0]["state_blob"])


def save_draft(src_path: str, message: str, meta: dict | None = None) -> None:
    with open(src_path, "rb") as f:
        blob = f.read()
    meta_json = json.dumps(meta, ensure_ascii=False) if meta is not None else None
    conn = db.connect()
    try:
        if meta_json is None:
            conn.execute(
                "INSERT INTO drafts (id, state_blob, updated_at) VALUES ('current', ?, datetime('now')) "
                "ON CONFLICT(id) DO UPDATE SET state_blob=excluded.state_blob, updated_at=datetime('now')",
                (blob,))
        else:
            conn.execute(
                "INSERT INTO drafts (id, state_blob, meta, updated_at) VALUES ('current', ?, ?, datetime('now')) "
                "ON CONFLICT(id) DO UPDATE SET state_blob=excluded.state_blob, meta=excluded.meta, updated_at=datetime('now')",
                (blob, meta_json))
        conn.commit()
    finally:
        conn.close()


def read_draft_meta() -> dict:
    rows = db.get("SELECT meta FROM drafts WHERE id = 'current'")
    if rows and rows[0]["meta"]:
        try:
            return json.loads(rows[0]["meta"])
        except Exception:
            return {}
    return {}


# ---- publish (freeze draft -> immutable snapshot) -------------------------

_PLAN_COLMAP = {
    "WEEK": "week", "DATE": "plan_date", "DAY": "day", "TECHNICIAN": "technician",
    "POS": "pos_id", "KATEGORIE": "category", "NAZEV_PROVOZOVNY": "name",
    "ULICE": "street", "CISLO": "house_number", "MESTO": "city", "OBLAST": "area",
    "POS_AREA": "pos_area", "PPT": "ppt", "REASON": "reason", "GROUP": "day_group",
}


def _materialise_published_rows(conn, version_id: str, snapshot_xlsx: str) -> int:
    """Best-effort: read MANAGER_PLAN_PUBLISHED from the snapshot and insert
    normalised rows for querying. Never fails the publish."""
    try:
        wb = openpyxl.load_workbook(snapshot_xlsx, read_only=True, data_only=True)
        try:
            sheet = "MANAGER_PLAN_PUBLISHED" if "MANAGER_PLAN_PUBLISHED" in wb.sheetnames else "MANAGER_PLAN"
            if sheet not in wb.sheetnames:
                return 0
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                return 0
            hidx = {str(h): i for i, h in enumerate(header)}
            cols = [(src, dst) for src, dst in _PLAN_COLMAP.items() if src in hidx]
            n = 0
            for row in it:
                if not row or all(v in (None, "") for v in row):
                    continue
                week = row[hidx["WEEK"]] if "WEEK" in hidx else None
                if week in (None, ""):
                    continue
                vals = {"snapshot_id": version_id, "year": 2026}
                for src, dst in cols:
                    vals[dst] = row[hidx[src]]
                fields = ", ".join(vals.keys())
                marks = ", ".join("?" for _ in vals)
                conn.execute(f"INSERT INTO published_plans ({fields}) VALUES ({marks})",
                             tuple(vals.values()))
                # lock the week
                conn.execute(
                    "INSERT INTO plan_lifecycle (year, week, status, snapshot_id) "
                    "VALUES (?, ?, 'Published', ?) ON CONFLICT(year, week) "
                    "DO UPDATE SET status='Published', snapshot_id=excluded.snapshot_id, updated_at=datetime('now')",
                    (vals.get("year", 2026), week, version_id))
                n += 1
            return n
        finally:
            wb.close()
    except Exception as e:  # noqa: BLE001 - publish must not fail on reporting
        print("published_plans materialisation skipped:", e)
        return 0


def publish_snapshot(local_snapshot_path: str, meta: dict) -> dict:
    """Commit `local_snapshot_path` as a new immutable snapshot, materialise
    normalised published rows, and return the index record."""
    version_id = _next_version_id()
    with open(local_snapshot_path, "rb") as f:
        blob = f.read()
    conn = db.connect()
    try:
        conn.execute(
            "INSERT INTO snapshots (id, message, published_week, published_by, "
            "engine_version, source_files, state_blob) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (version_id, meta.get("message", ""), meta.get("publishedWeek"),
             meta.get("publishedBy"), meta.get("engineVersion"),
             json.dumps(meta.get("sourceFiles"), ensure_ascii=False) if meta.get("sourceFiles") is not None else None,
             blob))
        _materialise_published_rows(conn, version_id, local_snapshot_path)
        conn.commit()
    finally:
        conn.close()
    return {"id": version_id, **meta}
