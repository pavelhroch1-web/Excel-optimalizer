"""The published SNAPSHOT - the complete, immutable planner state each new
Draft resumes from (the git-like model the product owner confirmed:
"navazovat na poslední publikovaný snapshot ... snapshot obsahuje kompletní
stav planneru včetně per-POS historie").

A snapshot is the single source of truth for everything that carries across
runs, so that "upload new exports a year from now, click Generate, get the
same result as if the planner had run continuously all year" holds. It is
NOT config (rules live in config_store.py) and NOT the fresh weekly upload
(RAW_DATA / SALESAPP_IMPORT); it is the accumulated STATE the engines build
up over time.

Which sheets belong here was decided by auditing every engine (Python and
Office Script) for what it reads that is neither config nor a fresh upload:

  PLANNING-CRITICAL (required for a reproducible plan):
    POS_MASTER              per-POS state: last-visit, weeksSinceLastVisit,
                            status, closed-since, manager overrides, notes
    VISIT_HISTORY_ACTUAL    accumulated campaign visits - Compliance dedups
                            new visits against it (by UID) AND recomputes
                            weeksSinceLastVisit from it, so it must persist
                            or neglect ages wrongly / visits double-count
    OTHER_VISIT_LOG         accumulated non-campaign visits (dedup + perf)
    OZ_VISIT_LOG            accumulated OZ visits (dedup + excluded from stats)
    COMPLIANCE_LOG          accumulated compliance results (Advisor / Perf /
                            Reporting read it)
    MANAGER_PLAN            current working plan - Planning carries locked
                            (published) weeks over unchanged from it
    MANAGER_PLAN_PUBLISHED  the published plan Compliance matches visits to
    PLAN_LIFECYCLE          per-week status (Published -> Active -> Closed)
                            and which weeks are locked

  DERIVED (reconstructable by re-running Advisor/Performance/Reporting over
  the sheets above - kept so a snapshot is a complete restore point, not
  because planning needs them):
    ADVISOR_LOG, TECHNICIAN_PERFORMANCE_LOG, TECHNICIAN_PERFORMANCE_SUMMARY,
    TECHNICIAN_TOP_ISSUES, DASHBOARD, POS_MAP_DATA

Audit notes (verified against the engines, 2026-07):
  - No engine writes CONTROL, so there is no mutable run-state hiding in it;
    it is pure config.
  - POS_STATUS_IMPORT and the legacy VISIT_HISTORY (non-actual) sheet are
    read by no engine - deliberately excluded.
  - Planning reads only POS_MASTER columns set by Import/Compliance; it never
    reads back a previous run's plannerStatus/assignedWeek/businessScore, so
    there is no per-POS planning state to carry beyond the sheets above.

Bootstrap: today load_snapshot() reads the proven scaffold workbook (a
one-time seed from the current Excel). Once publishing is built, it reads
the latest published snapshot instead - same shape, so nothing downstream
changes.
"""
from __future__ import annotations

import datetime
import os

import openpyxl

# Accumulated state the engines build up and read across runs.
PLANNING_CRITICAL_SHEETS = (
    "POS_MASTER",
    "VISIT_HISTORY_ACTUAL",
    "OTHER_VISIT_LOG",
    "OZ_VISIT_LOG",
    "COMPLIANCE_LOG",
    "MANAGER_PLAN",
    "MANAGER_PLAN_PUBLISHED",
    "PLAN_LIFECYCLE",
)

# Reconstructable report outputs - included so a snapshot is a full restore
# point, but a Draft does not depend on them being present.
DERIVED_SHEETS = (
    "ADVISOR_LOG",
    "TECHNICIAN_PERFORMANCE_LOG",
    "TECHNICIAN_PERFORMANCE_SUMMARY",
    "TECHNICIAN_TOP_ISSUES",
    "DASHBOARD",
    "POS_MAP_DATA",
)

SNAPSHOT_SHEETS = PLANNING_CRITICAL_SHEETS + DERIVED_SHEETS

DEFAULT_SEED_WORKBOOK = os.environ.get(
    "CONFIG_SEED_WORKBOOK",
    os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "workbook",
        "FieldForceOptimizer_V11_scaffold.xlsx",
    ),
)


def _cell_to_json(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return "" if v is None else v


def _read_sheet(wb, name: str) -> list[list]:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_to_json(v) for v in row])
    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    return rows


def load_snapshot(seed_workbook: str | None = None, include_derived: bool = True) -> dict[str, list[list]]:
    """Returns the complete accumulated planner state as a {sheetName: rows}
    dict - every sheet a new Draft resumes from. Missing sheets are simply
    omitted (the pipeline falls back to an empty header for those)."""
    path = seed_workbook or DEFAULT_SEED_WORKBOOK
    wanted = SNAPSHOT_SHEETS if include_derived else PLANNING_CRITICAL_SHEETS
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return {name: _read_sheet(wb, name) for name in wanted if name in wb.sheetnames}
    finally:
        wb.close()
