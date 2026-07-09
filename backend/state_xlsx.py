"""Serialize an engine state dict ({sheetName: [[cell,...],...]}) to a plain
.xlsx and back. Used to persist a Draft and each immutable published
Snapshot as a real workbook file - one the manager can also just download
and open. No styling, no formulas: these are internal working/state files,
not the presentation workbook.
"""
from __future__ import annotations

import datetime

import openpyxl


def save_state(state: dict[str, list[list]], path: str) -> None:
    """Writes every sheet in `state` to a new workbook at `path`."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in state.items():
        ws = wb.create_sheet(title=name[:31])
        for row in rows:
            ws.append([_xl(v) for v in row])
    if not wb.sheetnames:  # never leave a truly empty workbook
        wb.create_sheet(title="EMPTY")
    wb.save(path)


def load_state(path: str) -> dict[str, list[list]]:
    """Reads every sheet of the workbook at `path` back into a state dict."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        state: dict[str, list[list]] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [[_json(v) for v in row] for row in ws.iter_rows(values_only=True)]
            while rows and all(v == "" for v in rows[-1]):
                rows.pop()
            state[name] = rows
        return state
    finally:
        wb.close()


def _xl(v):
    # openpyxl can't store tz-aware / ISO strings as dates automatically; keep
    # values as-is (str/number), which round-trips cleanly for our tables.
    if v is None:
        return ""
    return v


def _json(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return "" if v is None else v
