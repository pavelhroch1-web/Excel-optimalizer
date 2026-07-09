"""Enrich MANAGER_PLAN rows for display, and let the manager manually edit
the still-Draft plan (remove a POS, add a POS, change a row's technician).

This is pure CRUD on top of Planning Engine's already-generated output - no
selection/scoring logic lives here. Matches the product owner's explicit
ask (2026-07-11): "Chci také možnost do výsledku ručně zasáhnout: odebrat
POS, přidat jiný POS, případně změnit technika." Only weeks that are still
Draft (not yet Published/Active/Closed in PLAN_LIFECYCLE) may be edited -
same boundary Publish Engine itself enforces, so a manual edit can never
touch a plan that's already gone out to the field.
"""
from __future__ import annotations

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from desktop_client.engines.dates_logic import work_days

MANAGER_PLAN_HEADER = [
    "WEEK", "DATE", "DAY", "TECHNICIAN", "POS", "KATEGORIE", "NAZEV_PROVOZOVNY",
    "ULICE", "CISLO", "MESTO", "OBLAST", "POS_AREA", "PPT", "LOS_ACTIVITY",
    "LOT_ACTIVITY", "REASON", "GPS_GROUP",
]

# Same tags PlanningEngine.ts/planning_engine.py stamp into REASON - see
# core.ts's reason-tagging comment ("PREMIUM |", "GPS BONUS |", "NEARBY |")
# and PlanningEngine.ts's CADENCE_RULES-driven "MANDATORY (<ruleId>) |" tag.
# Purely a display translation - the underlying tags are never changed.
_REASON_LABELS = {
    "CORE": "CORE kategorie (garantovaná návštěva)",
    "PREMIUM": "Vysoká hodnota (top 20 % PPT)",
    "NEARBY": "Blízko jiných navštívených POS",
    "GPS BONUS": "Efektivní trasa (geo shluk)",
}


def friendly_reason(raw: str) -> str:
    if not raw:
        return ""
    parts = [p.strip() for p in raw.split("|") if p.strip()]
    labels = []
    for p in parts:
        if p in _REASON_LABELS:
            labels.append(_REASON_LABELS[p])
        elif p.startswith("MANDATORY"):
            inner = p[len("MANDATORY"):].strip(" ()")
            labels.append(f"Povinná návštěva ({inner})" if inner else "Povinná návštěva")
        else:
            labels.append(p)
    return ", ".join(labels)


def _header_index(ws) -> dict[str, int]:
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value not in (None, "")}


def get_locked_weeks(wb, control_year: int) -> set[int]:
    ws = wb["PLAN_LIFECYCLE"]
    idx = _header_index(ws)
    locked = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx["year"] - 1] != control_year:
            continue
        if row[idx["status"] - 1] in ("Published", "Active", "Closed"):
            locked.add(int(row[idx["week"] - 1]))
    return locked


def get_control_year(wb) -> int:
    ws = wb["CONTROL"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and str(row[0]).strip().upper() == "YEAR":
            try:
                return int(row[1])
            except (TypeError, ValueError):
                pass
    import datetime
    return datetime.date.today().year


def read_enriched_draft(path: str) -> list[dict]:
    """MANAGER_PLAN rows joined with POS_MASTER fields the manager needs to
    judge each pick (last visit, weeks since, terminal type, market), plus a
    human-readable REASON."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        pm_ws = wb["POS_MASTER"]
        pm_idx = _header_index(pm_ws)
        pos_by_id: dict[str, dict] = {}
        for row in pm_ws.iter_rows(min_row=2, values_only=True):
            pid = row[pm_idx["posId"] - 1]
            if not pid:
                continue
            pos_by_id[str(pid)] = {
                "terminalType": row[pm_idx["terminalType"] - 1],
                "market": row[pm_idx["market"] - 1],
                "classification": row[pm_idx["classification"] - 1],
                "lastRealVisitDate": row[pm_idx["lastRealVisitDate"] - 1],
                "weeksSinceLastVisit": row[pm_idx["weeksSinceLastVisit"] - 1],
            }

        mp_ws = wb["MANAGER_PLAN"]
        mp_idx = _header_index(mp_ws)
        rows = []
        for row in mp_ws.iter_rows(min_row=2, values_only=True):
            pos_id = row[mp_idx["POS"] - 1]
            if pos_id in (None, ""):
                continue
            base = {h: row[mp_idx[h] - 1] for h in MANAGER_PLAN_HEADER if h in mp_idx}
            extra = pos_by_id.get(str(pos_id), {})
            base["terminalType"] = extra.get("terminalType")
            base["market"] = extra.get("market")
            base["classification"] = extra.get("classification")
            base["lastRealVisitDate"] = extra.get("lastRealVisitDate")
            base["weeksSinceLastVisit"] = extra.get("weeksSinceLastVisit")
            base["REASON_FRIENDLY"] = friendly_reason(base.get("REASON") or "")
            rows.append(base)
        return rows
    finally:
        wb.close()


def remove_pos(path: str, week: int, pos_id: str, technician: str) -> int:
    """Deletes matching MANAGER_PLAN row(s). Returns how many were removed."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        control_year = get_control_year(wb)
        if week in get_locked_weeks(wb, control_year):
            raise ValueError(f"Týden {week} je již publikovaný - nelze ho ručně upravovat.")
        ws = wb["MANAGER_PLAN"]
        idx = _header_index(ws)
        to_delete = []
        for r in range(2, ws.max_row + 1):
            if (
                ws.cell(r, idx["WEEK"]).value == week
                and str(ws.cell(r, idx["POS"]).value) == str(pos_id)
                and ws.cell(r, idx["TECHNICIAN"]).value == technician
            ):
                to_delete.append(r)
        for r in reversed(to_delete):
            ws.delete_rows(r, 1)
        wb.save(path)
        return len(to_delete)
    finally:
        wb.close()


def change_technician(path: str, week: int, pos_id: str, old_technician: str, new_technician: str) -> int:
    """Reassigns matching MANAGER_PLAN row(s) to a different technician."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        control_year = get_control_year(wb)
        if week in get_locked_weeks(wb, control_year):
            raise ValueError(f"Týden {week} je již publikovaný - nelze ho ručně upravovat.")
        ws = wb["MANAGER_PLAN"]
        idx = _header_index(ws)
        changed = 0
        for r in range(2, ws.max_row + 1):
            if (
                ws.cell(r, idx["WEEK"]).value == week
                and str(ws.cell(r, idx["POS"]).value) == str(pos_id)
                and ws.cell(r, idx["TECHNICIAN"]).value == old_technician
            ):
                ws.cell(r, idx["TECHNICIAN"], new_technician)
                changed += 1
        wb.save(path)
        return changed
    finally:
        wb.close()


def _cz_date(d) -> str:
    return f"{d.day}. {d.month}. {d.year}"


def add_pos(path: str, week: int, day: str, technician: str, pos_id: str) -> dict:
    """Appends a new MANAGER_PLAN row for pos_id, filling POS_MASTER's real
    address/category/PPT - same fields Planning Engine itself stamps, so a
    manually-added row looks identical to an algorithm-picked one except for
    its REASON."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        control_year = get_control_year(wb)
        if week in get_locked_weeks(wb, control_year):
            raise ValueError(f"Týden {week} je již publikovaný - nelze ho ručně upravovat.")

        pm_ws = wb["POS_MASTER"]
        pm_idx = _header_index(pm_ws)
        pos_row = None
        for row in pm_ws.iter_rows(min_row=2, values_only=True):
            if str(row[pm_idx["posId"] - 1]) == str(pos_id):
                pos_row = row
                break
        if pos_row is None:
            raise ValueError(f"POS {pos_id} nebyl nalezen v POS_MASTER.")

        days = work_days(control_year, week)
        day_row = next((d for d in days if d.day == day), None)
        if day_row is None:
            raise ValueError(f"Den {day} není platný pracovní den v týdnu {week}.")

        ws = wb["MANAGER_PLAN"]
        idx = _header_index(ws)
        for r in range(2, ws.max_row + 1):
            if (
                ws.cell(r, idx["WEEK"]).value == week
                and str(ws.cell(r, idx["POS"]).value) == str(pos_id)
                and ws.cell(r, idx["TECHNICIAN"]).value == technician
            ):
                raise ValueError("Tento POS už je tomuto technikovi v tomto týdnu naplánován.")

        new_row = {
            "WEEK": week,
            "DATE": _cz_date(day_row.date),
            "DAY": day_row.day,
            "TECHNICIAN": technician,
            "POS": str(pos_id),
            "KATEGORIE": pos_row[pm_idx["category"] - 1],
            "NAZEV_PROVOZOVNY": pos_row[pm_idx["nazev"] - 1],
            "ULICE": pos_row[pm_idx["street"] - 1],
            "CISLO": pos_row[pm_idx["houseNumber"] - 1],
            "MESTO": pos_row[pm_idx["city"] - 1],
            "OBLAST": pos_row[pm_idx["area"] - 1],
            "POS_AREA": pos_row[pm_idx["posArea"] - 1],
            "PPT": pos_row[pm_idx["ppt"] - 1],
            "LOS_ACTIVITY": pos_row[pm_idx["currentLosActivity"] - 1],
            "LOT_ACTIVITY": pos_row[pm_idx["currentLotActivity"] - 1],
            "REASON": "Ruční přidání manažerem",
            "GPS_GROUP": "",
        }
        last_row = ws.max_row
        while last_row > 1 and ws.cell(last_row, 1).value in (None, ""):
            last_row -= 1
        target_row = last_row + 1
        for h in MANAGER_PLAN_HEADER:
            if h in idx:
                ws.cell(target_row, idx[h], new_row.get(h, ""))
        wb.save(path)
        new_row["terminalType"] = pos_row[pm_idx["terminalType"] - 1]
        new_row["market"] = pos_row[pm_idx["market"] - 1]
        new_row["classification"] = pos_row[pm_idx["classification"] - 1]
        new_row["lastRealVisitDate"] = pos_row[pm_idx["lastRealVisitDate"] - 1]
        new_row["weeksSinceLastVisit"] = pos_row[pm_idx["weeksSinceLastVisit"] - 1]
        new_row["REASON_FRIENDLY"] = friendly_reason(new_row["REASON"])
        return new_row
    finally:
        wb.close()
